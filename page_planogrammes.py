"""
Page Planogrammes — interface Streamlit complète.
Données sauvegardées dans MongoDB (collections planogrammes + produits_lib).
"""

import streamlit as st
import pandas as pd
import datetime
import io
import re
from pathlib import Path

try:
    import pdfplumber
    _PDFPLUMBER_OK = True
except ImportError:
    _PDFPLUMBER_OK = False

from planogrammes_storage import (
    load_planogrammes, save_planogramme, delete_planogramme, duplicate_planogramme,
    load_produits, save_produit, delete_produit,
)

# ── Palettes couleurs disponibles pour les slots ────────
SLOT_COLORS = {
    "Aucune":   "",
    "Bleu":     "#1a2f47",
    "Vert":     "#12302a",
    "Ambre":    "#2e2a14",
    "Corail":   "#2e1a1a",
    "Violet":   "#1f1a3a",
    "Sarcelle": "#12302a",
    "Rose":     "#2a1a2a",
    "Gris":     "#2a2a2a",
}

SLOT_COLORS_DISPLAY = {
    "":         "#2e3450",
    "#1a2f47":  "#1a2f47",
    "#12302a":  "#12302a",
    "#2e2a14":  "#2e2a14",
    "#2e1a1a":  "#2e1a1a",
    "#1f1a3a":  "#1f1a3a",
    "#2a1a2a":  "#2a1a2a",
    "#2a2a2a":  "#2a2a2a",
}


# ════════════════════════════════════════════════════════
# CALCUL VALEUR TOTALE
# ════════════════════════════════════════════════════════

def _calcul_valeur_totale(plano: dict) -> float:
    """
    Calcule le coût d'achat TTC total du planogramme :
    somme de (prix_achat × (1 + TVA/100) × quantité) pour chaque slot rempli.
    Le prix_achat et la TVA sont récupérés depuis la bibliothèque de produits
    en faisant le lien par le nom du produit.
    Fallback : si le produit n'est pas dans la bibliothèque, utilise le prix
    du slot directement.
    """
    # Construire un index nom → produit depuis la bibliothèque
    produits = _get_produits()
    index_nom = {}
    for pr in produits:
        nom_clean = pr.get("nom", "").strip().lower()
        if nom_clean:
            index_nom[nom_clean] = pr

    total = 0.0
    for s in plano.get("slots", {}).values():
        if not s.get("product"):
            continue
        try:
            qty = float(s.get("qty", 0) or 0)
        except (ValueError, TypeError):
            qty = 0.0
        if qty <= 0:
            continue

        # Chercher le produit dans la bibliothèque
        nom_slot = s.get("product", "").strip().lower()
        prod_lib = index_nom.get(nom_slot)

        if prod_lib and prod_lib.get("prix_achat") and prod_lib.get("tva") is not None:
            try:
                prix_achat = float(prod_lib["prix_achat"])
                tva        = float(prod_lib["tva"])
                total += prix_achat * (1 + tva / 100) * qty
                continue
            except (ValueError, TypeError):
                pass

        # Fallback : prix saisi manuellement dans le slot
        try:
            prix = float(s.get("price", 0) or 0)
            total += prix * qty
        except (ValueError, TypeError):
            pass

    return total


# ════════════════════════════════════════════════════════
# HELPERS SESSION STATE
# ════════════════════════════════════════════════════════

def _init_state():
    defaults = {
        "pg_current_id":    None,   # _id du planogramme en cours d'édition
        "pg_current":       None,   # dict planogramme en mémoire
        "pg_view":          "list", # "list" | "editor" | "library"
        "pg_sel_slot":      None,   # "r-c" du slot sélectionné
        "pg_dirty":         False,  # modifications non sauvegardées
        "pg_search_lib":    "",
        "pg_planogrammes":  None,   # cache liste
        "pg_produits":      None,   # cache bibliothèque
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v


def _reload_planogrammes():
    st.session_state.pg_planogrammes = load_planogrammes()


def _reload_produits():
    st.session_state.pg_produits = load_produits()


def _get_planogrammes():
    if st.session_state.pg_planogrammes is None:
        _reload_planogrammes()
    return st.session_state.pg_planogrammes or []


def _get_produits():
    if st.session_state.pg_produits is None:
        _reload_produits()
    return st.session_state.pg_produits or []


def _new_slots(rows, cols):
    return {f"{r}-{c}": {"product": "", "price": "", "qty": "", "color": ""}
            for r in range(rows) for c in range(cols)}


def _set_current(plano: dict):
    st.session_state.pg_current    = plano
    st.session_state.pg_current_id = plano.get("_id")
    st.session_state.pg_sel_slot   = None
    st.session_state.pg_dirty      = False
    st.session_state.pg_view       = "editor"


def _mark_dirty():
    st.session_state.pg_dirty = True


# ════════════════════════════════════════════════════════
# COMPOSANTS UI RÉUTILISABLES
# ════════════════════════════════════════════════════════

def _slot_color_badge(color: str) -> str:
    bg = color if color else "#2e3450"
    return f'<span style="display:inline-block;width:12px;height:12px;border-radius:50%;background:{bg};border:1px solid rgba(255,255,255,0.2);vertical-align:middle;margin-right:4px"></span>'


def _render_grid_preview(plano: dict, editable: bool = True):
    """Affiche la grille du distributeur sous forme de tableau HTML cliquable."""
    rows    = plano.get("rows", 1)
    cols    = plano.get("cols", 1)
    slots   = plano.get("slots", {})
    labels  = plano.get("row_labels", [f"R{i+1}" for i in range(rows)])
    sel     = st.session_state.pg_sel_slot

    # Calcul largeur colonne
    col_w = max(60, min(110, int(900 / (cols + 1))))

    html = f"""
    <style>
    .vm-grid {{ border-collapse: collapse; width: 100%; font-family: -apple-system, sans-serif; }}
    .vm-grid th {{
        background: #1e2130; color: #9ba3b8; font-size: 10px; font-weight: 600;
        padding: 4px 2px; text-align: center; border: 1px solid rgba(255,255,255,0.06);
        width: {col_w}px; min-width: 44px;
    }}
    .vm-grid td.label {{
        background: #262b3d; color: #9ba3b8; font-size: 10px; font-weight: 700;
        padding: 5px 6px; text-align: center; border: 1px solid rgba(255,255,255,0.08);
        white-space: nowrap; min-width: 28px;
    }}
    .vm-grid td.slot {{
        border: 1px solid rgba(255,255,255,0.08); padding: 4px 3px; text-align: center;
        vertical-align: top; font-size: 10px; min-width: 44px; width: {col_w}px;
        cursor: pointer; transition: outline 0.1s;
    }}
    .vm-grid td.slot:hover {{ outline: 1.5px solid #4d9de0; outline-offset: -1px; }}
    .vm-grid td.slot.selected {{ outline: 2px solid #4d9de0; outline-offset: -1px; }}
    .slot-product {{ font-weight: 600; color: #e8eaf0; line-height: 1.3; word-break: break-word; }}
    .slot-price {{ color: #4d9de0; margin-top: 2px; }}
    .slot-qty {{ color: #9ba3b8; }}
    .slot-empty {{ color: #3d4460; font-size: 12px; }}
    </style>
    <table class="vm-grid">
    <thead><tr><th style="width:28px"></th>"""

    for c in range(cols):
        html += f"<th>C{c+1}</th>"
    html += "</tr></thead><tbody>"

    for r in range(rows):
        lbl = labels[r] if r < len(labels) else f"R{r+1}"
        html += f'<tr><td class="label">{lbl}</td>'
        for c in range(cols):
            key = f"{r}-{c}"
            s   = slots.get(key, {})
            bg  = s.get("color", "") or "#1e2130"
            is_sel = (sel == key)
            cls = "slot selected" if is_sel else "slot"
            html += f'<td class="{cls}" style="background:{bg}">'
            if s.get("product"):
                html += f'<div class="slot-product">{s["product"]}</div>'
                if s.get("price"):
                    try:
                        html += f'<div class="slot-price">{float(s["price"]):.2f} €</div>'
                    except Exception:
                        html += f'<div class="slot-price">{s["price"]} €</div>'
                if s.get("qty"):
                    html += f'<div class="slot-qty">×{s["qty"]}</div>'
            else:
                html += '<div class="slot-empty">—</div>'
            html += "</td>"
        html += "</tr>"

    html += "</tbody></table>"
    st.markdown(html, unsafe_allow_html=True)


# ════════════════════════════════════════════════════════
# VUE : LISTE DES PLANOGRAMMES
# ════════════════════════════════════════════════════════

def _view_list():
    planogrammes = _get_planogrammes()

    # ── Header ──
    hc1, hc2 = st.columns([5, 3])
    with hc1:
        st.markdown("### 📋 Tous les planogrammes")
    with hc2:
        bc1, bc2, bc3 = st.columns(3)
        with bc1:
            if st.button("+ Nouveau", use_container_width=True, type="primary"):
                st.session_state.pg_view = "new"
                st.rerun()
        with bc2:
            if st.button("↑ Import PDF", use_container_width=True):
                st.session_state.pg_view = "import_pdf"
                st.rerun()
        with bc3:
            if st.button("📦 Bibliothèque", use_container_width=True):
                st.session_state.pg_view = "library"
                st.rerun()

    if not planogrammes:
        st.info("Aucun planogramme. Créez-en un avec le bouton **+ Nouveau**.")
        return

    # ── Grille de cartes ──
    cols_per_row = 3
    for i in range(0, len(planogrammes), cols_per_row):
        row_planos = planogrammes[i:i + cols_per_row]
        cols = st.columns(cols_per_row)
        for j, p in enumerate(row_planos):
            with cols[j]:
                filled = sum(1 for s in p.get("slots", {}).values() if s.get("product"))
                total  = p.get("rows", 0) * p.get("cols", 0)
                upd    = p.get("updated_at", "—")

                with st.container(border=True):
                    st.markdown(f"**{p['nom']}**")
                    st.caption(
                        f"{'Double' if p.get('type')=='double' else 'Simple'} · "
                        f"{p.get('rows',0)} × {p.get('cols',0)} · "
                        f"{filled}/{total} remplis"
                    )
                    valeur = _calcul_valeur_totale(p)
                    st.caption(f"💰 Coût achat TTC : **{valeur:,.2f} €** · Modifié le {upd}")

                    bc1, bc2, bc3 = st.columns(3)
                    with bc1:
                        if st.button("✏️ Éditer", key=f"edit_{p['_id']}", use_container_width=True):
                            _set_current(p)
                            st.rerun()
                    with bc2:
                        if st.button("⎘ Copier", key=f"dup_{p['_id']}", use_container_width=True):
                            new_id = duplicate_planogramme(p["_id"])
                            if new_id:
                                _reload_planogrammes()
                                st.toast(f"Copie créée.")
                                st.rerun()
                    with bc3:
                        if st.button("🗑️", key=f"del_{p['_id']}", use_container_width=True):
                            st.session_state[f"confirm_del_{p['_id']}"] = True
                            st.rerun()

                    # Confirmation suppression
                    if st.session_state.get(f"confirm_del_{p['_id']}"):
                        st.warning(f"Supprimer **{p['nom']}** ?")
                        cc1, cc2 = st.columns(2)
                        with cc1:
                            if st.button("Oui, supprimer", key=f"yes_{p['_id']}", type="primary"):
                                delete_planogramme(p["_id"])
                                st.session_state.pop(f"confirm_del_{p['_id']}", None)
                                _reload_planogrammes()
                                st.rerun()
                        with cc2:
                            if st.button("Annuler", key=f"no_{p['_id']}"):
                                st.session_state.pop(f"confirm_del_{p['_id']}", None)
                                st.rerun()


# ════════════════════════════════════════════════════════
# VUE : NOUVEAU PLANOGRAMME
# ════════════════════════════════════════════════════════

def _view_new():
    st.markdown("### ➕ Nouveau planogramme")

    with st.form("form_new_plano"):
        nom  = st.text_input("Nom", placeholder="Double BF Octobre 2025")
        c1, c2, c3 = st.columns(3)
        with c1:
            type_dist = st.selectbox("Type", ["simple", "double"])
        with c2:
            rows = st.number_input("Rangées", min_value=1, max_value=20, value=5)
        with c3:
            cols = st.number_input("Colonnes", min_value=1, max_value=24, value=7)

        submitted = st.form_submit_button("Créer", type="primary", use_container_width=True)

    if submitted:
        if not nom.strip():
            st.error("Le nom est obligatoire.")
            return
        plano = {
            "nom":        nom.strip(),
            "type":       type_dist,
            "rows":       int(rows),
            "cols":       int(cols),
            "row_labels": [f"R{i+1}" for i in range(int(rows))],
            "slots":      _new_slots(int(rows), int(cols)),
        }
        new_id = save_planogramme(plano)
        plano["_id"] = new_id
        _reload_planogrammes()
        _set_current(plano)
        st.rerun()

    if st.button("← Retour"):
        st.session_state.pg_view = "list"
        st.rerun()


# ════════════════════════════════════════════════════════
# VUE : ÉDITEUR
# ════════════════════════════════════════════════════════

def _view_editor():
    p = st.session_state.pg_current
    if not p:
        st.session_state.pg_view = "list"
        st.rerun()
        return

    # ── Barre d'outils éditeur ──
    tc1, tc2, tc3, tc4, tc5, tc6, tc7 = st.columns([3, 1, 1, 1, 1, 1, 1])
    with tc1:
        st.markdown(f"### ✏️ {p['nom']}")
    with tc2:
        if st.button("← Liste", use_container_width=True):
            st.session_state.pg_view = "list"
            st.session_state.pg_sel_slot = None
            st.rerun()
    with tc3:
        dirty_label = "💾 Sauvegarder*" if st.session_state.pg_dirty else "💾 Sauvegarder"
        if st.button(dirty_label, use_container_width=True, type="primary"):
            save_planogramme(p)
            _reload_planogrammes()
            st.session_state.pg_dirty = False
            st.toast("✅ Sauvegardé dans MongoDB.")
    with tc4:
        if st.button("⎘ Dupliquer", use_container_width=True):
            new_id = duplicate_planogramme(p["_id"])
            if new_id:
                _reload_planogrammes()
                st.toast("Copie créée.")
    with tc5:
        if st.button("↓ Excel", use_container_width=True):
            _do_export_excel(p)
    with tc6:
        if st.button("↓ PDF", use_container_width=True):
            _do_export_pdf(p)
    with tc7:
        if st.button("🗑️ Suppr.", use_container_width=True):
            st.session_state["confirm_del_editor"] = True

    if st.session_state.get("confirm_del_editor"):
        st.warning(f"Supprimer **{p['nom']}** définitivement ?")
        cc1, cc2, _ = st.columns([1, 1, 4])
        with cc1:
            if st.button("Oui, supprimer", type="primary"):
                delete_planogramme(p["_id"])
                st.session_state.pop("confirm_del_editor", None)
                st.session_state.pg_current    = None
                st.session_state.pg_current_id = None
                st.session_state.pg_view       = "list"
                _reload_planogrammes()
                st.rerun()
        with cc2:
            if st.button("Annuler"):
                st.session_state.pop("confirm_del_editor", None)
                st.rerun()

    st.divider()

    # ── Deux colonnes : paramètres | grille ──
    left, right = st.columns([1, 3])

    with left:
        # Paramètres généraux
        with st.expander("⚙️ Paramètres", expanded=True):
            new_nom = st.text_input("Nom", value=p["nom"], key="ed_nom")
            if new_nom != p["nom"]:
                p["nom"] = new_nom; _mark_dirty()

            new_type = st.selectbox("Type", ["simple", "double"],
                                    index=0 if p.get("type") == "simple" else 1, key="ed_type")
            if new_type != p.get("type"):
                p["type"] = new_type; _mark_dirty()

            new_cols = st.number_input("Colonnes", 1, 24, value=p["cols"], key="ed_cols")
            if int(new_cols) != p["cols"]:
                old_cols = p["cols"]
                p["cols"] = int(new_cols)
                # Ajouter ou retirer les colonnes dans les slots
                for r in range(p["rows"]):
                    for c in range(int(new_cols)):
                        k = f"{r}-{c}"
                        if k not in p["slots"]:
                            p["slots"][k] = {"product": "", "price": "", "qty": "", "color": ""}
                _mark_dirty()

        # Rangées
        with st.expander("📏 Rangées", expanded=True):
            bc1, bc2 = st.columns(2)
            with bc1:
                if st.button("+ Rangée", use_container_width=True):
                    r = p["rows"]
                    p["rows"] += 1
                    p["row_labels"].append(f"R{p['rows']}")
                    for c in range(p["cols"]):
                        p["slots"][f"{r}-{c}"] = {"product": "", "price": "", "qty": "", "color": ""}
                    _mark_dirty(); st.rerun()
            with bc2:
                if st.button("− Rangée", use_container_width=True, disabled=(p["rows"] <= 1)):
                    r = p["rows"] - 1
                    for c in range(p["cols"]):
                        p["slots"].pop(f"{r}-{c}", None)
                    p["rows"] -= 1
                    if p["row_labels"]:
                        p["row_labels"].pop()
                    _mark_dirty(); st.rerun()

            labels = p.get("row_labels", [])
            for i in range(p["rows"]):
                cur = labels[i] if i < len(labels) else f"R{i+1}"
                new_lbl = st.text_input(f"R{i+1}", value=cur, key=f"lbl_{i}", label_visibility="collapsed")
                if new_lbl != cur:
                    while len(p["row_labels"]) <= i:
                        p["row_labels"].append(f"R{len(p['row_labels'])+1}")
                    p["row_labels"][i] = new_lbl
                    _mark_dirty()

        # Import CSV rapide
        with st.expander("↑ Import CSV"):
            st.caption("Format : une ligne = une rangée. Cellules séparées par `;` : `Produit;prix;qté`")
            csv_file = st.file_uploader("Fichier CSV", type=["csv"], key="ed_csv_import", label_visibility="collapsed")
            if csv_file and st.button("Importer", key="do_csv_import"):
                _import_csv_to_plano(p, csv_file.read())
                _mark_dirty(); st.rerun()

    with right:
        # Métriques du planogramme
        filled_count = sum(1 for s in p.get("slots", {}).values() if s.get("product"))
        total_count  = p.get("rows", 0) * p.get("cols", 0)
        valeur_totale = _calcul_valeur_totale(p)

        mc1, mc2, mc3, mc4 = st.columns(4)
        mc1.metric("Type",        "Double" if p.get("type") == "double" else "Simple")
        mc2.metric("Emplacements", f"{filled_count}/{total_count}")
        mc3.metric("Coût achat TTC", f"{valeur_totale:,.2f} €")
        mc4.metric("Rangées × Col.", f"{p['rows']} × {p['cols']}")
        st.caption("Cliquez sur un emplacement pour le modifier.")

        # Grille visuelle
        _render_grid_preview(p)

        # Sélecteur de slot via selectbox (alternative cliquable)
        st.markdown("**Sélectionner un emplacement :**")
        sc1, sc2, sc3 = st.columns(3)
        with sc1:
            sel_row = st.selectbox(
                "Rangée", range(p["rows"]),
                format_func=lambda i: p["row_labels"][i] if i < len(p["row_labels"]) else f"R{i+1}",
                key="sel_row"
            )
        with sc2:
            sel_col = st.selectbox("Colonne", range(p["cols"]),
                                   format_func=lambda c: f"C{c+1}", key="sel_col")
        with sc3:
            if st.button("✏️ Modifier cet emplacement", use_container_width=True):
                st.session_state.pg_sel_slot = f"{sel_row}-{sel_col}"
                st.rerun()

        # Éditeur de slot
        sel = st.session_state.pg_sel_slot
        if sel and sel in p["slots"]:
            s = p["slots"][sel]
            r_i, c_i = sel.split("-")
            row_lbl = p["row_labels"][int(r_i)] if int(r_i) < len(p["row_labels"]) else f"R{int(r_i)+1}"

            st.divider()
            st.markdown(f"**Emplacement : {row_lbl} · C{int(c_i)+1}**")

            # Appliquer depuis bibliothèque (hors form pour éviter conflits)
            produits = _get_produits()
            if produits:
                prod_names = ["— Choisir dans la bibliothèque —"] + [pr["nom"] for pr in produits]
                # Clé inclut le slot pour forcer reset à chaque changement de slot
                choice = st.selectbox("Bibliothèque de produits", prod_names,
                                      key=f"lib_pick_{sel}")
                if choice != "— Choisir dans la bibliothèque —":
                    prod_obj = next((pr for pr in produits if pr["nom"] == choice), None)
                    if prod_obj and st.button("↙️ Appliquer ce produit", type="primary", key=f"apply_lib_{sel}"):
                        # Utiliser prix_ttc en priorité, sinon prix_ht
                        prix_slot = prod_obj.get("prix_ttc", "") or prod_obj.get("prix_ht", "") or prod_obj.get("prix", "")
                        try:
                            prix_slot = f"{float(prix_slot):.2f}" if prix_slot else ""
                        except Exception:
                            pass
                        p["slots"][sel] = {
                            "product": prod_obj["nom"],
                            "price":   prix_slot,
                            "qty":     prod_obj.get("quantite", ""),
                            "color":   prod_obj.get("couleur", ""),
                        }
                        _mark_dirty(); st.rerun()

            # Formulaire d'édition manuelle — la clé du form inclut le slot
            # pour forcer Streamlit à recréer les widgets à chaque slot différent
            with st.form(key=f"form_slot_{sel}"):
                color_names = list(SLOT_COLORS.keys())
                cur_color = s.get("color", "")
                cur_color_name = next((n for n, v in SLOT_COLORS.items() if v == cur_color), "Aucune")

                ec1, ec2, ec3 = st.columns(3)
                with ec1:
                    new_prod = st.text_input("Produit", value=s.get("product", ""))
                with ec2:
                    new_price = st.text_input("Prix (€)", value=s.get("price", ""))
                with ec3:
                    new_qty = st.text_input("Quantité ×", value=s.get("qty", ""))

                new_color_name = st.selectbox(
                    "Couleur de fond", color_names,
                    index=color_names.index(cur_color_name)
                )
                new_color = SLOT_COLORS[new_color_name]

                fc1, fc2 = st.columns([2, 1])
                with fc1:
                    apply = st.form_submit_button(
                        "✅ Appliquer", use_container_width=True, type="primary"
                    )
                with fc2:
                    clear = st.form_submit_button("🗑️ Vider", use_container_width=True)

            if apply:
                p["slots"][sel] = {
                    "product": new_prod,
                    "price":   new_price,
                    "qty":     new_qty,
                    "color":   new_color,
                }
                _mark_dirty(); st.rerun()

            if clear:
                p["slots"][sel] = {"product": "", "price": "", "qty": "", "color": ""}
                _mark_dirty(); st.rerun()


# ════════════════════════════════════════════════════════
# VUE : BIBLIOTHÈQUE PRODUITS
# ════════════════════════════════════════════════════════

def _view_library():
    produits = _get_produits()

    hc1, hc2, hc3, hc4 = st.columns([3, 2, 2, 2])
    with hc1:
        st.markdown("### 📦 Bibliothèque de produits")
    with hc2:
        if st.button("+ Ajouter un produit", use_container_width=True, type="primary"):
            st.session_state.pg_view = "new_product"
            st.rerun()
    with hc3:
        if st.button("↑ Import Excel", use_container_width=True):
            st.session_state.pg_view = "import_produits_excel"
            st.rerun()
    with hc4:
        if st.button("← Retour aux planogrammes", use_container_width=True):
            st.session_state.pg_view = "list"
            st.rerun()

    # Afficher résultat de l'import si disponible
    if st.session_state.get("_excel_import_result"):
        added, updated, skipped = st.session_state.pop("_excel_import_result")
        st.success(f"✅ Import terminé : **{added}** ajoutés · **{updated}** mis à jour · **{skipped}** ignorés")

    sc1, sc2 = st.columns([3, 2])
    with sc1:
        search = st.text_input("🔍 Rechercher", placeholder="Nom, code ou catégorie…", key="lib_search")
    with sc2:
        cats = sorted(set(p.get("categorie", "") for p in produits if p.get("categorie")))
        cat_filter = st.selectbox("Catégorie", ["Toutes"] + cats, key="lib_cat_filter")

    filtered = [p for p in produits
                if (not search or search.lower() in p["nom"].lower()
                    or search.lower() in p.get("code", "").lower()
                    or search.lower() in p.get("categorie", "").lower())
                and (cat_filter == "Toutes" or p.get("categorie", "") == cat_filter)]

    if filtered:
        st.caption(f"{len(filtered)} produit(s) · Stock total estimé : **{sum(_marge_produit(p)[2] for p in filtered if p.get('prix_ttc')):.2f} €** valeur TTC")

    if not filtered:
        st.info("Aucun produit." + (" Aucun résultat pour cette recherche." if search else " Ajoutez-en un ou importez un fichier Excel !"))
        return

    # Tableau récapitulatif
    rows = []
    for pr in filtered:
        marge_ht, marge_pct, _ = _marge_produit(pr)
        rows.append({
            "Code":          pr.get("code", "—"),
            "Nom":           pr["nom"],
            "Catégorie":     pr.get("categorie", "—"),
            "P.U. HT":       f"{float(pr['prix_ht']):.2f} €"  if pr.get("prix_ht")  else "—",
            "Prix achat":    f"{float(pr['prix_achat']):.2f} €" if pr.get("prix_achat") else "—",
            "TVA":           f"{float(pr['tva']):.1f} %"       if pr.get("tva")      else "—",
            "Prix TTC":      f"{float(pr['prix_ttc']):.2f} €" if pr.get("prix_ttc") else "—",
            "Marge HT":      f"{marge_ht:.2f} €" if marge_ht is not None else "—",
            "Marge %":       f"{marge_pct:.1f} %" if marge_pct is not None else "—",
            "Couleur":       pr.get("couleur", ""),
        })

    df_lib = pd.DataFrame(rows)

    # Affichage tableau + boutons par ligne
    st.dataframe(
        df_lib.drop(columns=["Couleur"]),
        use_container_width=True,
        hide_index=True,
        height=min(600, 38 + len(df_lib) * 35),
    )

    st.divider()
    st.caption("Actions sur un produit :")
    ac1, ac2, ac3 = st.columns([3, 1, 1])
    with ac1:
        prod_names = [f"{p.get('code','')} — {p['nom']}" for p in filtered]
        selected_name = st.selectbox("Sélectionner un produit", prod_names, key="lib_action_select", label_visibility="collapsed")
        selected_pr = filtered[prod_names.index(selected_name)] if selected_name else None
    with ac2:
        if selected_pr and st.button("✏️ Modifier", use_container_width=True):
            st.session_state["edit_product_id"] = selected_pr["_id"]
            st.session_state.pg_view = "new_product"
            st.rerun()
    with ac3:
        if selected_pr and st.button("🗑️ Supprimer", use_container_width=True):
            delete_produit(selected_pr["_id"])
            _reload_produits()
            st.rerun()


# ════════════════════════════════════════════════════════
# VUE : FORMULAIRE PRODUIT (création / édition)
# ════════════════════════════════════════════════════════

def _view_new_product():
    edit_id  = st.session_state.get("edit_product_id")
    produits = _get_produits()
    existing = next((p for p in produits if p["_id"] == edit_id), None) if edit_id else None

    title = "✏️ Modifier le produit" if existing else "➕ Ajouter un produit"
    st.markdown(f"### {title}")

    def _v(field, default=""):
        return existing.get(field, default) if existing else default

    with st.form(key=f"form_product_{edit_id or 'new'}"):
        c1, c2 = st.columns([2, 1])
        with c1:
            nom  = st.text_input("Nom du produit *", value=_v("nom"))
        with c2:
            code = st.text_input("Code", value=_v("code"), placeholder="ex: REDBULL25CL")

        c3, c4 = st.columns(2)
        with c3:
            cat = st.text_input("Catégorie", value=_v("categorie"), placeholder="Boissons, Barres, Accessoires…")
        with c4:
            color_names = list(SLOT_COLORS.keys())
            cur_color_name = next((n for n, v in SLOT_COLORS.items() if v == _v("couleur")), "Aucune")
            col_choice = st.selectbox("Couleur d'affichage", color_names, index=color_names.index(cur_color_name))

        st.markdown("**Prix**")
        p1, p2, p3 = st.columns(3)
        with p1:
            prix_ht    = st.text_input("P.U. HT (€)",        value=_v("prix_ht"),    placeholder="3.51")
        with p2:
            prix_achat = st.text_input("Prix achat unitaire (€)", value=_v("prix_achat"), placeholder="1.13")
        with p3:
            tva        = st.text_input("TVA (%)",              value=_v("tva"),       placeholder="5.5")

        # Calcul auto Prix TTC
        prix_ttc_val = ""
        try:
            ht  = float(prix_ht.replace(",", ".")) if prix_ht else None
            t   = float(str(tva).replace(",", ".")) if tva else None
            if ht and t is not None:
                prix_ttc_val = f"{ht * (1 + t/100):.4f}"
        except Exception:
            pass
        prix_ttc = st.text_input("Prix TTC (€) — calculé auto", value=prix_ttc_val, placeholder="3.70")

        submitted = st.form_submit_button("Enregistrer", type="primary", use_container_width=True)

    if submitted:
        if not nom.strip():
            st.error("Le nom est obligatoire.")
            return
        produit = {
            "nom":        nom.strip(),
            "code":       code.strip().upper(),
            "categorie":  cat.strip(),
            "couleur":    SLOT_COLORS[col_choice],
            "prix_ht":    prix_ht.strip().replace(",", "."),
            "prix_achat": prix_achat.strip().replace(",", "."),
            "tva":        str(tva).strip().replace(",", "."),
            "prix_ttc":   prix_ttc.strip().replace(",", "."),
        }
        if existing:
            produit["_id"] = existing["_id"]
        save_produit(produit)
        st.session_state.pop("edit_product_id", None)
        _reload_produits()
        st.toast("✅ Produit enregistré.")
        st.session_state.pg_view = "library"
        st.rerun()

    if st.button("← Annuler"):
        st.session_state.pop("edit_product_id", None)
        st.session_state.pg_view = "library"
        st.rerun()


# ════════════════════════════════════════════════════════
# IMPORT CSV
# ════════════════════════════════════════════════════════

def _import_csv_to_plano(p: dict, raw: bytes):
    try:
        text  = raw.decode("utf-8-sig").strip()
        lines = text.split("\n")
        rows  = len(lines)
        cols  = max(len(l.split(";")) for l in lines)
        p["rows"]       = rows
        p["cols"]       = cols
        p["row_labels"] = [f"R{i+1}" for i in range(rows)]
        p["slots"]      = {}
        for r, line in enumerate(lines):
            cells = line.split(";")
            for c in range(cols):
                cell = cells[c].strip() if c < len(cells) else ""
                price_m = re.search(r"(\d+[.,]\d{2})\s*€?", cell)
                qty_m   = re.search(r"[xX×]\s*(\d+)", cell)
                price   = price_m.group(1).replace(",", ".") if price_m else ""
                qty     = qty_m.group(1) if qty_m else ""
                product = re.sub(r"\d+[.,]\d{2}\s*€?", "", cell)
                product = re.sub(r"[xX×]\s*\d+", "", product).strip()
                p["slots"][f"{r}-{c}"] = {"product": product, "price": price, "qty": qty, "color": ""}
        st.toast(f"CSV importé : {rows} rangées × {cols} colonnes.")
    except Exception as e:
        st.error(f"Erreur import CSV : {e}")


# ════════════════════════════════════════════════════════
# EXPORTS
# ════════════════════════════════════════════════════════

def _do_export_excel(p: dict):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = p["nom"][:31]

        # En-tête
        header = ["Rangée"] + [f"C{c+1}" for c in range(p["cols"])]
        ws.append(header)
        for cell in ws[1]:
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.fill      = PatternFill("solid", fgColor="1F4E79")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for r in range(p["rows"]):
            lbl = p["row_labels"][r] if r < len(p["row_labels"]) else f"R{r+1}"
            row_data = [lbl]
            for c in range(p["cols"]):
                s = p["slots"].get(f"{r}-{c}", {})
                if s.get("product"):
                    cell_val = s["product"]
                    if s.get("qty"):   cell_val += f" ×{s['qty']}"
                    if s.get("price"):
                        try: cell_val += f" {float(s['price']):.2f}€"
                        except: cell_val += f" {s['price']}€"
                    row_data.append(cell_val)
                else:
                    row_data.append("")
            ws.append(row_data)

        # Style corps
        for row in ws.iter_rows(min_row=2):
            for i, cell in enumerate(row):
                cell.border    = border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                if i == 0:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill("solid", fgColor="E8EAF0")

        # Largeurs colonnes
        ws.column_dimensions["A"].width = 10
        for c in range(1, p["cols"] + 1):
            ws.column_dimensions[get_column_letter(c + 1)].width = 18

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        safe_name = re.sub(r"[^\w\s-]", "_", p["nom"])
        st.download_button(
            label="⬇️ Télécharger Excel",
            data=buf.getvalue(),
            file_name=f"{safe_name}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        st.error(f"Erreur export Excel : {e}")


def _hex_to_rgb(hex_color: str):
    """Convertit #rrggbb en tuple (r, g, b) float 0-1 pour reportlab."""
    hex_color = (hex_color or "").strip().lstrip("#")
    if len(hex_color) == 6:
        r = int(hex_color[0:2], 16) / 255
        g = int(hex_color[2:4], 16) / 255
        b = int(hex_color[4:6], 16) / 255
        return (r, g, b)
    return None


def _do_export_pdf(p: dict):
    """Génère un vrai fichier PDF avec reportlab et le propose en téléchargement."""
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT

        # ── Orientation : paysage si beaucoup de colonnes ──
        page_size = landscape(A4)
        page_w, page_h = page_size
        margin = 15 * mm

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf,
            pagesize=page_size,
            leftMargin=margin, rightMargin=margin,
            topMargin=14 * mm, bottomMargin=14 * mm,
        )

        styles = getSampleStyleSheet()

        # ── Tailles de police adaptatives selon le nombre de colonnes ──
        # Plus il y a de colonnes, plus on réduit pour que tout rentre
        cols = p["cols"]
        if cols <= 8:
            fs_cell, fs_price, fs_qty, fs_empty = 7, 7, 7, 9
            row_h_mm = 18
        elif cols <= 12:
            fs_cell, fs_price, fs_qty, fs_empty = 6, 6, 6, 7
            row_h_mm = 16
        elif cols <= 16:
            fs_cell, fs_price, fs_qty, fs_empty = 5, 5, 5, 6
            row_h_mm = 14
        else:
            fs_cell, fs_price, fs_qty, fs_empty = 4, 4, 4, 5
            row_h_mm = 12

        leading_cell  = fs_cell + 2
        leading_price = fs_price + 2

        title_style = ParagraphStyle(
            "Title", parent=styles["Normal"],
            fontSize=14, fontName="Helvetica-Bold",
            textColor=colors.HexColor("#1F4E79"),
            spaceAfter=3,
        )
        sub_style = ParagraphStyle(
            "Sub", parent=styles["Normal"],
            fontSize=9, fontName="Helvetica",
            textColor=colors.HexColor("#666666"),
            spaceAfter=8,
        )
        cell_style = ParagraphStyle(
            "Cell", parent=styles["Normal"],
            fontSize=fs_cell, fontName="Helvetica-Bold",
            alignment=TA_CENTER, leading=leading_cell,
        )
        price_style = ParagraphStyle(
            "Price", parent=styles["Normal"],
            fontSize=fs_price, fontName="Helvetica",
            textColor=colors.HexColor("#185FA5"),
            alignment=TA_CENTER, leading=leading_price,
        )
        qty_style = ParagraphStyle(
            "Qty", parent=styles["Normal"],
            fontSize=fs_qty, fontName="Helvetica",
            textColor=colors.HexColor("#888888"),
            alignment=TA_CENTER, leading=leading_price,
        )
        empty_style = ParagraphStyle(
            "Empty", parent=styles["Normal"],
            fontSize=fs_empty, fontName="Helvetica",
            textColor=colors.HexColor("#CCCCCC"),
            alignment=TA_CENTER,
        )

        # ── Calcul largeur colonnes ──
        usable_w = page_w - 2 * margin
        label_w  = 14 * mm
        data_col_w = (usable_w - label_w) / p["cols"]
        col_widths = [label_w] + [data_col_w] * p["cols"]

        # ── Construction du tableau ──
        # Ligne d'en-tête
        header_row = [Paragraph("", cell_style)] + [
            Paragraph(f"C{c+1}", cell_style) for c in range(p["cols"])
        ]
        table_data  = [header_row]
        cell_styles_cmds = [
            # En-tête bleue
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
            ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
            ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",   (0, 0), (-1, 0), 7),
            ("ALIGN",      (0, 0), (-1, 0), "CENTER"),
            ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
            ("GRID",       (0, 0), (-1, -1), 0.3, colors.HexColor("#DDDDDD")),
            ("ROWBACKGROUNDS", (0, 0), (0, -1), [colors.HexColor("#E8EAF0")]),
            ("FONTNAME",   (0, 1), (0, -1), "Helvetica-Bold"),
            ("FONTSIZE",   (0, 1), (0, -1), 8),
            ("ALIGN",      (0, 1), (0, -1), "CENTER"),
        ]

        for r in range(p["rows"]):
            lbl = p["row_labels"][r] if r < len(p["row_labels"]) else f"R{r+1}"
            row_data = [Paragraph(lbl, cell_style)]
            for c in range(p["cols"]):
                s = p["slots"].get(f"{r}-{c}", {})
                if s.get("product"):
                    parts = [Paragraph(s["product"], cell_style)]
                    if s.get("price"):
                        try:    parts.append(Paragraph(f"{float(s['price']):.2f} €", price_style))
                        except: parts.append(Paragraph(f"{s['price']} €", price_style))
                    if s.get("qty"):
                        parts.append(Paragraph(f"×{s['qty']}", qty_style))
                    row_data.append(parts)
                else:
                    row_data.append(Paragraph("—", empty_style))

                # Couleur de fond du slot
                bg_hex = s.get("color", "")
                if bg_hex and bg_hex.startswith("#") and len(bg_hex) == 7:
                    try:
                        rgb = _hex_to_rgb(bg_hex)
                        if rgb:
                            bg_color = colors.Color(*rgb)
                            cell_styles_cmds.append(
                                ("BACKGROUND", (c + 1, r + 1), (c + 1, r + 1), bg_color)
                            )
                    except Exception:
                        pass

            table_data.append(row_data)

        # Hauteur de ligne
        row_h = 18 * mm
        row_heights = [6 * mm] + [row_h_mm * mm] * p["rows"]

        table = Table(table_data, colWidths=col_widths, rowHeights=row_heights)
        table.setStyle(TableStyle(cell_styles_cmds))

        # ── Assemblage ──
        today = datetime.date.today().strftime("%d/%m/%Y")
        type_label = "Double" if p.get("type") == "double" else "Simple"
        story = [
            Paragraph(p["nom"], title_style),
            Paragraph(
                f"{type_label} — {p['rows']} rangées × {p['cols']} colonnes — Exporté le {today}",
                sub_style
            ),
            table,
        ]

        doc.build(story)
        buf.seek(0)

        safe_name = re.sub(r"[^\w\s-]", "_", p["nom"])
        st.download_button(
            label="⬇️ Télécharger le PDF",
            data=buf.getvalue(),
            file_name=f"{safe_name}.pdf",
            mime="application/pdf",
        )

    except ImportError:
        st.error("La librairie `reportlab` n'est pas installée. Ajoutez `reportlab` dans requirements.txt.")
    except Exception as e:
        st.error(f"Erreur export PDF : {e}")



# ════════════════════════════════════════════════════════
# PARSEUR PDF
# ════════════════════════════════════════════════════════

def _parse_planogramme_pdf(file_bytes: bytes) -> dict:
    """
    Parse un PDF de planogramme. Gère deux formats :
    - Format BF  : ligne quantités seules (X12, X9…) / ligne produits / ligne prix
    - Format FP  : produit + quantité dans la même cellule (ex: "Red Bull X9") / ligne prix
    """
    import io as _io
    with pdfplumber.open(_io.BytesIO(file_bytes)) as pdf:
        all_rows = []
        title = ""

        for page in pdf.pages:
            tables = page.extract_tables()
            if not tables:
                continue
            table = tables[0]

            # Chercher le titre dans les premières lignes
            for row in table[:5]:
                for cell in (row or []):
                    if cell and re.search(
                        r'VENDING|PLANOGRAMM|FITNESSPARK|BASIC.?FIT|COMBOPLUS',
                        str(cell), re.I
                    ) and len(str(cell)) > 4:
                        title = str(cell).strip()
                        break

            # ── Helpers ──────────────────────────────────────
            def is_price_row(row):
                cells = [c for c in (row or []) if c and str(c).strip()]
                if not cells:
                    return False
                n_prices = sum(1 for c in cells if re.search(r'\d+[,\.]\d{2}\s*€', str(c)))
                return n_prices >= max(1, len(cells) * 0.5)

            def is_qty_only_row(row):
                """Format BF : toutes les cellules non vides sont X12, X10…"""
                cells = [c for c in (row or []) if c and str(c).strip()]
                if len(cells) < 2:
                    return False
                n_qty = sum(1 for c in cells if re.match(r'^X\d+$', str(c).strip()))
                return n_qty >= max(2, len(cells) * 0.5)

            def extract_product_qty(cell_text):
                """Extrait (produit, qty) d'une cellule qui contient les deux."""
                text = str(cell_text or '').replace('\n', ' ').strip()
                if not text or text.lower() == 'none':
                    return None, None
                # Quantité en fin : "Red Bull X9" ou "Evian x8"
                m = re.search(r'[Xx](\d+)\s*$', text)
                if m:
                    product = text[:m.start()].strip()
                    return product, m.group(1)
                # Quantité en début : "X9 Red Bull"
                m2 = re.search(r'^[Xx](\d+)\s+(.+)', text)
                if m2:
                    return m2.group(2).strip(), m2.group(1)
                return text, ''

            def extract_price(cell_text):
                m = re.search(r'(\d+[,\.]\d{2})', str(cell_text or ''))
                return m.group(1).replace(',', '.') if m else ''

            # ── Parcours du tableau ───────────────────────────
            i = 2  # skip lignes d'en-tête
            while i < len(table):
                row = table[i]

                # Ignorer les lignes de prix et les lignes de labels purs
                if is_price_row(row):
                    i += 1
                    continue

                # ── Format BF : ligne de quantités pures ──
                if is_qty_only_row(row):
                    qty_row   = row
                    prod_row  = table[i + 1] if i + 1 < len(table) else []
                    price_row = table[i + 2] if i + 2 < len(table) else []
                    n = max(len(qty_row), len(prod_row or []), len(price_row or []))
                    slots = []
                    for c in range(n):
                        qty   = str(qty_row[c] if c < len(qty_row) else '').strip()
                        prod  = str((prod_row or [])[c] if c < len(prod_row or []) else '').replace('\n', ' ').strip()
                        price = extract_price((price_row or [])[c] if c < len(price_row or []) else '')
                        if re.match(r'^X\d+$', qty) and prod and prod.lower() not in ('none', ''):
                            slots.append({'product': prod, 'price': price, 'qty': qty[1:]})
                        else:
                            slots.append(None)
                    if any(s for s in slots):
                        all_rows.append({'slots': slots})
                    i += 3
                    continue

                # ── Format FP : produit + quantité dans la même cellule ──
                cells = [c for c in (row or []) if c and str(c).strip() and str(c).strip().lower() != 'none']
                n_embedded = sum(1 for c in cells if re.search(r'[Xx]\d+', str(c)))
                if cells and n_embedded >= max(1, len(cells) * 0.4):
                    price_row = table[i + 1] if i + 1 < len(table) else []
                    n = max(len(row or []), len(price_row or []))
                    slots = []
                    for c in range(n):
                        cell  = (row or [])[c] if c < len(row or []) else ''
                        price = extract_price((price_row or [])[c] if c < len(price_row or []) else '')
                        product, qty = extract_product_qty(cell)
                        if product and product.lower() not in ('none', ''):
                            slots.append({'product': product, 'price': price, 'qty': qty or ''})
                        else:
                            slots.append(None)
                    if any(s for s in slots):
                        all_rows.append({'slots': slots})
                    i += 2
                    continue

                i += 1

        # Supprimer les lignes sans aucun produit réel
        all_rows = [r for r in all_rows if any(s and s.get('product') for s in r['slots'])]

        if not all_rows:
            raise ValueError(
                "Aucune rangée de produits détectée. "
                "Formats supportés : BF (lignes X12/produits/prix) et FP (produit+quantité/prix)."
            )

        type_dist = 'double' if 'DOUBLE' in title.upper() else 'simple'
        return {'title': title or 'Planogramme importé', 'type': type_dist, 'rows': all_rows}


def _build_plano_from_parsed(parsed: dict, nom: str) -> dict:
    """Convertit le résultat du parseur en structure MongoDB planogramme."""
    rows_data = parsed["rows"]
    n_rows    = len(rows_data)
    n_cols    = max(len(r["slots"]) for r in rows_data)

    slots = {}
    for r, row in enumerate(rows_data):
        for c in range(n_cols):
            s = row["slots"][c] if c < len(row["slots"]) else None
            if s and s.get("product"):
                slots[f"{r}-{c}"] = {
                    "product": s["product"],
                    "price":   s.get("price", ""),
                    "qty":     s.get("qty", ""),
                    "color":   "",
                }
            else:
                slots[f"{r}-{c}"] = {"product": "", "price": "", "qty": "", "color": ""}

    return {
        "nom":        nom,
        "type":       parsed.get("type", "simple"),
        "rows":       n_rows,
        "cols":       n_cols,
        "row_labels": [f"R{i+1}" for i in range(n_rows)],
        "slots":      slots,
    }


# ════════════════════════════════════════════════════════
# VUE : IMPORT PDF
# ════════════════════════════════════════════════════════

def _view_import_pdf():
    st.markdown("### 📄 Importer un planogramme depuis un PDF")

    if st.button("← Retour"):
        st.session_state.pg_view = "list"
        st.rerun()

    if not _PDFPLUMBER_OK:
        st.error("La librairie `pdfplumber` n'est pas installée. Ajoutez `pdfplumber` dans `requirements.txt`.")
        return

    st.info(
        "**Format supporté** : planogrammes NXT Level / Basic-Fit avec la structure "
        "Quantités (X12, X10…) → Produits → Prix par rangée."
    )

    uploaded = st.file_uploader(
        "Choisir un fichier PDF", type=["pdf"],
        key="pdf_import_uploader", label_visibility="collapsed"
    )

    if not uploaded:
        return

    # Parser le PDF dès l'upload
    if st.session_state.get("pdf_parsed_name") != uploaded.name:
        with st.spinner("Analyse du PDF en cours…"):
            try:
                parsed = _parse_planogramme_pdf(uploaded.read())
                st.session_state["pdf_parsed"]      = parsed
                st.session_state["pdf_parsed_name"] = uploaded.name
                st.session_state["pdf_import_nom"]  = parsed["title"]
            except Exception as e:
                st.error(f"Erreur lors de la lecture du PDF : {e}")
                return

    parsed = st.session_state.get("pdf_parsed")
    if not parsed:
        return

    # ── Aperçu du contenu détecté ──
    total_slots = sum(
        sum(1 for s in row["slots"] if s and s.get("product"))
        for row in parsed["rows"]
    )
    n_cols = max(len(r["slots"]) for r in parsed["rows"])

    st.success(
        f"✅ **{len(parsed['rows'])} rangées** · **{n_cols} colonnes** · "
        f"**{total_slots} emplacements** détectés"
    )

    # Tableau d'aperçu
    preview_rows = []
    for i, row in enumerate(parsed["rows"]):
        filled = [s for s in row["slots"] if s and s.get("product")]
        preview_rows.append({
            "Rangée":    f"R{i+1}",
            "Remplis":   len(filled),
            "Exemple 1": filled[0]["product"] if len(filled) > 0 else "—",
            "Exemple 2": filled[1]["product"] if len(filled) > 1 else "—",
            "Exemple 3": filled[2]["product"] if len(filled) > 2 else "—",
        })
    st.dataframe(
        pd.DataFrame(preview_rows),
        use_container_width=True, hide_index=True
    )

    st.divider()

    # ── Options avant import ──
    with st.form("form_pdf_import"):
        nom = st.text_input(
            "Nom du planogramme",
            value=st.session_state.get("pdf_import_nom", parsed["title"])
        )
        c1, c2 = st.columns(2)
        with c1:
            type_dist = st.selectbox(
                "Type de distributeur",
                ["simple", "double"],
                index=0 if parsed.get("type") == "simple" else 1
            )
        with c2:
            st.markdown(f"**Colonnes détectées :** {n_cols}")

        submitted = st.form_submit_button(
            "✅ Importer ce planogramme", type="primary", use_container_width=True
        )

    if submitted:
        if not nom.strip():
            st.error("Le nom est obligatoire.")
            return

        parsed["type"] = type_dist
        plano = _build_plano_from_parsed(parsed, nom.strip())
        new_id = save_planogramme(plano)
        plano["_id"] = new_id

        # Nettoyage session
        for k in ["pdf_parsed", "pdf_parsed_name", "pdf_import_nom"]:
            st.session_state.pop(k, None)

        _reload_planogrammes()
        _set_current(plano)
        st.toast(f"✅ Planogramme '{nom}' importé avec succès !")
        st.rerun()


# ════════════════════════════════════════════════════════
# HELPER MARGE
# ════════════════════════════════════════════════════════

def _marge_produit(pr: dict):
    """Retourne (marge_ht, marge_pct, valeur_ttc) ou (None, None, 0)."""
    try:
        ht     = float(pr.get("prix_ht", 0) or 0)
        achat  = float(pr.get("prix_achat", 0) or 0)
        ttc    = float(pr.get("prix_ttc", 0) or 0)
        if ht and achat:
            marge_ht  = ht - achat
            marge_pct = (marge_ht / ht * 100) if ht else 0
            return round(marge_ht, 4), round(marge_pct, 2), round(ttc, 4)
    except Exception:
        pass
    return None, None, 0.0


# ════════════════════════════════════════════════════════
# VUE : IMPORT PRODUITS DEPUIS EXCEL
# ════════════════════════════════════════════════════════

def _view_import_produits_excel():
    st.markdown("### ↑ Importer des produits depuis Excel")

    if st.button("← Retour à la bibliothèque"):
        st.session_state.pg_view = "library"
        st.rerun()

    st.info(
        "**Format attendu :** colonnes `Code`, `Nom`, `P.U. HT`, `Prix unit. achat`, `TVA`, `Prix TTC`. "
        "Les colonnes supplémentaires sont ignorées."
    )

    uploaded = st.file_uploader(
        "Fichier Excel (.xlsx)", type=["xlsx"],
        key="excel_prod_uploader", label_visibility="collapsed"
    )
    if not uploaded:
        return

    try:
        df = pd.read_excel(uploaded)
    except Exception as e:
        st.error(f"Impossible de lire le fichier : {e}")
        return

    # Mapping flexible des colonnes
    col_map = {}
    for col in df.columns:
        c = str(col).strip().lower()
        if "code" in c:                          col_map["code"]       = col
        elif "nom" in c:                          col_map["nom"]        = col
        elif "p.u" in c or "pu ht" in c or (c == "p.u. ht"):
                                                  col_map["prix_ht"]    = col
        elif "achat" in c:                        col_map["prix_achat"] = col
        elif "tva" in c:                          col_map["tva"]        = col
        elif "ttc" in c:                          col_map["prix_ttc"]   = col

    if "nom" not in col_map:
        st.error("Colonne `Nom` introuvable dans le fichier.")
        return

    # Aperçu
    st.success(f"✅ {len(df)} produit(s) détectés")
    cols_found = {k: v for k, v in col_map.items()}
    st.caption(f"Colonnes mappées : {cols_found}")
    st.dataframe(df.head(10), use_container_width=True, hide_index=True)

    # Options
    with st.form("form_import_excel_prod"):
        cat_default = st.text_input("Catégorie par défaut", placeholder="ex: Boissons, Barres…")
        mode = st.radio(
            "Mode d'import",
            ["Ajouter les nouveaux seulement (ne pas écraser)", "Écraser si même code"],
            index=1
        )
        submitted = st.form_submit_button("✅ Importer", type="primary", use_container_width=True)

    # Guard anti-doublon : on exécute l'import une seule fois grâce au flag
    if submitted and not st.session_state.get("_excel_import_done"):
        st.session_state["_excel_import_done"] = True

        produits_existants = _get_produits()
        codes_existants = {p.get("code", "").upper(): p["_id"]
                           for p in produits_existants if p.get("code")}

        added = updated = skipped = 0

        for _, row in df.iterrows():
            nom = str(row.get(col_map.get("nom", ""), "")).strip()
            if not nom or nom.lower() == "nan":
                continue

            code = str(row.get(col_map.get("code", ""), "")).strip().upper() if "code" in col_map else ""
            if code == "NAN":
                code = ""

            def _fval(field, _row=row):
                if field not in col_map: return ""
                val = _row.get(col_map[field], "")
                try:
                    f = float(val)
                    return f"{f:.4f}" if f else ""
                except Exception:
                    return ""

            # Calcul TTC auto si manquant
            prix_ttc = _fval("prix_ttc")
            if not prix_ttc:
                try:
                    ht  = float(_fval("prix_ht") or 0)
                    tva = float(_fval("tva") or 0)
                    if ht:
                        prix_ttc = f"{ht * (1 + tva/100):.4f}"
                except Exception:
                    pass

            produit = {
                "nom":        nom,
                "code":       code,
                "categorie":  cat_default.strip() if cat_default.strip() else "",
                "couleur":    "",
                "prix_ht":    _fval("prix_ht"),
                "prix_achat": _fval("prix_achat"),
                "tva":        _fval("tva"),
                "prix_ttc":   prix_ttc,
            }

            if mode == "Ajouter les nouveaux seulement (ne pas écraser)":
                if code and code in codes_existants:
                    skipped += 1
                    continue
                save_produit(produit)
                added += 1

            else:  # Écraser si même code
                if code and code in codes_existants:
                    produit["_id"] = codes_existants[code]
                    save_produit(produit)
                    updated += 1
                else:
                    save_produit(produit)
                    added += 1

        _reload_produits()
        st.session_state["_excel_import_result"] = (added, updated, skipped)
        st.session_state.pg_view = "library"
        st.rerun()

    # Nettoyage du flag après le rerun
    st.session_state.pop("_excel_import_done", None)


# ════════════════════════════════════════════════════════
# POINT D'ENTRÉE PRINCIPAL
# ════════════════════════════════════════════════════════

def render():
    _init_state()

    view = st.session_state.pg_view

    if view == "list":
        _view_list()
    elif view == "new":
        _view_new()
    elif view == "editor":
        _view_editor()
    elif view == "library":
        _view_library()
    elif view == "new_product":
        _view_new_product()
    elif view == "import_pdf":
        _view_import_pdf()
    elif view == "import_produits_excel":
        _view_import_produits_excel()
    else:
        st.session_state.pg_view = "list"
        st.rerun()
