"""
Application Streamlit - Outil Réappro
  ├── 📦 Suivi Réapprovisionneurs
  └── 🗂️  Planogrammes
"""

import io
import streamlit as st
import pandas as pd
import datetime

from planning_parser import get_today_day_str, get_default_jour_analyse, JOURS
import page_machines
import page_no_audit
import page_cr
import page_planogrammes
import page_inventaires
import page_controle_reception
import page_indefinis
import page_quartix
import page_picklist
import page_rapport_employe
from mongo_storage import load_plannings_from_mongo
from chargement_parser import parse_chargement_csv, croiser_planning_chargement
from excel_export import generer_excel

# ────────────────────────────────────────────────────────
# CONFIG PAGE
# ────────────────────────────────────────────────────────
DARK_CSS = """<style>
/* ── DARK MODE OVERRIDE ── */
.stApp,
[data-testid="stAppViewContainer"],
[data-testid="stMain"] {
    background-color: #0F1923 !important;
    color: #E8EEF7 !important;
}
[data-testid="block-container"],
.block-container {
    background-color: #0F1923 !important;
}
[data-testid="stSidebar"] {
    background-color: #1A2535 !important;
}
.main-title { color: #7BC4E8 !important; }
.subtitle   { color: #8FA8C8 !important; }
p, span, label, div { color: #E8EEF7 !important; }
.stButton > button[kind="secondary"] {
    color: #7BC4E8 !important;
    border-color: #2A4A6B !important;
    background-color: transparent !important;
}
.stButton > button[kind="secondary"]:hover {
    background-color: #1E2E42 !important;
    border-color: #E8922A !important;
    color: #E8922A !important;
}
[data-testid="metric-container"] {
    background-color: #1A2535 !important;
    border: 1px solid #2A4A6B !important;
    border-radius: 0.5rem !important;
    padding: 0.5rem !important;
}
[data-testid="metric-container"] label,
[data-testid="metric-container"] [data-testid="stMetricLabel"] { color: #7BC4E8 !important; }
[data-testid="metric-container"] [data-testid="stMetricValue"] { color: #E8EEF7 !important; }
[data-testid="metric-container"] [data-testid="stMetricDelta"] { color: #8FA8C8 !important; }
.stTextInput input,
.stNumberInput input,
.stTextArea textarea,
[data-testid="stSelectbox"] > div > div,
[data-baseweb="select"] > div,
[data-baseweb="input"] > div {
    background-color: #1A2535 !important;
    color: #E8EEF7 !important;
    border-color: #2A4A6B !important;
}
[data-baseweb="popover"],
[data-baseweb="menu"],
ul[data-baseweb="menu"] {
    background-color: #1E2E42 !important;
    border-color: #2A4A6B !important;
}
[data-baseweb="menu"] li { color: #E8EEF7 !important; }
[data-baseweb="menu"] li:hover { background-color: #2A4A6B !important; }
[data-testid="stDataFrame"] > div,
.stDataFrame { background-color: #1A2535 !important; }
[data-testid="stDataFrame"] > div > div { border-color: #2A4A6B !important; }
[data-testid="stExpander"] {
    border-color: #2A4A6B !important;
    background-color: #1A2535 !important;
}
[data-testid="stExpander"] summary {
    color: #E8EEF7 !important;
    background-color: #1A2535 !important;
}
[data-testid="stExpander"] summary:hover { background-color: #1E2E42 !important; }
[data-testid="stTabs"] [role="tablist"] { border-bottom-color: #2A4A6B !important; }
[data-testid="stTabs"] button[role="tab"] { color: #8FA8C8 !important; }
[data-testid="stTabs"] button[role="tab"][aria-selected="true"] {
    color: #E8922A !important;
    border-bottom-color: #E8922A !important;
}
hr { border-color: rgba(123,196,232,0.2) !important; }
[data-testid="stFileUploader"],
[data-testid="stFileUploaderDropzone"],
section[data-testid="stFileUploaderDropzone"] {
    background-color: #1A2535 !important;
    border-color: #2A4A6B !important;
}
[data-testid="stFileUploader"] label,
[data-testid="stFileUploader"] small,
[data-testid="stFileUploaderDropzoneInstructions"],
[data-testid="stFileUploaderDropzoneInstructions"] span,
[data-testid="stFileUploaderDropzoneInstructions"] small,
[data-testid="stFileUploaderDropzone"] span,
[data-testid="stFileUploaderDropzone"] small,
[data-testid="stFileUploaderDropzone"] p {
    color: #E8EEF7 !important;
}
[data-testid="stFileUploaderDropzone"] button,
[data-testid="stFileUploaderDropzone"] button span {
    background-color: #1E2E42 !important;
    color: #7BC4E8 !important;
    border-color: #2A4A6B !important;
}
[data-testid="stFileUploaderDropzone"] button:hover,
[data-testid="stFileUploaderDropzone"] button:hover span {
    background-color: #2A4A6B !important;
    border-color: #7BC4E8 !important;
    color: #E8EEF7 !important;
}
[data-testid="stAlert"] {
    background-color: #1E2E42 !important;
    border-color: #2A4A6B !important;
}
.stCaption, [data-testid="stCaptionContainer"] { color: #8FA8C8 !important; }
.stDownloadButton > button {
    background-color: #1E2E42 !important;
    color: #7BC4E8 !important;
    border-color: #2A4A6B !important;
}
.stDownloadButton > button:hover {
    background-color: #2A4A6B !important;
    border-color: #7BC4E8 !important;
}
[data-baseweb="tag"] { background-color: #2A4A6B !important; }
[data-baseweb="tag"] span { color: #E8EEF7 !important; }
</style>"""

st.set_page_config(
    page_title="Distriprot Data",
    page_icon="📦",
    layout="wide",
    initial_sidebar_state="collapsed",
)

st.markdown("""
<style>
/* Cache le bouton toggle de la sidebar */
[data-testid="collapsedControl"] { display: none !important; }

/* Titres */
.main-title { font-size:2rem; font-weight:800; color:#1B3D6F; margin-bottom:0.2rem; }
.subtitle   { font-size:1rem; color:#555; margin-bottom:1.5rem; }

/* Bouton primaire (nav actif + actions principales) */
.stButton > button[kind="primary"] {
    background-color: #E8922A !important;
    border-color: #E8922A !important;
    color: #ffffff !important;
}
.stButton > button[kind="primary"]:hover {
    background-color: #D07A1A !important;
    border-color: #D07A1A !important;
    color: #ffffff !important;
}

/* Bouton secondaire (nav inactif) */
.stButton > button[kind="secondary"] {
    color: #1B3D6F !important;
    border-color: #7BC4E8 !important;
}
.stButton > button[kind="secondary"]:hover {
    background-color: #EEF5FC !important;
    border-color: #E8922A !important;
    color: #E8922A !important;
}

/* Métriques */
[data-testid="metric-container"] label { color: #1B3D6F; }

/* Empêche le retour à la ligne dans les boutons nav */
.stButton > button { white-space: nowrap !important; overflow: hidden; text-overflow: ellipsis; }

/* Remonte le contenu au maximum en haut */
[data-testid="stHeader"] { display: none !important; }
.block-container { padding-top: 0.5rem !important; }

/* Bouton toggle thème */
.theme-toggle-btn > button {
    background-color: transparent !important;
    border: 1px solid #7BC4E8 !important;
    color: #1B3D6F !important;
    font-size: 1.1rem !important;
    height: 2rem !important;
    width: 2.6rem !important;
    min-height: 0 !important;
    border-radius: 50% !important;
    padding: 0.2rem 0.5rem !important;
}
.theme-toggle-btn > button:hover {
    background-color: #EEF5FC !important;
    border-color: #E8922A !important;
}
</style>
""", unsafe_allow_html=True)

if st.session_state.get("dark_mode", False):
    st.markdown(DARK_CSS, unsafe_allow_html=True)

# ────────────────────────────────────────────────────────
# SESSION STATE
# ────────────────────────────────────────────────────────
if "page" not in st.session_state:
    st.session_state.page = "machines"
if "results" not in st.session_state:
    st.session_state.results = {}
if "jour_analyse" not in st.session_state:
    st.session_state.jour_analyse = get_default_jour_analyse()
if "chargement_bytes" not in st.session_state:
    st.session_state["chargement_bytes"] = None
if "excel_bytes" not in st.session_state:
    st.session_state["excel_bytes"] = None
if "dark_mode" not in st.session_state:
    st.session_state.dark_mode = False

# ────────────────────────────────────────────────────────
# NAVIGATION — barre en haut de la page principale
# ────────────────────────────────────────────────────────
import os as _os


# ────────────────────────────────────────────────────────
# HELPERS : MACHINES SANS CHARGEMENT (bas de page Tournées)
# ────────────────────────────────────────────────────────

def _parse_long_charg(file_bytes: bytes) -> pd.DataFrame:
    """
    Parse a long-period chargement CSV.
    Returns df with columns: Machine, Salle_csv, Date, is_fait.
    Same format as the daily chargement CSV but covering a longer period.
    """
    df = None
    for sep in [";", ",", "\t"]:
        for enc in ["utf-8-sig", "utf-8", "latin-1"]:
            try:
                _df = pd.read_csv(
                    io.BytesIO(file_bytes), sep=sep, dtype=str,
                    encoding=enc, on_bad_lines="skip",
                )
                _df.columns = [
                    c.strip().strip('"').replace("<br/>", " ") for c in _df.columns
                ]
                _df = _df.loc[:, _df.columns.str.strip() != ""]
                if len(_df.columns) >= 4:
                    lc = [c.lower() for c in _df.columns]
                    if (
                        any("machine" in c for c in lc)
                        and any("statut" in c or "status" in c for c in lc)
                        and any("date" in c for c in lc)
                    ):
                        df = _df
                        break
            except Exception:
                continue
        if df is not None:
            break

    if df is None:
        raise ValueError(
            "Format CSV non reconnu. Le fichier doit contenir les colonnes "
            "Machine, Statut et Date."
        )

    col_machine = col_statut = col_date = col_valeur = col_client = None
    for col in df.columns:
        cl = col.lower().strip()
        if "machine" in cl and col_machine is None:
            col_machine = col
        if ("statut" in cl or "status" in cl) and col_statut is None:
            col_statut = col
        if "date" in cl and col_date is None:
            col_date = col
        if (
            ("val" in cl and "ref" in cl)
            or cl in ("valeur", "montant", "prix", "amount", "value")
        ) and col_valeur is None:
            col_valeur = col
        if (
            "tiers" in cl or cl == "client" or "salle" in cl
        ) and col_client is None:
            col_client = col

    if not all([col_machine, col_statut, col_date]):
        raise ValueError(
            f"Colonnes requises non trouvées (Machine, Statut, Date). "
            f"Colonnes disponibles : {list(df.columns)}"
        )

    date_raw = df[col_date].astype(str).str.strip().str.strip('"')
    dates = pd.to_datetime(date_raw, format="%d/%m/%Y %H:%M", errors="coerce")
    mask = dates.isna()
    if mask.any():
        dates = dates.copy()
        dates.loc[mask] = pd.to_datetime(date_raw[mask], dayfirst=True, errors="coerce")

    out = pd.DataFrame()
    out["Machine"]   = df[col_machine].astype(str).str.strip().str.strip('"')
    out["Salle_csv"] = (
        df[col_client].astype(str).str.strip().str.strip('"') if col_client else ""
    )
    out["Date"]      = dates
    out["_statut"]   = (
        df[col_statut].astype(str).str.strip().str.strip('"').str.capitalize()
    )
    if col_valeur:
        out["_valeur"] = pd.to_numeric(
            df[col_valeur]
            .astype(str)
            .str.strip()
            .str.strip('"')
            .str.replace(",", ".", regex=False),
            errors="coerce",
        ).fillna(0.0)
    else:
        out["_valeur"] = 1.0

    out["is_fait"] = out["_statut"].isin(["Fait", "Annulé"]) & (out["_valeur"] != 0.0)
    return out[["Machine", "Salle_csv", "Date", "is_fait"]].dropna(subset=["Date"])


@st.cache_data(show_spinner=False, ttl=300)
def _load_salles_for_sc() -> pd.DataFrame:
    """Load machines from MongoDB: Client, Code, Approvisionneur."""
    from mongo_storage import _get_client as _get_mc
    try:
        client  = _get_mc()
        db_name = st.secrets["mongo"]["db_name"]
        docs = list(
            client[db_name]["machines"].find(
                {}, {"_id": 0, "Client": 1, "Code": 1, "Approvisionneur": 1}
            )
        )
        df = pd.DataFrame(docs)
        if df.empty:
            return df
        for c in ["Client", "Code", "Approvisionneur"]:
            if c not in df.columns:
                df[c] = ""
        return df.fillna("").drop_duplicates(subset=["Code"])
    except Exception:
        return pd.DataFrame(columns=["Client", "Code", "Approvisionneur"])


def _compute_sans_charg(
    charg_df: pd.DataFrame,
    salles_df: pd.DataFrame,
    plannings: dict,
    seuil_jours: int,
) -> pd.DataFrame:
    """
    Find salles without a valid chargement for more than seuil_jours days.
    Returns df: Salle, Code machine, Réappro, Dernier chargement, Jours sans chargement.
    """
    today = datetime.date.today()
    ref   = datetime.datetime.combine(today, datetime.time())

    # Build reverse map: machine_code → reappro name
    machine_to_reappro: dict[str, str] = {}
    for reappro, days in plannings.items():
        for jour, salles in days.items():
            for entry in salles:
                if isinstance(entry, (list, tuple)) and len(entry) >= 2:
                    machine_to_reappro[str(entry[1]).strip()] = reappro

    fait_df    = charg_df[charg_df["is_fait"]]
    last_charg = fait_df.groupby("Machine")["Date"].max()

    rows = []
    for _, row in salles_df.iterrows():
        code  = str(row.get("Code", "")).strip()
        salle = str(row.get("Client", "")).strip()
        appro = str(row.get("Approvisionneur", "")).strip()
        if not code:
            continue

        reappro = machine_to_reappro.get(code, appro)

        if code in last_charg.index:
            last  = last_charg[code]
            jours = (ref - last).days
            if jours <= seuil_jours:
                continue
            last_str = last.strftime("%d/%m/%Y")
        else:
            last     = None
            jours    = None
            last_str = "Jamais"

        rows.append({
            "Salle":                 salle,
            "Code machine":          code,
            "Réappro":               reappro,
            "Dernier chargement":    last_str,
            "Jours sans chargement": jours,
        })

    return pd.DataFrame(rows)


def _build_sc_excel_by_zone(df: pd.DataFrame) -> bytes:
    """
    Export Excel des machines sans chargement avec une feuille par zone (OUEST, IDF, …).
    La zone est déduite via la collection 'reappros' (code → zone).
    Dans chaque feuille de zone, les réappros sont séparés par une ligne de titre colorée.
    Une feuille 'Toutes zones' récapitule tout.
    """
    from mongo_storage import _get_client as _get_mc
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    # ── Mapping réappro code → zone (large) et prenom ────────────────────────
    zone_map: dict[str, str] = {}
    prenom_map: dict[str, str] = {}
    try:
        client  = _get_mc()
        db_name = st.secrets["mongo"]["db_name"]
        docs = list(
            client[db_name]["reappros"].find(
                {}, {"_id": 0, "code": 1, "zone": 1, "prenom": 1}
            )
        )
        for d in docs:
            code = str(d.get("code", "")).strip()
            if not code:
                continue
            if d.get("zone"):
                zone_map[code] = str(d["zone"]).strip()
            if d.get("prenom"):
                prenom_map[code] = str(d["prenom"]).strip()
    except Exception:
        pass

    export_df = df.drop(columns=["Tag"], errors="ignore").copy()
    export_df["Zone"] = export_df["Réappro"].map(zone_map).fillna("Non assigné")

    want = [
        "Zone", "Salle", "Code machine", "Réappro",
        "Dernier chargement", "Jours sans chargement",
        "Dernière vente", "Commentaire",
    ]
    cols = [c for c in want if c in export_df.columns]
    export_df = export_df[cols].sort_values(
        ["Zone", "Réappro", "Jours sans chargement"],
        ascending=[True, True, False],
        na_position="last",
    )

    # ── Style constants ───────────────────────────────────────────────────────
    HDR_FILL = PatternFill("solid", fgColor="1B3D6F")
    HDR_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=9)
    SEP_FILL = PatternFill("solid", fgColor="2E75B6")
    SEP_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=9)
    ALIGN_C  = Alignment(horizontal="center", vertical="center")
    ALIGN_L  = Alignment(horizontal="left",   vertical="center", wrap_text=False)

    WIDTH_HINTS = {
        "Zone": 20, "Salle": 38, "Code machine": 13,
        "Réappro": 12, "Dernier chargement": 16,
        "Jours sans chargement": 20, "Dernière vente": 15, "Commentaire": 40,
    }

    def _write_header(ws, display_cols):
        for ci, col in enumerate(display_cols, 1):
            cell = ws.cell(1, ci, col)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = ALIGN_C
        ws.row_dimensions[1].height = 18

    def _write_sep_row(ws, row_idx, reappro_code, display_cols):
        prenom = prenom_map.get(reappro_code, "")
        label = f"  {reappro_code}" + (f"  —  {prenom}" if prenom else "")
        n_cols = len(display_cols)
        for ci in range(1, n_cols + 1):
            cell = ws.cell(row_idx, ci, label if ci == 1 else "")
            cell.fill = SEP_FILL
            cell.font = SEP_FONT
            cell.alignment = ALIGN_L
        ws.row_dimensions[row_idx].height = 16

    def _write_data_row(ws, row_idx, row, display_cols):
        for ci, col in enumerate(display_cols, 1):
            val = row[col]
            if val is None or (isinstance(val, float) and str(val) == "nan"):
                val = ""
            c = ws.cell(row_idx, ci, val)
            c.alignment = ALIGN_L

    def _set_col_widths(ws, display_cols):
        for ci, col in enumerate(display_cols, 1):
            ws.column_dimensions[get_column_letter(ci)].width = WIDTH_HINTS.get(col, 16)

    def _write_sheet_all(ws, frame: pd.DataFrame):
        display_cols = list(frame.columns)
        _write_header(ws, display_cols)
        for ri, (_, row) in enumerate(frame.reset_index(drop=True).iterrows(), 2):
            _write_data_row(ws, ri, row, display_cols)
        _set_col_widths(ws, display_cols)

    def _write_sheet_zone(ws, frame: pd.DataFrame):
        display_cols = [c for c in frame.columns if c != "Zone"]
        _write_header(ws, display_cols)
        current_row = 2
        for reappro_code, grp in frame.groupby("Réappro", sort=False):
            _write_sep_row(ws, current_row, str(reappro_code), display_cols)
            current_row += 1
            for _, row in grp.iterrows():
                _write_data_row(ws, current_row, row, display_cols)
                current_row += 1
        _set_col_widths(ws, display_cols)

    # ── Build workbook ───────────────────────────────────────────────────────
    wb = Workbook()
    wb.remove(wb.active)

    ws_all = wb.create_sheet("Toutes zones")
    _write_sheet_all(ws_all, export_df)
    ws_all.freeze_panes = "A2"

    zones = sorted(
        export_df["Zone"].unique(),
        key=lambda z: (z == "Non assigné", z.lower()),
    )
    for zone in zones:
        zone_df = export_df[export_df["Zone"] == zone]
        sheet_name = str(zone)[:31]
        ws = wb.create_sheet(sheet_name)
        _write_sheet_zone(ws, zone_df)
        ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _render_bilan_semaine():
    """Contenu de l'onglet Bilan Semaine — accessible avec ou sans analyse lancée."""
    from mongo_storage import (list_weeks_justifications_nf, load_justifications_nf_week,
                               update_justification_nf_date)
    from collections import defaultdict as _dd

    _JOURS = {0: "Lundi", 1: "Mardi", 2: "Mercredi",
              3: "Jeudi", 4: "Vendredi", 5: "Samedi", 6: "Dimanche"}

    weeks_dispo = list_weeks_justifications_nf()

    if not weeks_dispo:
        st.info(
            "Aucun bilan enregistré pour le moment.\n\n"
            "👉 Allez dans l'onglet **❌ Non Faites**, saisissez des justifications "
            "et cliquez **💾 Enregistrer** pour alimenter ce bilan."
        )
        return

    def _week_label(iso_year, iso_week):
        mon = datetime.date.fromisocalendar(iso_year, iso_week, 1)
        fri = datetime.date.fromisocalendar(iso_year, iso_week, 5)
        return f"S{iso_week} · {iso_year}   ({mon.strftime('%d/%m')} → {fri.strftime('%d/%m/%Y')})"

    col_week, col_jour = st.columns([3, 2])
    with col_week:
        selected_week = st.selectbox(
            "Semaine",
            options=weeks_dispo,
            format_func=lambda x: _week_label(*x),
            key="bilan_sem_week_select",
        )
    with col_jour:
        filtre_jour = st.selectbox(
            "Jour",
            options=["Tous les jours", "Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi"],
            key="bilan_sem_jour_select",
        )

    bilan_docs = load_justifications_nf_week(*selected_week)

    if not bilan_docs:
        st.info("Aucune donnée enregistrée pour cette semaine.")
        return

    total_jours    = len({d["date_analyse"] for d in bilan_docs})
    total_reappros = len({d["reappro"]      for d in bilan_docs})
    total_salles   = sum(len(d["salles"])   for d in bilan_docs)
    total_justif   = sum(
        1 for d in bilan_docs
        for s in d["salles"] if s.get("justification", "").strip()
    )
    k1, k2, k3, k4 = st.columns(4)
    k1.metric("📅 Jours enregistrés",  total_jours)
    k2.metric("👤 Réappros concernés", total_reappros)
    k3.metric("❌ Salles non faites",  total_salles)
    k4.metric("📝 Avec justification", total_justif)

    st.divider()

    # ── Groupement par date ───────────────────────────────────────────────────
    by_date = _dd(list)
    for doc in bilan_docs:
        by_date[doc["date_analyse"]].append(doc)

    # Appliquer le filtre jour
    dates_a_afficher = sorted(by_date.keys())
    if filtre_jour != "Tous les jours":
        dates_a_afficher = [
            d for d in dates_a_afficher
            if _JOURS[datetime.date.fromisoformat(d).weekday()] == filtre_jour
        ]

    if not dates_a_afficher:
        st.info(f"Aucune salle non faite enregistrée pour **{filtre_jour}** cette semaine.")
    else:
        _CSS = """
        <style>
        .bilan-table { width:100%; border-collapse:collapse; font-size:14px; margin-bottom:6px; }
        .bilan-table th {
            background:#f0f0f0; padding:8px 12px; text-align:left;
            border:1px solid #ccc; font-weight:700; color:#333;
        }
        .bilan-table td { padding:7px 12px; border:1px solid #e0e0e0; vertical-align:middle; }
        .bilan-table tr.group-start td { border-top:3px solid #444 !important; }
        .bilan-table .td-reappro { font-weight:700; background:#fafafa; }
        .bilan-table .td-justif  { color:#C0392B; font-weight:600; }
        .bilan-table .td-justif-center { color:#C0392B; font-weight:600;
                                         text-align:center; vertical-align:middle; }
        </style>
        """

        for date_str in dates_a_afficher:
            date_obj    = datetime.date.fromisoformat(date_str)
            jour_fr     = _JOURS[date_obj.weekday()]
            date_disp   = date_obj.strftime("%d/%m/%Y")
            day_docs    = by_date[date_str]
            nb_sal_jour = sum(len(d["salles"]) for d in day_docs)

            st.markdown(
                f"### 📆 {jour_fr} {date_disp} "
                f"<span style='color:#C0392B;font-size:0.85em;'>"
                f"— {nb_sal_jour} salle{'s' if nb_sal_jour > 1 else ''} non faite{'s' if nb_sal_jour > 1 else ''}"
                f"</span>",
                unsafe_allow_html=True,
            )

            html_rows = []
            for doc in day_docs:
                reappro  = doc["reappro"]
                salles   = doc["salles"]
                n        = len(salles)
                justifs  = [s.get("justification", "").strip() for s in salles]
                same_j   = len(set(justifs)) == 1

                for i, s in enumerate(salles):
                    group_cls = ' class="group-start"' if i == 0 else ""
                    row = f"<tr{group_cls}>"
                    if i == 0:
                        row += f'<td rowspan="{n}" class="td-reappro">{reappro}</td>'
                    row += f'<td>{s["client"]}</td>'
                    row += f'<td>{s["machine"]}</td>'
                    if same_j and i == 0:
                        val = justifs[0] if justifs[0] else "—"
                        row += f'<td rowspan="{n}" class="td-justif-center">{val}</td>'
                    elif not same_j:
                        val = justifs[i] if justifs[i] else "—"
                        row += f'<td class="td-justif">{val}</td>'
                    row += "</tr>"
                    html_rows.append(row)

            table_html = (
                _CSS
                + '<table class="bilan-table">'
                + "<thead><tr>"
                + "<th>Réappro</th><th>Client / Salle</th><th>Machine</th><th>Justification</th>"
                + "</tr></thead>"
                + "<tbody>" + "".join(html_rows) + "</tbody>"
                + "</table>"
            )
            st.markdown(table_html, unsafe_allow_html=True)
            st.divider()

    # ── Export Excel ──────────────────────────────────────────────────────────
    def _build_bilan_excel(docs_par_date: dict, dates: list, week_label: str) -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Bilan semaine"

        # Styles
        hdr_font   = Font(bold=True, color="FFFFFF", size=11)
        hdr_fill   = PatternFill("solid", fgColor="2E4057")
        day_font   = Font(bold=True, color="FFFFFF", size=11)
        day_fill   = PatternFill("solid", fgColor="C0392B")
        reap_font  = Font(bold=True, size=10)
        reap_fill  = PatternFill("solid", fgColor="F2F2F2")
        justif_font = Font(color="C0392B", bold=True, size=10)
        center_al  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_al    = Alignment(vertical="center", wrap_text=True)

        thin  = Side(style="thin",   color="CCCCCC")
        thick = Side(style="medium", color="444444")
        thin_border  = Border(left=thin,  right=thin,  top=thin,  bottom=thin)
        thick_top    = Border(left=thin,  right=thin,  top=thick, bottom=thin)

        # Titre
        ws.merge_cells("A1:D1")
        ws["A1"] = f"Bilan semaine — {week_label}"
        ws["A1"].font = Font(bold=True, size=13, color="2E4057")
        ws["A1"].alignment = center_al
        ws.row_dimensions[1].height = 22

        current_row = 2

        for date_str in dates:
            date_obj  = datetime.date.fromisoformat(date_str)
            jour_fr   = _JOURS[date_obj.weekday()]
            date_disp = date_obj.strftime("%d/%m/%Y")
            day_docs  = docs_par_date[date_str]

            # Ligne jour
            ws.merge_cells(f"A{current_row}:D{current_row}")
            cell = ws[f"A{current_row}"]
            cell.value = f"{jour_fr}  {date_disp}"
            cell.font      = day_font
            cell.fill      = day_fill
            cell.alignment = center_al
            ws.row_dimensions[current_row].height = 18
            current_row += 1

            # En-têtes colonnes
            for col_i, hdr in enumerate(["Réappro", "Client / Salle", "Machine", "Justification"], 1):
                c = ws.cell(row=current_row, column=col_i, value=hdr)
                c.font = hdr_font
                c.fill = hdr_fill
                c.alignment = center_al
                c.border = thin_border
            ws.row_dimensions[current_row].height = 16
            current_row += 1

            for doc in day_docs:
                reappro = doc["reappro"]
                salles  = doc["salles"]
                n       = len(salles)
                justifs = [s.get("justification", "").strip() for s in salles]
                same_j  = len(set(justifs)) == 1
                start_r = current_row

                for i, s in enumerate(salles):
                    is_first = (i == 0)
                    brd = thick_top if is_first else thin_border

                    # Réappro (fusionné verticalement si > 1 salle)
                    if is_first:
                        c = ws.cell(row=current_row, column=1, value=reappro)
                        c.font = reap_font; c.fill = reap_fill
                        c.alignment = center_al; c.border = brd
                        if n > 1:
                            ws.merge_cells(
                                start_row=start_r, start_column=1,
                                end_row=start_r + n - 1, end_column=1
                            )

                    # Client / Salle
                    c = ws.cell(row=current_row, column=2, value=s["client"])
                    c.alignment = left_al; c.border = brd if is_first else thin_border

                    # Machine
                    c = ws.cell(row=current_row, column=3, value=s["machine"])
                    c.alignment = center_al; c.border = brd if is_first else thin_border

                    # Justification
                    if same_j and is_first:
                        val = justifs[0] if justifs[0] else ""
                        c = ws.cell(row=current_row, column=4, value=val)
                        c.font = justif_font; c.alignment = center_al
                        c.border = brd
                        if n > 1:
                            ws.merge_cells(
                                start_row=start_r, start_column=4,
                                end_row=start_r + n - 1, end_column=4
                            )
                    elif not same_j:
                        val = justifs[i] if justifs[i] else ""
                        c = ws.cell(row=current_row, column=4, value=val)
                        c.font = justif_font; c.alignment = left_al
                        c.border = brd if is_first else thin_border

                    ws.row_dimensions[current_row].height = 15
                    current_row += 1

            current_row += 1  # ligne vide entre les jours

        # Largeurs colonnes
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 28

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    iso_y, iso_w = selected_week
    xl_label = _week_label(iso_y, iso_w)
    xl_dates = dates_a_afficher  # respecte le filtre jour actif
    xl_bytes = _build_bilan_excel(by_date, xl_dates, xl_label)
    fname = f"bilan_S{iso_w}_{iso_y}"
    if filtre_jour != "Tous les jours":
        fname += f"_{filtre_jour}"
    fname += ".xlsx"

    st.download_button(
        label="📥 Exporter en Excel",
        data=xl_bytes,
        file_name=fname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="bilan_export_xl",
    )

    # ── Correction de date ────────────────────────────────────────────────────
    with st.expander("🔧 Corriger la date d'une entrée"):
        st.caption("Utile si une tournée a été enregistrée avec la mauvaise date (ex : Mardi au lieu de Lundi).")

        # Charger tous les docs pour construire la liste des entrées corrigeables
        all_docs = []
        for yw in weeks_dispo:
            all_docs.extend(load_justifications_nf_week(*yw))

        if not all_docs:
            st.info("Aucune entrée en base.")
        else:
            entrees = {
                f"{d['reappro']} — {d['jour']} {d['date_analyse']}": (d["reappro"], d["date_analyse"])
                for d in all_docs
            }
            choix = st.selectbox("Entrée à corriger", list(entrees.keys()),
                                 key="corr_entree_select")
            reappro_corr, old_date_corr = entrees[choix]

            new_date_corr = st.date_input(
                "Nouvelle date",
                value=datetime.date.fromisoformat(old_date_corr),
                key="corr_new_date",
            )

            if st.button("✅ Appliquer la correction", key="btn_corr_date"):
                ok = update_justification_nf_date(
                    reappro_corr, old_date_corr, new_date_corr.isoformat()
                )
                if ok:
                    st.success(f"Date corrigée : {old_date_corr} → {new_date_corr.isoformat()}")
                    st.rerun()
                else:
                    st.error("Entrée introuvable ou erreur MongoDB.")


def _render_machines_sans_chargement(plannings: dict):
    """Section at the bottom of Tournées page: salles without chargement for > N days."""
    st.divider()
    st.markdown("### 🔴 Machines sans chargement")
    st.caption(
        "Salles n'ayant reçu aucun chargement valide depuis plus d'une semaine. "
        "Déposez un export chargement CSV couvrant une longue période."
    )

    col_ul1, col_ul2, col_seuil = st.columns([3, 3, 1])
    with col_ul1:
        st.caption("📦 Export chargement (longue période)")
        up = st.file_uploader(
            "Export chargement longue période (CSV)",
            type=["csv"],
            key="sc_uploader",
            label_visibility="collapsed",
        )
    with col_ul2:
        st.caption("📉 Export télémétrie (même fichier que No Audit)")
        up_telem = st.file_uploader(
            "Export télémétrie (CSV)",
            type=["csv"],
            key="sc_telem_uploader",
            label_visibility="collapsed",
        )
    with col_seuil:
        seuil = st.number_input(
            "Seuil jours", min_value=1, value=7, step=1, key="sc_seuil",
            help="Nombre minimum de jours sans chargement pour apparaître dans la liste",
            label_visibility="visible",
        )

    if up is not None:
        st.session_state["sc_bytes"] = up.getvalue()
    if up_telem is not None:
        st.session_state["sc_telem_bytes"] = up_telem.getvalue()

    col_btn, col_sort, _ = st.columns([1.2, 3, 2.8])
    with col_btn:
        calc = st.button(
            "🔍 Analyser",
            key="sc_calc",
            disabled=(st.session_state.get("sc_bytes") is None),
            use_container_width=True,
        )
    with col_sort:
        sort_desc = st.radio(
            "Tri urgence",
            ["🔴 Plus urgent d'abord", "🟢 Moins urgent d'abord"],
            key="sc_sort",
            horizontal=True,
            label_visibility="collapsed",
        )

    if calc and st.session_state.get("sc_bytes"):
        with st.spinner("Analyse en cours…"):
            try:
                from page_no_audit import _parse_telemetry
                df_raw    = _parse_long_charg(st.session_state["sc_bytes"])
                salles_df = _load_salles_for_sc()
                df_res    = _compute_sans_charg(df_raw, salles_df, plannings, int(seuil))

                # Dernière vente depuis la télémétrie (optionnel)
                if st.session_state.get("sc_telem_bytes"):
                    telem_df   = _parse_telemetry(st.session_state["sc_telem_bytes"])
                    last_vente = (
                        telem_df[telem_df["Price"] > 0]
                        .groupby("Salle")["Date"].max()
                    )
                    df_res["Dernière vente"] = df_res["Salle"].apply(
                        lambda s: (
                            last_vente[s].strftime("%d/%m/%Y")
                            if s in last_vente.index and not pd.isna(last_vente[s])
                            else "Jamais"
                        )
                    )

                st.session_state["sc_result"]     = df_res
                st.session_state["sc_seuil_used"] = int(seuil)
                # Reset working df so comments are re-loaded from DB on next render
                for _k in ("sc_working_df", "sc_version", "sc_last_sort"):
                    st.session_state.pop(_k, None)
            except Exception as e:
                st.error(f"❌ Erreur : {e}")

    df_sc = st.session_state.get("sc_result")
    if df_sc is None:
        if st.session_state.get("sc_bytes"):
            st.info("Cliquez sur **Analyser** pour calculer.")
        return

    seuil_used = st.session_state.get("sc_seuil_used", int(seuil))

    if df_sc.empty:
        st.success(
            f"✅ Aucune salle sans chargement depuis plus de {seuil_used} jours."
        )
        return

    # ── Load incidents from MongoDB ─────────────────────────────────────────
    def _load_sc_incidents() -> list:
        from mongo_storage import _get_client as _get_mc
        try:
            client  = _get_mc()
            db_name = st.secrets["mongo"]["db_name"]
            return list(
                client[db_name]["incidents"].find(
                    {"status": "actif"},
                    {"_id": 0, "salle": 1, "type": 1, "commentaire": 1, "since_date": 1},
                )
            )
        except Exception:
            return []

    today     = datetime.date.today()
    incidents = _load_sc_incidents()

    sv_map: dict[str, str] = {}   # salle → since_date str  (sans_ventes)
    na_map: dict[str, str] = {}   # salle → commentaire     (no_audit)
    sc_saved: dict[str, str] = {} # salle → commentaire     (sans_chargement already saved)
    for inc in incidents:
        s = inc.get("salle", "")
        t = inc.get("type", "")
        if t == "sans_ventes":
            sv_map[s] = inc.get("since_date", "")
        elif t == "no_audit":
            na_map[s] = inc.get("commentaire", "")
        elif t == "sans_chargement":
            sc_saved[s] = inc.get("commentaire", "")

    def _prefill(salle: str) -> str:
        # Priority 1: already saved comment for this section
        if sc_saved.get(salle):
            return sc_saved[salle]
        # Priority 2: auto-label from sans_ventes incident
        if salle in sv_map:
            sd = sv_map[salle]
            if sd:
                try:
                    sd_date = datetime.datetime.strptime(sd.split(" ")[0], "%d/%m/%Y").date()
                    return f"Sans ventes depuis {(today - sd_date).days} j"
                except Exception:
                    pass
            return "Sans ventes"
        # Priority 3: existing no_audit comment
        if na_map.get(salle):
            return na_map[salle]
        return ""

    def _tag(salle: str) -> str:
        if salle in sv_map:
            return "🔴 Sans ventes"
        if salle in na_map:
            return "🟡 No audit"
        return ""

    # ── Session-state working df (preserves edits across reruns) ────────────
    ascending      = "Moins" in sort_desc
    sc_working_key = "sc_working_df"
    sc_version_key = "sc_version"
    sc_sort_key    = "sc_last_sort"

    if sc_working_key not in st.session_state:
        base = df_sc.copy()
        base["Tag"]         = base["Salle"].apply(_tag)
        base["Commentaire"] = base["Salle"].apply(_prefill)
        base = base.sort_values(
            "Jours sans chargement", ascending=ascending, na_position="last"
        ).reset_index(drop=True)
        st.session_state[sc_working_key] = base
        st.session_state[sc_version_key] = 0
        st.session_state[sc_sort_key]    = ascending
    elif st.session_state.get(sc_sort_key) != ascending:
        # Re-sort in-place, preserving any edited comments
        wdf = st.session_state[sc_working_key].sort_values(
            "Jours sans chargement", ascending=ascending, na_position="last"
        ).reset_index(drop=True)
        st.session_state[sc_working_key] = wdf
        st.session_state[sc_version_key] = st.session_state.get(sc_version_key, 0) + 1
        st.session_state[sc_sort_key]    = ascending

    working_df = st.session_state[sc_working_key]

    st.caption(
        f"**{len(working_df)}** salle(s) sans chargement depuis > **{seuil_used}** jours"
    )

    # ── Data editor ──────────────────────────────────────────────────────────
    col_order = [
        "Tag", "Salle", "Code machine", "Réappro",
        "Dernier chargement", "Jours sans chargement", "Dernière vente", "Commentaire",
    ]
    col_order = [c for c in col_order if c in working_df.columns]

    col_cfg: dict = {
        c: st.column_config.Column(disabled=True)
        for c in col_order if c != "Commentaire"
    }
    col_cfg["Tag"] = st.column_config.TextColumn("Tag", disabled=True, width="small")
    col_cfg["Jours sans chargement"] = st.column_config.NumberColumn(
        "Jours sans chargement", format="%d j", disabled=True
    )
    col_cfg["Commentaire"] = st.column_config.TextColumn(
        "Commentaire", help="Saisissez une note (sauvegardée en base)", max_chars=300
    )

    edited_df = st.data_editor(
        working_df[col_order],
        use_container_width=True,
        hide_index=True,
        height=min(700, 38 + len(working_df) * 35),
        column_config=col_cfg,
        key=f"sc_editor_{st.session_state[sc_version_key]}",
    )

    # ── Actions: save + export ───────────────────────────────────────────────
    col_save, col_dl, _ = st.columns([1.5, 2, 3.5])

    with col_save:
        if st.button("💾 Sauvegarder", key="sc_save", use_container_width=True):
            from mongo_storage import _get_client as _get_mc
            try:
                client  = _get_mc()
                db_name = st.secrets["mongo"]["db_name"]
                col_db  = client[db_name]["incidents"]
                now     = datetime.datetime.utcnow()
                saved_n = 0
                for _, row in edited_df.iterrows():
                    commentaire = (row.get("Commentaire") or "").strip()
                    salle       = row["Salle"]
                    since_date  = str(row.get("Dernier chargement", ""))
                    if not commentaire:
                        continue
                    existing = col_db.find_one(
                        {"salle": salle, "type": "sans_chargement", "status": "actif"}
                    )
                    if existing:
                        col_db.update_one(
                            {"_id": existing["_id"]},
                            {"$set": {"commentaire": commentaire, "since_date": since_date}},
                        )
                    else:
                        col_db.insert_one({
                            "salle":       salle,
                            "type":        "sans_chargement",
                            "commentaire": commentaire,
                            "since_date":  since_date,
                            "created_at":  now,
                            "resolved_at": None,
                            "status":      "actif",
                        })
                    saved_n += 1
                st.session_state[sc_working_key] = edited_df.copy()
                st.toast(f"✅ {saved_n} commentaire(s) sauvegardé(s).")
            except Exception as e:
                st.error(f"❌ Erreur MongoDB : {e}")

    with col_dl:
        st.download_button(
            "📥 Exporter Excel (par zone)",
            data=_build_sc_excel_by_zone(edited_df),
            file_name=f"machines_sans_chargement_{datetime.date.today():%Y%m%d}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="sc_dl",
            use_container_width=True,
        )


# Logo — ligne dédiée, aligné à gauche
_logo_path = "assets/logo.png"
_logo_col, _theme_col, _ = st.columns([1, 1, 8])
with _logo_col:
    if _os.path.exists(_logo_path):
        st.image(_logo_path, width=90)
    else:
        st.markdown('<div style="font-size:1.3rem;font-weight:900;color:#1B3D6F;">DP</div>', unsafe_allow_html=True)
with _theme_col:
    _icon = "🌙" if not st.session_state.dark_mode else "☀️"
    st.markdown('<div class="theme-toggle-btn">', unsafe_allow_html=True)
    if st.button(_icon, key="toggle_dark_mode", help="Basculer mode sombre/clair"):
        st.session_state.dark_mode = not st.session_state.dark_mode
        st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)

# Barre de navigation — 11 boutons
nav_c1, nav_c2, nav_c3, nav_c4, nav_c5, nav_c6, nav_c7, nav_c8, nav_c9, nav_c10, nav_c11 = st.columns(11)
with nav_c1:
    if st.button(
        "🖥️ Machines",
        key="nav_machines",
        use_container_width=True,
        type="primary" if st.session_state.page == "machines" else "secondary",
    ):
        st.session_state.page = "machines"
        st.rerun()
with nav_c2:
    if st.button(
        "📦 Tournées",
        key="nav_suivi",
        use_container_width=True,
        type="primary" if st.session_state.page == "suivi" else "secondary",
    ):
        st.session_state.page = "suivi"
        st.rerun()
with nav_c3:
    if st.button(
        "📉 No Audit",
        key="nav_no_audit",
        use_container_width=True,
        type="primary" if st.session_state.page == "no_audit" else "secondary",
    ):
        st.session_state.page = "no_audit"
        st.rerun()
with nav_c4:
    if st.button(
        "🗂️ Planogrammes",
        key="nav_plano",
        use_container_width=True,
        type="primary" if st.session_state.page == "planogrammes" else "secondary",
    ):
        st.session_state.page = "planogrammes"
        st.rerun()
with nav_c5:
    if st.button(
        "📊 Inventaires",
        key="nav_inv",
        use_container_width=True,
        type="primary" if st.session_state.page == "inventaires" else "secondary",
    ):
        st.session_state.page = "inventaires"
        st.rerun()
with nav_c6:
    if st.button(
        "🛒 Commandes",
        key="nav_cmd",
        use_container_width=True,
        type="primary" if st.session_state.page == "commandes" else "secondary",
    ):
        st.session_state.page = "commandes"
        st.rerun()
with nav_c7:
    if st.button(
        "❓ Indéfinis",
        key="nav_indef",
        use_container_width=True,
        type="primary" if st.session_state.page == "indefinis" else "secondary",
    ):
        st.session_state.page = "indefinis"
        st.rerun()
with nav_c8:
    if st.button(
        "📝 CR",
        key="nav_cr",
        use_container_width=True,
        type="primary" if st.session_state.page == "cr" else "secondary",
    ):
        st.session_state.page = "cr"
        st.rerun()
with nav_c9:
    if st.button(
        "🗺️ Quartix",
        key="nav_quartix",
        use_container_width=True,
        type="primary" if st.session_state.page == "quartix" else "secondary",
    ):
        st.session_state.page = "quartix"
        st.rerun()
with nav_c10:
    if st.button(
        "📋 Picklist",
        key="nav_picklist",
        use_container_width=True,
        type="primary" if st.session_state.page == "picklist" else "secondary",
    ):
        st.session_state.page = "picklist"
        st.rerun()
with nav_c11:
    if st.button(
        "👤 Rapport",
        key="nav_rapport_employe",
        use_container_width=True,
        type="primary" if st.session_state.page == "rapport_employe" else "secondary",
    ):
        st.session_state.page = "rapport_employe"
        st.rerun()

st.divider()



# ════════════════════════════════════════════════════════
# PAGE : MACHINES
# ════════════════════════════════════════════════════════
if st.session_state.page == "machines":

    st.markdown('<div class="main-title">🖥️ Parc Machines</div>', unsafe_allow_html=True)
    st.markdown(
        '<div class="subtitle">Visualisez et gérez le parc de distributeurs automatiques en exploitation.</div>',
        unsafe_allow_html=True,
    )

    page_machines.render()


# ════════════════════════════════════════════════════════
# PAGE : NO AUDIT / SANS VENTES
# ════════════════════════════════════════════════════════
elif st.session_state.page == "no_audit":

    st.markdown('<div class="main-title">📉 No Audit / Sans Ventes</div>', unsafe_allow_html=True)
    st.markdown(
        '<div class="subtitle">Identifiez les salles absentes de la télémétrie ou sans ventes sur une période donnée.</div>',
        unsafe_allow_html=True,
    )

    page_no_audit.render()


# ════════════════════════════════════════════════════════
# PAGE : SUIVI RÉAPPROVISIONNEURS
# ════════════════════════════════════════════════════════
elif st.session_state.page == "suivi":

    @st.cache_data(show_spinner=False, ttl=300)
    def _get_plannings():
        return load_plannings_from_mongo()

    plannings, planning_errors = _get_plannings()

    st.markdown('<div class="main-title">📦 Suivi des Réapprovisionneurs</div>', unsafe_allow_html=True)
    st.markdown(
        f'<div class="subtitle">Analyse du <b>{st.session_state.jour_analyse}</b> — '
        f'{datetime.date.today().strftime("%d/%m/%Y")}</div>',
        unsafe_allow_html=True,
    )

    # ── Plannings chargés ─────────────────────────────────
    st.divider()
    st.markdown("### ⚙️ Plannings chargés")
    if plannings:
        st.success(f"✅ {len(plannings)} plannings chargés — `MongoDB Atlas`")
        with st.expander("Voir les réappros"):
            for emp, p in sorted(plannings.items()):
                total = sum(len(v) for v in p.values())
                st.caption(f"**{emp}** — {total} salles/semaine")
    else:
        st.error("❌ Aucun planning trouvé.")
        for k, v in planning_errors.items():
            st.error(f"{k}: {v}")
    if planning_errors and plannings:
        with st.expander(f"⚠️ {len(planning_errors)} erreur(s)"):
            for k, v in planning_errors.items():
                st.warning(f"{k}: {v}")

    with st.expander("📥 Mettre à jour les plannings"):
        st.caption("Déposer un ou plusieurs fichiers CSV nommés `{réappro}.csv`")
        plan_files = st.file_uploader(
            "Fichiers planning CSV",
            type=["csv"],
            accept_multiple_files=True,
            key="planning_uploader",
            label_visibility="collapsed",
        )
        if plan_files:
            col_imp, _ = st.columns([1, 4])
            with col_imp:
                if st.button("📥 Importer", use_container_width=True, key="btn_import_planning"):
                    from mongo_storage import upsert_planning
                    from planning_parser import parse_planning_file
                    ok, errs = [], []
                    for f in plan_files:
                        employe = f.name.removesuffix(".csv").strip()
                        try:
                            planning, semaine = parse_planning_file(f.read(), employe)
                            upsert_planning(employe, planning, semaine)
                            total = sum(len(v) for v in planning.values())
                            ok.append(f"{employe} ({total} salles, {semaine})")
                        except Exception as e:
                            errs.append(f"{employe} : {e}")
                    for msg in ok:
                        st.success(f"✅ {msg}")
                    for msg in errs:
                        st.error(f"❌ {msg}")
                    if ok:
                        _get_plannings.clear()
                        st.rerun()

    with st.expander("🗑️ Supprimer un planning"):
        if plannings:
            to_delete = st.multiselect(
                "Réappros à supprimer",
                options=sorted(plannings.keys()),
                key="planning_to_delete",
                label_visibility="collapsed",
            )
            if to_delete:
                col_del, _ = st.columns([1, 4])
                with col_del:
                    if st.button(
                        f"🗑️ Supprimer ({len(to_delete)})",
                        use_container_width=True,
                        key="btn_delete_planning",
                        type="primary",
                    ):
                        from mongo_storage import delete_planning
                        errs = []
                        for emp in to_delete:
                            try:
                                delete_planning(emp)
                            except Exception as e:
                                errs.append(f"{emp} : {e}")
                        deleted = [e for e in to_delete if e not in [err.split(" :")[0] for err in errs]]
                        for emp in deleted:
                            st.success(f"✅ {emp} supprimé")
                        for msg in errs:
                            st.error(f"❌ {msg}")
                        _get_plannings.clear()
                        st.rerun()
        else:
            st.caption("_Aucun planning à supprimer._")

    # ── Jour d'analyse ────────────────────────────────────
    st.divider()
    jour_col, _ = st.columns([1, 4])
    with jour_col:
        st.markdown("**Jour d'analyse**")
        jour_selectionne = st.selectbox(
            "Jour", JOURS,
            index=JOURS.index(st.session_state.jour_analyse)
            if st.session_state.jour_analyse in JOURS else 0,
            label_visibility="collapsed",
            key="jour_analyse_select",
        )
        if jour_selectionne != st.session_state.jour_analyse:
            st.session_state.jour_analyse = jour_selectionne
            st.rerun()

    # ── Export Excel ───────────────────────────────────────
    st.divider()
    st.markdown("### 📥 Export Excel")
    if st.session_state.results:
        col_exp, _ = st.columns([1, 4])
        with col_exp:
            if st.button("📊 Générer le fichier Excel", use_container_width=True):
                with st.spinner("Génération..."):
                    try:
                        st.session_state["excel_bytes"] = generer_excel(
                            st.session_state.results, st.session_state.jour_analyse
                        )
                        st.session_state["excel_jour"] = st.session_state.jour_analyse
                    except Exception as e:
                        st.error(f"Erreur Excel : {e}")

            # Bouton download TOUJOURS rendu en dehors du if st.button(...)
            if st.session_state.get("excel_bytes"):
                date_str = datetime.date.today().strftime("%Y%m%d")
                jour_export = st.session_state.get("excel_jour", st.session_state.jour_analyse)
                st.download_button(
                    label="⬇️ Télécharger Excel",
                    data=st.session_state["excel_bytes"],
                    file_name=f"suivi_reappro_{date_str}_{jour_export}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
    else:
        st.caption("_Analysez un fichier de chargement pour générer l'export._")

    st.divider()
    st.markdown("### 📂 Déposer le fichier de chargement machine du jour")
    uploaded = st.file_uploader(
        "Fichier export chargement (CSV)", type=["csv"],
        key="chargement_uploader", label_visibility="collapsed",
    )

    # Persister les bytes en session pour survivre aux st.rerun()
    if uploaded is not None:
        st.session_state["chargement_bytes"] = uploaded.getvalue()

    has_file = st.session_state.get("chargement_bytes") is not None

    col_btn, _ = st.columns([1, 5])
    with col_btn:
        lancer = st.button(
            "🚀 Lancer l'analyse", type="primary",
            use_container_width=True, disabled=(not has_file),
        )

    if lancer and has_file:
        with st.spinner("Analyse en cours..."):
            try:
                chargement = parse_chargement_csv(st.session_state["chargement_bytes"])
                results = croiser_planning_chargement(
                    plannings, chargement, st.session_state.jour_analyse
                )
                st.session_state.results = results
                st.session_state["excel_bytes"] = None  # Invalider l'ancien export
                st.toast(f"✅ Analyse terminée — {len(results)} réappros traités")
            except Exception as e:
                st.error(f"Erreur : {e}")
                st.stop()
        st.rerun()  # Force re-render : Export Excel voit maintenant st.session_state.results

    results = st.session_state.results
    jour    = st.session_state.jour_analyse

    if not results:
        # Sans analyse : seulement l'onglet Bilan Semaine
        (tab_bilan_sem,) = st.tabs(["📊 Bilan Semaine"])
        with tab_bilan_sem:
            _render_bilan_semaine()
        _render_machines_sans_chargement(plannings)
        st.stop()

    # Analyse disponible — KPIs
    st.divider()
    total_prev   = sum(len(d["salles_prevues"])    for d in results.values())
    total_fait   = sum(len(d["salles_faites"])     for d in results.values())
    total_nf     = sum(len(d["salles_non_faites"]) for d in results.values())
    total_joker  = sum(sum(1 for s in d["salles_faites"] if s["is_joker"]) for d in results.values())
    total_decale = sum(1 for d in results.values() if d.get("tournee_decalee"))
    taux_global  = round((total_fait / total_prev * 100) if total_prev > 0 else 0, 1)

    k1, k2, k3, k4, k5, k6 = st.columns(6)
    k1.metric("📋 Prévues",     total_prev)
    k2.metric("✅ Faites",      total_fait)
    k3.metric("❌ Non faites",  total_nf,
              delta=f"-{total_nf}" if total_nf else None, delta_color="inverse")
    k4.metric("🔄 Jokers",      total_joker)
    k5.metric("📅 Décalées",    total_decale)
    k6.metric("📈 Taux global", f"{taux_global}%")

    tab_recap, tab_nf, tab_jokers, tab_decale, tab_detail, tab_bilan_sem = st.tabs([
        "📋 Récapitulatif", "❌ Non Faites", "🔄 Jokers", "📅 Tournées Décalées",
        "🔍 Détail par réappro", "📊 Bilan Semaine",
    ])

    with tab_recap:
        rows = []
        for reappro, data in sorted(results.items()):
            nb_p  = len(data["salles_prevues"])
            nb_f  = len(data["salles_faites"])
            nb_nf = len(data["salles_non_faites"])
            nb_j  = sum(1 for s in data["salles_faites"] if s["is_joker"])
            taux  = round((nb_f / nb_p * 100) if nb_p > 0 else 0, 1)
            decale = "📅 " + data.get("tournee_decalee", {}).get("jour_detecte", "") if data.get("tournee_decalee") else ""
            rows.append({"Réappro": reappro, "Prévues": nb_p, "Faites": nb_f,
                         "Non Faites": nb_nf, "Jokers": nb_j, "Décalée": decale, "Taux (%)": taux})
        df_recap = pd.DataFrame(rows)
        def color_row(row):
            if row["Non Faites"] == 0:
                return ["background-color:#1E7E34; color:#FFFFFF; font-weight:600"] * len(row)
            elif row["Non Faites"] == row["Prévues"]:
                return ["background-color:#C0392B; color:#FFFFFF; font-weight:600"] * len(row)
            else:
                return ["background-color:#E67E22; color:#FFFFFF; font-weight:600"] * len(row)
        st.dataframe(
            df_recap.style.apply(color_row, axis=1),
            use_container_width=True, hide_index=True,
            height=min(700, 38 + len(df_recap) * 35),
        )

    with tab_nf:
        nf_rows = []
        nf_by_reappro = {}
        for reappro, data in sorted(results.items()):
            salles_nf = data["salles_non_faites"]
            if salles_nf:
                nf_by_reappro[reappro] = salles_nf
            for s in salles_nf:
                nf_rows.append({"Réappro": reappro, "Client / Salle": s["client"], "Machine": s["machine"]})

        if nf_rows:
            df_nf = pd.DataFrame(nf_rows)
            st.dataframe(
                df_nf.style.map(lambda _: "background-color:#C0392B; color:#FFFFFF; font-weight:600"),
                use_container_width=True, hide_index=True,
                height=min(700, 38 + len(df_nf) * 35),
            )

            # ── Justifications par réappro / par salle ─────────────────────
            from mongo_storage import save_justification_nf, load_justifications_nf
            # Date réelle de la tournée : retrouver la date calendaire du jour analysé
            _JOURS_IDX = {"Lundi": 0, "Mardi": 1, "Mercredi": 2, "Jeudi": 3, "Vendredi": 4}
            _today     = datetime.date.today()
            _jour_idx  = _JOURS_IDX.get(jour, _today.weekday())
            _delta     = _jour_idx - _today.weekday()
            if _delta > 0:   # le jour est "dans le futur" → c'est la semaine précédente
                _delta -= 7
            date_analyse = (_today + datetime.timedelta(days=_delta)).isoformat()

            # Pré-remplissage depuis MongoDB : { (reappro, machine): texte }
            existing_docs = load_justifications_nf(date_analyse)
            prefill = {}
            for doc in existing_docs:
                for s in doc.get("salles", []):
                    prefill[(doc["reappro"], s["machine"])] = s.get("justification", "")

            # Boutons généraux (remplissent toutes les salles d'un réappro)
            _PRESETS_GENERAL = [
                ("AM",              "AM (arrêt maladie)"),
                ("AT",              "AT (arrêt de travail)"),
                ("AI",              "AI (absence injustifiée)"),
                ("CP",              "CP (congés payés)"),
                ("WA",              "WA (WhatsApp)"),
                ("Véhicule garage", "Véhicule garage"),
            ]
            # Boutons spécifiques par salle
            _PRESETS_SALLE = [
                ("DA HS",           "DA HS"),
                ("Salle fermée",    "Salle fermée"),
                ("Tournée décalée", "Tournée décalée"),
            ]

            # ── Initialiser les valeurs depuis MongoDB (une seule fois par session) ──
            # On utilise des clés session_state SANS les lier à un widget (pas de key=)
            # pour pouvoir les modifier librement depuis les boutons preset.
            for _r, _snf in nf_by_reappro.items():
                for _s in _snf:
                    _vk = f"justif__{_r}__{_s['machine']}"
                    if _vk not in st.session_state:
                        st.session_state[_vk] = prefill.get((_r, _s["machine"]), "")

            st.divider()
            st.subheader("📝 Justifications des salles non faites")

            for reappro, salles_nf in sorted(nf_by_reappro.items()):

                # ── Traitement des presets (AVANT l'expander) ─────────────────────────
                # Les presets écrivent directement dans session_state — pas de conflit
                # puisque les text_inputs n'ont pas de key= liée.
                _had_pending = False

                pending_all = f"_pending_all__{reappro}"
                if pending_all in st.session_state:
                    _had_pending = True
                    val_all = st.session_state.pop(pending_all)
                    for s in salles_nf:
                        st.session_state[f"justif__{reappro}__{s['machine']}"] = val_all

                for s in salles_nf:
                    pending_s = f"_pending_s__{reappro}__{s['machine']}"
                    if pending_s in st.session_state:
                        _had_pending = True
                        val_s = st.session_state.pop(pending_s)
                        st.session_state[f"justif__{reappro}__{s['machine']}"] = val_s

                # Forcer l'expander ouvert si un preset vient d'être appliqué
                _exp_open_key = f"_exp_open__{reappro}"
                if _had_pending:
                    st.session_state[_exp_open_key] = True
                _is_expanded = st.session_state.pop(_exp_open_key, False)

                nb = len(salles_nf)
                with st.expander(
                    f"👤 {reappro} — {nb} salle{'s' if nb > 1 else ''} non faite{'s' if nb > 1 else ''}",
                    expanded=_is_expanded,
                ):
                    # ── En-têtes ────────────────────────────────────────────────────────
                    hdr_s, hdr_m, hdr_j, hdr_b = st.columns([3, 1, 3, 2])
                    hdr_s.markdown("**Client / Salle**")
                    hdr_m.markdown("**Machine**")
                    hdr_j.markdown("**Justification**")
                    hdr_b.markdown("**Cas spécifique**")
                    st.divider()

                    # ── Une ligne par salle ──────────────────────────────────────────────
                    for s in salles_nf:
                        col_s, col_m, col_j, col_b = st.columns([3, 1, 3, 2])
                        col_s.markdown(s["client"])
                        col_m.markdown(f"`{s['machine']}`")
                        val_key = f"justif__{reappro}__{s['machine']}"

                        with col_j:
                            # On utilise key= (obligatoire pour éviter les IDs dupliqués).
                            # Les presets sont traités AVANT l'expander (avant que le widget
                            # soit instancié), donc modifier session_state[val_key] est autorisé.
                            st.text_input(
                                "justification",
                                key=val_key,
                                placeholder="Justification...",
                                label_visibility="collapsed",
                            )

                        with col_b:
                            ps_cols = st.columns(len(_PRESETS_SALLE))
                            for pi, (label, valeur) in enumerate(_PRESETS_SALLE):
                                if ps_cols[pi].button(
                                    label,
                                    key=f"ps__{reappro}__{s['machine']}__{pi}",
                                    use_container_width=True,
                                ):
                                    st.session_state[f"_pending_s__{reappro}__{s['machine']}"] = valeur
                                    st.rerun()

                    st.divider()

                    # ── Barre du bas — Enregistrer + boutons généraux ────────────────────
                    col_save, col_gen = st.columns([1, 5])
                    with col_save:
                        enregistrer = st.button(
                            "💾 Enregistrer",
                            key=f"btn_justif__{reappro}",
                            type="primary",
                            use_container_width=True,
                        )
                    with col_gen:
                        pg_cols = st.columns(len(_PRESETS_GENERAL))
                        for pi, (label, valeur) in enumerate(_PRESETS_GENERAL):
                            if pg_cols[pi].button(
                                label,
                                key=f"pg__{reappro}__{pi}",
                                use_container_width=True,
                            ):
                                st.session_state[f"_pending_all__{reappro}"] = valeur
                                st.rerun()

                    if enregistrer:
                        salles_payload = [
                            {
                                "client": s["client"],
                                "machine": s["machine"],
                                "justification": st.session_state.get(
                                    f"justif__{reappro}__{s['machine']}", ""
                                ),
                            }
                            for s in salles_nf
                        ]
                        ok = save_justification_nf(
                            reappro=reappro,
                            date_analyse=date_analyse,
                            jour=jour,
                            salles=salles_payload,
                        )
                        if ok:
                            st.success(f"✅ Justifications enregistrées pour {reappro}.")
                        else:
                            st.error("Erreur lors de l'enregistrement. Vérifiez la connexion MongoDB.")
        else:
            st.success("🎉 Toutes les salles ont été faites !")

    with tab_jokers:
        joker_rows = []
        for reappro, data in sorted(results.items()):
            for s in data["salles_faites"]:
                if s["is_joker"]:
                    joker_rows.append({
                        "Réappro Prévu": reappro, "Client / Salle": s["client"],
                        "Machine": s["machine"], "Fait Par": s["employe_reel"],
                        "Valeur Ref": s["val_ref"], "Statut": s["statut"],
                    })
        if joker_rows:
            df_joker = pd.DataFrame(joker_rows)
            st.dataframe(
                df_joker.style.map(lambda _: "background-color:#E67E22; color:#FFFFFF; font-weight:600"),
                use_container_width=True, hide_index=True,
            )
        else:
            st.info("Aucun remplacement détecté.")

    with tab_decale:
        decale_rows = []
        for reappro, data in sorted(results.items()):
            td = data.get("tournee_decalee")
            if td:
                nb_matchees = len(td["salles_decalees"])
                decale_rows.append({
                    "Réappro": reappro, "Jour planifié": jour,
                    "Tournée détectée": td["jour_detecte"],
                    "Salles matchées": nb_matchees,
                    "Total fait": td["nb_machines_faites"],
                    "Salles (planning)": ", ".join(f"{c} [{m}]" for c, m in td["salles_decalees"][:5])
                                         + ("..." if nb_matchees > 5 else ""),
                })
        if decale_rows:
            st.info(
                "Ces réappros ont des salles non faites sur le jour analysé, mais ont quand même "
                "effectué des chargements correspondant à un **autre jour** de leur planning."
            )
            df_decale = pd.DataFrame(decale_rows)
            st.dataframe(
                df_decale.style.map(lambda _: "background-color:#6C3483; color:#FFFFFF; font-weight:600"),
                use_container_width=True, hide_index=True,
            )
            st.markdown("#### Détail des salles décalées")
            for reappro, data in sorted(results.items()):
                td = data.get("tournee_decalee")
                if not td:
                    continue
                with st.expander(f"🔍 {reappro} — tournée du {td['jour_detecte']} faite un {jour}"):
                    detail_dec = [{"Client / Salle": c, "Machine": m} for c, m in td["salles_decalees"]]
                    if detail_dec:
                        st.dataframe(
                            pd.DataFrame(detail_dec).style.map(
                                lambda _: "background-color:#6C3483; color:#FFFFFF; font-weight:600"
                            ),
                            use_container_width=True, hide_index=True,
                        )
                    if td["salles_hors_planning"]:
                        st.caption(f"⚠️ {len(td['salles_hors_planning'])} machine(s) faite(s) hors planning : "
                                   + ", ".join(td["salles_hors_planning"]))
        else:
            st.success("✅ Aucune tournée décalée détectée.")

    with tab_detail:
        selected = st.selectbox("Choisir un réappro", sorted(results.keys()))
        if selected:
            data = results[selected]
            nb_p  = len(data["salles_prevues"])
            nb_f  = len(data["salles_faites"])
            nb_nf = len(data["salles_non_faites"])
            nb_j  = sum(1 for s in data["salles_faites"] if s["is_joker"])
            taux  = round((nb_f / nb_p * 100) if nb_p > 0 else 0, 1)
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Prévues",    nb_p)
            c2.metric("Faites",     nb_f)
            c3.metric("Non Faites", nb_nf)
            c4.metric("Taux",       f"{taux}%")
            detail_rows = []
            for s in data["salles_non_faites"]:
                detail_rows.append({"Statut": "❌ Non Faite", "Client / Salle": s["client"],
                                     "Machine": s["machine"], "Fait Par": "", "Valeur Ref": ""})
            for s in data["salles_faites"]:
                if s["is_joker"]:
                    detail_rows.append({"Statut": "🔄 Joker", "Client / Salle": s["client"],
                                         "Machine": s["machine"], "Fait Par": s["employe_reel"],
                                         "Valeur Ref": s["val_ref"]})
            for s in data["salles_faites"]:
                if not s["is_joker"]:
                    detail_rows.append({"Statut": "✅ Fait", "Client / Salle": s["client"],
                                         "Machine": s["machine"], "Fait Par": s["employe_reel"],
                                         "Valeur Ref": s["val_ref"]})
            def color_detail(row):
                if "Non Faite" in str(row["Statut"]):
                    return ["background-color:#C0392B; color:#FFFFFF; font-weight:600"] * len(row)
                elif "Joker" in str(row["Statut"]):
                    return ["background-color:#E67E22; color:#FFFFFF; font-weight:600"] * len(row)
                else:
                    return ["background-color:#1E7E34; color:#FFFFFF; font-weight:600"] * len(row)
            if detail_rows:
                df_d = pd.DataFrame(detail_rows)
                st.dataframe(
                    df_d.style.apply(color_detail, axis=1),
                    use_container_width=True, hide_index=True,
                    height=min(700, 38 + len(df_d) * 35),
                )

    with tab_bilan_sem:
        _render_bilan_semaine()

    _render_machines_sans_chargement(plannings)



# ════════════════════════════════════════════════════════
# PAGE : PLANOGRAMMES
# ════════════════════════════════════════════════════════
elif st.session_state.page == "planogrammes":

    st.markdown('<div class="main-title">🗂️ Gestionnaire de Planogrammes</div>', unsafe_allow_html=True)
    st.markdown(
        '<div class="subtitle">Créez, modifiez et exportez vos planogrammes de distributeurs automatiques.</div>',
        unsafe_allow_html=True,
    )

    page_planogrammes.render()


# ════════════════════════════════════════════════════════
# PAGE : INVENTAIRES
# ════════════════════════════════════════════════════════
elif st.session_state.page == "inventaires":

    st.markdown('<div class="main-title">📊 Suivi des Inventaires</div>', unsafe_allow_html=True)
    st.markdown(
        '<div class="subtitle">Analysez les inventaires machines par réappro — '
        'contrôle des seuils, produits manquants, statuts.</div>',
        unsafe_allow_html=True,
    )

    page_inventaires.render()


# ════════════════════════════════════════════════════════
# PAGE : COMMANDES
# ════════════════════════════════════════════════════════
elif st.session_state.page == "commandes":

    st.markdown('<div class="main-title">🛒 Suivi des Commandes</div>', unsafe_allow_html=True)
    st.markdown(
        '<div class="subtitle">Importez un screenshot de mail de commande et injectez-le dans le fichier Excel de suivi.</div>',
        unsafe_allow_html=True,
    )

    page_controle_reception.render()


# ════════════════════════════════════════════════════════
# PAGE : INDÉFINIS
# ════════════════════════════════════════════════════════
elif st.session_state.page == "indefinis":

    st.markdown('<div class="main-title">❓ Détection des Indéfinis</div>', unsafe_allow_html=True)
    st.markdown(
        '<div class="subtitle">Identifiez les lignes mal paramétrées qui génèrent des ventes en INDÉFINI.</div>',
        unsafe_allow_html=True,
    )

    page_indefinis.render()


# ════════════════════════════════════════════════════════
# PAGE : CR
# ════════════════════════════════════════════════════════
elif st.session_state.page == "cr":

    st.markdown('<div class="main-title">📝 Compte Rendu Hebdomadaire</div>', unsafe_allow_html=True)
    st.markdown(
        '<div class="subtitle">Générez le compte rendu hebdomadaire par zone géographique.</div>',
        unsafe_allow_html=True,
    )

    page_cr.render()


# ════════════════════════════════════════════════════════
# PAGE : QUARTIX
# ════════════════════════════════════════════════════════
elif st.session_state.page == "quartix":

    st.markdown('<div class="main-title">🗺️ Trajets QUARTIX</div>', unsafe_allow_html=True)
    st.markdown(
        '<div class="subtitle">Visualisez le trajet complet d\'un réappro — tracé de route, arrêts et durées.</div>',
        unsafe_allow_html=True,
    )

    page_quartix.render()


# ════════════════════════════════════════════════════════
# PAGE : PICKLIST VS CHARGEMENT
# ════════════════════════════════════════════════════════
elif st.session_state.page == "picklist":

    page_picklist.render()


# ════════════════════════════════════════════════════════
# PAGE : RAPPORT EMPLOYÉ
# ════════════════════════════════════════════════════════
elif st.session_state.page == "rapport_employe":

    page_rapport_employe.render()
