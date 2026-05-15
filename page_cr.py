"""
Page CR — Compte Rendu Hebdomadaire

Génère un email de compte rendu par zone géographique.
Sources :
  - Collection MongoDB "reappros"  : répartition zones ↔ réappros (import Excel)
  - Collection MongoDB "machines"  : parc machines (champ Approvisionneur)
  - Collection MongoDB "incidents" : problèmes actifs non résolus
"""

import io
import datetime

import altair as alt
import pandas as pd
import streamlit as st

import re as _re

from mongo_storage import (
    _get_client,
    save_bilan_semaine, load_bilan_semaine, list_bilan_semaines, delete_bilan_semaine,
    list_inventaires_semaines, load_inventaires_semaine,
)
from page_inventaires import _parse_planning_for_reappro, WEEKDAY_TO_JOUR

# ────────────────────────────────────────────────────────
# CONSTANTES
# ────────────────────────────────────────────────────────

ZONES = ["IDF", "OUEST", "NORD ET CENTRE", "SUD OUEST", "SUD EST", "EST"]

SEP = "________________________________________"

SECTION_DEFAULTS: dict[str, str] = {
    "Livraisons / Fournisseurs": "Les livraisons ont été contrôlées et validées.",
    "Tournées":                   "Toutes les tournées se sont bien déroulées.",
    "Inventaire":                 "",
}


# ────────────────────────────────────────────────────────
# MONGODB — helpers
# ────────────────────────────────────────────────────────

def _get_col(name: str):
    client = _get_client()
    db_name = st.secrets["mongo"]["db_name"]
    return client[db_name][name]


@st.cache_data(show_spinner=False, ttl=300)
def _load_reappros_from_mongo() -> pd.DataFrame:
    """Retourne le DataFrame réappros {code, reappro, prenom, zone_geo, zone}."""
    try:
        docs = list(_get_col("reappros").find({}, {"_id": 0}))
        if not docs:
            return pd.DataFrame(columns=["code", "reappro", "prenom", "zone_geo", "zone"])
        return pd.DataFrame(docs)
    except Exception as e:
        st.error(f"❌ Impossible de charger les réappros : {e}")
        return pd.DataFrame()


def _load_incidents_for_zone(zone: str, reappros_df: pd.DataFrame) -> list[dict]:
    """
    Retourne la liste des incidents actifs pour une zone donnée.
    Chaque incident est enrichi avec le prénom du réappro responsable.
    """
    if reappros_df.empty:
        return []

    # 1. Codes réappros de la zone
    codes_zone = set(
        reappros_df[reappros_df["zone"] == zone]["code"].str.strip().tolist()
    )
    if not codes_zone:
        return []

    # 2. Salles dont l'Approvisionneur est dans les codes de la zone
    machines = list(
        _get_col("machines").find(
            {"Approvisionneur": {"$in": list(codes_zone)}},
            {"_id": 0, "Client": 1, "Approvisionneur": 1},
        )
    )
    # mapping salle → code_reappro
    salle_to_code: dict[str, str] = {
        m["Client"]: m.get("Approvisionneur", "") for m in machines
    }
    salles_zone = set(salle_to_code.keys())
    if not salles_zone:
        return []

    # 3. Incidents actifs pour ces salles (avec type et created_at)
    incidents_raw = list(
        _get_col("incidents").find(
            {"salle": {"$in": list(salles_zone)}, "status": "actif"},
            {"_id": 0, "salle": 1, "commentaire": 1, "type": 1, "created_at": 1, "since_date": 1},
        )
    )

    # 4. Enrichir avec prénom du réappro
    code_to_prenom: dict[str, str] = dict(
        zip(reappros_df["code"].str.strip(), reappros_df["prenom"].str.strip())
    )
    result = []
    for inc in incidents_raw:
        salle = inc.get("salle", "")
        code  = salle_to_code.get(salle, "")
        prenom = code_to_prenom.get(code, code)
        # Priorité : since_date (date métier) > created_at (fallback)
        created_at = inc.get("created_at")
        since_str  = inc.get("since_date") or (
            created_at.strftime("%d/%m/%Y") if created_at else None
        )
        result.append({
            "salle":       salle,
            "commentaire": inc.get("commentaire", ""),
            "type":        inc.get("type", ""),   # "no_audit" ou "sans_ventes"
            "since":       since_str,
            "prenom":      prenom,
            "code":        code,
        })

    # Tri : no_audit d'abord, puis sans_ventes ; dans chaque groupe tri par réappro puis salle
    result.sort(key=lambda x: (0 if x["type"] == "no_audit" else 1, x["code"], x["salle"]))
    return result


def _build_da_content(incidents: list[dict]) -> str:
    """
    Génère le texte de la section DA en distinguant :
      - Sans remontée télémétrie (no_audit)   → affiché en premier
      - Sans ventes (sans_ventes)              → affiché en second
    """
    if not incidents:
        return "Toutes les salles ont été traitées dans les groupes."

    no_audit   = [i for i in incidents if i["type"] == "no_audit"]
    sans_ventes = [i for i in incidents if i["type"] == "sans_ventes"]

    def _fmt_line(inc: dict) -> str:
        commentaire = inc["commentaire"].strip() if inc["commentaire"] else "—"
        since = f" (depuis le {inc['since']})" if inc.get("since") else ""
        return f"{inc['salle']}{since} : {commentaire} par {inc['prenom']}"

    lines = []

    if no_audit:
        lines.append("• Sans remontée télémétrie :")
        for inc in no_audit:
            lines.append(f"  {_fmt_line(inc)}")

    if no_audit and sans_ventes:
        lines.append("")  # ligne vide entre les deux groupes

    if sans_ventes:
        lines.append("• Sans ventes :")
        for inc in sans_ventes:
            lines.append(f"  {_fmt_line(inc)}")

    return "\n".join(lines)


def _generate_mail(
    zone: str,
    sections: list[tuple[str, str]],  # [(titre, contenu), ...]
) -> str:
    """Génère le texte complet du mail."""
    lines = ["Bonjour à tous,", ""]
    for i, (titre, contenu) in enumerate(sections, start=1):
        lines.append(f"{i}.{titre} :")
        lines.append("")
        lines.append(contenu)
        lines.append(SEP)
    lines.append("")
    lines.append("Bien cordialement,")
    return "\n".join(lines)


# ────────────────────────────────────────────────────────
# IMPORT EXCEL
# ────────────────────────────────────────────────────────

def _import_reappros(raw_bytes: bytes) -> tuple[int, list[str]]:
    """
    Importe le fichier Reappro Guide.xlsx dans MongoDB.
    Retourne (nb_lignes, zones_détectées).
    """
    df = pd.read_excel(io.BytesIO(raw_bytes), header=0)
    df.columns = [str(c).strip() for c in df.columns]

    # Mapping colonnes par position :
    # col 0 = Code, col 1 = prenom, col 2 = Zone Géographique, col 3 = zone, col 4 = Responsable
    cols = df.columns.tolist()
    rename = {
        cols[0]: "code",
        cols[1]: "prenom",
        cols[2]: "zone_geo",
        cols[3]: "zone",
        cols[4]: "responsable",
    }
    df = df.rename(columns=rename)[["code", "prenom", "zone_geo", "zone", "responsable"]]
    df = df.dropna(subset=["code", "zone"]).copy()
    for c in df.columns:
        df[c] = df[c].astype(str).str.strip()
    df = df[df["code"] != ""]

    docs = df.to_dict("records")
    col = _get_col("reappros")
    col.delete_many({})
    if docs:
        col.insert_many(docs)

    zones = sorted(df["zone"].unique().tolist())
    _load_reappros_from_mongo.clear()
    return len(docs), zones


# ────────────────────────────────────────────────────────
# INVENTAIRE AUTO-GENERATION
# ────────────────────────────────────────────────────────

def _build_inventaire_cr_text(
    done_records: list,
    plannings_mongo: dict,
    reappros_df: pd.DataFrame,
    zone: str,
    include_vendredi: bool = False,
) -> str:
    """
    Génère le texte de la section Inventaire du CR depuis les enregistrements BDD.
    done_records : [{"reappro": ..., "date": "dd/mm/yyyy", "code": ...}, ...]
    Cas gérés :
      - Tout fait                    → non mentionné
      - 0 fait toute la semaine      → "aucune salle faite de la semaine"
      - ≤ 2 salles faites la semaine → "aucune salle faite cette semaine sauf X, Y"
      - Quelques manquants par jour  → "X, Y n'ont pas été faites"
      - 0 fait ce jour               → "aucune salle n'a été faite"
      - 1 fait ce jour               → "aucune salle n'a été faite sauf X"
      - Joker                        → "fait par [prénom]"
    """
    if not done_records:
        return "Aucune donnée d'inventaire disponible."

    done_set = {(r["reappro"], r["date"], r["code"]) for r in done_records}

    if not reappros_df.empty and zone:
        zone_codes = set(reappros_df[reappros_df["zone"] == zone]["code"].str.strip())
    else:
        zone_codes = set(plannings_mongo.keys())

    code_to_prenom = (
        dict(zip(reappros_df["code"].str.strip(), reappros_df["prenom"].str.strip()))
        if not reappros_df.empty else {}
    )

    # ISO weeks depuis les dates des enregistrements
    dates = {r["date"] for r in done_records}
    iso_pairs_set = set()
    for d in dates:
        try:
            dt = datetime.datetime.strptime(d, "%d/%m/%Y")
            iso = dt.isocalendar()
            iso_pairs_set.add((int(iso[0]), int(iso[1])))
        except ValueError:
            pass
    iso_pairs = sorted(iso_pairs_set)
    if not iso_pairs:
        return "Aucune date trouvée dans les enregistrements."

    week_days = []
    for iso_year, iso_week in iso_pairs:
        monday = datetime.datetime.fromisocalendar(iso_year, iso_week, 1)
        for offset in range(5):
            day_dt  = monday + datetime.timedelta(days=offset)
            jour_fr = WEEKDAY_TO_JOUR.get(day_dt.weekday())
            if jour_fr and (include_vendredi or jour_fr != "Vendredi"):
                week_days.append((day_dt.strftime("%d/%m/%Y"), jour_fr))

    zone_plannings = {r: p for r, p in plannings_mongo.items() if r in zone_codes}
    if not zone_plannings:
        return "Aucun réappro trouvé pour cette zone dans les plannings."

    def _accord(n):
        return "n'ont pas été faites" if n > 1 else "n'a pas été faite"

    result_blocks = []

    for reappro in sorted(zone_plannings.keys()):
        planning = _parse_planning_for_reappro(zone_plannings[reappro])
        prenom   = code_to_prenom.get(reappro, reappro)

        # ── Collecter les données de toute la semaine ─────────────────────
        week_own    = []   # toutes salles faites par lui cette semaine
        week_joker  = []   # [(salle, prenom_joker)] cette semaine
        week_missing= 0
        day_data    = []   # [(jour_fr, done_own, done_joker, missing)]

        for date_str, jour_fr in week_days:
            if jour_fr not in planning or not planning[jour_fr]:
                continue
            jour_plan  = planning[jour_fr]
            done_own   = []
            done_joker = []
            missing    = []

            for code, info in sorted(jour_plan.items(), key=lambda x: x[1]["label"]):
                salle = info["label"]
                if (reappro, date_str, code) in done_set:
                    done_own.append(salle)
                else:
                    joker_r = next(
                        (r for r in plannings_mongo
                         if r != reappro and (r, date_str, code) in done_set),
                        None,
                    )
                    if joker_r:
                        done_joker.append((salle, code_to_prenom.get(joker_r, joker_r)))
                    else:
                        missing.append(salle)

            week_own.extend(done_own)
            week_joker.extend(done_joker)
            week_missing += len(missing)
            day_data.append((jour_fr, done_own, done_joker, missing))

        if not day_data:
            continue

        week_done = len(week_own) + len(week_joker)

        # ── Résumé hebdomadaire si presque rien fait ──────────────────────
        if week_done == 0 and week_missing > 0:
            result_blocks.append(f"{prenom} ({reappro}) :\n- aucune salle faite de la semaine")
            continue

        if week_done <= 2 and week_missing > 0:
            sauf_parts = week_own + [f"{s} (fait par {p})" for s, p in week_joker]
            result_blocks.append(
                f"{prenom} ({reappro}) :\n"
                f"- aucune salle faite cette semaine sauf {', '.join(sauf_parts)}"
            )
            continue

        # ── Détail jour par jour ──────────────────────────────────────────
        day_lines = []
        for jour_fr, done_own, done_joker, missing in day_data:
            done_count = len(done_own) + len(done_joker)

            if not missing and not done_joker:
                continue  # jour parfait

            if not missing:
                jk = ", ".join(f"{s} (fait par {p})" for s, p in done_joker)
                day_lines.append(f"- {jour_fr} : {jk}")
                continue

            joker_suffix = (
                " — " + ", ".join(f"{s} fait par {p}" for s, p in done_joker)
                if done_joker else ""
            )

            if done_count == 0:
                line = f"- {jour_fr} : aucune salle n'a été faite"
            elif len(done_own) == 1 and not done_joker:
                line = f"- {jour_fr} : aucune salle n'a été faite sauf {done_own[0]}"
            elif done_count == 1 and done_joker:
                s, p = done_joker[0]
                line = f"- {jour_fr} : aucune salle n'a été faite sauf {s} (fait par {p})"
            else:
                miss_str = ", ".join(missing)
                line = f"- {jour_fr} : {miss_str} {_accord(len(missing))}{joker_suffix}"

            day_lines.append(line)

        if day_lines:
            result_blocks.append(f"{prenom} ({reappro}) :\n" + "\n".join(day_lines))

    if not result_blocks:
        return "Tous les inventaires ont été réalisés conformément au planning."

    return "\n\n".join(result_blocks)


def _extract_joker_name(statut: str) -> str:
    """Extrait le prénom du joker depuis 'Fait par X' ou 'Fait par X (le DD/MM/YY)'."""
    return _re.sub(r"\s*\(le .+?\)$", "", statut[len("Fait par"):]).strip()


def _build_inventaire_cr_text_from_bilan(
    bilan_rows: list,
    reappros_df: pd.DataFrame,
    zone: str,
    include_vendredi: bool = False,
) -> str:
    """
    Génère le même texte que _build_inventaire_cr_text mais depuis les bilan_rows
    déjà calculés (collection bilan_semaine).
    Statuts : 'Fait' | 'Fait le DD/MM' | 'Fait par X' | 'Fait par X (le DD/MM)' | 'Non fait'
    """
    if not bilan_rows:
        return "Aucune donnée de bilan disponible."

    if not reappros_df.empty and zone:
        zone_codes = set(reappros_df[reappros_df["zone"] == zone]["code"].str.strip())
    else:
        zone_codes = {r["reappro"] for r in bilan_rows}

    code_to_prenom = (
        dict(zip(reappros_df["code"].str.strip(), reappros_df["prenom"].str.strip()))
        if not reappros_df.empty else {}
    )

    rows = [
        r for r in bilan_rows
        if r["reappro"] in zone_codes
        and (include_vendredi or r.get("jour") != "Vendredi")
    ]
    if not rows:
        return "Aucun réappro trouvé pour cette zone dans le bilan."

    def _accord(n):
        return "n'ont pas été faites" if n > 1 else "n'a pas été faite"

    result_blocks = []

    for reappro in sorted(zone_codes):
        sub = [r for r in rows if r["reappro"] == reappro]
        if not sub:
            continue

        prenom = code_to_prenom.get(reappro, reappro)

        def _unique_salles(rows, pred):
            """Unique cleaned salle names for rows matching pred."""
            seen, out = set(), []
            for r in rows:
                if pred(r):
                    s = r["salle"].strip().rstrip(",").strip()
                    if s not in seen:
                        seen.add(s)
                        out.append(s)
            return out

        # Salle done = statut starts with "Fait" (own OR joker — we don't care who did it)
        def _sauf_label(salle: str, row) -> str:
            """'salle' or 'salle (fait par Prénom)' depending on whether it was a joker."""
            if row["statut"].startswith("Fait par"):
                code = _extract_joker_name(row["statut"])
                name = code_to_prenom.get(code, code)
                return f"{salle} (fait par {name})"
            return salle

        week_done_rows = [r for r in sub if r["statut"].startswith("Fait")]
        week_missing   = _unique_salles(sub, lambda r: r["statut"] == "Non fait")

        # Deduplicate done rows by salle name, keeping first occurrence
        _seen_done: set = set()
        week_done_dedup = []
        for r in week_done_rows:
            s = r["salle"].strip().rstrip(",").strip()
            if s not in _seen_done:
                _seen_done.add(s)
                week_done_dedup.append((s, r))
        week_done = len(week_done_dedup)

        if week_done == 0 and week_missing:
            result_blocks.append(f"{prenom} ({reappro}) :\n- aucune salle faite de la semaine")
            continue

        if week_done <= 2 and week_missing:
            sauf = [_sauf_label(s, r) for s, r in week_done_dedup]
            result_blocks.append(
                f"{prenom} ({reappro}) :\n"
                f"- aucune salle faite cette semaine sauf {', '.join(sauf)}"
            )
            continue

        day_lines = []
        for jour_fr in _JOURS_ORDER:
            if jour_fr == "Vendredi" and not include_vendredi:
                continue
            day_sub = [r for r in sub if r.get("jour") == jour_fr]
            if not day_sub:
                continue

            # Deduplicate done rows for this day
            _seen_day: set = set()
            done_dedup = []
            for r in day_sub:
                if r["statut"].startswith("Fait"):
                    s = r["salle"].strip().rstrip(",").strip()
                    if s not in _seen_day:
                        _seen_day.add(s)
                        done_dedup.append((s, r))

            missing = _unique_salles(day_sub, lambda r: r["statut"] == "Non fait")

            if not missing:
                continue  # tout fait ce jour → on ne mentionne rien

            done_count = len(done_dedup)

            if done_count == 0:
                line = f"- {jour_fr} : aucune salle n'a été faite"
            elif done_count == 1:
                s, r = done_dedup[0]
                line = f"- {jour_fr} : aucune salle n'a été faite sauf {_sauf_label(s, r)}"
            else:
                line = f"- {jour_fr} : {', '.join(missing)} {_accord(len(missing))}"

            day_lines.append(line)

        if day_lines:
            result_blocks.append(f"{prenom} ({reappro}) :\n" + "\n".join(day_lines))

    if not result_blocks:
        return "Tous les inventaires ont été réalisés conformément au planning."

    return "\n\n".join(result_blocks)


# ────────────────────────────────────────────────────────
# BILAN SEMAINE (page CR, indépendant)
# ────────────────────────────────────────────────────────

def _parse_bilan_export(raw_bytes: bytes) -> pd.DataFrame:
    """Parse le nouvel export inventaires (Employé/Machine/Val.Ref) sans détail produit."""
    df = pd.read_csv(
        io.BytesIO(raw_bytes), sep=";", encoding="utf-8-sig",
        dtype=str, skipinitialspace=True,
    )
    # Nettoyer les noms de colonnes (guillemets, <br/>, espaces)
    df.columns = [
        c.strip().strip('"').replace("<br/>", " ").replace("\n", "")
        for c in df.columns
    ]
    # Supprimer colonne vide éventuelle (trailing semicolon)
    df = df.loc[:, df.columns.str.strip() != ""]

    required = ["Tiers", "Date début", "Statut", "Employé", "Machine", "Val. Ref"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Colonnes manquantes : {', '.join(missing)}")

    # Strip champs texte clés
    for col in ["Statut", "Employé", "Machine", "Tiers"]:
        df[col] = df[col].str.strip().str.strip('"')

    # Filtre : uniquement les inventaires "Fait"
    df = df[df["Statut"] == "Fait"].copy()

    # Parse date (format "dd/mm/yyyy HH:MM")
    df["_dt"] = pd.to_datetime(
        df["Date début"].str.strip().str.strip('"'),
        format="%d/%m/%Y %H:%M", errors="coerce",
    )
    df["Date"] = df["_dt"].dt.strftime("%d/%m/%Y")

    # Val. Ref → float
    df["Val_ref"] = pd.to_numeric(
        df["Val. Ref"].str.strip().str.strip('"').str.replace(",", ".", regex=False),
        errors="coerce",
    ).fillna(0.0)

    # Déduplique (Employé, Machine, Date) : garde le max Val_ref (ignore doublons à 0)
    df = (
        df.sort_values("Val_ref", ascending=False)
        .groupby(["Employé", "Machine", "Date"], as_index=False)
        .first()
    )

    return df[["Employé", "Machine", "Tiers", "Date", "_dt", "Val_ref"]].dropna(subset=["Date"])


def _planning_by_machine(planning_raw: dict) -> dict:
    """
    Construit {jour_fr: {machine_code: {"label": ..., "client_code": ...}}}
    depuis le planning brut MongoDB (même format que load_plannings_from_mongo).
    Clé = code machine (ex "2219M1"), pas le code client.
    """
    JOURS = {"Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi"}
    result = {}
    for jour, salles in planning_raw.items():
        if jour not in JOURS:
            continue
        result[jour] = {}
        for salle in salles:
            if len(salle) < 2:
                continue
            client_full = str(salle[0]).strip()
            machine     = str(salle[1]).strip()
            if not machine:
                continue
            if " - " in client_full:
                client_code = client_full.split(" - ")[0].strip()
                label       = client_full.split(" - ", 1)[1].strip()
            else:
                client_code = client_full
                label       = client_full
            result[jour][machine] = {"label": label, "client_code": client_code}
    return result


_JOURS_ORDER = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi"]
_WEEKDAY_TO_JOUR = {0: "Lundi", 1: "Mardi", 2: "Mercredi", 3: "Jeudi", 4: "Vendredi"}


def _build_bilan_cr(df: pd.DataFrame, plannings_mongo: dict) -> list:
    """
    Croise le DataFrame d'inventaires avec les plannings MongoDB.
    Groupé par jour de semaine (Lundi…Vendredi) — jamais dupliqué même si le fichier
    couvre plusieurs semaines. La semaine ISO la plus représentée est la semaine de
    référence ; si une salle a été faite un autre lundi/mardi, le statut indique la
    date réelle ("Fait le DD/MM/YY" / "Fait par X (le DD/MM/YY)").
    """
    df = df.copy()
    df["_iso_year"] = df["_dt"].dt.isocalendar().year.astype("Int64")
    df["_iso_week"] = df["_dt"].dt.isocalendar().week.astype("Int64")

    # Semaine de référence = ISO week avec le plus d'inventaires
    week_counts = df.groupby(["_iso_year", "_iso_week"]).size()
    if week_counts.empty:
        return []
    ref_year, ref_week = week_counts.idxmax()
    ref_monday = datetime.datetime.fromisocalendar(int(ref_year), int(ref_week), 1)
    ref_dates  = {
        _WEEKDAY_TO_JOUR[i]: (ref_monday + datetime.timedelta(days=i)).strftime("%d/%m/%Y")
        for i in range(5)
    }

    # Index par (employe, machine) sur TOUTE la semaine — indépendant du jour réel
    # Un réappro peut faire ses salles sur un jour différent de celui planifié
    done_by_emp: dict = {}   # {(employe, machine): (actual_date, val)}
    done_by_any: dict = {}   # {machine: (employe, actual_date, val)} — pour jokers

    for _, row in df.iterrows():
        emp     = row["Employé"]
        machine = row["Machine"]
        date    = row["Date"]
        val     = row["Val_ref"]

        key_emp = (emp, machine)
        # Priorité à la date de la semaine de référence si plusieurs entrées
        if key_emp not in done_by_emp or date in ref_dates.values():
            done_by_emp[key_emp] = (date, val)

        if machine not in done_by_any or date in ref_dates.values():
            done_by_any[machine] = (emp, date, val)

    rows = []
    for reappro, planning_raw in sorted(plannings_mongo.items()):
        planning  = _planning_by_machine(planning_raw)
        seen_mach = set()  # évite de compter deux fois la même machine

        for jour_fr in _JOURS_ORDER:
            if jour_fr not in planning:
                continue
            ref_date = ref_dates[jour_fr]
            for machine, info in sorted(planning[jour_fr].items(),
                                        key=lambda x: x[1]["label"]):
                if machine in seen_mach:
                    continue
                seen_mach.add(machine)

                own_key = (reappro, machine)

                if own_key in done_by_emp:
                    actual_date, val = done_by_emp[own_key]
                    fait_par = reappro
                    statut   = "Fait" if actual_date == ref_date else f"Fait le {actual_date}"
                elif machine in done_by_any:
                    joker_emp, actual_date, val = done_by_any[machine]
                    fait_par = joker_emp
                    statut   = (f"Fait par {joker_emp}" if actual_date == ref_date
                                else f"Fait par {joker_emp} (le {actual_date})")
                else:
                    statut   = "Non fait"
                    fait_par = None
                    val      = None

                rows.append({
                    "reappro":  reappro,
                    "jour":     jour_fr,
                    "ref_date": ref_date,
                    "machine":  machine,
                    "salle":    info["label"],
                    "statut":   statut,
                    "fait_par": fait_par,
                    "val":      val,
                })
    return rows


def _export_bilan_cr_excel(rows: list, code_to_prenom: dict) -> bytes:
    """Génère un Excel avec 1 onglet par réappro."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    def _fill(h): return PatternFill("solid", fgColor=h.lstrip("#"))
    def _font(bold=False, color="000000", size=9):
        return Font(bold=bold, color=color, name="Arial", size=size)
    def _align(h="left"):
        return Alignment(horizontal=h, vertical="center", wrap_text=False)
    _s   = Side(style="thin", color="CCCCCC")
    _brd = Border(left=_s, right=_s, top=_s, bottom=_s)

    C_DARK  = "1F4E79"; C_MID = "2E75B6"; C_SUB = "BDD7EE"
    C_OK_BG = "D4EDDA"; C_OK_FG = "1E7E34"
    C_NOK_BG = "BBDEFB"; C_NOK_FG = "1565C0"
    C_JOK_BG = "E8DAEF"; C_JOK_FG = "6C3483"
    C_WHITE  = "FFFFFF"

    HDRS   = ["Date", "Jour", "Salle", "Machine", "Statut", "Valeur HT"]
    WIDTHS = [11,      10,     40,      10,         22,       12]

    wb = Workbook()
    wb.remove(wb.active)

    reappros = sorted({r["reappro"] for r in rows})

    for reappro in reappros:
        prenom  = code_to_prenom.get(reappro, reappro)
        ws      = wb.create_sheet(reappro[:31])
        ws.freeze_panes = "A3"

        N = len(HDRS)
        ws.merge_cells(f"A1:{get_column_letter(N)}1")
        tc = ws.cell(1, 1, f"Bilan semaine — {prenom} ({reappro})")
        tc.fill = _fill(C_DARK); tc.font = _font(True, C_WHITE, 12)
        tc.alignment = _align("center"); ws.row_dimensions[1].height = 24

        for ci, (h, w) in enumerate(zip(HDRS, WIDTHS), 1):
            c = ws.cell(2, ci, h)
            c.fill = _fill(C_MID); c.font = _font(True, C_WHITE, 9)
            c.alignment = _align("center")
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[2].height = 18

        sub = [r for r in rows if r["reappro"] == reappro]
        ri  = 3
        cur = None

        for r in sub:
            if r["jour"] != cur:
                cur      = r["jour"]
                day_rows = [x for x in sub if x["jour"] == r["jour"]]
                nb_f = sum(1 for x in day_rows if not x["statut"].startswith("Non"))
                nb_t = len(day_rows)
                txt  = f"  {r['jour']}  {r['ref_date']}   |   {nb_f} / {nb_t} faits"
                ws.merge_cells(f"A{ri}:{get_column_letter(N)}{ri}")
                sep = ws.cell(ri, 1, txt)
                sep.fill = _fill(C_SUB); sep.font = _font(True, C_DARK, 10)
                sep.alignment = _align("left"); ws.row_dimensions[ri].height = 20
                ri += 1

            st = r["statut"]
            if st.startswith("Non"):
                bg, fg = C_NOK_BG, C_NOK_FG
            elif st.startswith("Fait par"):
                bg, fg = C_JOK_BG, C_JOK_FG
            else:  # Fait / Fait le XX
                bg, fg = C_OK_BG, C_OK_FG

            ht   = round(r["val"], 2) if r["val"] is not None else ""
            vals = [r["ref_date"], r["jour"], r["salle"], r["machine"], st, ht]

            for ci, v in enumerate(vals, 1):
                c = ws.cell(ri, ci, v)
                c.fill = _fill(bg); c.border = _brd
                c.font = _font(bold=(ci == 5), color=(fg if ci == 5 else "000000"), size=9)
                c.alignment = _align("center" if ci in (1, 2, 4, 5, 6) else "left")
                if ci == 6 and isinstance(v, float):
                    c.number_format = "#,##0.00 €"
            ws.row_dimensions[ri].height = 16
            ri += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _section_bilan_cr(plannings_mongo: dict, reappros_df: pd.DataFrame):
    """Section Bilan semaine dans la page CR — indépendante de tout autre upload."""
    st.markdown("---")
    st.markdown("### 📅 Bilan semaine")

    if not plannings_mongo:
        st.warning("⚠️ Plannings non disponibles.")
        return

    # ── Source : BDD ou fichier ───────────────────────────────────────────────
    saved_bilans = list_bilan_semaines()
    bilan_rows: list = []

    if saved_bilans:
        sem_labels = [d.get("label") or f"S{d['iso_week']} {d['iso_year']}" for d in saved_bilans]
        st.caption("**Charger un bilan sauvegardé :**")
        col_sel, col_load, col_del = st.columns([3, 1, 1])
        with col_sel:
            sel_idx = st.selectbox(
                "Semaine BDD", range(len(sem_labels)),
                format_func=lambda i: sem_labels[i],
                key="cr_bilan_bdd_sel", label_visibility="collapsed",
            )
        with col_load:
            if st.button("📂 Charger", key="cr_bilan_bdd_btn", use_container_width=True):
                sel = saved_bilans[sel_idx]
                doc = load_bilan_semaine(sel["iso_year"], sel["iso_week"])
                st.session_state["cr_bilan_rows_bdd"] = doc.get("rows", [])
                st.session_state["cr_bilan_src_label"] = sem_labels[sel_idx]
        with col_del:
            if st.button("🗑️ Supprimer", key="cr_bilan_bdd_del", use_container_width=True,
                         type="secondary"):
                sel = saved_bilans[sel_idx]
                delete_bilan_semaine(sel["iso_year"], sel["iso_week"])
                st.session_state.pop("cr_bilan_rows_bdd", None)
                st.success(f"✅ Bilan S{sel['iso_week']} {sel['iso_year']} supprimé.")
                st.rerun()

        if "cr_bilan_rows_bdd" in st.session_state and st.session_state["cr_bilan_rows_bdd"]:
            bilan_rows = st.session_state["cr_bilan_rows_bdd"]
            st.success(f"📂 Bilan {st.session_state.get('cr_bilan_src_label', '')} chargé depuis la BDD.")

        st.markdown("<div style='margin:8px 0;text-align:center;color:#aaa'>— ou —</div>",
                    unsafe_allow_html=True)

    st.caption("**Uploader un nouveau fichier :**")
    up = st.file_uploader(
        "Export inventaires (CSV)", type=["csv"],
        key="cr_bilan_upload", label_visibility="collapsed",
    )
    if up is not None:
        try:
            with st.spinner("Analyse…"):
                df_bilan   = _parse_bilan_export(up.read())
                bilan_rows = _build_bilan_cr(df_bilan, plannings_mongo)
                # reset BDD cache so the fresh file takes precedence
                st.session_state.pop("cr_bilan_rows_bdd", None)
        except Exception as e:
            st.error(f"❌ {e}")
            return

    if not bilan_rows:
        st.info("Déposez un fichier CSV ou chargez un bilan depuis la BDD.")
        return

    if not bilan_rows:
        st.warning("Aucune correspondance entre le fichier et les plannings.")
        return

    code_to_prenom = (
        dict(zip(reappros_df["code"].str.strip(), reappros_df["prenom"].str.strip()))
        if not reappros_df.empty else {}
    )

    # ── Filtres ───────────────────────────────────────────────────────────────
    # Construire les listes de choix depuis reappros_df (zone = responsable)
    all_reappros_in_bilan = sorted({r["reappro"] for r in bilan_rows})

    # Responsables disponibles pour les réappros présents dans le bilan
    if not reappros_df.empty:
        resp_map = (
            reappros_df
            .dropna(subset=["code", "responsable"])
            .assign(code=lambda d: d["code"].str.strip(),
                    responsable=lambda d: d["responsable"].str.strip())
            .set_index("code")["responsable"]
            .to_dict()
        )
        all_responsables = sorted({
            resp_map[r] for r in all_reappros_in_bilan if r in resp_map
        })
    else:
        resp_map = {}
        all_responsables = []

    fc1, fc2 = st.columns(2)
    with fc1:
        sel_resp = st.multiselect(
            "Filtrer par responsable",
            options=all_responsables,
            default=[],
            key="cr_bilan_f_resp",
            placeholder="Tous les responsables",
        )
    with fc2:
        reappro_options = all_reappros_in_bilan
        if sel_resp:
            reappro_options = [r for r in all_reappros_in_bilan
                               if resp_map.get(r) in sel_resp]
        sel_reappros = st.multiselect(
            "Filtrer par réappro",
            options=reappro_options,
            default=[],
            key="cr_bilan_f_reappro",
            placeholder="Tous les réappros",
        )

    only_nf = st.checkbox(
        "Afficher uniquement les salles non faites",
        value=False,
        key="cr_bilan_f_nf",
    )

    # Appliquer les filtres
    active_reappros = sel_reappros if sel_reappros else (
        reappro_options if sel_resp else all_reappros_in_bilan
    )
    filtered_rows = [r for r in bilan_rows if r["reappro"] in active_reappros]
    if only_nf:
        filtered_rows = [r for r in filtered_rows if r["statut"] == "Non fait"]

    # ── KPIs ─────────────────────────────────────────────────────────────────
    nb_plan  = len(filtered_rows)
    nb_fait  = sum(1 for r in filtered_rows if r["statut"].startswith("Fait") and not r["statut"].startswith("Fait par"))
    nb_joker = sum(1 for r in filtered_rows if r["statut"].startswith("Fait par"))
    nb_nf    = sum(1 for r in filtered_rows if r["statut"] == "Non fait")
    pct      = round((nb_fait + nb_joker) / nb_plan * 100, 1) if nb_plan else 0

    k1, k2, k3, k4, k5 = st.columns(5)
    k1.metric("📋 Planifiées",  nb_plan)
    k2.metric("✅ Faites",      nb_fait,  delta=f"{pct}%")
    k3.metric("👥 Fait par autre", nb_joker)
    k4.metric("🔵 Non faites",  nb_nf,
              delta=f"-{nb_nf}" if nb_nf else None, delta_color="inverse")
    k5.metric("👥 Réappros",    len({r["reappro"] for r in filtered_rows}))

    # ── Graphiques ────────────────────────────────────────────────────────────
    _COLOR_SCALE = alt.Scale(
        domain=["✅ Fait", "🔀 Fait par autre", "🔵 Non fait"],
        range=["#2e7d32", "#7b1fa2", "#1565c0"],
    )

    chart_df = pd.DataFrame([{
        "prenom":  code_to_prenom.get(r["reappro"], r["reappro"]),
        "reappro": r["reappro"],
        "jour":    r["jour"],
        "cat": (
            "🔀 Fait par autre" if r["statut"].startswith("Fait par")
            else "✅ Fait"       if r["statut"].startswith("Fait")
            else "🔵 Non fait"
        ),
    } for r in filtered_rows])

    # — Donut global —
    donut_src = chart_df.groupby("cat").size().reset_index(name="n")
    donut_src["pct_label"] = donut_src["n"].apply(
        lambda v: f"{round(v / donut_src['n'].sum() * 100, 1)}%"
    )
    donut = (
        alt.Chart(donut_src)
        .mark_arc(innerRadius=52, outerRadius=90)
        .encode(
            theta=alt.Theta("n:Q"),
            color=alt.Color("cat:N", scale=_COLOR_SCALE,
                            legend=alt.Legend(title=None, orient="bottom",
                                              labelFontSize=11)),
            tooltip=[
                alt.Tooltip("cat:N",       title="Statut"),
                alt.Tooltip("n:Q",         title="Nb salles"),
                alt.Tooltip("pct_label:N", title="Part"),
            ],
        )
        .properties(title=alt.TitleParams("Composition globale", fontSize=13),
                    width=210, height=210)
    )

    # — Barres verticales empilées (normalisées) par réappro —
    joker_rows = [
        {"prenom": code_to_prenom.get(r["reappro"], r["reappro"]),
         "joker":  _extract_joker_name(r["statut"])}
        for r in filtered_rows if r["statut"].startswith("Fait par")
    ]
    joker_by_prenom = (
        pd.DataFrame(joker_rows).groupby("prenom")["joker"]
        .apply(lambda s: ", ".join(sorted(set(s))))
        .reset_index(name="jokers")
        if joker_rows else pd.DataFrame(columns=["prenom", "jokers"])
    )

    bar_src = chart_df.groupby(["prenom", "cat"]).size().reset_index(name="n")
    bar_src = bar_src.merge(joker_by_prenom, on="prenom", how="left")
    bar_src["jokers"] = bar_src.apply(
        lambda row: row["jokers"] if row["cat"] == "🔀 Fait par autre" else "", axis=1
    ).fillna("")
    totals    = chart_df.groupby("prenom").size().rename("total")
    done_cnt  = (
        chart_df[chart_df["cat"] != "🔵 Non fait"]
        .groupby("prenom").size().rename("done")
    )
    pct_order = (
        pd.concat([totals, done_cnt], axis=1)
        .fillna(0)
        .assign(pct=lambda d: d["done"] / d["total"])
        .sort_values("pct", ascending=False)
        .index.tolist()
    )
    lolli_chart = (
        alt.Chart(bar_src)
        .mark_bar()
        .encode(
            x=alt.X("prenom:N", sort=pct_order,
                    axis=alt.Axis(title=None, labelAngle=-40, labelFontSize=11)),
            y=alt.Y("n:Q", stack="normalize",
                    axis=alt.Axis(format="%", title=None, tickCount=4)),
            color=alt.Color("cat:N", scale=_COLOR_SCALE,
                            legend=alt.Legend(title=None, orient="top",
                                              labelFontSize=11)),
            tooltip=[
                alt.Tooltip("prenom:N", title="Réappro"),
                alt.Tooltip("cat:N",    title="Statut"),
                alt.Tooltip("n:Q",      title="Nb salles"),
                alt.Tooltip("jokers:N", title="Fait par"),
            ],
        )
        .properties(
            title=alt.TitleParams("Taux de complétion par réappro", fontSize=13),
            height=220,
        )
    )

    # — Non faites par jour —
    day_src = (
        chart_df[chart_df["cat"] == "🔵 Non fait"]
        .groupby("jour").size().reset_index(name="non_fait")
    )
    # Ensure all days present even with 0
    day_src = (
        pd.DataFrame({"jour": _JOURS_ORDER})
        .merge(day_src, on="jour", how="left")
        .fillna(0)
        .astype({"non_fait": int})
    )
    max_nf = int(day_src["non_fait"].max()) if not day_src.empty else 1

    day_bars = (
        alt.Chart(day_src)
        .mark_bar(cornerRadiusTopLeft=4, cornerRadiusTopRight=4)
        .encode(
            x=alt.X("jour:N", sort=_JOURS_ORDER, axis=alt.Axis(title=None, labelFontSize=13)),
            y=alt.Y("non_fait:Q", axis=alt.Axis(title="Non faites", tickMinStep=1)),
            color=alt.condition(
                alt.datum.non_fait >= max_nf,
                alt.value("#c62828"),
                alt.value("#1565c0"),
            ),
            tooltip=[
                alt.Tooltip("jour:N",     title="Jour"),
                alt.Tooltip("non_fait:Q", title="Non faites"),
            ],
        )
    )
    day_text = (
        alt.Chart(day_src)
        .mark_text(dy=-8, fontSize=13, fontWeight="bold", color="#333")
        .encode(
            x=alt.X("jour:N", sort=_JOURS_ORDER),
            y=alt.Y("non_fait:Q"),
            text=alt.Text("non_fait:Q"),
        )
    )
    day_chart = (
        alt.layer(day_bars, day_text)
        .properties(
            title=alt.TitleParams("Non faites par jour", fontSize=13),
            width=460, height=260,
        )
    )

    # ── Onglets ───────────────────────────────────────────────────────────────
    tab_lolli, tab_donut, tab_jour = st.tabs([
        "📊 Complétion par réappro",
        "🍩 Répartition globale",
        "📅 Non faites par jour",
    ])

    with tab_lolli:
        st.altair_chart(lolli_chart, use_container_width=True)

    with tab_donut:
        c_left, c_right = st.columns([1, 2])
        with c_left:
            st.altair_chart(donut.properties(width=260, height=260),
                            use_container_width=False)
        with c_right:
            detail = donut_src.rename(columns={
                "cat": "Statut", "n": "Nb salles", "pct_label": "Part"
            })
            st.dataframe(detail[["Statut", "Nb salles", "Part"]],
                         hide_index=True, use_container_width=False)

    with tab_jour:
        st.altair_chart(day_chart, use_container_width=False)

    # ── Export + Sauvegarde BDD ───────────────────────────────────────────────
    col_exp, col_name, col_save = st.columns([1, 2, 1])
    with col_exp:
        st.download_button(
            "📥 Exporter Excel",
            data=_export_bilan_cr_excel(filtered_rows, code_to_prenom),
            file_name=f"bilan_semaine_{datetime.date.today():%Y%m%d}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="cr_bilan_dl",
        )
    with col_name:
        ref_dt_default = datetime.datetime.strptime(bilan_rows[0]["ref_date"], "%d/%m/%Y")
        iso_cal_default = ref_dt_default.isocalendar()
        default_label = f"S{iso_cal_default[1]} {iso_cal_default[0]}"
        save_label = st.text_input(
            "Nom de la sauvegarde",
            value=default_label,
            key="cr_bilan_save_label",
            label_visibility="collapsed",
            placeholder="Ex : S15 2026 — semaine normale",
        )
    with col_save:
        if st.button("💾 Sauvegarder en BDD", key="cr_bilan_save",
                     use_container_width=True):
            if not save_label.strip():
                st.warning("Donnez un nom à la sauvegarde.")
            else:
                try:
                    ref_dt  = datetime.datetime.strptime(bilan_rows[0]["ref_date"], "%d/%m/%Y")
                    iso_cal = ref_dt.isocalendar()
                    iso_y, iso_w = int(iso_cal[0]), int(iso_cal[1])
                    save_bilan_semaine(bilan_rows, iso_y, iso_w, label=save_label.strip())
                    st.session_state["cr_bilan_rows_bdd"]  = bilan_rows
                    st.session_state["cr_bilan_src_label"] = save_label.strip()
                    st.success(f"✅ « {save_label.strip()} » sauvegardé en BDD !")
                    st.rerun()
                except Exception as e:
                    st.error(f"❌ Erreur lors de la sauvegarde : {e}")

    st.divider()

    # ── Accordéons par réappro ────────────────────────────────────────────────
    for reappro in sorted({r["reappro"] for r in filtered_rows}):
        sub    = [r for r in filtered_rows if r["reappro"] == reappro]
        nb_f   = sum(1 for r in sub if r["statut"].startswith("Fait") and not r["statut"].startswith("Fait par"))
        nb_jok = sum(1 for r in sub if r["statut"].startswith("Fait par"))
        nb_t   = len(sub)
        nb_a   = nb_t - nb_f - nb_jok
        pct_r  = round((nb_f + nb_jok) / nb_t * 100, 1) if nb_t else 0
        prenom = code_to_prenom.get(reappro, reappro)

        titre = (
            f"**{prenom}** ({reappro}) — {nb_f + nb_jok}/{nb_t}"
            f" · ✅ {nb_f} · 🔀 {nb_jok} · 🔵 {nb_a} abs · {pct_r}%"
        )
        with st.expander(titre, expanded=False):
            for jour_fr in _JOURS_ORDER:
                day_rows = [r for r in sub if r["jour"] == jour_fr]
                if not day_rows:
                    continue
                nb_df = sum(1 for r in day_rows if not r["statut"].startswith("Non"))
                nb_dt = len(day_rows)
                ref_date = day_rows[0]["ref_date"]

                color = "#2e7d32" if nb_df == nb_dt else ("#1565c0" if nb_df == 0 else "#e65100")
                icon  = "✅" if nb_df == nb_dt else ("🔵" if nb_df == 0 else "🟠")
                st.markdown(
                    f'<div style="border-left:4px solid {color};padding:6px 12px;'
                    f'border-radius:4px;margin-bottom:4px">'
                    f'<b>{icon} {jour_fr} {ref_date}</b> — {nb_df}/{nb_dt} faits</div>',
                    unsafe_allow_html=True,
                )

                df_display = pd.DataFrame([{
                    "Machine": r["machine"],
                    "Salle":   r["salle"],
                    "Statut":  r["statut"],
                    "Val. HT": f"{r['val']:.2f} €" if r["val"] else "",
                } for r in day_rows])

                def _row_style(row):
                    if row["Statut"].startswith("Non"):
                        return ["background-color:#BBDEFB"] * len(row)
                    elif row["Statut"].startswith("Fait par"):
                        return ["background-color:#E8DAEF"] * len(row)
                    else:
                        return ["background-color:#D4EDDA"] * len(row)

                st.dataframe(
                    df_display.style.apply(_row_style, axis=1),
                    hide_index=True, use_container_width=True,
                    height=min(300, 38 + len(df_display) * 35),
                )


# ────────────────────────────────────────────────────────
# RENDER
# ────────────────────────────────────────────────────────

def render():
    reappros_df = _load_reappros_from_mongo()
    no_reappros = reappros_df.empty

    # Plannings (pour la génération auto inventaire)
    plannings_mongo: dict = {}
    try:
        from mongo_storage import load_plannings_from_mongo

        @st.cache_data(show_spinner=False, ttl=300)
        def _get_plannings_cr():
            return load_plannings_from_mongo()

        plannings_mongo, _ = _get_plannings_cr()
    except Exception:
        pass

    if no_reappros:
        st.warning("⚠️ Aucune répartition des zones en base. Importez le fichier ci-dessous.")

    # ── Sélecteurs ───────────────────────────────────────
    col_zone, col_date, _ = st.columns([2, 2, 4])
    with col_zone:
        zone = st.selectbox("Zone", ZONES, key="cr_zone", disabled=no_reappros)

    # Vider le contenu DA en cache si la zone a changé depuis le dernier rendu
    if st.session_state.get("_cr_last_zone") != zone:
        for k in list(st.session_state.keys()):
            if k.startswith("cr_da_content_"):
                del st.session_state[k]
        st.session_state["_cr_last_zone"] = zone
    with col_date:
        date_rapport = st.date_input(
            "Date du rapport",
            value=datetime.date.today(),
            key="cr_date",
        )

    # ── Objet ────────────────────────────────────────────
    objet_default = f"COMPTE RENDU {zone}"
    objet = st.text_input("📧 Objet du mail", value=objet_default, key=f"cr_objet_{zone}")

    st.divider()

    # ── Sections modulables ───────────────────────────────
    st.markdown("#### Sections à inclure")

    # --- Section DA (auto) ---
    col_chk_da, _ = st.columns([3, 7])
    with col_chk_da:
        include_da = st.checkbox("Distributeur Automatique (DA)", value=True, key="cr_chk_da")

    da_content = ""
    if include_da:
        if not no_reappros:
            with st.spinner("Chargement des incidents DA..."):
                incidents = _load_incidents_for_zone(zone, reappros_df)
            da_content = _build_da_content(incidents)
        else:
            da_content = "Toutes les salles ont été traitées dans les groupes."
        da_content = st.text_area(
            "Contenu DA",
            value=da_content,
            height=150,
            key=f"cr_da_content_{zone}",
            label_visibility="collapsed",
        )

    st.markdown("---")

    # --- Sections libres standard ---
    sections_standard = [
        ("Livraisons / Fournisseurs", True),
        ("Tournées", True),
        ("Inventaire", False),
    ]

    section_contents: dict[str, str] = {}
    for titre, default_checked in sections_standard:
        col_chk, _ = st.columns([3, 7])
        with col_chk:
            checked = st.checkbox(titre, value=default_checked, key=f"cr_chk_{titre}")

        if checked:
            if titre == "Inventaire":
                # ── Source 1 : inventaires_semaine (page Inventaires) ───────
                semaines_dispo  = list_inventaires_semaines()
                bilan_semaines  = list_bilan_semaines()

                if semaines_dispo:
                    st.caption("**Depuis la page Inventaires :**")
                    sem_labels = [
                        f"S{d['iso_week']} {d['iso_year']}"
                        + (f"  (sauvé le {d['saved_at'][:10]})" if d.get("saved_at") else "")
                        for d in semaines_dispo
                    ]
                    col_sel, col_vend, col_gen_inv = st.columns([3, 2, 2])
                    with col_sel:
                        sel_idx = st.selectbox(
                            "Semaine inv", range(len(sem_labels)),
                            format_func=lambda i: sem_labels[i],
                            key=f"cr_inv_sem_{zone}",
                            label_visibility="collapsed",
                        )
                    with col_vend:
                        include_vend = st.checkbox(
                            "Inclure vendredi",
                            value=False,
                            key=f"cr_inv_vend_{zone}",
                        )
                    with col_gen_inv:
                        if st.button("🔄 Générer", key=f"cr_inv_gen_{zone}",
                                     use_container_width=True):
                            with st.spinner("Chargement…"):
                                sel_doc      = semaines_dispo[sel_idx]
                                inv_doc      = load_inventaires_semaine(sel_doc["iso_year"], sel_doc["iso_week"])
                                done_records = inv_doc.get("done", [])
                                generated    = _build_inventaire_cr_text(
                                    done_records, plannings_mongo, reappros_df, zone,
                                    include_vendredi=include_vend,
                                )
                            st.session_state[f"cr_inv_text_{zone}"] = generated
                            st.session_state[f"cr_txt_{titre}_{zone}"] = generated
                            st.rerun()

                # ── Source 2 : bilan_semaine (page CR) ──────────────────────
                if bilan_semaines:
                    if semaines_dispo:
                        st.markdown(
                            "<div style='margin:6px 0;text-align:center;color:#aaa;font-size:0.85rem'>— ou depuis le bilan semaine —</div>",
                            unsafe_allow_html=True,
                        )
                    else:
                        st.caption("**Depuis le bilan semaine :**")

                    bilan_labels = [d.get("label") or f"S{d['iso_week']} {d['iso_year']}" for d in bilan_semaines]
                    col_bsel, col_bvend, col_bgen = st.columns([3, 2, 2])
                    with col_bsel:
                        bsel_idx = st.selectbox(
                            "Semaine bilan", range(len(bilan_labels)),
                            format_func=lambda i: bilan_labels[i],
                            key=f"cr_bilan_inv_sem_{zone}",
                            label_visibility="collapsed",
                        )
                    with col_bvend:
                        binclude_vend = st.checkbox(
                            "Inclure vendredi",
                            value=False,
                            key=f"cr_bilan_inv_vend_{zone}",
                        )
                    with col_bgen:
                        if st.button("🔄 Générer", key=f"cr_bilan_inv_gen_{zone}",
                                     use_container_width=True):
                            with st.spinner("Chargement…"):
                                bsel_doc   = bilan_semaines[bsel_idx]
                                bilan_doc  = load_bilan_semaine(bsel_doc["iso_year"], bsel_doc["iso_week"])
                                generated  = _build_inventaire_cr_text_from_bilan(
                                    bilan_doc.get("rows", []),
                                    reappros_df, zone,
                                    include_vendredi=binclude_vend,
                                )
                            st.session_state[f"cr_inv_text_{zone}"] = generated
                            st.session_state[f"cr_txt_{titre}_{zone}"] = generated
                            st.rerun()

                if not semaines_dispo and not bilan_semaines:
                    st.caption(
                        "⚠️ Aucune donnée sauvegardée. "
                        "Utilisez **💾 Sauvegarder en BDD** depuis la page Inventaires ou le Bilan semaine ci-dessous."
                    )

                inv_default = st.session_state.get(f"cr_inv_text_{zone}",
                                                    SECTION_DEFAULTS.get("Inventaire", ""))
                section_contents[titre] = st.text_area(
                    titre,
                    value=inv_default,
                    height=220,
                    key=f"cr_txt_{titre}_{zone}",
                    label_visibility="collapsed",
                )
            else:
                default_txt = SECTION_DEFAULTS.get(titre, "")
                section_contents[titre] = st.text_area(
                    titre,
                    value=default_txt,
                    height=120,
                    key=f"cr_txt_{titre}",
                    label_visibility="collapsed",
                )
        st.markdown("---")

    # --- Section "Autre" optionnelle ---
    col_chk_autre, _ = st.columns([3, 7])
    with col_chk_autre:
        include_autre = st.checkbox("Autre section", value=False, key="cr_chk_autre")
    autre_titre = ""
    autre_content = ""
    if include_autre:
        autre_titre = st.text_input(
            "Titre de la section",
            placeholder="Ex : Formations, RH, Matériel...",
            key="cr_autre_titre",
        )
        autre_content = st.text_area(
            "Contenu",
            height=120,
            key="cr_autre_txt",
            label_visibility="collapsed",
        )

    st.divider()

    # ── Génération ────────────────────────────────────────
    col_gen, _ = st.columns([2, 6])
    with col_gen:
        generer = st.button(
            "📋 Générer le mail",
            type="primary",
            use_container_width=True,
            key="cr_generer",
        )

    if generer:
        # Construire la liste ordonnée des sections cochées
        sections: list[tuple[str, str]] = []
        if include_da:
            sections.append(("Distributeur Automatique (DA)", da_content))
        for titre in ["Livraisons / Fournisseurs", "Tournées", "Inventaire"]:
            if titre in section_contents:
                sections.append((titre, section_contents[titre]))
        if include_autre and autre_titre.strip():
            sections.append((autre_titre.strip(), autre_content))

        mail_text = _generate_mail(zone, sections)
        st.session_state["cr_mail_result"] = mail_text
        st.session_state["cr_mail_objet"]  = objet

    # ── Affichage du résultat ─────────────────────────────
    if "cr_mail_result" in st.session_state:
        st.markdown("### 📬 Mail généré")
        st.markdown(
            f"**Objet :** `{st.session_state.get('cr_mail_objet', objet)}`"
        )
        st.text_area(
            "Contenu du mail",
            value=st.session_state["cr_mail_result"],
            height=400,
            key="cr_mail_display",
            label_visibility="collapsed",
        )
        # Bouton copier via st.code (sélectionnable facilement)
        with st.expander("📋 Version copiable (sélectionner tout avec Ctrl+A)"):
            st.code(st.session_state["cr_mail_result"], language=None)

    # ════════════════════════════════════════════════════
    # IMPORT RÉAPPROS (bas de page)
    # ════════════════════════════════════════════════════
    st.divider()
    st.markdown("### ⚙️ Mise à jour de la répartition des zones")
    st.caption("Importer le fichier `Reappro Guide.xlsx` pour mettre à jour les zones et réappros.")

    uploaded = st.file_uploader(
        "Reappro Guide.xlsx",
        type=["xlsx"],
        key="cr_reappro_uploader",
        label_visibility="collapsed",
    )

    if uploaded:
        col_imp, _ = st.columns([2, 6])
        with col_imp:
            if st.button(
                "📥 Mettre à jour la répartition",
                type="primary",
                use_container_width=True,
                key="cr_import_btn",
            ):
                with st.spinner("Import en cours..."):
                    try:
                        nb, zones_det = _import_reappros(uploaded.read())
                        st.success(
                            f"✅ **{nb}** réappros importés — "
                            f"zones détectées : {', '.join(zones_det)}"
                        )
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ Erreur lors de l'import : {e}")

    # ── Tableau récap réappros en base ───────────────────
    if not reappros_df.empty:
        with st.expander(f"👁️ Voir les {len(reappros_df)} réappros en base"):
            st.dataframe(
                reappros_df[["code", "prenom", "zone_geo", "zone", "responsable"]].rename(columns={
                    "code": "Code", "prenom": "Prénom",
                    "zone_geo": "Zone Géo", "zone": "Zone", "responsable": "Responsable",
                }),
                use_container_width=True,
                hide_index=True,
            )

    _section_bilan_cr(plannings_mongo, reappros_df)
