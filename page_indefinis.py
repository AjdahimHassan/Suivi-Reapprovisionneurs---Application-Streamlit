import streamlit as st
import pandas as pd
import io
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from planogrammes_storage import load_produits
from mongo_storage import load_plannings_from_mongo

# ─────────────────────────────────────────────────────────────────────────────
# ONGLET 1 — INDÉFINIS
# ─────────────────────────────────────────────────────────────────────────────

def _detect_sep(raw: str) -> str:
    first_line = raw.split("\n")[0]
    return ";" if first_line.count(";") >= first_line.count(",") else ","


def parse_ventes(file_bytes: bytes) -> pd.DataFrame:
    for enc in ("utf-8", "latin-1", "cp1252"):
        try:
            raw = file_bytes.decode(enc)
            break
        except Exception:
            continue
    sep = _detect_sep(raw)
    df = pd.read_csv(io.StringIO(raw), sep=sep, quotechar='"', dtype=str)
    df.columns = [c.strip().replace('"', "") for c in df.columns]

    rename = {}
    for col in df.columns:
        cu = col.upper()
        if "CODE" in cu and "PRODUIT" in cu:
            rename[col] = "CODE_PRODUIT"
        elif cu in ("PU", "PRIX UNITAIRE", "PRIX_UNITAIRE", "PRIXUNITAIRE"):
            rename[col] = "PU"
        elif "LDP" in cu or "LIGNE" in cu:
            rename[col] = "LDP"
        elif "MACHINE" in cu:
            rename[col] = "CODE_MACHINE"
        elif "CLIENT" in cu and "CODE" in cu:
            rename[col] = "CODE_CLIENT"
    df = df.rename(columns=rename)

    if "CODE_PRODUIT" not in df.columns:
        raise ValueError("Colonne CODE_PRODUIT introuvable dans le fichier de ventes.")

    if "PU" in df.columns:
        df["PU"] = pd.to_numeric(
            df["PU"].astype(str).str.replace(",", ".").str.strip(), errors="coerce"
        )

    if "LDP" in df.columns:
        df["LDP"] = pd.to_numeric(
            df["LDP"].astype(str).str.replace(",", ".").str.strip(), errors="coerce"
        )

    df["CODE_PRODUIT"] = df["CODE_PRODUIT"].astype(str).str.strip().str.upper()
    return df


def parse_planno(file_bytes: bytes) -> pd.DataFrame:
    for enc in ("utf-8", "latin-1", "cp1252"):
        try:
            raw = file_bytes.decode(enc)
            break
        except Exception:
            continue
    sep = _detect_sep(raw)
    lines = raw.splitlines()

    header_idx = 0
    for i, line in enumerate(lines):
        if line.strip().lower().startswith("code") or line.strip().lower().startswith('"code'):
            header_idx = i
            break

    body = "\n".join(lines[header_idx:])
    df = pd.read_csv(io.StringIO(body), sep=sep, quotechar='"', dtype=str)
    df.columns = [c.strip().replace('"', "") for c in df.columns]

    rename = {}
    for col in df.columns:
        cu = col.upper()
        if "CODE" in cu and "PRODUIT" in cu:
            rename[col] = "CODE_PRODUIT"
        elif cu == "CODE":
            rename[col] = "CODE_PRODUIT"
        elif "LIBELLE" in cu or "LIBELLÉ" in cu or "NOM" in cu:
            rename[col] = "LIBELLE"
        elif "LDP" in cu or ("LIGNE" in cu and "PRIX" in cu):
            rename[col] = "LDP"
        elif cu in ("PU", "PRIX UNITAIRE", "PRIX_UNITAIRE"):
            rename[col] = "PU"
        elif "NIV" in cu and "HAUT" in cu:
            rename[col] = "NIV_HAUT"
        elif "UNITE" in cu or "UNITÉ" in cu:
            rename[col] = "UNITE"
    df = df.rename(columns=rename)

    if "CODE_PRODUIT" not in df.columns:
        raise ValueError("Colonne CODE_PRODUIT introuvable dans le planogramme.")

    if "PU" in df.columns:
        df["PU"] = pd.to_numeric(
            df["PU"].astype(str).str.replace(",", ".").str.strip(), errors="coerce"
        )
    if "LDP" in df.columns:
        df["LDP"] = pd.to_numeric(
            df["LDP"].astype(str).str.replace(",", ".").str.strip(), errors="coerce"
        )

    df["CODE_PRODUIT"] = df["CODE_PRODUIT"].astype(str).str.strip().str.upper()
    df = df[df["CODE_PRODUIT"].notna() & (df["CODE_PRODUIT"] != "") & (df["CODE_PRODUIT"] != "NAN")]
    return df


def analyser_indefinis(df_ventes: pd.DataFrame, df_planno: pd.DataFrame) -> list:
    indefinis = df_ventes[df_ventes["CODE_PRODUIT"] == "INDEFINI"].copy()
    if indefinis.empty or "PU" not in indefinis.columns:
        return []

    codes_vendus = set(
        df_ventes[df_ventes["CODE_PRODUIT"] != "INDEFINI"]["CODE_PRODUIT"].unique()
    )

    resultats = []
    prix_traites = set()

    for _, row in indefinis.iterrows():
        prix = row.get("PU")
        if pd.isna(prix) or prix in prix_traites:
            continue
        prix_traites.add(prix)

        if "PU" in df_planno.columns:
            meme_prix = df_planno[abs(df_planno["PU"] - prix) < 0.01].copy()
        else:
            meme_prix = pd.DataFrame()

        suspects = meme_prix[~meme_prix["CODE_PRODUIT"].isin(codes_vendus)]
        autres_vendus = meme_prix[meme_prix["CODE_PRODUIT"].isin(codes_vendus)]

        resultats.append({
            "prix": prix,
            "indefini_rows": indefinis[abs(indefinis["PU"] - prix) < 0.01],
            "planno_meme_prix": meme_prix,
            "suspects": suspects,
            "autres_vendus": autres_vendus,
        })

    return resultats


# ─────────────────────────────────────────────────────────────────────────────
# ONGLET 2 — CONTRÔLE DES PRIX
# ─────────────────────────────────────────────────────────────────────────────

_TOLERANCE_HT = 0.05


def parse_ventes_prix(file_bytes: bytes) -> pd.DataFrame:
    for enc in ("utf-8", "latin-1", "cp1252"):
        try:
            raw = file_bytes.decode(enc)
            break
        except Exception:
            continue

    df = pd.read_csv(io.StringIO(raw), sep=";", quotechar='"', dtype=str)
    df.columns = [c.strip().replace('"', "") for c in df.columns]

    # Filtre "Audit Telemetrie" si la colonne existe
    col_type = next(
        (c for c in df.columns if "type" in c.lower() and "che" in c.lower()), None
    )
    if col_type:
        df = df[df[col_type].astype(str).str.strip().str.lower() == "audit telemetrie"]

    col_map = {
        "Code DA":          "Machine",
        "Stock Destination": "Code_client",
        "Nom client":       "Salle",
        "Code produit":     "Code",
        "Libellé produit":  "Produit",
        "Quantité":         "Quantité",
        "Montant HT":       "Montant_HT",
        "Date":             "Date",
    }
    df = df.rename(columns={k: v for k, v in col_map.items() if k in df.columns})

    for num_col in ("Quantité", "Montant_HT"):
        if num_col in df.columns:
            df[num_col] = pd.to_numeric(
                df[num_col].astype(str).str.replace(",", ".").str.strip(), errors="coerce"
            )

    df = df[df["Quantité"].notna() & (df["Quantité"] > 0)]
    df = df[df["Montant_HT"].notna()]

    df["PU_HT"] = (df["Montant_HT"] / df["Quantité"]).round(3)

    cols = [c for c in ("Date", "Machine", "Code_client", "Salle", "Code", "Produit", "Quantité", "Montant_HT", "PU_HT") if c in df.columns]
    return df[cols].reset_index(drop=True)


def analyser_prix(df: pd.DataFrame, produits: list, tolerance: float = _TOLERANCE_HT):
    # Construit {code: [prix1, prix2, ...]} pour gérer les produits
    # qui ont plusieurs prix valides selon la machine (ex: Red Bull 2,84 € ou 3,03 €)
    lib = {}
    for p in produits:
        code = str(p.get("code", "")).strip().upper()
        prix = p.get("prix_ht", None)
        if code and prix:
            try:
                prix_float = round(float(str(prix).replace(",", ".")), 3)
                lib.setdefault(code, [])
                if prix_float not in lib[code]:
                    lib[code].append(prix_float)
            except (ValueError, TypeError):
                pass

    anomalies = []
    non_ref = set()

    for _, row in df.iterrows():
        code = str(row.get("Code", "")).strip().upper()
        pu = row.get("PU_HT")
        if pd.isna(pu):
            continue
        if code not in lib:
            non_ref.add(code)
            continue
        prix_valides = lib[code]
        # OK si le prix réel est proche d'AU MOINS UN prix valide
        ecarts = [round(pu - p, 3) for p in prix_valides]
        ecart_min = min(ecarts, key=abs)
        if abs(ecart_min) > tolerance:
            # Affiche le prix valide le plus proche comme référence
            attendu = prix_valides[ecarts.index(ecart_min)]
            anomalies.append({
                "Date": row.get("Date", ""),
                "Machine": row.get("Machine", ""),
                "Code_client": str(row.get("Code_client", "")).strip().upper(),
                "Salle": row.get("Salle", ""),
                "Code": code,
                "Produit": row.get("Produit", ""),
                "Prix attendu (HT)": attendu,
                "Prix réel (HT)": pu,
                "Écart (€)": ecart_min,
            })

    anomalies_df = pd.DataFrame(anomalies)
    if not anomalies_df.empty:
        anomalies_df = anomalies_df.sort_values(
            "Écart (€)", key=abs, ascending=False
        ).reset_index(drop=True)

    return anomalies_df, non_ref


# ─────────────────────────────────────────────────────────────────────────────
# PLANNO AUDIT (configuration machine ligne par ligne)
# ─────────────────────────────────────────────────────────────────────────────

def parse_planno_audit(file_bytes: bytes) -> pd.DataFrame:
    """
    Parse le fichier CSV d'audit planno (Audit IUC180).
    Colonnes attendues : SOURCE AUDIT, CODE MACHINE, CODE CLIENT,
                         MODELE MACHINE, CODE PRODUIT, PU, LDP, PAYMENT
    """
    for enc in ("utf-8", "latin-1", "cp1252"):
        try:
            raw = file_bytes.decode(enc)
            break
        except Exception:
            continue

    df = pd.read_csv(io.StringIO(raw), sep=";", quotechar='"', dtype=str)
    df.columns = [c.strip().replace('"', "") for c in df.columns]

    # Renommage normalisé
    rename = {
        "SOURCE AUDIT":   "SOURCE_AUDIT",
        "CODE MACHINE":   "CODE_MACHINE",
        "CODE CLIENT":    "CODE_CLIENT",
        "MODELE MACHINE": "MODELE",
        "CODE PRODUIT":   "CODE_PRODUIT",
        "PU":             "PU",
        "LDP":            "LDP",
        "PAYMENT":        "PAYMENT",
    }
    df = df.rename(columns={k: v for k, v in rename.items() if k in df.columns})

    if "PU" in df.columns:
        df["PU"] = pd.to_numeric(
            df["PU"].astype(str).str.replace(",", ".").str.strip(), errors="coerce"
        )
    if "LDP" in df.columns:
        df["LDP"] = pd.to_numeric(
            df["LDP"].astype(str).str.strip(), errors="coerce"
        )
    if "CODE_MACHINE" in df.columns:
        df["CODE_MACHINE"] = df["CODE_MACHINE"].astype(str).str.strip().str.upper()
    if "CODE_PRODUIT" in df.columns:
        df["CODE_PRODUIT"] = df["CODE_PRODUIT"].astype(str).str.strip().str.upper()

    return df.reset_index(drop=True)


# ─────────────────────────────────────────────────────────────────────────────
# EXPORT PAR RÉAPPRO
# ─────────────────────────────────────────────────────────────────────────────

def build_reappro_mapping(plannings: dict) -> dict:
    """
    Construit un dictionnaire {code_client_upper: reappro} depuis les plannings MongoDB.
    Le code client est la partie avant " - " dans le champ client du planning.
    Ex: "FPPY01F - FP TARBES" → code = "FPPY01F"
    """
    mapping = {}
    for reappro, jours in plannings.items():
        for salles in jours.values():
            for client, _ in salles:
                code = client.split(" - ")[0].strip().upper()
                if code:
                    mapping[code] = reappro
    return mapping


def generer_export_reappros(anomalies_df: pd.DataFrame, reappro_map: dict, filtrer: list = None) -> bytes:
    """
    Génère un fichier Excel avec une feuille Résumé + une feuille par réappro.
    Si `filtrer` est fourni (liste de codes réappro), seuls ceux-ci sont inclus.
    """
    """
    Génère un fichier Excel avec une feuille Résumé + une feuille par réappro.
    Chaque feuille liste les salles problématiques avec le détail des produits.
    """
    # Styles communs
    HDR_FILL   = PatternFill("solid", fgColor="1F4E79")
    HDR_FONT   = Font(bold=True, color="FFFFFF")
    SALLE_FILL = PatternFill("solid", fgColor="D9E1F2")
    SALLE_FONT = Font(bold=True, color="1F4E79")
    FILL_RED   = PatternFill("solid", fgColor="F5C6CB")
    FILL_ORA   = PatternFill("solid", fgColor="FFD8A8")
    FILL_YEL   = PatternFill("solid", fgColor="FFF3CD")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _ecart_fill(ecart):
        a = abs(ecart)
        if a >= 0.50: return FILL_RED
        if a >= 0.20: return FILL_ORA
        return FILL_YEL

    def _apply_hdr(ws, headers, row=1):
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.fill = HDR_FILL
            c.font = HDR_FONT
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border

    # Assigner chaque anomalie à un réappro
    df = anomalies_df.copy()
    df["Réappro"] = df["Code_client"].apply(
        lambda x: reappro_map.get(str(x).strip().upper(), "Non assigné")
    )
    # Filtrer sur la sélection si fournie
    if filtrer:
        df = df[df["Réappro"].isin(filtrer)]

    wb = openpyxl.Workbook()

    # ── Feuille Résumé ──────────────────────────────────────────────────────
    ws_res = wb.active
    ws_res.title = "Résumé"
    hdrs = ["Réappro", "Salles KO", "Nb anomalies"]
    _apply_hdr(ws_res, hdrs)

    resume = (
        df.groupby("Réappro")
        .agg(Salles_KO=("Salle", "nunique"), Nb_anomalies=("Code", "count"))
        .reset_index()
        .sort_values("Nb_anomalies", ascending=False)
    )
    for r_idx, row in enumerate(resume.itertuples(index=False), start=2):
        ws_res.cell(r_idx, 1, row.Réappro)
        ws_res.cell(r_idx, 2, row.Salles_KO)
        ws_res.cell(r_idx, 3, row.Nb_anomalies)
        for col in range(1, 4):
            ws_res.cell(r_idx, col).border = border

    ws_res.column_dimensions["A"].width = 18
    ws_res.column_dimensions["B"].width = 14
    ws_res.column_dimensions["C"].width = 16

    # ── Une feuille par réappro ──────────────────────────────────────────────
    COLS_XLS = ["Salle", "Machine", "Code", "Produit", "Prix attendu (HT)", "Prix réel (HT)", "Écart (€)"]
    COL_W    = [35, 18, 22, 40, 20, 18, 14]

    for reappro in sorted(df["Réappro"].unique()):
        sub = df[df["Réappro"] == reappro].sort_values(["Salle", "Produit"])
        ws = wb.create_sheet(title=reappro[:31])  # Excel limite à 31 caractères
        _apply_hdr(ws, COLS_XLS)

        current_row = 2
        current_salle = None

        for _, r in sub.iterrows():
            salle = r.get("Salle", "")
            # Ligne de séparation entre salles
            if salle != current_salle:
                current_salle = salle
                for col in range(1, 8):
                    c = ws.cell(current_row, col)
                    c.fill = SALLE_FILL
                    c.font = SALLE_FONT
                    c.border = border
                ws.cell(current_row, 1, salle)
                ws.merge_cells(
                    start_row=current_row, start_column=1,
                    end_row=current_row, end_column=7
                )
                current_row += 1

            vals = [
                salle,
                r.get("Machine", ""),
                r.get("Code", ""),
                r.get("Produit", ""),
                r.get("Prix attendu (HT)", ""),
                r.get("Prix réel (HT)", ""),
                r.get("Écart (€)", ""),
            ]
            fill = _ecart_fill(r.get("Écart (€)", 0))
            for col, val in enumerate(vals, 1):
                c = ws.cell(current_row, col, val)
                c.fill = fill
                c.border = border
                if col in (5, 6):
                    c.number_format = "0.000 €"
                    c.alignment = Alignment(horizontal="center")
                elif col == 7:
                    c.number_format = '+0.000 €;-0.000 €;0.000 €'
                    c.alignment = Alignment(horizontal="center")
            current_row += 1

        for i, (col_letter, w) in enumerate(zip("ABCDEFG", COL_W), 1):
            ws.column_dimensions[col_letter].width = w

    # Retourner les bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────────────────────────────────────
# EXPORT MACHINE (fichier de ventes par machine avec surlignage)
# ─────────────────────────────────────────────────────────────────────────────

def _generer_export_machine(
    df: pd.DataFrame,
    prix_ko_ttc: dict,
    machine: str,
    salle: str,
) -> bytes:
    """
    Génère un fichier Excel soigné pour les ventes d'une machine.
    Les lignes dont le PU correspond à un prix KO (TTC) sont surlignées en rouge.
    Retourne les bytes prêts pour st.download_button.
    """
    # ── Styles ──────────────────────────────────────────────────────────────
    TITLE_FILL = PatternFill("solid", fgColor="1F4E79")
    TITLE_FONT = Font(bold=True, color="FFFFFF", size=13)
    HDR_FILL   = PatternFill("solid", fgColor="2E75B6")
    HDR_FONT   = Font(bold=True, color="FFFFFF", size=11)
    KO_FILL    = PatternFill("solid", fgColor="F5C6CB")
    KO_FONT    = Font(color="7B1A1A")
    YEL_FILL   = PatternFill("solid", fgColor="FFF3CD")
    YEL_FONT   = Font(color="664D00")
    ALT_FILL   = PatternFill("solid", fgColor="F2F7FC")
    thin       = Side(style="thin", color="B8CCE4")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    CENTER     = Alignment(horizontal="center", vertical="center")
    LEFT       = Alignment(horizontal="left", vertical="center")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = machine[:31]

    cols = list(df.columns)
    nb_cols = len(cols)

    # ── LDP en doublon ───────────────────────────────────────────────────────
    ldp_doublons = set()
    if "LDP" in df.columns:
        ldp_counts = df["LDP"].value_counts()
        ldp_doublons = set(ldp_counts[ldp_counts > 1].index)

    # ── Ligne titre ──────────────────────────────────────────────────────────
    titre = f"Machine : {machine}   |   {salle}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=nb_cols)
    tc = ws.cell(1, 1, titre)
    tc.fill  = TITLE_FILL
    tc.font  = TITLE_FONT
    tc.alignment = CENTER
    ws.row_dimensions[1].height = 28

    # ── Ligne en-tête colonnes ────────────────────────────────────────────────
    for col_idx, col_name in enumerate(cols, 1):
        c = ws.cell(2, col_idx, col_name)
        c.fill      = HDR_FILL
        c.font      = HDR_FONT
        c.alignment = CENTER
        c.border    = border
    ws.row_dimensions[2].height = 22

    # ── Données ───────────────────────────────────────────────────────────────
    def _is_ko(code_produit, pu_val):
        code = str(code_produit).upper()
        if code not in prix_ko_ttc or pu_val is None or pd.isna(pu_val):
            return False
        return any(abs(float(pu_val) - p) < 0.06 for p in prix_ko_ttc[code])

    for row_idx, (_, row) in enumerate(df.iterrows(), start=3):
        code_val  = row.get("CODE_PRODUIT", "")
        pu_val    = row.get("PU", None)
        ldp_val   = row.get("LDP")
        is_red    = _is_ko(code_val, pu_val) or str(code_val).upper() == "INDEFINI"
        is_yellow = (not is_red) and (ldp_val in ldp_doublons)

        if is_red:
            row_fill = KO_FILL
            row_font = KO_FONT
        elif is_yellow:
            row_fill = YEL_FILL
            row_font = YEL_FONT
        else:
            row_fill = ALT_FILL if row_idx % 2 == 0 else None
            row_font = Font(size=10)

        for col_idx, col_name in enumerate(cols, 1):
            val = row.get(col_name, "")
            c = ws.cell(row_idx, col_idx, val)
            c.border = border
            if row_fill:
                c.fill = row_fill
            if row_font:
                c.font = row_font

            # Formats par colonne
            if col_name == "PU" and val is not None and not (isinstance(val, float) and pd.isna(val)):
                c.number_format = "0.00 €"
                c.alignment = CENTER
            elif col_name == "Prix attendu (TTC)" and val is not None and val != "":
                c.alignment = CENTER
            elif col_name == "LDP":
                c.number_format = "0"
                c.alignment = CENTER
            else:
                c.alignment = LEFT

        ws.row_dimensions[row_idx].height = 18

    # ── Largeurs colonnes ────────────────────────────────────────────────────
    col_widths = {
        "CODE_PRODUIT":    28,
        "LDP":             8,
        "PU":              14,
        "Prix attendu (TTC)": 20,
    }
    for col_idx, col_name in enumerate(cols, 1):
        letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[letter].width = col_widths.get(col_name, 18)

    # ── Figer la ligne d'en-tête ─────────────────────────────────────────────
    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────────────────────────────────────
# EXPORT PNG MACHINE
# ─────────────────────────────────────────────────────────────────────────────

def _generer_png_machine(
    df: pd.DataFrame,
    prix_ko_ttc: dict,
    machine: str,
    salle: str,
) -> bytes:
    """
    Génère une image PNG du tableau de ventes d'une machine.
    Lignes KO = fond rouge, header = bleu foncé.
    """
    # Couleurs
    C_HEADER  = "#1F4E79"
    C_HDR_TXT = "white"
    C_TITLE   = "#1F4E79"
    C_KO_BG   = "#F5C6CB"
    C_KO_TXT  = "#7B1A1A"
    C_ALT     = "#EBF2FA"
    C_WHITE   = "white"
    C_BORDER  = "#B8CCE4"

    cols = list(df.columns)
    n_rows = len(df)
    n_cols = len(cols)

    # Largeurs relatives par colonne
    col_w_map = {
        "CODE_PRODUIT":    3.5,
        "LDP":             1.0,
        "PU":              1.6,
        "Prix attendu (TTC)": 2.2,
    }
    col_widths = [col_w_map.get(c, 2.0) for c in cols]
    total_w = sum(col_widths)

    # Dimensions figure
    row_h    = 0.42   # hauteur par ligne de données (inch)
    hdr_h    = 0.50   # hauteur header
    title_h  = 0.55   # hauteur titre
    fig_h    = title_h + hdr_h + n_rows * row_h + 0.2
    fig_w    = max(total_w * 1.1, 8)

    fig, ax = plt.subplots(figsize=(fig_w, fig_h))
    ax.axis("off")
    fig.patch.set_facecolor("white")

    # ── Titre ──────────────────────────────────────────────────────────────
    fig.text(
        0.5, 1 - (title_h / 2) / fig_h,
        f"Machine : {machine}   |   {salle}",
        ha="center", va="center",
        fontsize=13, fontweight="bold", color=C_HDR_TXT,
        bbox=dict(boxstyle="square,pad=0", facecolor=C_TITLE, edgecolor="none"),
        transform=fig.transFigure,
    )

    # Calcul des positions x normalisées
    margin = 0.02
    usable = 1 - 2 * margin
    x_positions = []
    cum = 0
    for w in col_widths:
        x_positions.append(margin + usable * cum / total_w)
        cum += w
    x_positions.append(margin + usable)  # bord droit

    def _row_y(row_idx):
        """Y du bas de la ligne (normalized)."""
        top = 1 - title_h / fig_h
        return top - (hdr_h + row_idx * row_h) / fig_h

    def _draw_cell(left, right, y_bottom, height, bg, text, txt_color, bold=False, fontsize=9.5, halign="center"):
        norm_h = height / fig_h
        rect = mpatches.FancyBboxPatch(
            (left, y_bottom), right - left, norm_h,
            boxstyle="square,pad=0",
            facecolor=bg, edgecolor=C_BORDER, linewidth=0.5,
            transform=fig.transFigure, clip_on=False,
        )
        fig.add_artist(rect)
        tx = (left + right) / 2 if halign == "center" else left + 0.008
        fig.text(
            tx, y_bottom + norm_h / 2,
            str(text) if text is not None and not (isinstance(text, float) and pd.isna(text)) else "",
            ha=halign, va="center",
            fontsize=fontsize, fontweight="bold" if bold else "normal",
            color=txt_color,
            transform=fig.transFigure, clip_on=False,
        )

    # ── Header colonnes ────────────────────────────────────────────────────
    top = 1 - title_h / fig_h
    for i, col_name in enumerate(cols):
        _draw_cell(
            x_positions[i], x_positions[i + 1],
            top - hdr_h / fig_h, hdr_h,
            C_HEADER, col_name, C_HDR_TXT,
            bold=True, fontsize=10,
        )

    # ── LDP en doublon ────────────────────────────────────────────────────
    ldp_doublons_png = set()
    if "LDP" in df.columns:
        ldp_counts = df["LDP"].value_counts()
        ldp_doublons_png = set(ldp_counts[ldp_counts > 1].index)

    # ── Couleur par ligne ─────────────────────────────────────────────────
    def _row_color(row):
        code = str(row.get("CODE_PRODUIT", "")).upper()
        pu   = row.get("PU", None)
        ldp  = row.get("LDP")
        if code == "INDEFINI":
            return "red"
        if code in prix_ko_ttc and pu is not None and not (isinstance(pu, float) and pd.isna(pu)):
            if any(abs(float(pu) - p) < 0.06 for p in prix_ko_ttc[code]):
                return "red"
        if ldp in ldp_doublons_png:
            return "yellow"
        return "normal"

    # ── Lignes de données ──────────────────────────────────────────────────
    for r_idx, (_, row) in enumerate(df.iterrows()):
        color  = _row_color(row)
        bg     = C_KO_BG if color == "red" else ("#FFF3CD" if color == "yellow" else (C_ALT if r_idx % 2 == 0 else C_WHITE))
        txt_c  = C_KO_TXT if color == "red" else ("#664D00" if color == "yellow" else "#1A1A1A")
        bold   = color == "red"
        y_bot  = _row_y(r_idx + 1)  # +1 pour passer l'header

        for i, col_name in enumerate(cols):
            val = row.get(col_name, "")
            # Formatage des valeurs numériques
            if col_name == "PU" and isinstance(val, float) and not pd.isna(val):
                val = f"{val:.2f} €"
            elif col_name == "LDP" and isinstance(val, float) and not pd.isna(val):
                val = str(int(val))

            halign = "left" if col_name == "CODE_PRODUIT" else "center"
            _draw_cell(
                x_positions[i], x_positions[i + 1],
                y_bot, row_h,
                bg, val, txt_c,
                bold=bold, fontsize=9, halign=halign,
            )

    plt.tight_layout(pad=0)
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────────────────────────────────────
# RENDU PRINCIPAL
# ─────────────────────────────────────────────────────────────────────────────

def render():
    st.title("Indéfinis & Contrôle des prix")

    tab1, tab2 = st.tabs(["🔍 Indéfinis", "💰 Contrôle des prix"])

    # ──────────────────────────────────────────────────────────────────────────
    # TAB 1 — INDÉFINIS
    # ──────────────────────────────────────────────────────────────────────────
    with tab1:
        st.markdown("### Détection des produits indéfinis")
        st.caption(
            "Importez le fichier de ventes (export machine) et le planogramme (CSV). "
            "L'outil détecte les lignes INDÉFINI et identifie les produits suspects "
            "au même prix dans le planogramme."
        )

        col_a, col_b = st.columns(2)
        with col_a:
            f_ventes = st.file_uploader(
                "Fichier de ventes (CSV machine)", type=["csv"], key="indef_ventes"
            )
        with col_b:
            f_planno = st.file_uploader(
                "Planogramme (CSV)", type=["csv"], key="indef_planno"
            )

        if not f_ventes and not f_planno:
            st.info("Importez les deux fichiers pour lancer l'analyse.")
        else:
            df_ventes = None
            if f_ventes:
                try:
                    df_ventes = parse_ventes(f_ventes.read())
                except Exception as e:
                    st.error(f"Erreur lecture ventes : {e}")

            df_planno = None
            if f_planno:
                try:
                    df_planno = parse_planno(f_planno.read())
                except Exception as e:
                    st.error(f"Erreur lecture planogramme : {e}")

            if df_ventes is not None and df_planno is not None:
                missing_v = [c for c in ("CODE_PRODUIT", "PU") if c not in df_ventes.columns]
                missing_p = [c for c in ("CODE_PRODUIT", "PU") if c not in df_planno.columns]
                if missing_v:
                    st.error(f"Colonnes manquantes dans ventes : {missing_v}")
                elif missing_p:
                    st.error(f"Colonnes manquantes dans planogramme : {missing_p}")
                else:
                    resultats = analyser_indefinis(df_ventes, df_planno)

                    nb_indefinis = len(df_ventes[df_ventes["CODE_PRODUIT"] == "INDEFINI"])
                    nb_suspects = sum(len(r["suspects"]) for r in resultats)
                    nb_planno = len(df_planno)

                    k1, k2, k3 = st.columns(3)
                    k1.metric("Lignes INDÉFINI", nb_indefinis)
                    k2.metric("Lignes suspectes", nb_suspects)
                    k3.metric("Lignes planno analysées", nb_planno)

                    st.divider()

                    if not resultats:
                        st.success("Aucune ligne INDÉFINI trouvée dans le fichier de ventes.")
                    else:
                        for r in resultats:
                            prix = r["prix"]
                            suspects = r["suspects"]
                            autres = r["autres_vendus"]

                            st.markdown(
                                f"<div style='background:#1a1010;border-left:4px solid #e74c3c;"
                                f"padding:12px 16px;border-radius:6px;margin-bottom:8px;'>"
                                f"<b style='color:#e74c3c'>INDÉFINI</b> &nbsp;|&nbsp; "
                                f"Prix : <b>{prix:.2f} €</b> &nbsp;|&nbsp; "
                                f"{len(suspects)} suspect(s) &nbsp;|&nbsp; {len(autres)} référence(s)"
                                f"</div>",
                                unsafe_allow_html=True,
                            )

                            c1, c2 = st.columns(2)
                            with c1:
                                st.markdown("**🔴 Suspects** — produits au même prix sans vente")
                                if suspects.empty:
                                    st.caption("Aucun suspect.")
                                else:
                                    cols_s = [c for c in ("CODE_PRODUIT", "LIBELLE", "LDP", "PU", "NIV_HAUT") if c in suspects.columns]
                                    st.dataframe(suspects[cols_s], use_container_width=True, hide_index=True)

                            with c2:
                                st.markdown("**🟢 Références** — produits au même prix ayant vendu")
                                if autres.empty:
                                    st.caption("Aucune référence trouvée.")
                                else:
                                    cols_r = [c for c in ("CODE_PRODUIT", "LIBELLE", "LDP", "PU") if c in autres.columns]
                                    st.dataframe(autres[cols_r], use_container_width=True, hide_index=True)

                            st.markdown("---")

                    with st.expander("Toutes les ventes"):
                        st.dataframe(df_ventes, use_container_width=True, hide_index=True)

                    with st.expander("Tout le planogramme"):
                        st.dataframe(df_planno, use_container_width=True, hide_index=True)

    # ──────────────────────────────────────────────────────────────────────────
    # TAB 2 — CONTRÔLE DES PRIX
    # ──────────────────────────────────────────────────────────────────────────
    with tab2:
        st.markdown("### Vérification des prix de vente HT")
        st.caption(
            "Importez l'export ERP (Audit Télémétrie, CSV séparé par ';'). "
            "Les prix réels (Montant HT ÷ Quantité) sont comparés aux prix HT "
            f"de la bibliothèque produits. Tolérance : ±{_TOLERANCE_HT} €."
        )

        f_erp = st.file_uploader(
            "Export ERP Audit Télémétrie (.csv)", type=["csv"], key="prix_erp"
        )

        # Réinitialiser les résultats si le fichier change
        if not f_erp:
            st.session_state.pop("prix_resultats", None)
            st.info("Importez le fichier ERP pour lancer l'analyse.")
        else:
            if st.button("🔍 Analyser", key="btn_analyser_prix"):
                try:
                    df_ventes_prix = parse_ventes_prix(f_erp.read())
                except Exception as e:
                    st.error(f"Erreur lecture fichier ERP : {e}")
                    df_ventes_prix = None

                if df_ventes_prix is not None and df_ventes_prix.empty:
                    st.warning("Aucune ligne exploitable dans le fichier.")
                    df_ventes_prix = None

                if df_ventes_prix is not None:
                    with st.spinner("Chargement de la bibliothèque produits…"):
                        produits = load_produits()

                    if not produits:
                        st.warning(
                            "La bibliothèque produits est vide. "
                            "Ajoutez des produits dans Planogrammes → Bibliothèque."
                        )
                    else:
                        anomalies_df, non_ref = analyser_prix(df_ventes_prix, produits)
                        with st.spinner("Chargement du planning…"):
                            plannings, errs = load_plannings_from_mongo()
                        reappro_map = build_reappro_mapping(plannings) if not errs else {}

                        # Construire lib_ttc : {code: [prix_ttc1, prix_ttc2, ...]}
                        # depuis le champ prix_ttc stocké en bibliothèque
                        lib_ttc = {}
                        for p in produits:
                            code = str(p.get("code", "")).strip().upper()
                            ttc_raw = p.get("prix_ttc", None)
                            if code and ttc_raw:
                                try:
                                    ttc = round(float(str(ttc_raw).replace(",", ".")), 2)
                                    lib_ttc.setdefault(code, [])
                                    if ttc not in lib_ttc[code]:
                                        lib_ttc[code].append(ttc)
                                except (ValueError, TypeError):
                                    pass

                        # Stocker dans session_state pour survivre aux reruns
                        st.session_state["prix_resultats"] = {
                            "anomalies_df": anomalies_df,
                            "non_ref": non_ref,
                            "nb_lignes": len(df_ventes_prix),
                            "reappro_map": reappro_map,
                            "planning_errs": errs,
                            "lib_ttc": lib_ttc,
                        }

            # Afficher les résultats depuis session_state (persiste entre reruns)
            if "prix_resultats" in st.session_state:
                res          = st.session_state["prix_resultats"]
                anomalies_df = res["anomalies_df"]
                non_ref      = res["non_ref"]
                nb_lignes    = res["nb_lignes"]
                reappro_map  = res["reappro_map"]
                errs         = res["planning_errs"]
                lib_ttc      = res.get("lib_ttc", {})

                nb_anomalies  = len(anomalies_df)
                nb_non_ref    = len(non_ref)
                nb_salles_ko  = anomalies_df["Salle"].nunique() if not anomalies_df.empty else 0
                pct_ok        = round((nb_lignes - nb_anomalies) / nb_lignes * 100, 1) if nb_lignes else 0.0

                k1, k2, k3, k4, k5 = st.columns(5)
                k1.metric("Lignes analysées", nb_lignes)
                k2.metric("Conformes", f"{pct_ok} %")
                k3.metric("Anomalies de prix", nb_anomalies)
                k4.metric("Salles concernées", nb_salles_ko)
                k5.metric("Codes non référencés", nb_non_ref)

                st.divider()

                if anomalies_df.empty:
                    st.success("Tous les produits référencés sont vendus au bon prix HT.")
                else:
                    st.markdown(
                        f"**{nb_anomalies} ligne(s) avec un écart supérieur à ±{_TOLERANCE_HT} €**"
                    )

                    def _color_row(row):
                        ecart = abs(row["Écart (€)"])
                        if ecart >= 0.50:
                            bg = "background-color: #f5c6cb; color: #7b1a1a;"
                        elif ecart >= 0.20:
                            bg = "background-color: #ffd8a8; color: #6b3a00;"
                        else:
                            bg = "background-color: #fff3cd; color: #664d00;"
                        return [bg] * len(row)

                    # ── Export par réappro ──────────────────────────────────
                    if errs:
                        st.warning(
                            "Planning indisponible, l'export par réappro est désactivé. "
                            f"Erreur : {list(errs.values())[0]}"
                        )
                    else:
                        df_tmp = anomalies_df.copy()
                        df_tmp["Réappro"] = df_tmp["Code_client"].apply(
                            lambda x: reappro_map.get(str(x).strip().upper(), "Non assigné")
                        )
                        tous_reappros = sorted(df_tmp["Réappro"].unique().tolist())

                        st.markdown("**Sélectionner les réappros à exporter :**")
                        col_sel, col_btn = st.columns([4, 1])
                        with col_sel:
                            selection = st.multiselect(
                                label="Réappros",
                                options=tous_reappros,
                                default=[],
                                placeholder="Ajouter un réappro…",
                                label_visibility="collapsed",
                                key="export_reappro_select",
                            )
                        with col_btn:
                            if st.button("Tout sélectionner", key="btn_tout_select"):
                                selection = tous_reappros

                        col_dl1, col_dl2 = st.columns(2)
                        with col_dl1:
                            if selection:
                                excel_sel = generer_export_reappros(anomalies_df, reappro_map, filtrer=selection)
                                fname = (
                                    f"controle_prix_{'_'.join(selection)}.xlsx"
                                    if len(selection) <= 3
                                    else "controle_prix_selection.xlsx"
                                )
                                st.download_button(
                                    label=f"📥 Exporter la sélection ({len(selection)} réappro(s))",
                                    data=excel_sel,
                                    file_name=fname,
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    key="dl_selection",
                                )
                            else:
                                st.button("📥 Exporter la sélection", disabled=True, key="dl_selection_off")
                        with col_dl2:
                            excel_all = generer_export_reappros(anomalies_df, reappro_map)
                            st.download_button(
                                label="📥 Exporter tous les réappros",
                                data=excel_all,
                                file_name="controle_prix_reappros.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="dl_all",
                            )

                    st.divider()

                    vue_machine, vue_complet = st.tabs(["🏭 Détail par machine", "📋 Détail complet"])

                    with vue_machine:
                        machines = anomalies_df["Machine"].unique() if "Machine" in anomalies_df.columns else []
                        for machine in sorted(machines):
                            sous_df = anomalies_df[anomalies_df["Machine"] == machine]
                            salle = sous_df["Salle"].iloc[0] if "Salle" in sous_df.columns and not sous_df.empty else ""
                            label = f"{machine} — {salle}" if salle else machine
                            ss_key = f"ventes_machine_{machine}"

                            with st.expander(f"**{label}** — {len(sous_df)} anomalie(s)", expanded=False):
                                # Tableau des anomalies ERP
                                styled_m = sous_df.style.apply(_color_row, axis=1).format(
                                    {
                                        "Prix attendu (HT)": "{:.3f} €",
                                        "Prix réel (HT)": "{:.3f} €",
                                        "Écart (€)": "{:+.3f} €",
                                    }
                                )
                                st.dataframe(styled_m, use_container_width=True, hide_index=True)

                                st.divider()

                                # ── Upload fichier ventes spécifique à cette machine ──
                                f_ventes_machine = st.file_uploader(
                                    f"📂 Importer le fichier de ventes pour {machine}",
                                    type=["csv"],
                                    key=f"upload_{machine}",
                                )
                                if f_ventes_machine:
                                    try:
                                        st.session_state[ss_key] = parse_planno_audit(f_ventes_machine.read())
                                    except Exception as e:
                                        st.error(f"Erreur lecture fichier : {e}")

                                if ss_key in st.session_state:
                                    df_ventes_m = st.session_state[ss_key]

                                    # Construire l'ensemble des prix TTC "mauvais" par produit
                                    # en convertissant Prix réel HT → TTC via les TVA courantes
                                    _TVA_RATES = [0.055, 0.10, 0.20]
                                    prix_ko_ttc = {}  # {code: set_of_bad_ttc_prices}
                                    for _, anom_row in sous_df.iterrows():
                                        code = str(anom_row["Code"]).upper()
                                        pu_ht = anom_row.get("Prix réel (HT)", None)
                                        if pu_ht is None:
                                            continue
                                        approx_ttc = {round(pu_ht * (1 + r), 2) for r in _TVA_RATES}
                                        prix_ko_ttc.setdefault(code, set()).update(approx_ttc)

                                    def _is_ligne_ko(code_produit, pu_val):
                                        code = str(code_produit).upper()
                                        if code not in prix_ko_ttc or pu_val is None:
                                            return False
                                        return any(
                                            abs(pu_val - p) < 0.06
                                            for p in prix_ko_ttc[code]
                                        )

                                    # Colonnes : CODE_PRODUIT | LDP | PU | Prix attendu (TTC)
                                    cols_base = [c for c in ("CODE_PRODUIT", "LDP", "PU") if c in df_ventes_m.columns]
                                    df_display = df_ventes_m[cols_base].copy().reset_index(drop=True)

                                    # LDP en doublon (même numéro de ligne qui apparaît 2+ fois)
                                    ldp_doublons = set()
                                    if "LDP" in df_display.columns:
                                        ldp_counts = df_display["LDP"].value_counts()
                                        ldp_doublons = set(ldp_counts[ldp_counts > 1].index)

                                    def _color_ventes_m(row):
                                        code = str(row.get("CODE_PRODUIT", "")).upper()
                                        ldp  = row.get("LDP")
                                        if code == "INDEFINI" or _is_ligne_ko(code, row.get("PU")):
                                            return ["background-color: #f5c6cb; color: #7b1a1a;"] * len(row)
                                        if ldp in ldp_doublons:
                                            return ["background-color: #fff3cd; color: #664d00;"] * len(row)
                                        return [""] * len(row)

                                    def _prix_attendu_ttc(code_produit, pu_val, machine_code, ldp_val=None):
                                        """
                                        Pour les lignes KO, retourne le prix TTC attendu
                                        depuis la bibliothèque produits, en tenant compte
                                        du type de machine (BF → prix bas, FP → prix haut).
                                        Retourne None pour les doublons LDP et les INDEFINI.
                                        """
                                        if ldp_val in ldp_doublons:
                                            return None
                                        code = str(code_produit).upper()
                                        if code == "INDEFINI" or not _is_ligne_ko(code, pu_val):
                                            return None
                                        ttc_list = lib_ttc.get(code, [])
                                        if not ttc_list:
                                            return None
                                        if len(ttc_list) == 1:
                                            return f"{ttc_list[0]:.2f} €"
                                        # Plusieurs prix valides → sélectionner selon le préfixe de la salle
                                        ttc_sorted = sorted(ttc_list)
                                        prefix = str(machine_code).upper().strip()
                                        if prefix.startswith("BF"):
                                            return f"{ttc_sorted[0]:.2f} €"
                                        if prefix.startswith("FP"):
                                            return f"{ttc_sorted[-1]:.2f} €"
                                        return " / ".join(f"{v:.2f} €" for v in ttc_sorted)

                                    df_display["Prix attendu (TTC)"] = df_display.apply(
                                        lambda row: _prix_attendu_ttc(
                                            row.get("CODE_PRODUIT", ""), row.get("PU", None),
                                            salle, row.get("LDP"),
                                        ),
                                        axis=1,
                                    )

                                    st.markdown("**📋 Ventes de la machine** — lignes surlignées = prix à corriger")
                                    fmt = {}
                                    if "PU" in df_display.columns:
                                        fmt["PU"] = "{:.2f} €"
                                    styled_v = df_display.style.apply(_color_ventes_m, axis=1).format(fmt)
                                    st.dataframe(styled_v, use_container_width=True, hide_index=True)

                                    # ── Résumé écrit des anomalies ───────────────
                                    lignes_resume = []

                                    # 1. Prix erronés (KO)
                                    ko_rows = df_display[df_display.apply(
                                        lambda r: _is_ligne_ko(r.get("CODE_PRODUIT", ""), r.get("PU")), axis=1
                                    )]
                                    if not ko_rows.empty:
                                        lignes_resume.append("🔴 Erreurs de prix détectées :")
                                        for _, kr in ko_rows.iterrows():
                                            code    = kr.get("CODE_PRODUIT", "")
                                            ldp     = kr.get("LDP")
                                            pu      = kr.get("PU")
                                            attendu = kr.get("Prix attendu (TTC)", "")
                                            ldp_str = f"LDP {int(ldp)}" if ldp is not None and not (isinstance(ldp, float) and pd.isna(ldp)) else "LDP ?"
                                            pu_str  = f"{pu:.2f} €" if pu is not None and not (isinstance(pu, float) and pd.isna(pu)) else "? €"
                                            attendu_str = f" → attendu : {attendu}" if attendu else ""
                                            lignes_resume.append(f"  - {code} ({ldp_str}) : prix constaté {pu_str}{attendu_str}")

                                    # 2. Produits INDEFINI
                                    indef_rows = df_display[
                                        df_display["CODE_PRODUIT"].astype(str).str.upper() == "INDEFINI"
                                    ] if "CODE_PRODUIT" in df_display.columns else pd.DataFrame()
                                    if not indef_rows.empty:
                                        lignes_resume.append("🔴 Produit(s) INDÉFINI :")
                                        for _, ir in indef_rows.iterrows():
                                            ldp = ir.get("LDP")
                                            pu  = ir.get("PU")
                                            ldp_str = f"LDP {int(ldp)}" if ldp is not None and not (isinstance(ldp, float) and pd.isna(ldp)) else "LDP ?"
                                            pu_str  = f"{pu:.2f} €" if pu is not None and not (isinstance(pu, float) and pd.isna(pu)) else "? €"
                                            lignes_resume.append(f"  - {ldp_str} : produit INDÉFINI vendu à {pu_str}")

                                    # 3. LDP en doublon
                                    if ldp_doublons:
                                        lignes_resume.append("🟡 Doublon(s) de Ligne De Prix détecté(s) :")
                                        for ldp_val in sorted(ldp_doublons):
                                            dup_rows = df_display[df_display["LDP"] == ldp_val]
                                            codes   = ", ".join(dup_rows["CODE_PRODUIT"].astype(str).tolist())
                                            ldp_str = f"LDP {int(ldp_val)}" if not (isinstance(ldp_val, float) and pd.isna(ldp_val)) else "LDP ?"
                                            lignes_resume.append(f"  - {ldp_str} apparaît {len(dup_rows)} fois ({codes})")

                                    if lignes_resume:
                                        html_lines = []
                                        for l in lignes_resume:
                                            if l.startswith("  -"):
                                                html_lines.append(
                                                    f"<div style='margin:3px 0 3px 16px;color:#dddddd;font-size:0.93em'>"
                                                    f"{l[3:].strip()}</div>"
                                                )
                                            else:
                                                html_lines.append(
                                                    f"<div style='margin:10px 0 4px 0;color:#ffffff;font-size:1em'>"
                                                    f"{l}</div>"
                                                )

                                        st.markdown(
                                            "<div style='background:#1e1e2e;border-left:4px solid #e74c3c;"
                                            "padding:14px 18px;border-radius:6px;margin:12px 0;'>"
                                            + "".join(html_lines)
                                            + "</div>",
                                            unsafe_allow_html=True,
                                        )
                                    else:
                                        st.success("Aucune anomalie détectée sur ce fichier de ventes.")

                                    # ── Exports Excel + PNG ─────────────────────
                                    excel_m = _generer_export_machine(
                                        df_display, prix_ko_ttc, machine, salle
                                    )
                                    png_m = _generer_png_machine(
                                        df_display, prix_ko_ttc, machine, salle
                                    )
                                    fname_base = f"ventes_{machine}_{salle.replace(' ', '_')}"
                                    col_xl, col_png = st.columns(2)
                                    with col_xl:
                                        st.download_button(
                                            label=f"📥 Exporter {machine} (Excel)",
                                            data=excel_m,
                                            file_name=f"{fname_base}.xlsx",
                                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                            key=f"dl_machine_{machine}",
                                        )
                                    with col_png:
                                        st.download_button(
                                            label=f"🖼️ Exporter {machine} (PNG)",
                                            data=png_m,
                                            file_name=f"{fname_base}.png",
                                            mime="image/png",
                                            key=f"dl_machine_png_{machine}",
                                        )

                    with vue_complet:
                        styled_c = anomalies_df.style.apply(_color_row, axis=1).format(
                            {
                                "Prix attendu (HT)": "{:.3f} €",
                                "Prix réel (HT)": "{:.3f} €",
                                "Écart (€)": "{:+.3f} €",
                            }
                        )
                        st.dataframe(styled_c, use_container_width=True, hide_index=True)

                if non_ref:
                    with st.expander(
                        f"Codes produits non référencés dans la bibliothèque ({nb_non_ref})"
                    ):
                        for code in sorted(non_ref):
                            st.markdown(f"- `{code}`")
