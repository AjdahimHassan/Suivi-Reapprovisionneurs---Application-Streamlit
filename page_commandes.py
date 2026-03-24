"""
Page Commandes — Suivi des commandes fournisseurs
Lecture d'un screenshot de mail → extraction Gemini → ajout dans le fichier Excel de suivi
"""

import streamlit as st
import base64
import io
import json
import datetime
import re
import os
import requests
from openpyxl import load_workbook
from pymongo import MongoClient
import gridfs

# ──────────────────────────────────────────────────────────────────────────────
# CONFIGURATION PRODUITS
# ──────────────────────────────────────────────────────────────────────────────

LIDIS_PRODUCTS = [
    "50CL Evian x24",
    "50CL Volvic Juicy Fraise x12",
    "50CL Volvic Juicy Exotique x24",
    "25CL Redbull x24",
    "25CL Redbull Zéro x24",
]

HIPRO_PRODUCTS = [
    "HIPRO A BOIRE 330 ML VANILLE",
    "HIPRO A BOIRE 330 ML SAVEUR CHOCOLAT",
]

HEROIC_PRODUCTS = [
    "Heroic sport Fruits rouges 500ML x6",
    "Heroic sport Citron vert Menthe 500ML x6",
    "HEROIC SPORT SAVEUR TROPICAL 500ML x6",
]

NXT_PRODUCTS = [
    "RTD Protein Shake Strawberry 330ml",
    "RTD Protein Shake Vanilla 330ml",
    "RTD Protein Shake Chocolate 330ml",
    "RTD Protein Shake Milky Chocolate 500ml",
    "RTD Protein Shake Vanilla 500ml",
    "Crispy Protein Raspberry Toffee",
    "Pocket Protein Salty Caramel",
    "Pocket Protein Caramel Cookie",
    "Peanut Boost",
    "Pre-workout Shots Lemon-Lime 60ml",
    "Basic-Fit - Bidon - Orange",
    "Basic-Fit - Lock (Key)",
    "Basic-Fit - Vending Towel black",
]

FOURNISSEURS = ["LIDIS", "HIPRO", "HEROIC", "NXT LEVEL"]

PRODUITS_PAR_FOURNISSEUR = {
    "LIDIS": LIDIS_PRODUCTS,
    "HIPRO": HIPRO_PRODUCTS,
    "HEROIC": HEROIC_PRODUCTS,
    "NXT LEVEL": NXT_PRODUCTS,
}

# ──────────────────────────────────────────────────────────────────────────────
# HELPERS
# ──────────────────────────────────────────────────────────────────────────────

def get_gemini_api_key():
    try:
        return st.secrets["gemini"]["api_key"]
    except (KeyError, FileNotFoundError):
        return os.environ.get("GEMINI_API_KEY", "")


def _get_mongo_db():
    uri = st.secrets["mongo"]["uri"]
    client = MongoClient(uri, serverSelectionTimeoutMS=5000)
    db_name = st.secrets["mongo"]["db_name"]
    return client[db_name]


def save_excel_to_mongo(file_bytes: bytes):
    """Sauvegarde le fichier Excel dans MongoDB (collection suivi_excel)."""
    db = _get_mongo_db()
    col = db["suivi_excel"]
    col.replace_one(
        {"nom": "suivi_commandes"},
        {"nom": "suivi_commandes", "data": file_bytes},
        upsert=True,
    )


def load_excel_from_mongo():
    """Charge le fichier Excel depuis MongoDB. Retourne bytes ou None."""
    try:
        db = _get_mongo_db()
        col = db["suivi_excel"]
        doc = col.find_one({"nom": "suivi_commandes"})
        if doc:
            return bytes(doc["data"])
        return None
    except Exception:
        return None


def get_uvc_multiplier(product_name):
    m = re.search(r"x(\d+)$", product_name.strip(), re.IGNORECASE)
    return int(m.group(1)) if m else None


# Multiplicateurs fixes pour les produits sans suffixe xNN
MULTIPLICATEURS_FIXES = {
    "Heroic sport Fruits rouges 500ML x6": 6,
    "Heroic sport Citron vert Menthe 500ML x6": 6,
    "HEROIC SPORT SAVEUR TROPICAL 500ML x6": 6,
    "HIPRO A BOIRE 330 ML VANILLE": 12,
    "HIPRO A BOIRE 330 ML SAVEUR CHOCOLAT": 12,
}

def get_uvc_formula_or_value(product_name, qty_col, row_num):
    mult = get_uvc_multiplier(product_name)
    if not mult:
        mult = MULTIPLICATEURS_FIXES.get(product_name)
    return f"={qty_col}{row_num}*{mult}" if mult else None


def find_last_data_row(ws, header_row):
    last = header_row
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        # Ignore les lignes qui n'ont que des 0 ou None (lignes vides avec formules)
        values = [c.value for c in row]
        non_empty = [v for v in values if v is not None and v != 0 and v != ""]
        if non_empty:
            last = row[0].row
    return last


def analyze_screenshot_with_gemini(image_bytes, mime, fournisseur):
    api_key = get_gemini_api_key()
    if not api_key:
        raise ValueError("Clé GEMINI_API_KEY introuvable dans les secrets Streamlit.")

    produits_list = "\n".join(f"- {p}" for p in PRODUITS_PAR_FOURNISSEUR[fournisseur])
    today = datetime.date.today().strftime("%d/%m/%Y")

    prompt = f"""Tu es un assistant qui extrait des informations de commande depuis des screenshots de mails.
Tu dois retourner UNIQUEMENT un JSON valide, sans backticks, sans texte avant ou apres.

Le fournisseur est : {fournisseur}

Liste des produits connus pour ce fournisseur :
{produits_list}

Tu dois retourner ce format JSON :
{{
  "date_commande": "DD/MM/YYYY",
  "depot": "Nom du depot ou box mentionne dans le mail",
  "produits": [
    {{"nom": "Nom exact du produit dans la liste", "quantite": nombre_entier}},
    ...
  ],
  "notes": "Toute information utile non capturee ci-dessus"
}}

Regles :
- Pour chaque produit mentionne, trouve le meilleur match dans la liste des produits connus.
- Si la quantite est en packs/palettes, convertis en nombre de packs. Pour LIDIS : 1 palette Evian = 84 packs. Les quantites mentionnees sont toujours en nombre de packs sauf si "palette" est explicitement mentionne.
- Pour LIDIS, le multiplicateur UVC est celui indique dans le nom du produit (x24, x12). Exemple : "50CL Volvic Juicy Fraise x12" signifie 12 unites par pack, pas 24.
- Pour HEROIC : le document peut etre un bon de commande avec un tableau. Utilise la colonne "QUANTITES COMMANDEES EN COLIS" (pas UVC, pas NBR DE PALETTES). Si cette colonne n'est pas visible, prends le nombre de colis commandes.
- Si la date n'est pas mentionnee, utilise aujourd'hui : {today}.
- Si le depot n'est pas clairement mentionne, mets "Non precise".
- Retourne UNIQUEMENT le JSON, rien d'autre."""

    b64 = base64.standard_b64encode(image_bytes).decode("utf-8")

    payload = {
        "contents": [
            {
                "parts": [
                    {"inline_data": {"mime_type": mime, "data": b64}},
                    {"text": prompt},
                ]
            }
        ]
    }

    url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-3-flash-preview:generateContent?key={api_key}"
    resp = requests.post(url, json=payload, timeout=30)
    resp.raise_for_status()

    raw = resp.json()["candidates"][0]["content"]["parts"][0]["text"].strip()
    raw = re.sub(r"^```json\s*", "", raw)
    raw = re.sub(r"\s*```$", "", raw)
    return json.loads(raw)


def add_commande_to_excel(wb, fournisseur, date_commande, depot, produits):
    from openpyxl.styles import Font, Border, Side, PatternFill
    
    if fournisseur not in wb.sheetnames:
        st.error(f"Feuille '{fournisseur}' introuvable dans le fichier Excel.")
        return

    ws = wb[fournisseur]
    last_row = find_last_data_row(ws, header_row=2)
    next_row = last_row + 1

    try:
        dt = datetime.datetime.strptime(date_commande, "%d/%m/%Y")
    except Exception:
        dt = datetime.datetime.today()

    is_nxt = fournisseur == "NXT LEVEL"
    max_col = ws.max_column or 16

    thin = Side(style="thin")
    thick = Side(style="medium")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    bold_font = Font(bold=True)

    for i, produit in enumerate(produits):
        row_num = next_row + i
        nom = produit.get("nom", "")
        quantite = produit.get("quantite", 0)

        if i == 0:
            cell_date = ws.cell(row=row_num, column=1)
            cell_date.value = dt
            cell_date.number_format = "DD/MM/YYYY"
            # Nettoyer le depot : supprimer "Home Box ", "home box ", etc.
            depot_clean = re.sub(r"(?i)home\s+box\s*[-–]?\s*", "", depot).strip()
            ws.cell(row=row_num, column=4).value = depot_clean

        if is_nxt:
            ws.cell(row=row_num, column=5).value = quantite
            ws.cell(row=row_num, column=7).value = nom
        else:
            ws.cell(row=row_num, column=6).value = quantite
            uvc = get_uvc_formula_or_value(nom, "F", row_num)
            if uvc:
                ws.cell(row=row_num, column=5).value = uvc
            ws.cell(row=row_num, column=7).value = nom

        # Appliquer gras + bordures fines sur toutes les colonnes de la ligne
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = bold_font
            cell.border = thin_border

    # Bordures épaisses extérieures sur le bloc inséré
    nb = len(produits)
    for col in range(1, max_col + 1):
        for row_num in range(next_row, next_row + nb):
            cell = ws.cell(row=row_num, column=col)
            left   = thick if col == 1 else thin
            right  = thick if col == max_col else thin
            top    = thick if row_num == next_row else thin
            bottom = thick if row_num == next_row + nb - 1 else thin
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)


# ──────────────────────────────────────────────────────────────────────────────
# RENDER
# ──────────────────────────────────────────────────────────────────────────────

def render():
    st.markdown("### 📸 Ajouter une commande depuis un screenshot")

    col_four, _ = st.columns([2, 4])
    with col_four:
        fournisseur = st.selectbox("Fournisseur", FOURNISSEURS, key="commande_fournisseur")

    st.divider()

    st.markdown("**📎 Screenshot du mail de commande**")
    screenshot = st.file_uploader(
        "Image du mail",
        type=["png", "jpg", "jpeg", "webp"],
        key="commande_screenshot",
        label_visibility="collapsed",
    )

    if screenshot:
        col_img, _ = st.columns([2, 3])
        with col_img:
            st.image(screenshot, caption="Apercu du screenshot", use_container_width=True)

    st.divider()

    if screenshot and st.button("🔍 Analyser le screenshot", type="primary", key="btn_analyze"):
        with st.spinner(f"Analyse du mail {fournisseur} en cours..."):
            try:
                screenshot.seek(0)
                image_bytes = screenshot.read()
                mime = screenshot.type or "image/png"
                result = analyze_screenshot_with_gemini(image_bytes, mime, fournisseur)
                st.session_state["commande_result"] = result
                st.session_state["commande_fournisseur_used"] = fournisseur
                st.success("Analyse terminee — verifiez et corrigez si besoin")
            except Exception as e:
                st.error(f"Erreur lors de l'analyse : {e}")
                return

    if "commande_result" not in st.session_state:
        return

    result = st.session_state["commande_result"]
    fournisseur_used = st.session_state.get("commande_fournisseur_used", fournisseur)

    st.markdown("### ✏️ Vérification et correction de la commande")

    col_date, col_depot = st.columns(2)
    with col_date:
        date_val = st.text_input(
            "📅 Date de commande",
            value=result.get("date_commande", datetime.date.today().strftime("%d/%m/%Y")),
            key="commande_date",
        )
    with col_depot:
        depot_val = st.text_input(
            "📍 Dépôt / Box",
            value=result.get("depot", ""),
            key="commande_depot",
        )

    st.markdown("**Produits détectés :**")

    produits_raw = result.get("produits", [])
    produits_list_ref = PRODUITS_PAR_FOURNISSEUR[fournisseur_used]

    edited_produits = []
    for i, p in enumerate(produits_raw):
        col_nom, col_qty, _ = st.columns([4, 2, 1])
        with col_nom:
            idx = produits_list_ref.index(p["nom"]) if p.get("nom") in produits_list_ref else 0
            nom_sel = st.selectbox(
                f"Produit {i+1}",
                produits_list_ref,
                index=idx,
                key=f"commande_produit_{i}",
                label_visibility="collapsed",
            )
        with col_qty:
            qty_sel = st.number_input(
                f"Qte {i+1}",
                min_value=0,
                value=int(p.get("quantite", 0)),
                step=1,
                key=f"commande_qty_{i}",
                label_visibility="collapsed",
            )
        edited_produits.append({"nom": nom_sel, "quantite": qty_sel})

    if st.button("➕ Ajouter un produit", key="btn_add_produit"):
        produits_raw.append({"nom": produits_list_ref[0], "quantite": 0})
        result["produits"] = produits_raw
        st.session_state["commande_result"] = result
        st.rerun()

    if result.get("notes"):
        st.caption(f"ℹ️ Note : {result['notes']}")

    st.divider()

    st.markdown("### 📥 Fichier Excel de suivi")

    # Charger depuis MongoDB
    excel_bytes_mongo = load_excel_from_mongo()

    if excel_bytes_mongo:
        st.success("✅ Fichier Excel chargé depuis la base de données")
        with st.expander("🔄 Remplacer le fichier Excel"):
            new_file = st.file_uploader(
                "Nouveau fichier Excel",
                type=["xlsx"],
                key="commande_excel_replace",
                label_visibility="collapsed",
            )
            if new_file and st.button("💾 Sauvegarder le nouveau fichier", key="btn_save_new"):
                save_excel_to_mongo(new_file.read())
                st.success("✅ Nouveau fichier sauvegardé !")
                st.rerun()
        excel_bytes_to_use = excel_bytes_mongo
    else:
        st.warning("⚠️ Aucun fichier Excel en base. Uploade-le une première fois ci-dessous.")
        excel_file = st.file_uploader(
            "Fichier Excel de suivi des commandes",
            type=["xlsx"],
            key="commande_excel",
            label_visibility="collapsed",
        )
        if excel_file:
            raw = excel_file.read()
            save_excel_to_mongo(raw)
            st.success("✅ Fichier sauvegardé en base pour les prochaines fois !")
            excel_bytes_to_use = raw
        else:
            excel_bytes_to_use = None

    if excel_bytes_to_use and st.button("💾 Injecter dans le fichier Excel", type="primary", key="btn_inject"):
        with st.spinner("Injection en cours..."):
            try:
                wb = load_workbook(io.BytesIO(excel_bytes_to_use))
                add_commande_to_excel(wb, fournisseur_used, date_val, depot_val, edited_produits)

                output = io.BytesIO()
                wb.save(output)
                output.seek(0)

                # Sauvegarder le fichier mis à jour en MongoDB
                save_excel_to_mongo(output.getvalue())

                st.success(f"✅ Commande ajoutée dans la feuille **{fournisseur_used}** — {len(edited_produits)} produit(s)")
                st.download_button(
                    label="⬇️ Télécharger le fichier mis à jour",
                    data=output.getvalue(),
                    file_name="SPECIMEN Suivi commandes marchandises DA - MARS AVRIL.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

                del st.session_state["commande_result"]
                st.session_state.pop("commande_fournisseur_used", None)

            except Exception as e:
                st.error(f"Erreur : {e}")
