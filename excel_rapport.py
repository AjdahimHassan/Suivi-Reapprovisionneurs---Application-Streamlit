"""
Export Excel employé — openpyxl, une feuille par section.
Interface : generate_excel_report(employe, d, sections) -> bytes
  employe  : code employé (str)
  d        : dict retourné par _export_data()
  sections : set de clés actives ("chargement", "non_faites", "picklist",
             "pointage", "quartix")
"""

import io
import datetime

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill,
)
from openpyxl.utils import get_column_letter

# ── Palette ───────────────────────────────────────────────────────────────────
_BLUE   = "1a6ead"
_BLUE2  = "155e96"
_ORANGE = "f0a500"
_GREEN  = "2eaa6e"
_RED    = "d94040"
_AMBER  = "b07a00"
_LIGHT  = "f4f7fb"
_HEAD   = "dce8f5"
_WHITE  = "FFFFFF"
_DARK   = "1a3a5c"
_MUTED  = "8a9bb0"

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color=_DARK, size=10, italic=False):
    return Font(name="Calibri", bold=bold, color=color, size=size, italic=italic)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _thin_border(sides="bottom"):
    s = Side(style="thin", color="D0D8E8")
    b = Border()
    if "bottom" in sides: b.bottom = s
    if "top"    in sides: b.top    = s
    if "left"   in sides: b.left   = s
    if "right"  in sides: b.right  = s
    return b

def _thick_border():
    s = Side(style="medium", color=_BLUE)
    return Border(bottom=s)


# ── Generic helpers ────────────────────────────────────────────────────────────

def _set_col_widths(ws, widths):
    """widths: list of widths for cols A, B, C…"""
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _section_title(ws, row, title, ncols=6):
    """Blue header spanning ncols columns."""
    ws.row_dimensions[row].height = 20
    c = ws.cell(row=row, column=1, value=title.upper())
    c.font      = _font(bold=True, color=_WHITE, size=10)
    c.fill      = _fill(_BLUE)
    c.alignment = _align(h="left", v="center")
    c.border    = Border(bottom=Side(style="medium", color=_ORANGE))
    if ncols > 1:
        ws.merge_cells(
            start_row=row, start_column=1,
            end_row=row,   end_column=ncols,
        )
    return row + 1


def _table_header(ws, row, labels, colors=None):
    """Dark-blue header row for a data table."""
    ws.row_dimensions[row].height = 16
    for col, lbl in enumerate(labels, 1):
        bg  = colors[col-1] if colors and col-1 < len(colors) else _BLUE2
        c   = ws.cell(row=row, column=col, value=lbl)
        c.font      = _font(bold=True, color=_WHITE, size=9)
        c.fill      = _fill(bg)
        c.alignment = _align(h="center", v="center")
        c.border    = Border(bottom=Side(style="thin", color=_ORANGE))
    return row + 1


def _kpi_block(ws, row, kpis, ncols_total=6):
    """
    Write a block of KPI pairs (label, value, color).
    Each KPI occupies 2 columns (label | value); pairs are laid out in a row.
    Returns next free row.
    """
    ws.row_dimensions[row].height = 14
    ws.row_dimensions[row + 1].height = 24
    col = 1
    for label, value, hex_col in kpis:
        if col > ncols_total:
            break
        # label cell
        lc = ws.cell(row=row, column=col, value=label)
        lc.font      = _font(size=8, color=_MUTED, italic=True)
        lc.alignment = _align(h="left", v="bottom")
        lc.fill      = _fill(_LIGHT)
        # value cell
        vc = ws.cell(row=row + 1, column=col, value=value)
        vc.font      = _font(bold=True, size=16, color=hex_col or _DARK)
        vc.alignment = _align(h="left", v="center")
        vc.fill      = _fill(_LIGHT)
        vc.border    = Border(bottom=Side(style="thin", color="D0D8E8"))
        col += 1
    return row + 3   # label row + value row + spacer


def _data_row(ws, row, values, bold_cols=None, color_map=None, alt=False):
    """Write one data row. color_map: {col_idx: hex_color}."""
    ws.row_dimensions[row].height = 14
    bg = _LIGHT if alt else _WHITE
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = _font(bold=(bold_cols and col in bold_cols), size=9,
                            color=(color_map or {}).get(col, _DARK))
        c.fill      = _fill(bg)
        c.alignment = _align(v="center", wrap=True)
        c.border    = _thin_border("bottom")
    return row + 1


def _spacer(ws, row, height=6):
    ws.row_dimensions[row].height = height
    return row + 1


# ── Sheet builders ─────────────────────────────────────────────────────────────

def _sheet_resume(wb, employe, d):
    ws = wb.active
    ws.title = "Résumé"
    _set_col_widths(ws, [22, 28, 18, 28, 18, 18])
    ws.sheet_view.showGridLines = False

    # ── Title banner ──────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 36
    c = ws.cell(row=1, column=1,
                value=f"DISTRIPROT — Rapport de performance hebdomadaire")
    c.font      = _font(bold=True, color=_WHITE, size=14)
    c.fill      = _fill(_BLUE)
    c.alignment = _align(h="left", v="center")
    ws.merge_cells("A1:F1")

    # sub-banner: employee + week
    ws.row_dimensions[2].height = 22
    sub = ws.cell(row=2, column=1,
                  value=f"{d['prenom']}  ({employe})  ·  {d['semaine']}  —  Généré le {d['gen_date']}")
    sub.font      = _font(bold=True, color=_WHITE, size=11)
    sub.fill      = _fill(_BLUE2)
    sub.alignment = _align(h="left", v="center")
    ws.merge_cells("A2:F2")

    row = 4
    # ── Meta ──────────────────────────────────────────────────────────────────
    row = _section_title(ws, row, "Informations générales", ncols=6)
    meta_rows = [
        ("Employé",         employe),
        ("Prénom",          d.get("prenom", "")),
        ("Zone",            d.get("zone", "—")),
        ("Responsable",     d.get("resp", "—")),
        ("Semaine",         d.get("semaine", "")),
        ("Généré le",       d.get("gen_date", "")),
    ]
    for i, (lbl, val) in enumerate(meta_rows):
        ws.row_dimensions[row].height = 14
        lc = ws.cell(row=row, column=1, value=lbl)
        lc.font      = _font(bold=True, size=9, color=_DARK)
        lc.fill      = _fill(_HEAD if i % 2 == 0 else _WHITE)
        lc.alignment = _align()
        vc = ws.cell(row=row, column=2, value=val)
        vc.font      = _font(size=9)
        vc.fill      = _fill(_HEAD if i % 2 == 0 else _WHITE)
        vc.alignment = _align()
        row += 1

    row = _spacer(ws, row, 8)

    # ── Chargement KPIs ───────────────────────────────────────────────────────
    row = _section_title(ws, row, "Chargement — synthèse", ncols=6)
    taux_col = _GREEN if d["taux"] >= 90 else (_AMBER if d["taux"] >= 75 else _RED)
    row = _kpi_block(ws, row, [
        ("Taux de complétion", f"{d['taux']}%",    taux_col),
        ("Planifiées",         d["total_prevu"],   _DARK),
        ("Faites",             d["total_fait"],    _GREEN),
        ("Non faites",         d["total_nf"],      _RED   if d["total_nf"]  else _GREEN),
        ("Jokers",             d["nb_joker"],      _AMBER if d["nb_joker"]  else _DARK),
    ], ncols_total=5)

    # ── Picklist KPIs (if available) ──────────────────────────────────────────
    ps = d.get("pick_stats")
    if ps:
        row = _spacer(ws, row, 6)
        row = _section_title(ws, row, "Picklist — synthèse", ncols=6)
        pc_col = _GREEN if ps["pct"] >= 75 else (_AMBER if ps["pct"] >= 50 else _RED)
        row = _kpi_block(ws, row, [
            ("Conformité",   f"{ps['pct']}%", pc_col),
            ("Conformes",    ps["conf"],      _GREEN),
            ("Non chargés",  ps["nc"],        _RED   if ps["nc"]    else _DARK),
            ("Insuffisants", ps["insuf"],     _AMBER if ps["insuf"] else _DARK),
            ("Surplus",      ps["surplus"],   _AMBER if ps["surplus"] else _DARK),
        ], ncols_total=5)

    # ── Pointage KPIs (if available) ─────────────────────────────────────────
    p = d.get("ptg")
    if p:
        row = _spacer(ws, row, 6)
        row = _section_title(ws, row, "Pointage — synthèse", ncols=6)
        row = _kpi_block(ws, row, [
            ("Jours analysés", p["nb_jours"],       _DARK),
            ("Présences",      p["nb_presents"],    _GREEN),
            ("Congés",         p["nb_conges"],      _DARK),
            ("Absences",       p["nb_autres_abs"],  _RED   if p["nb_autres_abs"]  else _DARK),
            ("Sans badge",     p["nb_sans_badge"],  _AMBER if p["nb_sans_badge"]  else _DARK),
        ], ncols_total=5)

    # ── Quartix KPIs (if available) ───────────────────────────────────────────
    q = d.get("qtx")
    if q:
        row = _spacer(ws, row, 6)
        row = _section_title(ws, row, "Quartix — synthèse", ncols=6)
        row = _kpi_block(ws, row, [
            ("Trajets",       q["total_trajets"],      _DARK),
            ("Distance",      f"{q['total_km']:.0f} km", _DARK),
            ("Hors horaires", q["nb_late"],  _RED  if q["nb_late"]  else _GREEN),
            ("Weekend",       q["nb_wkend"], _RED  if q["nb_wkend"] else _GREEN),
        ], ncols_total=4)


def _sheet_chargement(wb, d, sections):
    ws = wb.create_sheet("Chargement")
    _set_col_widths(ws, [18, 12, 12, 12, 14, 16])
    ws.sheet_view.showGridLines = False

    row = 1
    row = _section_title(ws, row, "Performance hebdomadaire — Chargement", ncols=6)

    # KPIs
    taux_col = _GREEN if d["taux"] >= 90 else (_AMBER if d["taux"] >= 75 else _RED)
    row = _kpi_block(ws, row, [
        ("Taux",      f"{d['taux']}%", taux_col),
        ("Planifiées", d["total_prevu"], _DARK),
        ("Faites",     d["total_fait"],  _GREEN),
        ("Non faites", d["total_nf"],    _RED if d["total_nf"] else _GREEN),
        ("Jokers",     d["nb_joker"],    _AMBER if d["nb_joker"] else _DARK),
    ], ncols_total=5)

    # Day table
    row = _section_title(ws, row, "Détail par jour", ncols=5)
    row = _table_header(ws, row, ["Jour", "Prévues", "Faites", "Non faites", "Décalée"])
    for i, (jour, data) in enumerate(d.get("weekly", {}).items()):
        nf    = len(data["salles_non_faites"])
        nf_c  = _RED if nf > 0 else _GREEN
        dec   = "Oui" if data.get("tournee_decalee") else ""
        row = _data_row(ws, row,
                        [jour,
                         len(data["salles_prevues"]),
                         len(data["salles_faites"]),
                         nf,
                         dec],
                        bold_cols={1},
                        color_map={4: nf_c},
                        alt=(i % 2 == 1))

    # NF list
    nf_list = d.get("nf_list", [])
    if "non_faites" in sections and nf_list:
        row = _spacer(ws, row, 8)
        row = _section_title(ws, row, "Salles non faites & justifications", ncols=4)
        row = _table_header(ws, row, ["Jour", "Machine / Salle", "Statut", "Motif"])
        for i, r in enumerate(nf_list):
            statut = r.get("statut", "Non justifié")
            sc = _AMBER if statut == "Justifié" else _RED
            row = _data_row(ws, row,
                            [r.get("jour", ""),
                             r.get("machine", ""),
                             statut,
                             r.get("motif", "")],
                            color_map={3: sc},
                            alt=(i % 2 == 1))


def _sheet_picklist(wb, d):
    ps = d.get("pick_stats")
    if not ps:
        return
    ws = wb.create_sheet("Picklist")
    _set_col_widths(ws, [14, 32, 14, 12, 12, 14, 14])
    ws.sheet_view.showGridLines = False

    row = 1
    row = _section_title(ws, row, "Conformité Picklist", ncols=7)
    pc_col = _GREEN if ps["pct"] >= 75 else (_AMBER if ps["pct"] >= 50 else _RED)
    row = _kpi_block(ws, row, [
        ("Conformité",   f"{ps['pct']}%", pc_col),
        ("Conformes",    ps["conf"],      _GREEN),
        ("Total lignes", ps["total"],     _DARK),
        ("Non chargés",  ps["nc"],        _RED   if ps["nc"]    else _DARK),
        ("Insuffisants", ps["insuf"],     _AMBER if ps["insuf"] else _DARK),
        ("Surplus",      ps["surplus"],   _AMBER if ps["surplus"] else _DARK),
    ], ncols_total=6)

    row = _section_title(ws, row, "Détail par machine / salle", ncols=7)
    row = _table_header(ws, row,
                        ["Code", "Client / Salle", "Conformité %",
                         "Conformes", "Total", "Non chargés", "Insuffisants"])
    for i, m in enumerate(d.get("pick_machines", [])):
        pct  = m.get("pct", 0)
        pc   = _GREEN if pct >= 75 else (_AMBER if pct >= 50 else _RED)
        nc_c = _RED  if m.get("non_charges",  0) else _DARK
        in_c = _AMBER if m.get("insuffisants", 0) else _DARK
        row = _data_row(ws, row,
                        [m.get("code", ""),
                         m.get("nom", ""),
                         f"{pct}%",
                         m.get("conformes", 0),
                         m.get("total", 0),
                         m.get("non_charges", 0),
                         m.get("insuffisants", 0)],
                        color_map={3: pc, 6: nc_c, 7: in_c},
                        alt=(i % 2 == 1))


def _min_to_time(minutes):
    if minutes is None:
        return "—"
    try:
        h, m = divmod(int(minutes), 60)
        return f"{h:02d}:{m:02d}"
    except Exception:
        return "—"


def _sheet_pointage(wb, d):
    p = d.get("ptg")
    if not p:
        return
    ws = wb.create_sheet("Pointage")
    _set_col_widths(ws, [22, 18, 22, 18])
    ws.sheet_view.showGridLines = False

    row = 1
    row = _section_title(ws, row, "Pointage — Analyse hebdomadaire", ncols=4)
    row = _kpi_block(ws, row, [
        ("Jours analysés", p["nb_jours"],      _DARK),
        ("Présences",      p["nb_presents"],   _GREEN),
        ("Congés payés",   p["nb_conges"],     _DARK),
        ("Absences",       p["nb_autres_abs"], _RED   if p["nb_autres_abs"]  else _DARK),
    ], ncols_total=4)

    row = _spacer(ws, row, 6)
    row = _section_title(ws, row, "Horaires moyens", ncols=4)
    avg_rows = [
        ("Arrivée moyenne",  _min_to_time(p.get("avg_entree_min"))),
        ("Départ moyen",     _min_to_time(p.get("avg_sortie_min"))),
        ("Travail moyen",    _min_to_time(p.get("avg_travail"))),
        ("Pause moyenne",    _min_to_time(p.get("avg_pause"))),
        ("Sans badge",       p.get("nb_sans_badge", 0)),
    ]
    for i, (lbl, val) in enumerate(avg_rows):
        ws.row_dimensions[row].height = 14
        lc = ws.cell(row=row, column=1, value=lbl)
        lc.font      = _font(bold=True, size=9)
        lc.fill      = _fill(_HEAD if i % 2 == 0 else _WHITE)
        lc.alignment = _align()
        vc = ws.cell(row=row, column=2, value=val)
        vc.font      = _font(size=9,
                             color=_AMBER if (lbl == "Sans badge" and val) else _DARK)
        vc.fill      = _fill(_HEAD if i % 2 == 0 else _WHITE)
        vc.alignment = _align()
        row += 1


def _sheet_quartix(wb, d):
    q = d.get("qtx")
    if not q:
        return
    ws = wb.create_sheet("Quartix")
    _set_col_widths(ws, [14, 14, 28, 24, 12])
    ws.sheet_view.showGridLines = False

    row = 1
    row = _section_title(ws, row, "Quartix — Véhicule & Tournée", ncols=5)
    row = _kpi_block(ws, row, [
        ("Période",        f"{q.get('date_min','')} → {q.get('date_max','')}",  _DARK),
        ("Trajets",        q.get("total_trajets", 0), _DARK),
        ("Distance totale",f"{q.get('total_km', 0):.0f} km", _DARK),
        ("Hors horaires",  q.get("nb_late", 0),  _RED  if q.get("nb_late",  0) else _GREEN),
        ("Weekend",        q.get("nb_wkend", 0), _RED  if q.get("nb_wkend", 0) else _GREEN),
    ], ncols_total=5)

    # Late trips
    late_trips = q.get("late_trips", [])
    if late_trips:
        row = _spacer(ws, row, 6)
        cutoff = q.get("cutoff", "")
        row = _section_title(ws, row,
                             f"Trajets hors horaires (après {cutoff})", ncols=5)
        row = _table_header(ws, row,
                            ["Date", "Jour", "Horaires", "Trajet", "Distance"])
        for i, t in enumerate(late_trips):
            date_val = t["date"]
            if hasattr(date_val, "strftime"):
                date_val = date_val.strftime("%d/%m/%Y")
            row = _data_row(ws, row,
                            [date_val,
                             t.get("jour", ""),
                             f"{t.get('heure_dep','')} → {t.get('heure_arr','')}",
                             f"{t.get('lieu_dep','')} → {t.get('lieu_arr','')}",
                             f"{t.get('distance_km', 0):.1f} km"],
                            color_map={3: _AMBER},
                            alt=(i % 2 == 1))

    # Weekend trips
    wknd_trips = q.get("weekend_trips", [])
    if wknd_trips:
        row = _spacer(ws, row, 6)
        row = _section_title(ws, row, "Activité weekend", ncols=5)
        row = _table_header(ws, row, ["Date", "Jour", "Heure départ", "Distance"])
        for i, t in enumerate(wknd_trips):
            date_val = t["date"]
            if hasattr(date_val, "strftime"):
                date_val = date_val.strftime("%d/%m/%Y")
            row = _data_row(ws, row,
                            [date_val,
                             t.get("jour", ""),
                             t.get("heure_dep", ""),
                             f"{t.get('distance_km', 0):.1f} km"],
                            color_map={3: _RED},
                            alt=(i % 2 == 1))

    if not late_trips and not wknd_trips:
        row = _spacer(ws, row, 6)
        c = ws.cell(row=row, column=1, value="✅ Aucune anomalie détectée cette semaine")
        c.font      = _font(bold=True, color=_GREEN, size=10)
        c.alignment = _align()
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=5)


# ── Public function ────────────────────────────────────────────────────────────

def generate_excel_report(employe: str, d: dict, sections: set) -> bytes:
    """
    Génère un fichier Excel multi-feuilles.
    Retourne les bytes du fichier .xlsx.
    """
    wb = Workbook()

    _sheet_resume(wb, employe, d)

    if "chargement" in sections:
        _sheet_chargement(wb, d, sections)

    if "picklist" in sections and d.get("pick_stats"):
        _sheet_picklist(wb, d)

    if "pointage" in sections and d.get("ptg"):
        _sheet_pointage(wb, d)

    if "quartix" in sections and d.get("qtx"):
        _sheet_quartix(wb, d)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
