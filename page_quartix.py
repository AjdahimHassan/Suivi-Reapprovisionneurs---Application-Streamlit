"""
Page QUARTIX — Visualisation des trajets réappros
  • Import Excel QUARTIX (format multi-feuilles)
  • Onglet 1 — Carte & Trajets : sélecteur véhicule + journée, carte Folium,
    KPIs distance/dépôt, stats multi-véhicules, cache admin
  • Onglet 2 — Analyse Passages Dépôt : par réappro, tableau matin/après-midi
    avec heure d'arrivée et durée au dépôt (tolérance 1 km)
"""

import re
import numpy as np
import pandas as pd
import streamlit as st
import folium
from streamlit_folium import st_folium
from geopy.distance import geodesic
import io as _io

from mongo_storage import (
    load_quartix_vehicle,
    load_all_quartix_vehicles,
    upsert_quartix_vehicle,
    upsert_quartix_vehicle_info,
    load_plannings_from_mongo,
    load_geocode_cache,
    save_geocode_entry,
    load_routes_cache,
    save_route_entry,
    clear_geocode_cache,
    clear_routes_cache,
)

# ── Constantes ────────────────────────────────────────────────────────────────

import requests as _requests
DEPOT_RADIUS_M = 1000   # rayon autour du dépôt pour détecter un passage (1 km)
MIN_STOP_MIN   = 5      # durée minimale d'arrêt pour compter comme un passage (min)
BEFORE_WORK_H  = 11     # heure limite "avant tournée" — départ doit être < 11h
FIRST_N_TRIPS  = 5      # on cherche le passage dépôt parmi les N premiers trajets
LAST_N_TRIPS   = 5      # on cherche le passage dépôt parmi les N derniers trajets (hors dernier)

COLS = [
    "Trajet", "Conducteur", "Départ", "Lieu de départ", "Lieu d'arrivée",
    "Arrivée", "Durée du sous-trajet", "Durée du trajet", "Arrêt moteur en marche",
    "Kilométrage professionnel", "Kilométrage privé", "Distance totale", "Vitesse moy.",
]

# Palette Distriprot
C_BLUE_DARK  = "#1B3D6F"
C_BLUE_LIGHT = "#7BC4E8"
C_ORANGE     = "#E8922A"
C_GREEN      = "#1E7E34"
C_RED        = "#C0392B"
C_PURPLE     = "#9B59B6"


# ── Géocodage — API adresse.data.gouv.fr ─────────────────────────────────────

_GEOCODE_FR_URL       = "https://api-adresse.data.gouv.fr/search/"
_GEOCODE_FR_BATCH_URL = "https://api-adresse.data.gouv.fr/search/csv/"


def _prep_addr_for_api(addr: str) -> str:
    """Supprime le suffixe ', France' superflu pour l'API française."""
    return re.sub(r",?\s*France\s*$", "", addr, flags=re.IGNORECASE).strip()


_GEOCODE_MIN_SCORE = 0.5  # valeur par défaut — modifiable via le slider dans l'UI


def _get_score_threshold() -> float:
    """Lit le seuil depuis le session_state si le slider a été affiché, sinon valeur par défaut."""
    import streamlit as _st
    return float(_st.session_state.get("q_score_threshold", _GEOCODE_MIN_SCORE))


def _geocode_single_fr(addr: str) -> tuple | None:
    """Géocode une adresse via api-adresse.data.gouv.fr. Retourne (lat, lon) ou None."""
    clean = _prep_addr_for_api(addr)
    try:
        resp = _requests.get(
            _GEOCODE_FR_URL,
            params={"q": clean, "limit": 1, "type": "housenumber,street"},
            timeout=8,
        )
        features = resp.json().get("features", [])
        if features and features[0]["properties"].get("score", 0) >= _get_score_threshold():
            lon, lat = features[0]["geometry"]["coordinates"]
            return (float(lat), float(lon))
    except Exception:
        pass
    return None


# ── Expansion des abréviations françaises (utilisée pour la correction UI) ────

_ADDR_ABBREVS = [
    (re.compile(r'\bAv\.\s*'),   'Avenue '),
    (re.compile(r'\bBd\.?\s*'),  'Boulevard '),
    (re.compile(r'\bTrav\.\s*'), 'Traverse '),
    (re.compile(r'\bPl\.\s*'),   'Place '),
    (re.compile(r'\bImp\.\s*'),  'Impasse '),
    (re.compile(r'\bCrs\.\s*'),  'Cours '),
    (re.compile(r'\bAll\.\s*'),  'Allée '),
    (re.compile(r'\bRte\.\s*'),  'Route '),
    (re.compile(r'\bl([A-ZÉÀÂÊÎÙÛ])'), r"l'\1"),
    (re.compile(r'\bChe\.\s*'),  'Chemin '),
    (re.compile(r'\bSt\.\s*'),   'Saint '),
    (re.compile(r'\bSte\.\s*'),  'Sainte '),
    (re.compile(r'\bVla\.\s*'),  'Villa '),
    (re.compile(r'\bSq\.\s*'),   'Square '),
]


def _expand_address(addr: str) -> str:
    for pattern, replacement in _ADDR_ABBREVS:
        addr = pattern.sub(replacement, addr)
    return addr.strip()


_ARRET_MOTEUR_RE = re.compile(
    r"(?:Véhicule\s+à\s+l.arrêt[^,]*,\s*moteur\s+en\s+marche\s*)+",
    re.IGNORECASE,
)


def _clean_addr(addr: str) -> str:
    cleaned = _ARRET_MOTEUR_RE.sub("", addr).strip()
    return cleaned if cleaned else addr.strip()


# ── Géocodage ────────────────────────────────────────────────────────────────

def _get_osrm_route(coord_a: tuple, coord_b: tuple, routes_cache: dict) -> list:
    key = (tuple(coord_a), tuple(coord_b))
    if key in routes_cache:
        return routes_cache[key]
    try:
        url = (
            f"https://router.project-osrm.org/route/v1/driving/"
            f"{coord_a[1]},{coord_a[0]};{coord_b[1]},{coord_b[0]}"
            f"?overview=full&geometries=geojson"
        )
        resp = _requests.get(url, timeout=5)
        data = resp.json()
        if data.get("code") == "Ok":
            raw   = data["routes"][0]["geometry"]["coordinates"]
            route = [[c[1], c[0]] for c in raw]
            routes_cache[key] = route
            save_route_entry(coord_a, coord_b, route)
            return route
    except Exception:
        pass
    route = [list(coord_a), list(coord_b)]
    routes_cache[key] = route
    save_route_entry(coord_a, coord_b, route)
    return route


def _geocode_all(addresses: list[str], cache: dict) -> dict:
    """Géocode les adresses manquantes via api-adresse.data.gouv.fr (appel batch)."""
    missing = [a for a in addresses if a and str(a).strip() and a not in cache]
    if not missing:
        return cache

    _postcode_re = re.compile(r'\b(\d{5})\b')

    def _split_addr(raw: str):
        clean = _prep_addr_for_api(raw)
        m = _postcode_re.search(clean)
        return clean, (m.group(1) if m else "")

    bar = st.progress(0, text=f"Géocodage de {len(missing)} adresse(s)…")
    try:
        splits   = [_split_addr(a) for a in missing]
        df_req   = pd.DataFrame({"adresse": [s[0] for s in splits],
                                  "postcode": [s[1] for s in splits]})
        csv_bytes = df_req.to_csv(index=False).encode("utf-8")

        resp = _requests.post(
            _GEOCODE_FR_BATCH_URL,
            files={"data": ("addr.csv", csv_bytes, "text/csv")},
            data={"columns": "adresse", "postcode": "postcode",
                  "result_type": "housenumber,street"},
            timeout=60,
        )
        resp.raise_for_status()

        df_res = pd.read_csv(_io.StringIO(resp.text))

        for i, orig in enumerate(missing):
            row   = df_res.iloc[i] if i < len(df_res) else None
            lat   = row.get("latitude")     if row is not None else None
            lon   = row.get("longitude")    if row is not None else None
            score = row.get("result_score", 0) if row is not None else 0
            score = score or 0

            if lat is not None and lon is not None and not pd.isna(lat) and not pd.isna(lon) and float(score) >= _get_score_threshold():
                coords: tuple | None = (float(lat), float(lon))
            else:
                coords = None

            cache[orig] = coords
            save_geocode_entry(orig, coords)
            bar.progress((i + 1) / len(missing))

    except Exception:
        # Fallback : un par un si le batch échoue
        for i, addr in enumerate(missing):
            coords = _geocode_single_fr(addr)
            cache[addr] = coords
            if coords:
                save_geocode_entry(addr, coords)
            bar.progress((i + 1) / len(missing))

    bar.empty()
    return cache


# ── Helpers analytiques ───────────────────────────────────────────────────────

def _parse_min(val) -> float:
    if val is None:
        return 0.0
    if isinstance(val, float) and np.isnan(val):
        return 0.0
    if hasattr(val, "total_seconds"):
        return val.total_seconds() / 60
    s = str(val).strip()
    if re.match(r"^\d{1,2}:\d{2}(:\d{2})?$", s):
        parts = s.split(":")
        return int(parts[0]) * 60 + int(parts[1])
    h = re.search(r"(\d+)\s*[hH]", s)
    m = re.search(r"(\d+)\s*[mM]in", s)
    total = 0.0
    if h:
        total += int(h.group(1)) * 60
    if m:
        total += int(m.group(1))
    return total


def _to_km(val) -> float:
    try:
        return float(str(val).replace(",", ".").replace(" ", "").replace("km", ""))
    except Exception:
        return 0.0


def _auto_detect_depot(df_all: pd.DataFrame) -> str | None:
    all_deps = df_all["Lieu de départ"].dropna().astype(str)
    if all_deps.empty:
        return None
    return all_deps.value_counts().idxmax()


def _stop_duration_at(grp: pd.DataFrame, i: int) -> float:
    if i <= 0:
        return 0.0
    prev_arr = grp.loc[i - 1, "_arr"]
    this_dep = grp.loc[i, "_dep"]
    if pd.notna(prev_arr) and pd.notna(this_dep):
        return max(0.0, (this_dep - prev_arr).total_seconds() / 60)
    return 0.0


def _check_depot_visit_day(
    grp: pd.DataFrame,
    cache: dict,
    depot_coords,
) -> tuple[bool, bool]:
    if depot_coords is None or grp.empty:
        return False, False

    grp = grp.sort_values("_dep").reset_index(drop=True)
    n   = len(grp)

    before_tour = False
    after_tour  = False

    def _near_depot(addr: str) -> bool:
        coords = cache.get(addr)
        return bool(coords and geodesic(coords, depot_coords).meters <= DEPOT_RADIUS_M)

    def _stop_after_arrival(i: int) -> float:
        if i + 1 >= n:
            return 0.0
        arr = grp.loc[i, "_arr"]
        nxt = grp.loc[i + 1, "_dep"]
        if pd.notna(arr) and pd.notna(nxt):
            return max(0.0, (nxt - arr).total_seconds() / 60)
        return 0.0

    for i in range(min(FIRST_N_TRIPS, n)):
        row = grp.loc[i]
        if row["_dep"].hour < BEFORE_WORK_H:
            if _near_depot(str(row["Lieu de départ"])) and _stop_duration_at(grp, i) >= MIN_STOP_MIN:
                before_tour = True
                break
        if pd.notna(row["_arr"]) and row["_arr"].hour < BEFORE_WORK_H:
            if _near_depot(str(row["Lieu d'arrivée"])) and _stop_after_arrival(i) >= MIN_STOP_MIN:
                before_tour = True
                break

    if n >= 2:
        start = max(0, n - LAST_N_TRIPS)
        for i in range(start, n - 1):
            row = grp.loc[i]
            if _near_depot(str(row["Lieu de départ"])) and _stop_duration_at(grp, i) >= MIN_STOP_MIN:
                after_tour = True
                break
            if _near_depot(str(row["Lieu d'arrivée"])) and _stop_after_arrival(i) >= MIN_STOP_MIN:
                after_tour = True
                break

    return before_tour, after_tour


def _count_depot_visits(
    df_vehicle: pd.DataFrame,
    cache: dict,
    depot_coords,
) -> tuple[int, int]:
    if depot_coords is None or df_vehicle.empty:
        return 0, 0
    before_count = 0
    after_count  = 0
    for _, grp in df_vehicle.groupby(df_vehicle["_dep"].dt.date):
        b, a = _check_depot_visit_day(grp, cache, depot_coords)
        if b:
            before_count += 1
        if a:
            after_count  += 1
    return before_count, after_count


# ── Style carte ───────────────────────────────────────────────────────────────

def _stop_style(minutes: float, is_first: bool, is_last: bool) -> tuple[str, int]:
    if is_first:
        return C_GREEN, 11
    if is_last:
        return C_RED, 11
    if minutes < 2:
        return C_BLUE_LIGHT, 5
    if minutes < 15:
        return C_BLUE_DARK, 8
    if minutes < 45:
        return C_ORANGE, 12
    return C_PURPLE, 17


# ── Nouveau : parse feuille résumé ────────────────────────────────────────────

def _parse_vehicle_names(xls: pd.ExcelFile) -> dict[str, str]:
    """Tente de lire la feuille résumé pour construire {plate: full_name}."""
    try:
        df_recap = xls.parse(xls.sheet_names[0], header=None)
        vehicle_sheets = xls.sheet_names[1:]
        name_map: dict[str, str] = {}
        for _, row in df_recap.iterrows():
            cells = [str(c).strip() for c in row if pd.notna(c) and str(c).strip()]
            for plate in vehicle_sheets:
                if plate in cells:
                    candidates = [c for c in cells if c != plate and len(c) > len(plate)]
                    if candidates:
                        name_map[plate] = max(candidates, key=len)
                    break
        return name_map
    except Exception:
        return {}


# ── Nouveau : détection passages dépôt ───────────────────────────────────────

def _near_depot(addr: str, cache: dict, depot_coords: tuple) -> bool:
    coords = cache.get(addr)
    return bool(coords and geodesic(coords, depot_coords).meters <= DEPOT_RADIUS_M)


def _find_depot_passages(
    df_vehicle: pd.DataFrame,
    cache: dict,
    depot_coords: tuple,
    min_gap_hours: float = 0.0,
) -> list[dict]:
    """
    Retourne au plus 2 passages dépôt par jour, classés par séquence :
      - "matin"      = 1er passage avec durée >= MIN_STOP_MIN (passage avant tournée)
      - "après-midi" = passage suivant (au moins min_gap_hours après le matin) avec
                       durée >= MIN_STOP_MIN ; si aucun n'atteint MIN_STOP_MIN, on
                       prend le premier passage suivant disponible quelle que soit sa durée.
    """
    result: list[dict] = []

    for date, grp in df_vehicle.groupby(df_vehicle["_dep"].dt.date):
        grp = grp.sort_values("_dep").reset_index(drop=True)
        n   = len(grp)

        # ── 1. Collecte brute de tous les passages dépôt de la journée ────────
        day_visits: list[dict] = []

        for i in range(n):
            row      = grp.iloc[i]
            dep_addr = str(row["Lieu de départ"])
            arr_addr = str(row["Lieu d'arrivée"])

            # Cas A — arrivée au dépôt
            if _near_depot(arr_addr, cache, depot_coords):
                arrive_time = row["_arr"]
                if pd.isna(arrive_time):
                    continue
                next_row    = grp.iloc[i + 1] if i < n - 1 else None
                depart_time = next_row["_dep"] if next_row is not None and pd.notna(next_row["_dep"]) else None
                duration_min = (
                    max(0.0, (depart_time - arrive_time).total_seconds() / 60)
                    if depart_time is not None else 0.0
                )
                day_visits.append({
                    "date":         date,
                    "arrive_time":  arrive_time,
                    "depart_time":  depart_time,
                    "duration_min": duration_min,
                    "address":      arr_addr,
                    "trips_before": i + 1,
                    "trips_after":  max(0, n - 1 - i),
                })

            # Cas B — départ depuis le dépôt
            elif _near_depot(dep_addr, cache, depot_coords):
                depart_time = row["_dep"]
                if pd.isna(depart_time):
                    continue
                if i == 0:
                    arrive_time  = depart_time
                    duration_min = 0.0
                else:
                    prev_arr = grp.iloc[i - 1]["_arr"]
                    if pd.isna(prev_arr):
                        continue
                    arrive_time  = prev_arr
                    duration_min = max(0.0, (depart_time - prev_arr).total_seconds() / 60)
                day_visits.append({
                    "date":         date,
                    "arrive_time":  arrive_time,
                    "depart_time":  depart_time,
                    "duration_min": duration_min,
                    "address":      dep_addr,
                    "trips_before": i,
                    "trips_after":  n - i,
                })

        # ── 2. Déduplication sur arrive_time — garder la durée la plus longue ─
        seen: dict = {}
        for idx, v in enumerate(day_visits):
            key = v["arrive_time"]
            if key not in seen or v["duration_min"] > day_visits[seen[key]]["duration_min"]:
                seen[key] = idx
        day_visits = sorted(
            [day_visits[i] for i in seen.values()],
            key=lambda v: v["arrive_time"],
        )

        # ── 3. Classification par séquence ────────────────────────────────────
        # "matin" = 1er passage avec durée >= MIN_STOP_MIN
        matin_idx = next(
            (i for i, v in enumerate(day_visits) if v["duration_min"] >= MIN_STOP_MIN),
            None,
        )
        if matin_idx is None:
            continue  # aucun passage assez long → journée ignorée

        result.append({**day_visits[matin_idx], "period": "matin"})

        # "après-midi" = passage suivant (respectant le délai min) avec durée >= MIN_STOP_MIN,
        # sinon le 1er passage suivant disponible (durée quelconque)
        matin_time = day_visits[matin_idx]["arrive_time"]
        import datetime as _dt
        min_gap_td = _dt.timedelta(hours=min_gap_hours)
        remaining  = [
            v for v in day_visits[matin_idx + 1:]
            if v["arrive_time"] >= matin_time + min_gap_td
        ]
        apmidi = next((v for v in remaining if v["duration_min"] >= MIN_STOP_MIN), None)
        if apmidi is None and remaining:
            apmidi = remaining[0]
        if apmidi is not None:
            result.append({**apmidi, "period": "après-midi"})

    return sorted(result, key=lambda p: (p["date"], p["arrive_time"]))


# ── Onglet 1 : Carte & Trajets (code existant) ────────────────────────────────

def _render_tab_carte() -> None:

    # ── Upload ────────────────────────────────────────────
    col_up, _ = st.columns([2, 3])
    with col_up:
        uploaded = st.file_uploader(
            "📂 Importer un export Excel QUARTIX",
            type=["xls", "xlsx"],
            key="quartix_uploader",
        )

    if not uploaded:
        st.markdown(
            f"<div style='background:#f0f4fa;border-left:5px solid {C_BLUE_DARK};"
            f"padding:16px 20px;border-radius:8px;margin-top:20px'>"
            f"<b style='color:{C_BLUE_DARK}'>👆 Importez un fichier Excel QUARTIX</b>"
            f"<p style='margin:6px 0 0;color:#555;font-size:14px'>"
            f"Le fichier doit contenir une feuille résumé + une feuille par véhicule "
            f"(format standard QUARTIX, extension <code>.xls</code> ou <code>.xlsx</code>)."
            f"</p></div>",
            unsafe_allow_html=True,
        )
        return

    try:
        xls = pd.ExcelFile(uploaded)
    except Exception as e:
        st.error(f"Impossible de lire le fichier : {e}")
        return

    sheets = xls.sheet_names
    if len(sheets) <= 1:
        st.error("Fichier invalide : au moins 2 feuilles attendues (résumé + 1 feuille par véhicule).")
        return

    vehicle_sheets = sheets[1:]
    name_map       = _parse_vehicle_names(xls)

    # ── 3. Sélecteurs véhicule + journée ──────────────────
    col_v, col_d, _ = st.columns([2, 2, 1])
    with col_v:
        selected_vehicle = st.selectbox("🚗 Véhicule (plaque)", vehicle_sheets, key="q_vehicle")

    try:
        df_raw = pd.read_excel(xls, sheet_name=selected_vehicle, header=4, usecols="B:N")
        df_raw.columns = COLS
    except Exception as e:
        st.error(f"Erreur lecture feuille «{selected_vehicle}» : {e}")
        return

    df_raw["_dep"] = pd.to_datetime(df_raw["Départ"].astype(str).str.replace(r'\s+[A-Z]{2,5}$', '', regex=True),  errors="coerce")
    df_raw["_arr"] = pd.to_datetime(df_raw["Arrivée"].astype(str).str.replace(r'\s+[A-Z]{2,5}$', '', regex=True), errors="coerce")
    df_raw = df_raw.dropna(subset=["_dep"]).sort_values("_dep").reset_index(drop=True)

    for _col in ["Lieu de départ", "Lieu d'arrivée"]:
        df_raw[_col] = df_raw[_col].astype(str).map(_clean_addr)

    if df_raw.empty:
        st.warning("Aucune donnée valide dans cette feuille.")
        return

    available_days = sorted(df_raw["_dep"].dt.date.unique(), reverse=True)
    day_fmts = [d.strftime("%d/%m/%Y") for d in available_days]

    day_idx_key = f"q_day_idx_{selected_vehicle}"
    if day_idx_key not in st.session_state:
        st.session_state[day_idx_key] = 0

    with col_d:
        sel_day_fmt = st.selectbox(
            "📅 Journée",
            day_fmts,
            index=st.session_state[day_idx_key],
            key="q_day",
        )
        st.session_state[day_idx_key] = day_fmts.index(sel_day_fmt)

    selected_day = next(d for d in available_days if d.strftime("%d/%m/%Y") == sel_day_fmt)

    # ── 4. Panel véhicule : employé + dépôt ──────────────
    st.divider()

    vehicle_doc = load_quartix_vehicle(selected_vehicle)
    try:
        plannings, _ = load_plannings_from_mongo()
        employee_options = sorted(plannings.keys())
    except Exception:
        employee_options = []

    db_depot_address = vehicle_doc.get("depot_address", "") if vehicle_doc else ""
    db_depot_coords  = vehicle_doc.get("depot_coords",  None) if vehicle_doc else None
    db_employe       = vehicle_doc.get("employe",        "") if vehicle_doc else ""
    is_approximated  = False

    if not db_depot_address:
        approx = _auto_detect_depot(df_raw)
        display_depot = approx or "— Non défini —"
        is_approximated = bool(approx)
    else:
        display_depot = db_depot_address

    edit_key = f"edit_depot_{selected_vehicle}"
    if edit_key not in st.session_state:
        st.session_state[edit_key] = False

    with st.container():
        st.markdown("#### 🚗 Informations véhicule")

        col_emp, col_dep, col_btn = st.columns([2, 4, 1])

        with col_emp:
            emp_idx = 0
            if db_employe and db_employe in employee_options:
                emp_idx = employee_options.index(db_employe) + 1
            selected_employee = st.selectbox(
                "👤 Employé",
                [""] + employee_options,
                index=emp_idx,
                key="q_employee",
            )

        with col_dep:
            if st.session_state[edit_key]:
                new_depot_input = st.text_input(
                    "📍 Adresse dépôt (modifiable)",
                    value=db_depot_address if db_depot_address else display_depot,
                    placeholder="Ex : 3 Rue des Abattoirs, 38120 Saint-Égrève",
                    key="q_depot_input",
                )
            else:
                badge = " 🔸 *approximé*" if is_approximated else ""
                st.markdown(f"**📍 Dépôt :**&nbsp; {display_depot}{badge}")

        with col_btn:
            st.markdown("<div style='margin-top:28px'>", unsafe_allow_html=True)
            if not st.session_state[edit_key]:
                if st.button("✏️", key="q_btn_edit", help="Modifier les infos"):
                    st.session_state[edit_key] = True
                    st.rerun()
            else:
                if st.button("✏️ Annuler", key="q_btn_cancel"):
                    st.session_state[edit_key] = False
                    st.rerun()
            st.markdown("</div>", unsafe_allow_html=True)

        if st.session_state[edit_key]:
            col_save, _ = st.columns([2, 5])
            with col_save:
                if st.button("💾 Sauvegarder", type="primary", key="q_btn_save", use_container_width=True):
                    addr_to_save = st.session_state.get("q_depot_input", "").strip()
                    emp_to_save  = st.session_state.get("q_employee", "").strip()

                    if not addr_to_save:
                        st.warning("Veuillez saisir une adresse de dépôt.")
                    else:
                        with st.spinner("Géocodage de l'adresse…"):
                            tmp_cache = load_geocode_cache()
                            tmp_cache = _geocode_all([addr_to_save], tmp_cache)
                            coords = tmp_cache.get(addr_to_save)

                        if not coords:
                            st.error("❌ Adresse introuvable. Vérifiez l'orthographe et réessayez.")
                        else:
                            try:
                                upsert_quartix_vehicle(
                                    plate=selected_vehicle,
                                    employe=emp_to_save,
                                    depot_address=addr_to_save,
                                    depot_coords=list(coords),
                                )
                                st.success(f"✅ Sauvegardé : {addr_to_save}")
                                st.session_state[edit_key] = False
                                db_depot_address = addr_to_save
                                db_depot_coords  = list(coords)
                                db_employe       = emp_to_save
                                is_approximated  = False
                            except Exception as e:
                                st.error(f"❌ Erreur MongoDB : {e}")
                            st.rerun()

    depot_coords = db_depot_coords

    # ── 5. Filtrage journée + géocodage ───────────────────
    df = df_raw[df_raw["_dep"].dt.date == selected_day].copy().reset_index(drop=True)
    if df.empty:
        st.warning("Aucun trajet pour ce jour.")
        return

    addrs: set[str] = set()
    for c in ["Lieu de départ", "Lieu d'arrivée"]:
        addrs.update(df[c].dropna().astype(str).unique())

    cache = load_geocode_cache()
    cache = _geocode_all(list(addrs), cache)

    # ── 6. Construction des arrêts ordonnés ───────────────
    n = len(df)
    stops = []
    prev_arr_addr = None

    for i in range(n):
        row      = df.iloc[i]
        dep_addr = str(row["Lieu de départ"])
        arr_addr = str(row["Lieu d'arrivée"])
        next_row = df.iloc[i + 1] if i < n - 1 else None

        if dep_addr != prev_arr_addr:
            stops.append({
                "addr":     dep_addr,
                "coords":   cache.get(dep_addr),
                "time":     row["_dep"],
                "stop_min": 0.0,
                "trip_dur": row.get("Durée du sous-trajet"),
                "dist":     row.get("Distance totale"),
                "is_first": (len(stops) == 0),
                "is_last":  False,
            })

        stop_min = 0.0
        if next_row is not None and pd.notna(row["_arr"]) and pd.notna(next_row["_dep"]):
            stop_min = max(0.0, (next_row["_dep"] - row["_arr"]).total_seconds() / 60)

        stops.append({
            "addr":     arr_addr,
            "coords":   cache.get(arr_addr),
            "time":     row["_arr"],
            "stop_min": stop_min,
            "trip_dur": next_row.get("Durée du sous-trajet") if next_row is not None else None,
            "dist":     next_row.get("Distance totale")       if next_row is not None else None,
            "is_first": False,
            "is_last":  (i == n - 1),
        })

        prev_arr_addr = arr_addr

    valid = [s for s in stops if s["coords"]]
    if len(valid) < 2:
        st.error(
            "Pas assez d'adresses géocodées pour tracer le trajet. "
            "Corrigez les adresses ci-dessous pour débloquer la carte."
        )
        all_stop_addrs = list(dict.fromkeys(s["addr"] for s in stops))
        missing_addrs  = [a for a in all_stop_addrs if not cache.get(a)]
        if missing_addrs:
            st.markdown(
                f"<div style='background:#fff3cd;border-left:4px solid {C_ORANGE};"
                f"padding:10px 14px;border-radius:6px;margin-bottom:8px'>"
                f"<b>⚠️ {len(missing_addrs)} adresse(s) non reconnue(s)</b></div>",
                unsafe_allow_html=True,
            )
            for orig_addr in missing_addrs:
                fix_key = f"q_addr_fix_{hash(orig_addr)}"
                if fix_key not in st.session_state:
                    st.session_state[fix_key] = _expand_address(orig_addr)
                col_addr, col_btn = st.columns([5, 1])
                with col_addr:
                    st.text_input(f"❌ {orig_addr}", key=fix_key, label_visibility="visible")
                with col_btn:
                    st.markdown("<div style='margin-top:28px'>", unsafe_allow_html=True)
                    if st.button("Valider", key=f"q_btn_fix_{hash(orig_addr)}", use_container_width=True):
                        corrected = st.session_state[fix_key].strip()
                        if corrected:
                            with st.spinner(f"Géocodage de « {corrected} »…"):
                                coords = _geocode_single_fr(corrected)
                            if coords:
                                save_geocode_entry(orig_addr, coords)
                                if corrected != orig_addr:
                                    save_geocode_entry(corrected, coords)
                                st.success(f"✅ {corrected}")
                                st.rerun()
                            else:
                                st.error("Introuvable. Essayez une autre formulation.")
                    st.markdown("</div>", unsafe_allow_html=True)
        return

    # ── 7. KPIs ───────────────────────────────────────────
    lats = [s["coords"][0] for s in valid]
    lons = [s["coords"][1] for s in valid]
    total_km   = sum(_to_km(s["dist"]) for s in stops if s["dist"])
    nb_stops   = sum(1 for s in stops if not s["is_first"] and not s["is_last"] and s["stop_min"] >= 2)
    long_stops = sum(1 for s in stops if not s["is_first"] and not s["is_last"] and s["stop_min"] >= 20)

    before_ok, after_ok = _check_depot_visit_day(df, cache, depot_coords)

    st.divider()
    c1, c2, c3, c4, c5, c6 = st.columns(6)
    c1.metric("🚗 Trajets",          len(df))
    c2.metric("📍 Arrêts ≥ 2 min",   nb_stops)
    c3.metric("⏸️ Pauses ≥ 20 min", long_stops)
    c4.metric("📏 Distance totale",  f"{total_km:.0f} km" if total_km > 0 else "—")
    c5.metric(
        "🏠 Passage avant tournée",
        "✅ OK" if before_ok else ("❌ Pas OK" if depot_coords else "—"),
        help=f"Passage au dépôt parmi les {FIRST_N_TRIPS} premiers trajets, avant {BEFORE_WORK_H}h, arrêt ≥ {MIN_STOP_MIN}min"
             if depot_coords else "Dépôt non configuré — renseignez l'adresse ci-dessus",
    )
    c6.metric(
        "🏠 Passage après tournée",
        "✅ OK" if after_ok else ("❌ Pas OK" if depot_coords else "—"),
        help=f"Passage au dépôt parmi les {LAST_N_TRIPS} derniers trajets (hors dernier), arrêt ≥ {MIN_STOP_MIN}min"
             if depot_coords else "Dépôt non configuré — renseignez l'adresse ci-dessus",
    )

    # ── Navigation journée ◀ / ▶ ──────────────────────────
    nav_left, nav_mid, nav_right = st.columns([1, 6, 1])
    with nav_left:
        if st.button("◀ Jour préc.", key="q_day_prev", use_container_width=True,
                     disabled=st.session_state[day_idx_key] >= len(available_days) - 1):
            st.session_state[day_idx_key] += 1
            st.rerun()
    with nav_mid:
        st.markdown(
            f"<p style='text-align:center;color:#888;margin:0;padding-top:6px'>"
            f"{sel_day_fmt} — {st.session_state[day_idx_key] + 1} / {len(available_days)}</p>",
            unsafe_allow_html=True,
        )
    with nav_right:
        if st.button("Jour suiv. ▶", key="q_day_next", use_container_width=True,
                     disabled=st.session_state[day_idx_key] <= 0):
            st.session_state[day_idx_key] -= 1
            st.rerun()

    st.divider()

    # ── 8. Carte Folium ───────────────────────────────────
    center = [np.mean(lats), np.mean(lons)]
    m = folium.Map(location=center, zoom_start=12, tiles="CartoDB positron")

    if depot_coords:
        folium.Circle(
            location=list(depot_coords),
            radius=DEPOT_RADIUS_M,
            color=C_ORANGE,
            fill=True,
            fill_color=C_ORANGE,
            fill_opacity=0.06,
            weight=2,
            dash_array="6",
            tooltip=f"Zone dépôt — rayon {DEPOT_RADIUS_M}m",
        ).add_to(m)

    routes_cache = load_routes_cache()
    osrm_bar = st.empty()
    new_routes = sum(
        1 for i in range(len(valid) - 1)
        if (tuple(valid[i]["coords"]), tuple(valid[i+1]["coords"])) not in routes_cache
    )
    if new_routes:
        osrm_bar.info(f"🗺️ Calcul des routes ({new_routes} segment(s))…")
    for i in range(len(valid) - 1):
        seg = _get_osrm_route(valid[i]["coords"], valid[i + 1]["coords"], routes_cache)
        folium.PolyLine(seg, color=C_BLUE_DARK, weight=3, opacity=0.75).add_to(m)
    osrm_bar.empty()

    for num_label, s in enumerate(valid, start=1):
        color, radius = _stop_style(s["stop_min"], s["is_first"], s["is_last"])
        time_str = s["time"].strftime("%H:%M") if pd.notna(s["time"]) else "?"
        dur_str  = f"{int(s['stop_min'])}min" if s["stop_min"] >= 1 else ""
        trip_min = _parse_min(s["trip_dur"])

        label_first = "🚀 Départ" if s["is_first"] else ""
        label_last  = "🏁 Arrivée finale" if s["is_last"] else ""

        dist_depot_str = ""
        if depot_coords and s["coords"]:
            dm = geodesic(s["coords"], depot_coords).meters
            dist_depot_str = f"<br>🏠 Distance dépôt : <b>{dm:.0f} m</b>"

        popup_html = f"""
        <div style="font-family:sans-serif;font-size:13px;min-width:200px;max-width:300px;">
            {"<b style='color:" + C_GREEN + "'>" + label_first + "</b><br>" if label_first else ""}
            {"<b style='color:" + C_RED   + "'>" + label_last  + "</b><br>" if label_last  else ""}
            <b style="color:{C_BLUE_DARK}">#{num_label} — {s['addr']}</b>
            <hr style="margin:4px 0">
            🕐 <b>{time_str}</b>
            {"<br>⏸️ Arrêt : <b>" + dur_str + "</b>" if dur_str else ""}
            {"<br>🚗 Trajet suivant : " + str(int(trip_min)) + " min" if trip_min > 0 else ""}
            {"<br>📏 " + str(s['dist']) + " km" if s['dist'] else ""}
            {dist_depot_str}
        </div>
        """
        tooltip = f"#{num_label} · {time_str} — {s['addr'][:38]}{'…' if len(s['addr']) > 38 else ''}"
        if dur_str:
            tooltip += f"  ⏸ {dur_str}"

        diam      = max(24, radius * 2)
        font_size = max(9, diam // 3)
        folium.Marker(
            location=s["coords"],
            icon=folium.DivIcon(
                html=(
                    f'<div style="'
                    f'width:{diam}px;height:{diam}px;border-radius:50%;'
                    f'background:{color};border:2px solid white;'
                    f'text-align:center;line-height:{diam}px;'
                    f'font-size:{font_size}px;font-weight:bold;color:white;'
                    f'box-shadow:0 2px 5px rgba(0,0,0,.45);'
                    f'">{num_label}</div>'
                ),
                icon_size=(diam, diam),
                icon_anchor=(diam // 2, diam // 2),
            ),
            popup=folium.Popup(popup_html, max_width=320),
            tooltip=tooltip,
        ).add_to(m)

    m.get_root().html.add_child(folium.Element(f"""
    <div style="position:fixed;bottom:28px;left:28px;z-index:9999;
                background:white;padding:10px 16px;border-radius:10px;
                box-shadow:0 2px 12px rgba(0,0,0,.2);
                font-family:sans-serif;font-size:12px;line-height:2.0">
        <b style="color:{C_BLUE_DARK};font-size:13px">Légende</b><br>
        <span style="color:{C_GREEN};font-size:18px">●</span>&nbsp;Départ (1er arrêt)<br>
        <span style="color:{C_RED};font-size:18px">●</span>&nbsp;Arrivée finale<br>
        <span style="color:{C_BLUE_LIGHT};font-size:18px">●</span>&nbsp;Arrêt &lt; 2 min<br>
        <span style="color:{C_BLUE_DARK};font-size:18px">●</span>&nbsp;Arrêt 2 – 15 min<br>
        <span style="color:{C_ORANGE};font-size:18px">●</span>&nbsp;Pause 15 – 45 min<br>
        <span style="color:{C_PURPLE};font-size:18px">●</span>&nbsp;Longue pause &gt; 45 min<br>
        {"<span style='color:" + C_ORANGE + ";font-size:12px'>◯</span>&nbsp;Zone dépôt (1 km)<br>" if depot_coords else ""}
        <span style="color:#999;font-size:10px">Taille des cercles ∝ durée d'arrêt</span>
    </div>
    """))

    m.fit_bounds([
        [min(lats) - 0.005, min(lons) - 0.005],
        [max(lats) + 0.005, max(lons) + 0.005],
    ])

    st_folium(m, use_container_width=True, height=600, returned_objects=[])

    # ── 8b. Adresses de la journée ────────────────────────
    all_stop_addrs = list(dict.fromkeys(s["addr"] for s in stops))
    missing_addrs  = [a for a in all_stop_addrs if not cache.get(a)]
    found_addrs    = [a for a in all_stop_addrs if cache.get(a)]

    if missing_addrs:
        st.markdown(
            f"<div style='background:#fff3cd;border-left:4px solid {C_ORANGE};"
            f"padding:10px 14px;border-radius:6px;margin-bottom:8px'>"
            f"<b>⚠️ {len(missing_addrs)} adresse(s) non reconnue(s)</b> — "
            f"corrigez et validez pour les voir sur la carte.</div>",
            unsafe_allow_html=True,
        )
        for orig_addr in missing_addrs:
            fix_key = f"q_addr_fix_{hash(orig_addr)}"
            if fix_key not in st.session_state:
                st.session_state[fix_key] = _expand_address(orig_addr)
            col_addr, col_btn = st.columns([5, 1])
            with col_addr:
                st.text_input(
                    f"❌ {orig_addr}",
                    key=fix_key,
                    label_visibility="visible",
                )
            with col_btn:
                st.markdown("<div style='margin-top:28px'>", unsafe_allow_html=True)
                if st.button("Valider", key=f"q_btn_fix_{hash(orig_addr)}", use_container_width=True):
                    corrected = st.session_state[fix_key].strip()
                    if corrected:
                        with st.spinner(f"Géocodage de « {corrected} »…"):
                            coords = _geocode_single_fr(corrected)
                        if coords:
                            save_geocode_entry(orig_addr, coords)
                            if corrected != orig_addr:
                                save_geocode_entry(corrected, coords)
                            st.success(f"✅ Trouvé : {corrected}")
                            st.rerun()
                        else:
                            st.error("Introuvable. Essayez une autre formulation.")
                st.markdown("</div>", unsafe_allow_html=True)

    if found_addrs:
        with st.expander(f"✅ {len(found_addrs)} adresse(s) reconnue(s)", expanded=False):
            for a in found_addrs:
                coords = cache[a]
                st.caption(f"✅ {a}  —  `{coords[0]:.5f}, {coords[1]:.5f}`")

    # ── 9. Tableau détail ─────────────────────────────────
    st.divider()
    st.markdown(f"### 📋 Détail des trajets — **{selected_vehicle}** — {sel_day_fmt}")

    disp = df[["Lieu de départ", "Lieu d'arrivée", "Durée du sous-trajet", "Distance totale"]].copy()
    disp.insert(0, "Départ",  df["_dep"].dt.strftime("%H:%M"))
    disp.insert(1, "Arrivée", df["_arr"].dt.strftime("%H:%M").fillna("?"))
    st.dataframe(disp, use_container_width=True, hide_index=True)

    # ── 10. Statistiques multi-véhicules ──────────────────
    st.divider()
    st.markdown("### 📊 Statistiques passages dépôt — tous les véhicules du fichier")
    st.caption(
        "Calculé sur l'ensemble des jours disponibles dans le fichier. "
        "Seuls les véhicules avec un dépôt enregistré en base sont analysés."
    )

    all_veh_docs = load_all_quartix_vehicles()
    cache_stats  = load_geocode_cache()

    stat_rows = []
    for plate in vehicle_sheets:
        vdoc        = all_veh_docs.get(plate, {})
        dep_addr    = vdoc.get("depot_address", "—")
        dep_coords  = vdoc.get("depot_coords",  None)
        emp_code    = vdoc.get("employe",        "—")

        try:
            df_v = pd.read_excel(xls, sheet_name=plate, header=4, usecols="B:N")
            df_v.columns = COLS
            df_v["_dep"] = pd.to_datetime(df_v["Départ"].astype(str).str.replace(r'\s+[A-Z]{2,5}$', '', regex=True),  errors="coerce")
            df_v["_arr"] = pd.to_datetime(df_v["Arrivée"].astype(str).str.replace(r'\s+[A-Z]{2,5}$', '', regex=True), errors="coerce")
            df_v = df_v.dropna(subset=["_dep"]).reset_index(drop=True)
            nb_days = df_v["_dep"].dt.date.nunique()
        except Exception:
            nb_days = 0
            df_v    = pd.DataFrame()

        if dep_coords and not df_v.empty:
            bef, aft = _count_depot_visits(df_v, cache_stats, dep_coords)
            before_str = str(bef)
            after_str  = str(aft)
        else:
            before_str = "—"
            after_str  = "—"

        stat_rows.append({
            "Véhicule":                  plate,
            "Employé":                   emp_code,
            "Dépôt enregistré":          dep_addr,
            "Jours dans le fichier":     nb_days if nb_days else "—",
            "Passages avant tournée":    before_str,
            "Passages après tournée":    after_str,
        })

    df_stats = pd.DataFrame(stat_rows)

    def _style_stats(row):
        base      = "font-weight:500"
        bef       = row["Passages avant tournée"]
        aft       = row["Passages après tournée"]
        has_depot = row["Dépôt enregistré"] != "—"
        if not has_depot:
            return [f"{base}; color:#888"] * len(row)
        if bef == "0" or aft == "0":
            return [f"background-color:#fdf3e7; {base}"] * len(row)
        return [f"background-color:#eaf4ec; {base}"] * len(row)

    st.dataframe(
        df_stats.style.apply(_style_stats, axis=1),
        use_container_width=True,
        hide_index=True,
    )

    # ── 11. Méthodologie ──────────────────────────────────
    st.divider()
    with st.expander("ℹ️ Méthodologie — comment sont calculés ces chiffres ?"):
        st.markdown(f"""
### Passage au dépôt **avant** la tournée ✅

Un passage est comptabilisé si **toutes** ces conditions sont réunies :
1. 🔢 Le trajet figure parmi les **{FIRST_N_TRIPS} premiers** de la journée
2. 🕐 L'heure de départ est **avant {BEFORE_WORK_H}h00**
3. 📍 La localisation est à moins de **{DEPOT_RADIUS_M} m** du dépôt enregistré
4. ⏱️ L'arrêt dure **au moins {MIN_STOP_MIN} minutes**

---

### Passage au dépôt **après** la tournée ✅

Un passage est comptabilisé si **toutes** ces conditions sont réunies :
1. 🔢 Le trajet figure parmi les **{LAST_N_TRIPS} derniers** de la journée, **sauf le tout dernier**
2. 📍 La localisation est à moins de **{DEPOT_RADIUS_M} m** du dépôt enregistré
3. ⏱️ L'arrêt dure **au moins {MIN_STOP_MIN} minutes**

---

### Pourquoi ces seuils ?

| Paramètre | Valeur | Raison |
|---|---|---|
| **Rayon dépôt** | {DEPOT_RADIUS_M} m | Le géocodage via OpenStreetMap peut être imprécis de quelques centaines de mètres. 1 km absorbe ces écarts sans confondre le dépôt avec un client voisin. |
| **Arrêt minimum** | {MIN_STOP_MIN} min | Un simple passage devant le dépôt ne doit pas être compté. 5 min correspond au minimum pour une action réelle : chargement, signature, pause. |
| **Limite matin** | {BEFORE_WORK_H}h | Les tournées commencent généralement après 7h et rarement après 11h. |
| **Position séquence** | {FIRST_N_TRIPS} premiers / {LAST_N_TRIPS} derniers | On raisonne par position dans la journée plutôt que par heure fixe pour le retour. |
        """)

    # ── 12. Administration du cache ───────────────────────
    st.divider()
    with st.expander("🗑️ Administration du cache"):
        st.caption(
            "Le cache stocke les coordonnées GPS (géocodage) et les routes OSRM calculées. "
            "Supprimez-le si des adresses corrigées ne s'affichent pas correctement."
        )
        col_gc, col_rc, col_both, _ = st.columns([2, 2, 2, 3])
        with col_gc:
            if st.button("🗑️ Vider cache géocodage", use_container_width=True):
                n = clear_geocode_cache()
                st.success(f"✅ {n} adresse(s) supprimée(s).")
                st.rerun()
        with col_rc:
            if st.button("🗑️ Vider cache routes", use_container_width=True):
                n = clear_routes_cache()
                st.success(f"✅ {n} route(s) supprimée(s).")
                st.rerun()
        with col_both:
            if st.button("🗑️ Tout vider", type="primary", use_container_width=True):
                ng = clear_geocode_cache()
                nr = clear_routes_cache()
                st.success(f"✅ Cache vidé — {ng} adresse(s) et {nr} route(s) supprimées.")
                st.rerun()


# ── Onglet 2 : helpers ───────────────────────────────────────────────────────

def _build_export_excel(
    result: dict,
    available_days: list,
    day_summary: dict,
    df_raw: "pd.DataFrame | None" = None,
    late_hour: int = 17,
) -> bytes:
    """Génère un classeur Excel simple avec le récapitulatif des passages dépôt."""
    import io as _io_mod
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # ── Alertes par jour (weekend / heure tardive) ───────────────────────────
    alert_map: dict = {}
    if df_raw is not None:
        for d in available_days:
            day_df = df_raw[df_raw["_dep"].dt.date == d]
            notes  = []
            if d.weekday() >= 5:
                notes.append("Weekend")
            if not day_df.empty and day_df["_dep"].dt.hour.max() >= late_hour:
                notes.append(f"Trajet après {late_hour}h")
            alert_map[d] = " / ".join(notes) if notes else ""
    else:
        alert_map = {d: "" for d in available_days}

    plate    = result.get("plate", "")
    passages = result.get("passages", [])

    # ── Helpers styles ────────────────────────────────────────────────────────
    HDR_FILL = PatternFill("solid", fgColor="1B3D6F")
    HDR_FONT = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    DAT_FONT = Font(name="Calibri", size=10, color="000000")
    THIN     = Side(style="thin", color="CCCCCC")
    BORDER   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CENTER   = Alignment(horizontal="center", vertical="center")
    LEFT     = Alignment(horizontal="left",   vertical="center")

    def _hdr(ws, row, col, value, align=CENTER):
        c = ws.cell(row=row, column=col, value=value)
        c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = align; c.border = BORDER

    def _dat(ws, row, col, value, align=CENTER):
        c = ws.cell(row=row, column=col, value=value)
        c.font = DAT_FONT; c.alignment = align; c.border = BORDER

    def _auto_width(ws, min_w=8, max_w=50):
        for col in ws.columns:
            best = max((len(str(c.value or "")) for c in col), default=min_w)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(best + 3, max_w)

    def _dur(p):
        if p is None:
            return "—"
        dep = p.get("depart_time")
        return "?" if (dep is None or pd.isna(dep)) else int(p["duration_min"])

    def _trips(p, key):
        if p is None:
            return "—"
        v = p.get(key)
        return int(v) if v is not None else "—"

    # ── Workbook ──────────────────────────────────────────────────────────────
    wb = Workbook()

    # ── Feuille 1 : Récapitulatif par jour ───────────────────────────────────
    ws = wb.active
    ws.title = "Récapitulatif"

    headers = [
        "Date", "Jour",
        "Avant tournée — Arrivée", "Avant tournée — Durée (min)", "Avant tournée — Trajets avant", "Avant tournée — Trajets après",
        "Après tournée — Arrivée", "Après tournée — Durée (min)", "Après tournée — Trajets avant", "Après tournée — Trajets après",
        "Observations",
    ]
    for ci, h in enumerate(headers, 1):
        _hdr(ws, 1, ci, h)
    ws.freeze_panes = "A2"

    for ri, d in enumerate(available_days, 2):
        m   = day_summary[d]["matin"]
        apm = day_summary[d]["après-midi"]
        row = [
            d.strftime("%d/%m/%Y"),
            d.strftime("%A").capitalize(),
            m["arrive_time"].strftime("%H:%M") if m else "—",
            _dur(m),
            _trips(m,   "trips_before"),
            _trips(m,   "trips_after"),
            apm["arrive_time"].strftime("%H:%M") if apm else "—",
            _dur(apm),
            _trips(apm, "trips_before"),
            _trips(apm, "trips_after"),
            alert_map.get(d, ""),
        ]
        aligns = [LEFT, LEFT, CENTER, CENTER, CENTER, CENTER, CENTER, CENTER, CENTER, CENTER, LEFT]
        for ci, (val, aln) in enumerate(zip(row, aligns), 1):
            _dat(ws, ri, ci, val, align=aln)

    _auto_width(ws)

    # ── Feuille 2 : Détail de chaque passage ─────────────────────────────────
    ws2 = wb.create_sheet("Détail passages")

    headers2 = [
        "Date", "Jour", "Période",
        "Arrivée dépôt", "Départ dépôt", "Durée (min)",
        "Trajets avant", "Trajets après",
        "Adresse",
    ]
    for ci, h in enumerate(headers2, 1):
        _hdr(ws2, 1, ci, h, align=CENTER if ci != 9 else LEFT)
    ws2.freeze_panes = "A2"

    for ri, p in enumerate(passages, 2):
        dep_t  = p["depart_time"]
        no_dep = dep_t is None or pd.isna(dep_t)
        row2   = [
            p["date"].strftime("%d/%m/%Y"),
            p["date"].strftime("%A").capitalize(),
            "Avant tournée" if p["period"] == "matin" else "Après tournée",
            p["arrive_time"].strftime("%H:%M"),
            "—" if no_dep else dep_t.strftime("%H:%M"),
            "?" if no_dep else int(p["duration_min"]),
            int(p.get("trips_before", 0)) if p.get("trips_before") is not None else "—",
            int(p.get("trips_after",  0)) if p.get("trips_after")  is not None else "—",
            p["address"],
        ]
        aligns2 = [LEFT, LEFT, CENTER, CENTER, CENTER, CENTER, CENTER, CENTER, LEFT]
        for ci, (val, aln) in enumerate(zip(row2, aligns2), 1):
            _dat(ws2, ri, ci, val, align=aln)

    _auto_width(ws2)

    buf = _io_mod.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _render_export_button(
    result: dict,
    available_days: list,
    day_summary: dict,
    df_raw: "pd.DataFrame | None" = None,
) -> None:
    """Bouton export Excel + contrôle du délai minimum entre les deux passages dépôt."""
    col_gap, col_export = st.columns([3, 4])

    with col_gap:
        new_gap = st.number_input(
            "⏱️ Délai min entre les 2 passages (heures)",
            min_value=0.0, max_value=12.0,
            value=float(st.session_state.get("q_min_gap_hours", result.get("min_gap_hours", 0.0))),
            step=0.5,
            key="q_min_gap_hours",
            help="Le passage 'après tournée' ne sera retenu que s'il survient au moins X heures après le passage 'avant tournée'.",
        )
        if new_gap != result.get("min_gap_hours", 0.0):
            depot_coords = result.get("depot_coords")
            if depot_coords and df_raw is not None:
                _cache = load_geocode_cache()
                passages = _find_depot_passages(df_raw, _cache, tuple(depot_coords), min_gap_hours=new_gap)
                st.session_state["q_passages_result"]["passages"]      = passages
                st.session_state["q_passages_result"]["min_gap_hours"] = new_gap
                st.rerun()

    with col_export:
        plate = result.get("plate", "export")
        try:
            xlsx_bytes = _build_export_excel(result, available_days, day_summary, df_raw=df_raw)
            filename = (
                f"passages_depot_{plate}_"
                f"{available_days[0].strftime('%Y%m%d')}_"
                f"{available_days[-1].strftime('%Y%m%d')}.xlsx"
            )
            st.markdown("<div style='margin-top:28px'>", unsafe_allow_html=True)
            st.download_button(
                label="⬇️ Exporter en Excel",
                data=xlsx_bytes,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="q_export_xlsx",
                use_container_width=True,
            )
            st.markdown("</div>", unsafe_allow_html=True)
        except Exception as e:
            st.warning(f"Export indisponible : {e}")

def _render_passages_result(
    result: dict,
    available_days: list,
    df_raw: "pd.DataFrame | None" = None,
) -> None:
    """Affiche les métriques + tableau + détail pour un résultat d'analyse."""
    import datetime as _dt

    passages = result["passages"]
    nb_days  = result["nb_days"]
    df_pass  = pd.DataFrame(passages) if passages else pd.DataFrame(
        columns=["date", "period", "arrive_time", "depart_time", "duration_min", "address"]
    )

    days_matin  = set(df_pass[df_pass["period"] == "matin"]["date"].unique()) if not df_pass.empty else set()
    days_apmidi = set(df_pass[df_pass["period"] == "après-midi"]["date"].unique()) if not df_pass.empty else set()
    nb_matin    = len(days_matin)
    nb_apmidi   = len(days_apmidi)

    # Bandeau dépôt
    depot_coords = result.get("depot_coords")
    if depot_coords:
        lat, lon = depot_coords[0], depot_coords[1]
        maps_url = f"https://www.google.com/maps?q={lat},{lon}"
        coords_str = f"{lat:.5f}, {lon:.5f}"
    else:
        maps_url   = None
        coords_str = "coordonnées inconnues"

    st.markdown(
        f"<div style='background:{C_BLUE_DARK};padding:10px 16px;border-radius:8px;"
        f"margin:16px 0 4px'>"
        f"<span style='color:{C_BLUE_LIGHT};font-size:18px'>🏠</span>"
        f"&nbsp;<span style='color:white;font-size:14px'>"
        f"<b>Dépôt analysé :</b> {result['depot_addr']}"
        f" &nbsp;·&nbsp; tolérance {DEPOT_RADIUS_M} m</span>"
        f"</div>",
        unsafe_allow_html=True,
    )
    if maps_url:
        st.caption(
            f"📍 Coordonnées géocodées : `{coords_str}` — "
            f"[Vérifier sur Google Maps]({maps_url})"
        )

    # Seuil de confiance géocodage
    col_score, _ = st.columns([2, 3])
    with col_score:
        score_threshold = st.slider(
            "🎯 Seuil de confiance géocodage",
            min_value=0.1, max_value=1.0, value=_GEOCODE_MIN_SCORE, step=0.05,
            key="q_score_threshold",
            help="Score minimum retourné par l'API adresse.gouv.fr. "
                 "Plus il est élevé, moins d'adresses sont acceptées mais plus elles sont précises. "
                 "En dessous du seuil → adresse rejetée (apparaît dans les non-géocodées).",
        )

    # Métriques
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("📅 Jours analysés", nb_days)
    c2.metric("☀️ Passages matin",    nb_matin,
              delta=f"{nb_matin/nb_days*100:.0f} % des jours" if nb_days else None)
    c3.metric("🌇 Passages après-midi", nb_apmidi,
              delta=f"{nb_apmidi/nb_days*100:.0f} % des jours" if nb_days else None)
    c4.metric("📍 Total passages détectés", len(passages))

    st.divider()

    # Tableau par jour
    st.markdown("#### 📆 Récapitulatif par jour")
    st.caption("🟢 Matin ET après-midi  ·  🟡 Un seul passage  ·  🔴 Aucun passage détecté")

    day_summary: dict = {d: {"matin": None, "après-midi": None} for d in available_days}
    for p in passages:
        d   = p["date"]
        per = p["period"]
        if d in day_summary:
            existing = day_summary[d][per]
            if existing is None or p["arrive_time"] < existing["arrive_time"]:
                day_summary[d][per] = p

    def _fmt(p: dict | None) -> str:
        if p is None:
            return "—"
        t   = p["arrive_time"].strftime("%H:%M")
        dep = p.get("depart_time")
        if dep is None or pd.isna(dep):
            return f"{t}  (? min)"
        return f"{t}  ({int(p['duration_min'])} min)"

    def _fmt_trips(p: dict | None, key: str) -> str:
        if p is None:
            return "—"
        v = p.get(key)
        return str(int(v)) if v is not None else "—"

    rows = []
    for d in available_days:
        m   = day_summary[d]["matin"]
        apm = day_summary[d]["après-midi"]
        rows.append({
            "Date":              d.strftime("%d/%m/%Y"),
            "Jour":              d.strftime("%A").capitalize(),
            "🌅 Matin":          _fmt(m),
            "Av. matin":         _fmt_trips(m, "trips_before"),
            "Ap. matin":         _fmt_trips(m, "trips_after"),
            "🌆 Après-midi":     _fmt(apm),
            "Av. A-M":           _fmt_trips(apm, "trips_before"),
            "Ap. A-M":           _fmt_trips(apm, "trips_after"),
            "_m": m is not None,
            "_a": apm is not None,
        })

    df_t = pd.DataFrame(rows)

    def _col_row(row):
        if row["_m"] and row["_a"]:
            bg = "background-color:#d4edda"
        elif row["_m"] or row["_a"]:
            bg = "background-color:#fff3cd"
        else:
            bg = "background-color:#f8d7da"
        return [bg] * len(row)

    st.dataframe(
        df_t.style.apply(_col_row, axis=1).hide(axis="columns", subset=["_m", "_a"]),
        use_container_width=True,
        hide_index=True,
    )

    # Bouton export Excel
    _render_export_button(result, available_days, day_summary, df_raw=df_raw)

    # Détail dépliable
    st.divider()
    with st.expander(f"🔎 Détail de tous les passages ({len(passages)} détecté(s))", expanded=False):
        if df_pass.empty:
            st.info("Aucun passage détecté sur la période analysée.")
        else:
            det = df_pass.copy()
            det["Date"]            = det["date"].apply(lambda d: d.strftime("%d/%m/%Y"))
            det["Période"]         = det["period"].apply(lambda p: "Matin" if p == "matin" else "Après-midi")
            det["Arrivée"]         = det["arrive_time"].dt.strftime("%H:%M")
            det["Départ dépôt"]    = det["depart_time"].apply(lambda t: t.strftime("%H:%M") if pd.notna(t) else "—")
            det["Durée (min)"]     = det.apply(
                lambda r: "?" if (r["depart_time"] is None or pd.isna(r["depart_time"])) else f"{int(r['duration_min'])}",
                axis=1,
            )
            det["Trajets avant"]   = det.apply(lambda r: str(int(r["trips_before"])) if "trips_before" in r and pd.notna(r["trips_before"]) else "—", axis=1)
            det["Trajets après"]   = det.apply(lambda r: str(int(r["trips_after"])) if "trips_after" in r and pd.notna(r["trips_after"]) else "—", axis=1)
            det["Adresse"]         = det["address"]
            st.dataframe(
                det[["Date", "Période", "Arrivée", "Départ dépôt", "Durée (min)",
                      "Trajets avant", "Trajets après", "Adresse"]],
                use_container_width=True, hide_index=True,
            )

    # Adresses non géocodées
    failed = result.get("failed_addrs")
    if failed is None:
        st.info("ℹ️ Relancez **Analyser** pour voir les adresses non géocodées.")
    else:
        label = (f"⚠️ {len(failed)} adresse(s) non géocodée(s) — invisibles pour la détection"
                 if failed else "✅ Toutes les adresses ont été géocodées")
        with st.expander(label, expanded=bool(failed)):
            if not failed:
                st.caption("Toutes les adresses du fichier ont été géocodées avec succès.")
            else:
                st.caption(
                    "Ces adresses n'ont pas pu être localisées automatiquement. "
                    "Renseignez les coordonnées GPS manuellement pour les inclure dans l'analyse."
                )
                any_saved = False
                for addr in list(failed):
                    col_addr, col_input, col_btn = st.columns([3, 2, 1])
                    with col_addr:
                        st.markdown(f"`{addr}`")
                    with col_input:
                        coords_raw = st.text_input(
                            "Coordonnées GPS",
                            placeholder="lat, lon  ex: 50.638, 2.979",
                            key=f"q_manual_coords_{addr}",
                            label_visibility="collapsed",
                        )
                    with col_btn:
                        if st.button("✅ Valider", key=f"q_manual_save_{addr}",
                                     use_container_width=True):
                            coords_input = coords_raw.strip()
                            coords_parsed = None
                            # Essai parsing direct lat, lon
                            m = re.match(r"^\s*(-?\d+\.?\d*)\s*[,\s]\s*(-?\d+\.?\d*)\s*$",
                                         coords_input)
                            if m:
                                coords_parsed = (float(m.group(1)), float(m.group(2)))
                            else:
                                # Tentative géocodage via API
                                tmp = _geocode_single_fr(coords_input)
                                if tmp:
                                    coords_parsed = tmp
                            if coords_parsed:
                                save_geocode_entry(addr, coords_parsed)
                                # Retirer de la liste des échecs
                                new_failed = [a for a in
                                              st.session_state["q_passages_result"]["failed_addrs"]
                                              if a != addr]
                                st.session_state["q_passages_result"]["failed_addrs"] = new_failed
                                any_saved = True
                                st.success(f"Sauvegardé : {coords_parsed[0]:.5f}, {coords_parsed[1]:.5f}")
                            else:
                                st.error("Coordonnées invalides ou adresse introuvable.")

                # Bouton relancer l'analyse si au moins une adresse a été corrigée
                if any_saved or st.session_state.get("q_manual_rerun_ready"):
                    st.session_state["q_manual_rerun_ready"] = True
                    if st.button("🔄 Relancer l'analyse avec les corrections",
                                 key="q_manual_rerun", type="primary"):
                        cache = load_geocode_cache()
                        depot_coords = result.get("depot_coords")
                        if depot_coords and df_raw is not None:
                            passages = _find_depot_passages(
                                df_raw, cache, tuple(depot_coords),
                                min_gap_hours=result.get("min_gap_hours", 0.0),
                            )
                            all_addrs = (
                                set(df_raw["Lieu de départ"].dropna().astype(str))
                                | set(df_raw["Lieu d'arrivée"].dropna().astype(str))
                            )
                            new_failed = sorted(a for a in all_addrs if not cache.get(a))
                            st.session_state["q_passages_result"]["passages"]    = passages
                            st.session_state["q_passages_result"]["failed_addrs"] = new_failed
                            st.session_state.pop("q_manual_rerun_ready", None)
                            st.rerun()

    # Adresses géocodées — affichage + correction
    if df_raw is not None:
        _cache_disp = load_geocode_cache()
        _all_addrs_disp = sorted(
            set(df_raw["Lieu de départ"].dropna().astype(str).unique())
            | set(df_raw["Lieu d'arrivée"].dropna().astype(str).unique())
        )
        _found_disp = [(a, _cache_disp[a]) for a in _all_addrs_disp if _cache_disp.get(a)]
        if _found_disp:
            with st.expander(
                f"✅ {len(_found_disp)} adresse(s) géocodée(s) — cliquez pour vérifier et corriger",
                expanded=False,
            ):
                st.caption(
                    "Vérifiez que chaque point est bien positionné sur Google Maps. "
                    "Si les coordonnées sont fausses, saisissez les bonnes (lat, lon) et cliquez ✅."
                )
                for _addr, _coords in _found_disp:
                    _lat, _lon = _coords
                    _maps_url = f"https://www.google.com/maps?q={_lat},{_lon}"
                    _fix_key  = f"q_found_fix_{hash(_addr)}"
                    if _fix_key not in st.session_state:
                        st.session_state[_fix_key] = f"{_lat:.5f}, {_lon:.5f}"
                    _col_addr, _col_inp, _col_btn = st.columns([4, 3, 1])
                    with _col_addr:
                        st.markdown(f"`{_addr}`  \n[📍 Voir sur Google Maps]({_maps_url})")
                    with _col_inp:
                        st.text_input(
                            "Coordonnées GPS",
                            key=_fix_key,
                            label_visibility="collapsed",
                            placeholder="lat, lon",
                        )
                    with _col_btn:
                        st.markdown("<div style='margin-top:4px'>", unsafe_allow_html=True)
                        if st.button(
                            "✅",
                            key=f"q_found_save_{hash(_addr)}",
                            use_container_width=True,
                            help="Sauvegarder les coordonnées corrigées",
                        ):
                            _raw = st.session_state[_fix_key].strip()
                            _m = re.match(
                                r"^\s*(-?\d+\.?\d*)\s*[,\s]\s*(-?\d+\.?\d*)\s*$", _raw
                            )
                            if _m:
                                _new_coords = (float(_m.group(1)), float(_m.group(2)))
                                save_geocode_entry(_addr, _new_coords)
                                st.success(f"✅ {_addr[:40]} → {_new_coords[0]:.5f}, {_new_coords[1]:.5f}")
                                st.rerun()
                            else:
                                with st.spinner("Géocodage…"):
                                    _new_coords = _geocode_single_fr(_raw)
                                if _new_coords:
                                    save_geocode_entry(_addr, _new_coords)
                                    st.success(f"✅ {_addr[:40]} → {_new_coords[0]:.5f}, {_new_coords[1]:.5f}")
                                    st.rerun()
                                else:
                                    st.error("Coordonnées invalides ou adresse introuvable.")
                        st.markdown("</div>", unsafe_allow_html=True)

    # Méthodologie
    with st.expander("ℹ️ Méthodologie de détection des passages", expanded=False):
        st.markdown(f"""
Un **passage au dépôt** est détecté quand l'adresse d'**arrivée** ou de **départ**
est à moins de **{DEPOT_RADIUS_M} m** du dépôt renseigné.

La classification se fait par **ordre chronologique**, pas par heure de la journée :

- 🚀 **Avant tournée** : premier passage de la journée avec une durée d'arrêt ≥ {MIN_STOP_MIN} min
- 🏁 **Après tournée** : passage suivant avec une durée ≥ {MIN_STOP_MIN} min.
  Si aucun n'atteint {MIN_STOP_MIN} min, le premier passage trouvé après la tournée est quand même affiché avec sa durée réelle (ou **?** si le départ n'est pas enregistré).

Géocodage : **API adresse.data.gouv.fr** (gouvernement français) — gratuite, sans clé, optimisée France.
        """)


def _render_passages_cache_admin() -> None:
    """Boutons de gestion du cache et de la mémoire de l'onglet Passages."""
    st.divider()
    with st.expander("🗑️ Cache & Mémoire", expanded=False):
        col1, col2, col3, _ = st.columns([2, 2, 2, 1])
        with col1:
            if st.button("🗑️ Effacer l'analyse en mémoire", use_container_width=True,
                         key="q_clear_result"):
                st.session_state.pop("q_passages_result", None)
                st.success("✅ Résultat effacé.")
                st.rerun()
        with col2:
            if st.button("🗑️ Vider cache géocodage", use_container_width=True,
                         key="q_clear_geocache"):
                n = clear_geocode_cache()
                st.success(f"✅ {n} adresse(s) supprimée(s).")
                st.rerun()
        with col3:
            if st.button("🗑️ Tout vider", type="primary", use_container_width=True,
                         key="q_clear_all"):
                st.session_state.pop("q_passages_result", None)
                ng = clear_geocode_cache()
                st.success(f"✅ Résultat + {ng} adresse(s) géocodée(s) supprimés.")
                st.rerun()


# ── Onglet 2 : Analyse Passages Dépôt ────────────────────────────────────────

def _render_tab_passages() -> None:

    # En-tête
    st.markdown(
        f"<div style='background:linear-gradient(135deg,{C_BLUE_DARK},{C_BLUE_LIGHT});"
        f"padding:16px 20px;border-radius:10px;margin-bottom:20px'>"
        f"<span style='color:white;font-size:17px;font-weight:700'>📍 Analyse des Passages Dépôt</span>"
        f"<p style='color:rgba(255,255,255,0.85);margin:4px 0 0;font-size:13px'>"
        f"Sélectionnez un réappro, confirmez l'adresse dépôt, puis lancez l'analyse "
        f"pour voir les passages matin et après-midi sur toute la période du fichier.</p>"
        f"</div>",
        unsafe_allow_html=True,
    )

    # ── Récupération du fichier depuis l'onglet Carte ─────
    uploaded       = st.session_state.get("quartix_uploader")
    xls            = None
    vehicle_sheets: list[str] = []
    name_map:       dict[str, str] = {}

    if uploaded:
        try:
            xls    = pd.ExcelFile(uploaded)
            sheets = xls.sheet_names
            if len(sheets) > 1:
                vehicle_sheets = sheets[1:]
                name_map       = _parse_vehicle_names(xls)
        except Exception:
            xls = None

    # Si aucun fichier chargé : afficher résultat en cache si dispo + gestion cache
    if xls is None:
        st.info(
            "📂 Importez d'abord un fichier QUARTIX dans l'onglet **Carte & Trajets** "
            "pour lancer une nouvelle analyse."
        )
        result = st.session_state.get("q_passages_result")
        if result:
            st.markdown("---")
            st.caption("📋 Dernier résultat enregistré en mémoire :")
            available_days = [
                pd.Timestamp(d).date() for d in result.get("available_days", [])
            ] or sorted({p["date"] for p in result["passages"]})
            _render_passages_result(result, available_days)
        _render_passages_cache_admin()
        return

    # ── Sélecteur réappro ─────────────────────────────────
    options_labels = [
        f"{p}  —  {name_map[p]}" if p in name_map else p
        for p in vehicle_sheets
    ]
    options_map    = dict(zip(options_labels, vehicle_sheets))

    col_sel, col_info = st.columns([3, 2])
    with col_sel:
        def _on_vehicle_change():
            st.session_state.pop("q_passages_depot_input", None)

        sel_label = st.selectbox("🚗 Réappro à analyser", options_labels,
                                 key="q_passages_vehicle_label",
                                 on_change=_on_vehicle_change)
    selected_plate = options_map[sel_label]

    # ── Chargement données ────────────────────────────────
    try:
        df_raw = pd.read_excel(xls, sheet_name=selected_plate, header=4, usecols="B:N")
        df_raw.columns = COLS
    except Exception as e:
        st.error(f"Erreur lecture feuille «{selected_plate}» : {e}")
        return

    df_raw["_dep"] = pd.to_datetime(df_raw["Départ"].astype(str).str.replace(r'\s+[A-Z]{2,5}$', '', regex=True),  errors="coerce")
    df_raw["_arr"] = pd.to_datetime(df_raw["Arrivée"].astype(str).str.replace(r'\s+[A-Z]{2,5}$', '', regex=True), errors="coerce")
    df_raw = df_raw.dropna(subset=["_dep"]).sort_values("_dep").reset_index(drop=True)
    for col in ["Lieu de départ", "Lieu d'arrivée"]:
        df_raw[col] = df_raw[col].astype(str).map(_clean_addr)

    if df_raw.empty:
        st.warning("Aucune donnée valide dans cette feuille.")
        return

    nb_days_file   = df_raw["_dep"].dt.date.nunique()
    available_days = sorted(df_raw["_dep"].dt.date.unique())
    date_min       = available_days[0]
    date_max       = available_days[-1]

    with col_info:
        st.markdown(
            f"<div style='background:#f0f4fa;border-radius:8px;padding:10px 14px;"
            f"margin-top:24px;font-size:13px;color:{C_BLUE_DARK}'>"
            f"<b>Période :</b> {date_min.strftime('%d/%m/%Y')} → {date_max.strftime('%d/%m/%Y')}<br>"
            f"<b>Jours :</b> {nb_days_file} &nbsp;|&nbsp; <b>Trajets :</b> {len(df_raw)}"
            f"</div>",
            unsafe_allow_html=True,
        )

    # ── Adresse dépôt + bouton Analyser ──────────────────
    vehicle_doc      = load_quartix_vehicle(selected_plate)
    db_depot_address = vehicle_doc.get("depot_address", "") if vehicle_doc else ""
    auto_detect      = _auto_detect_depot(df_raw)
    default_depot    = db_depot_address or auto_detect or ""
    is_approximated  = bool(not db_depot_address and auto_detect)

    col_depot, col_btn = st.columns([5, 1])
    with col_depot:
        depot_input = st.text_input(
            "📍 Adresse du dépôt",
            value=default_depot,
            placeholder="Ex : 39 Av. de la Pépinière, 59320 Haubourdin  ou  50.638, 2.979",
            key="q_passages_depot_input",
            help="Adresse texte OU coordonnées GPS : latitude, longitude (ex : 50.6384, 2.9794)",
        )
        if is_approximated:
            st.caption("🔸 Adresse auto-détectée — vérifiez et corrigez si nécessaire avant d'analyser")
    with col_btn:
        st.markdown("<div style='margin-top:28px'>", unsafe_allow_html=True)
        run_btn = st.button("🔍 Analyser", type="primary", key="q_passages_run",
                            use_container_width=True)
        st.markdown("</div>", unsafe_allow_html=True)

    min_gap_hours = st.session_state.get("q_min_gap_hours", 0.0)

    # ── Lancement de l'analyse ────────────────────────────
    if run_btn:
        depot_addr_clean = depot_input.strip()
        if not depot_addr_clean:
            st.warning("Veuillez saisir une adresse ou des coordonnées GPS.")
            return

        # Détection coordonnées directes : "50.638, 2.979" ou "50.638 2.979"
        _coord_match = re.match(
            r"^\s*(-?\d+\.?\d*)\s*[,\s]\s*(-?\d+\.?\d*)\s*$",
            depot_addr_clean,
        )
        if _coord_match:
            try:
                depot_coords = (float(_coord_match.group(1)), float(_coord_match.group(2)))
                cache = load_geocode_cache()
            except ValueError:
                depot_coords = None
        else:
            with st.spinner("Géocodage du dépôt…"):
                cache = load_geocode_cache()
                cache = _geocode_all([depot_addr_clean], cache)
                depot_coords = cache.get(depot_addr_clean)

        if not depot_coords:
            st.error("❌ Adresse dépôt introuvable. Vérifiez l'orthographe ou entrez les coordonnées GPS.")
            return

        all_addrs: set[str] = set()
        for c in ["Lieu de départ", "Lieu d'arrivée"]:
            all_addrs.update(df_raw[c].dropna().astype(str).unique())

        with st.spinner("Géocodage des adresses du fichier…"):
            cache = _geocode_all(list(all_addrs), cache)

        failed_addrs = sorted(a for a in all_addrs if not cache.get(a))

        with st.spinner("Analyse des passages…"):
            passages = _find_depot_passages(df_raw, cache, tuple(depot_coords), min_gap_hours=min_gap_hours)

        st.session_state["q_passages_result"] = {
            "plate":          selected_plate,
            "depot_addr":     depot_addr_clean,
            "depot_coords":   depot_coords,
            "passages":       passages,
            "nb_days":        nb_days_file,
            "available_days": [d.isoformat() for d in available_days],
            "failed_addrs":   failed_addrs,
            "min_gap_hours":  min_gap_hours,
        }

    # ── Affichage des résultats ───────────────────────────
    result = st.session_state.get("q_passages_result")
    if result is None:
        st.info("👆 Sélectionnez un réappro, vérifiez l'adresse dépôt, puis cliquez **Analyser**.")
        _render_passages_cache_admin()
        return

    if result["plate"] != selected_plate:
        st.info("Le réappro sélectionné a changé. Cliquez **Analyser** pour recalculer.")
        _render_passages_cache_admin()
        return

    _render_passages_result(result, available_days, df_raw=df_raw)
    _render_passages_cache_admin()


# ── Onglet 3 : Analyse Hebdomadaire ──────────────────────────────────────────

_JOURS_FR = {0: "Lundi", 1: "Mardi", 2: "Mercredi", 3: "Jeudi",
             4: "Vendredi", 5: "Samedi", 6: "Dimanche"}

_MIN_TRIP_KM = 1.0  # seuil minimal — trajet < 1 km ignoré


def _normalize_plate(plate: str) -> str:
    """Normalise une plaque pour la comparaison (retire espaces/tirets, majuscules)."""
    return re.sub(r'[\s\-\.]', '', str(plate)).upper()


def _vehicle_label(plate: str, vehicles_db: dict) -> str:
    """Retourne 'PLAQUE — Prénom' si connu, sinon juste la plaque."""
    prenom = vehicles_db.get(plate, {}).get("prenom", "").strip()
    return f"{plate}  —  {prenom}" if prenom else plate


def _build_hebdo_excel(
    date_min,
    date_max,
    cutoff,
    late_trips: dict,
    weekend_trips: dict,
    vehicles_db: dict,
) -> bytes:
    """Génère un rapport Excel de l'analyse hebdomadaire en mémoire et retourne les bytes."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    HDR_FILL   = PatternFill("solid", fgColor="1B3D6F")
    HDR_FONT   = Font(color="FFFFFF", bold=True, size=11)
    WARN_FILL  = PatternFill("solid", fgColor="FDECEA")
    WKND_FILL  = PatternFill("solid", fgColor="FFF3E0")
    OK_FILL    = PatternFill("solid", fgColor="E8F5E9")
    THIN       = Side(style="thin", color="CCCCCC")
    BORDER     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CENTER     = Alignment(horizontal="center", vertical="center")
    WRAP       = Alignment(wrap_text=True, vertical="center")

    def _style_header_row(ws, row_idx, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = CENTER
            cell.border = BORDER

    def _auto_width(ws):
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    # ── Feuille Résumé ──────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Résumé"

    ws_sum["A1"] = "Rapport Hebdomadaire QUARTIX"
    ws_sum["A1"].font = Font(bold=True, size=14, color="1B3D6F")
    ws_sum.merge_cells("A1:E1")
    ws_sum["A1"].alignment = CENTER

    ws_sum["A2"] = f"Période : {date_min.strftime('%d/%m/%Y')} → {date_max.strftime('%d/%m/%Y')}"
    ws_sum["A2"].font = Font(italic=True, size=11)
    ws_sum.merge_cells("A2:E2")

    ws_sum["A3"] = f"Heure limite tournée : {cutoff.strftime('%H:%M')}"
    ws_sum["A3"].font = Font(italic=True, size=11)
    ws_sum.merge_cells("A3:E3")

    headers_sum = ["Véhicule", "Réappro", "Trajets hors horaires", "Activité weekend", "Statut"]
    for ci, h in enumerate(headers_sum, 1):
        ws_sum.cell(row=5, column=ci, value=h)
    _style_header_row(ws_sum, 5, len(headers_sum))

    all_plates = sorted(set(list(late_trips.keys()) + list(weekend_trips.keys()) +
                            [p for p in vehicles_db]))
    row = 6
    for plate in all_plates:
        prenom     = vehicles_db.get(plate, {}).get("prenom", "").strip()
        nb_late    = len(late_trips.get(plate, []))
        nb_wknd    = len(weekend_trips.get(plate, []))
        if nb_late == 0 and nb_wknd == 0:
            statut = "✅ RAS"
            fill   = OK_FILL
        else:
            parts  = []
            if nb_late:  parts.append("hors horaires")
            if nb_wknd:  parts.append("weekend")
            statut = "⚠️ " + " + ".join(parts)
            fill   = WARN_FILL
        for ci, val in enumerate([plate, prenom, nb_late or "", nb_wknd or "", statut], 1):
            c = ws_sum.cell(row=row, column=ci, value=val)
            c.fill = fill
            c.border = BORDER
            c.alignment = CENTER
        row += 1

    _auto_width(ws_sum)

    # ── Feuille Trajets hors horaires ───────────────────────────────────────
    ws_late = wb.create_sheet("Trajets hors horaires")
    headers_late = ["Véhicule", "Réappro", "Date", "Jour", "Heure départ", "Heure arrivée",
                    "Lieu de départ", "Lieu d'arrivée", "Distance (km)"]
    for ci, h in enumerate(headers_late, 1):
        ws_late.cell(row=1, column=ci, value=h)
    _style_header_row(ws_late, 1, len(headers_late))

    row = 2
    for plate, trips in late_trips.items():
        prenom = vehicles_db.get(plate, {}).get("prenom", "").strip()
        for t in trips:
            vals = [
                plate, prenom,
                t["date"].strftime("%d/%m/%Y"), t["jour"],
                t["heure_dep"], t["heure_arr"],
                t["lieu_dep"], t["lieu_arr"],
                t["distance_km"],
            ]
            for ci, val in enumerate(vals, 1):
                c = ws_late.cell(row=row, column=ci, value=val)
                c.fill = WARN_FILL
                c.border = BORDER
                c.alignment = CENTER if ci not in (7, 8) else WRAP
            row += 1
    _auto_width(ws_late)

    # ── Feuille Activité weekend ────────────────────────────────────────────
    ws_wknd = wb.create_sheet("Activité weekend")
    headers_wknd = ["Véhicule", "Réappro", "Date", "Jour", "Heure départ", "Distance (km)"]
    for ci, h in enumerate(headers_wknd, 1):
        ws_wknd.cell(row=1, column=ci, value=h)
    _style_header_row(ws_wknd, 1, len(headers_wknd))

    row = 2
    for plate, trips in weekend_trips.items():
        prenom = vehicles_db.get(plate, {}).get("prenom", "").strip()
        for t in trips:
            vals = [plate, prenom, t["date"].strftime("%d/%m/%Y"), t["jour"],
                    t["heure_dep"], t["distance_km"]]
            for ci, val in enumerate(vals, 1):
                c = ws_wknd.cell(row=row, column=ci, value=val)
                c.fill = WKND_FILL
                c.border = BORDER
                c.alignment = CENTER
            row += 1
    _auto_width(ws_wknd)

    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_hebdo_mail(
    resp_nom: str,
    date_min,
    date_max,
    cutoff,
    late_trips: dict,
    weekend_trips: dict,
    vehicles_db: dict,
) -> str:
    date_str_min = date_min.strftime("%d/%m/%Y")
    date_str_max = date_max.strftime("%d/%m/%Y")
    cutoff_str   = cutoff.strftime("%H:%M")

    lines = [
        f"Objet : Rapport hebdomadaire des trajets — semaine du {date_str_min} au {date_str_max}",
        "",
        "Bonjour,",
        "",
        f"Suite à l'analyse des données Quartix pour la semaine du {date_str_min} au {date_str_max},",
    ]

    if not late_trips and not weekend_trips:
        lines += [
            "aucune anomalie n'a été détectée sur cette période.",
            "",
            "Les tournées se sont déroulées dans les horaires et jours habituels.",
        ]
    else:
        lines.append("voici les points d'attention identifiés :")

        if late_trips:
            lines += ["", "─" * 60, f"TRAJETS HORS HORAIRES (après {cutoff_str})", "─" * 60]
            for plate, trips in late_trips.items():
                doc   = vehicles_db.get(plate, {})
                prenom = doc.get("prenom", "").strip()
                vlabel = f"{plate}" + (f" ({prenom})" if prenom else "")
                lines.append(f"\nVéhicule {vlabel} :")
                for t in trips:
                    lines.append(
                        f"  • {t['date'].strftime('%d/%m/%Y')} ({t['jour']}) : "
                        f"{t['heure_dep']} (heure de départ), {t['distance_km']} km"
                    )

        if weekend_trips:
            lines += ["", "─" * 60, "ACTIVITÉ WEEKEND", "─" * 60]
            for plate, days in weekend_trips.items():
                doc    = vehicles_db.get(plate, {})
                prenom = doc.get("prenom", "").strip()
                vlabel = f"{plate}" + (f" ({prenom})" if prenom else "")
                lines.append(f"\nVéhicule {vlabel} :")
                for d in days:
                    lines.append(
                        f"  • {d['jour']} {d['date'].strftime('%d/%m/%Y')} : "
                        f"{d['heure_dep']} (heure de départ), {d['distance_km']} km"
                    )

        lines += ["", "Cordialement,"]

    lines += [""]
    return "\n".join(lines)


def _render_hebdo_vehicle_editor(vehicle_sheets: list[str], vehicles_db: dict) -> dict:
    """
    Affiche une section éditable des informations conducteurs/zones.
    Retourne le vehicles_db mis à jour après une éventuelle sauvegarde.
    """
    with st.expander("⚙️ Informations conducteurs & zones (édition)", expanded=False):

        # ── Sous-section : import depuis le guide Excel ────
        st.markdown("**📥 Importer depuis le guide Excel (Reappro_Guide_Plaque)**")
        guide_file = st.file_uploader(
            "Fichier guide plaques (.xlsx)",
            type=["xls", "xlsx"],
            key="hebdo_guide_uploader",
        )
        if guide_file and st.button("⬇️ Importer", key="hebdo_guide_import"):
            try:
                df_guide = pd.read_excel(guide_file)
                df_guide.columns = [c.strip() for c in df_guide.columns]
                col_map = {c.lower(): c for c in df_guide.columns}

                plate_col  = col_map.get("plaque")
                prenom_col = col_map.get("prenom")
                zone_col   = col_map.get("zone") or col_map.get("zone géographique")
                resp_col   = col_map.get("responsable")

                if not plate_col:
                    st.error("Colonne 'Plaque' introuvable dans le guide.")
                else:
                    guide_index: dict[str, dict] = {}
                    for _, row in df_guide.iterrows():
                        raw = str(row[plate_col]).strip()
                        if raw and raw.lower() != "nan":
                            guide_index[_normalize_plate(raw)] = {
                                "prenom":      str(row[prenom_col]).strip() if prenom_col and pd.notna(row.get(prenom_col)) else "",
                                "zone":        str(row[zone_col]).strip()   if zone_col   and pd.notna(row.get(zone_col))   else "",
                                "responsable": str(row[resp_col]).strip()   if resp_col   and pd.notna(row.get(resp_col))   else "",
                            }

                    matched, unmatched = 0, []
                    for plate in vehicle_sheets:
                        info = guide_index.get(_normalize_plate(plate))
                        if info:
                            upsert_quartix_vehicle_info(
                                plate=plate,
                                prenom=info["prenom"],
                                zone=info["zone"],
                                responsable=info["responsable"],
                            )
                            matched += 1
                        else:
                            unmatched.append(plate)

                    st.success(f"✅ {matched} véhicule(s) mis à jour.")
                    if unmatched:
                        st.warning(f"Plaques non trouvées dans le guide : {', '.join(unmatched)}")
                    try:
                        vehicles_db = load_all_quartix_vehicles()
                    except Exception:
                        pass
                    st.rerun()
            except Exception as e:
                st.error(f"Erreur lecture guide : {e}")

        # ── Tableau éditable ───────────────────────────────
        st.markdown("**Modifiez directement les informations puis cliquez Sauvegarder.**")

        rows = []
        for plate in vehicle_sheets:
            doc = vehicles_db.get(plate, {})
            rows.append({
                "Plaque":       plate,
                "Conducteur":   doc.get("prenom", ""),
                "Zone":         doc.get("zone", ""),
                "Responsable":  doc.get("responsable", ""),
            })

        edited_df = st.data_editor(
            pd.DataFrame(rows),
            column_config={
                "Plaque":      st.column_config.TextColumn("Plaque", disabled=True, width="small"),
                "Conducteur":  st.column_config.TextColumn("Conducteur", width="medium"),
                "Zone":        st.column_config.TextColumn("Zone", width="medium"),
                "Responsable": st.column_config.TextColumn("Responsable", width="medium"),
            },
            hide_index=True,
            use_container_width=True,
            key="hebdo_veh_editor",
        )

        if st.button("💾 Sauvegarder les modifications", type="primary", key="hebdo_veh_save"):
            for _, row in edited_df.iterrows():
                upsert_quartix_vehicle_info(
                    plate=str(row["Plaque"]),
                    prenom=str(row["Conducteur"]).strip(),
                    zone=str(row["Zone"]).strip(),
                    responsable=str(row["Responsable"]).strip(),
                )
            st.success("✅ Informations sauvegardées.")
            try:
                vehicles_db = load_all_quartix_vehicles()
            except Exception:
                pass
            st.rerun()

    return vehicles_db


def _render_tab_hebdo() -> None:
    from datetime import time as _dtime

    st.markdown(
        f"<div style='background:linear-gradient(135deg,{C_BLUE_DARK},{C_BLUE_LIGHT});"
        f"padding:16px 20px;border-radius:10px;margin-bottom:20px'>"
        f"<span style='color:white;font-size:17px;font-weight:700'>📅 Analyse Hebdomadaire</span>"
        f"<p style='color:rgba(255,255,255,0.85);margin:4px 0 0;font-size:13px'>"
        f"Importez un rapport de trajets sur une semaine complète pour détecter les trajets "
        f"hors horaires et les activités weekend, puis générez un mail récapitulatif.</p>"
        f"</div>",
        unsafe_allow_html=True,
    )

    # ── 1. Upload ─────────────────────────────────────────
    col_up, _ = st.columns([2, 3])
    with col_up:
        uploaded = st.file_uploader(
            "📂 Importer un rapport Quartix (semaine complète)",
            type=["xls", "xlsx"],
            key="quartix_hebdo_uploader",
        )

    # Efface le mail généré si un nouveau fichier est chargé
    current_file_id = uploaded.name if uploaded else None
    if current_file_id != st.session_state.get("hebdo_last_file_id"):
        st.session_state["hebdo_last_file_id"] = current_file_id
        st.session_state.pop("hebdo_mail_text", None)

    if not uploaded:
        st.markdown(
            f"<div style='background:#f0f4fa;border-left:5px solid {C_BLUE_DARK};"
            f"padding:16px 20px;border-radius:8px;margin-top:20px'>"
            f"<b style='color:{C_BLUE_DARK}'>👆 Importez un fichier Excel QUARTIX (semaine)</b>"
            f"<p style='margin:6px 0 0;color:#555;font-size:14px'>"
            f"Format standard QUARTIX multi-feuilles : feuille résumé + une feuille par véhicule. "
            f"Idéalement du lundi au lundi (7 jours).</p></div>",
            unsafe_allow_html=True,
        )
        return

    try:
        xls = pd.ExcelFile(uploaded)
    except Exception as e:
        st.error(f"Impossible de lire le fichier : {e}")
        return

    sheets = xls.sheet_names
    if len(sheets) <= 1:
        st.error("Fichier invalide : au moins 2 feuilles attendues (résumé + 1 feuille par véhicule).")
        return

    vehicle_sheets = sheets[1:]

    # ── 2. Parsing de tous les véhicules ──────────────────
    all_dfs: dict[str, pd.DataFrame] = {}
    parse_errors: list[str] = []
    for sheet in vehicle_sheets:
        try:
            df_raw = pd.read_excel(xls, sheet_name=sheet, header=4, usecols="B:N")
            df_raw.columns = COLS
            df_raw["_dep"] = pd.to_datetime(
                df_raw["Départ"].astype(str).str.replace(r'\s+[A-Z]{2,5}$', '', regex=True),
                errors="coerce",
            )
            df_raw["_arr"] = pd.to_datetime(
                df_raw["Arrivée"].astype(str).str.replace(r'\s+[A-Z]{2,5}$', '', regex=True),
                errors="coerce",
            )
            df_raw = df_raw.dropna(subset=["_dep"]).sort_values("_dep").reset_index(drop=True)
            for col in ["Lieu de départ", "Lieu d'arrivée"]:
                df_raw[col] = df_raw[col].astype(str).map(_clean_addr)
            if not df_raw.empty:
                all_dfs[sheet] = df_raw
        except Exception as e:
            parse_errors.append(f"{sheet} : {e}")

    if parse_errors:
        st.warning("Erreurs de lecture pour certaines feuilles : " + " | ".join(parse_errors))

    if not all_dfs:
        st.error("Aucune donnée valide trouvée dans le fichier.")
        return

    all_dep_series = pd.concat([df["_dep"] for df in all_dfs.values()])
    date_min = all_dep_series.dt.date.min()
    date_max = all_dep_series.dt.date.max()

    # ── 3. Chargement BDD véhicules + éditeur ─────────────
    try:
        vehicles_db = load_all_quartix_vehicles()
    except Exception:
        vehicles_db = {}

    vehicles_db = _render_hebdo_vehicle_editor(vehicle_sheets, vehicles_db)

    # ── 4. Paramètres analyse ─────────────────────────────
    st.divider()
    col_cut, col_info = st.columns([2, 3])
    with col_cut:
        cutoff = st.time_input(
            "🕐 Heure limite de fin de tournée",
            value=_dtime(18, 0),
            key="hebdo_cutoff",
            help="Tout trajet ≥ 1 km dont le départ ou l'arrivée dépasse cette heure sera signalé.",
        )
    with col_info:
        nb_veh     = len(all_dfs)
        nb_trajets = sum(len(df) for df in all_dfs.values())
        st.markdown(
            f"<div style='background:#f0f4fa;border-radius:8px;padding:10px 14px;"
            f"margin-top:24px;font-size:13px;color:{C_BLUE_DARK}'>"
            f"<b>Période :</b> {date_min.strftime('%d/%m/%Y')} → {date_max.strftime('%d/%m/%Y')}"
            f"&nbsp;|&nbsp; <b>Véhicules :</b> {nb_veh}"
            f"&nbsp;|&nbsp; <b>Trajets :</b> {nb_trajets}</div>",
            unsafe_allow_html=True,
        )

    # ── 5. Analyse ────────────────────────────────────────
    late_trips:    dict[str, list] = {}
    weekend_trips: dict[str, list] = {}

    for plate, df in all_dfs.items():
        # Trajets hors horaires (jours de semaine, distance ≥ 1 km)
        weekday_df = df[(df["_dep"].dt.weekday < 5) &
                        (df["Distance totale"].apply(_to_km) >= _MIN_TRIP_KM)].copy()
        mask_late = weekday_df["_dep"].dt.time > cutoff
        mask_arr  = weekday_df["_arr"].notna() & (weekday_df["_arr"].dt.time > cutoff)
        late_df   = weekday_df[mask_late | mask_arr]
        if not late_df.empty:
            late_trips[plate] = [
                {
                    "date":        row["_dep"].date(),
                    "jour":        _JOURS_FR[row["_dep"].weekday()],
                    "heure_dep":   row["_dep"].strftime("%H:%M"),
                    "heure_arr":   row["_arr"].strftime("%H:%M") if pd.notna(row["_arr"]) else "—",
                    "lieu_dep":    str(row["Lieu de départ"]),
                    "lieu_arr":    str(row["Lieu d'arrivée"]),
                    "distance_km": _to_km(row["Distance totale"]),
                }
                for _, row in late_df.iterrows()
            ]

        # Activité weekend — trajets individuels ≥ 1 km
        wkend_df = df[(df["_dep"].dt.weekday >= 5) &
                      (df["Distance totale"].apply(_to_km) >= _MIN_TRIP_KM)]
        if not wkend_df.empty:
            weekend_trips[plate] = [
                {
                    "date":        row["_dep"].date(),
                    "jour":        _JOURS_FR[row["_dep"].weekday()],
                    "heure_dep":   row["_dep"].strftime("%H:%M"),
                    "distance_km": _to_km(row["Distance totale"]),
                }
                for _, row in wkend_df.iterrows()
            ]

    # ── 6. Affichage résultats ────────────────────────────
    st.divider()
    st.markdown("### 🔍 Résultats de l'analyse")

    for plate in vehicle_sheets:
        if plate not in all_dfs:
            continue
        label_base = _vehicle_label(plate, vehicles_db)
        has_late   = plate in late_trips
        has_wkend  = plate in weekend_trips
        tags       = (" ⚠️" if has_late else "") + (" 🗓️" if has_wkend else "")
        label      = f"🚗 {label_base}{tags}"

        with st.expander(label, expanded=(has_late or has_wkend)):
            if not has_late and not has_wkend:
                st.success("✅ Aucune anomalie détectée")
            else:
                if has_late:
                    st.markdown(f"**⚠️ Trajets hors horaires** (après {cutoff.strftime('%H:%M')})")
                    late_rows = [
                        {
                            "Date":          t["date"].strftime("%d/%m/%Y"),
                            "Jour":          t["jour"],
                            "Heure départ":  t["heure_dep"],
                            "Distance (km)": t["distance_km"],
                        }
                        for t in late_trips[plate]
                    ]
                    st.dataframe(pd.DataFrame(late_rows), use_container_width=True, hide_index=True)

                if has_wkend:
                    st.markdown("**🗓️ Activité weekend**")
                    wk_rows = [
                        {
                            "Date":          d["date"].strftime("%d/%m/%Y"),
                            "Jour":          d["jour"],
                            "Heure départ":  d["heure_dep"],
                            "Distance (km)": d["distance_km"],
                        }
                        for d in weekend_trips[plate]
                    ]
                    st.dataframe(pd.DataFrame(wk_rows), use_container_width=True, hide_index=True)

    # ── 7. Export Excel ───────────────────────────────────
    st.divider()
    st.markdown("### 📊 Rapport Excel")
    excel_bytes = _build_hebdo_excel(
        date_min      = date_min,
        date_max      = date_max,
        cutoff        = cutoff,
        late_trips    = late_trips,
        weekend_trips = weekend_trips,
        vehicles_db   = vehicles_db,
    )
    fname = f"rapport_hebdo_quartix_{date_min.strftime('%Y%m%d')}_{date_max.strftime('%Y%m%d')}.xlsx"
    st.download_button(
        label     = "⬇️ Télécharger le rapport Excel",
        data      = excel_bytes,
        file_name = fname,
        mime      = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type      = "primary",
        key       = "hebdo_dl_excel",
    )

    # ── 8. Génération mail ────────────────────────────────
    st.divider()
    st.markdown("### 📧 Générer un mail de signalement")

    # Génère (ou régénère) le mail à chaque clic, même si un mail existait déjà
    if st.button("📝 Générer le mail", type="primary", key="hebdo_gen_mail"):
        st.session_state["hebdo_mail_text"] = _build_hebdo_mail(
            resp_nom      = "",
            date_min      = date_min,
            date_max      = date_max,
            cutoff        = cutoff,
            late_trips    = late_trips,
            weekend_trips = weekend_trips,
            vehicles_db   = vehicles_db,
        )

    if st.session_state.get("hebdo_mail_text"):
        st.text_area("📋 Texte du mail (copiez-collez dans votre messagerie)",
                     value=st.session_state["hebdo_mail_text"], height=440, key="hebdo_mail_area")


# ── Point d'entrée de la page ─────────────────────────────────────────────────

def render() -> None:
    tab_carte, tab_passages, tab_hebdo = st.tabs([
        "🗺️  Carte & Trajets",
        "📍  Analyse Passages Dépôt",
        "📅  Analyse hebdomadaire",
    ])

    with tab_carte:
        _render_tab_carte()

    with tab_passages:
        _render_tab_passages()

    with tab_hebdo:
        _render_tab_hebdo()
