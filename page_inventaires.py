"""
Page Inventaires — Analyse des inventaires machines par réappro.

Fonctionnalités :
  1. Upload d'un export CSV d'inventaires
  2. Détection automatique du type de machine (BF Simple / BF Double / FP / WUF)
     — Double détecté par présence de Volvic (VOLVICEXOTIC50CL / VOLVICFRAISE50CL)
  3. Comparaison du montant HT inventorié aux seuils min/max par type
  4. Croisement avec le planning (depuis MongoDB) pour détecter les inventaires non faits
  5. Détail des produits manquants (quantité = 0) avec nom du produit

Seuils :
  BF Simple     : min 380 € — max 480 €
  BF Double     : min 700 € — max 957,50 €
  FP IDF        : min 300 € — max 409 €
  FP Province   : min 300 € — max 412 €
  WUF / Autre   : min 280 € — max 400 €
"""

import io
import datetime
import pandas as pd
import streamlit as st
from planogrammes_storage import (
    load_plannos_theoriques,
    load_machine_planno_links,
    save_machine_planno_link,
)

# ─────────────────────────────────────────
# Constantes
# ─────────────────────────────────────────
SEUILS = {
    "BF Simple":   {"min": 380,  "max": 480.00,  "label": "Basic Fit Simple"},
    "BF Double":   {"min": 700,  "max": 957.50,  "label": "Basic Fit Double"},
    "FP IDF":      {"min": 300,  "max": 409.00,  "label": "Fitness Park IDF"},
    "FP Province": {"min": 300,  "max": 412.00,  "label": "Fitness Park Province"},
    "WUF":         {"min": 280,  "max": 400.00,  "label": "WUF"},
    "Autre":       {"min": 280,  "max": 450.00,  "label": "Autre"},
}

PRODUITS_DOUBLE = {"VOLVICEXOTIC50CL", "VOLVICFRAISE50CL"}

IDF_RESSOURCES = {"RIDF1", "RIDF2", "RIDF3", "RIDF4", "RIDF5", "RIDF6", "RIDF7", "RIDF8"}

WEEKDAY_TO_JOUR = {0: "Lundi", 1: "Mardi", 2: "Mercredi", 3: "Jeudi", 4: "Vendredi"}

COLOR_OK     = "#1E7E34"
COLOR_BAD    = "#C0392B"   # rouge  — problèmes planno (absents / hors planno)
COLOR_SEUIL  = "#1565C0"   # bleu   — hors seuils (sous min ou au-dessus max)
COLOR_ORANGE = "#E67E22"   # orange — (conservé pour compat interne)
COLOR_GREY   = "#6C757D"
WHITE        = "#FFFFFF"


# ══════════════════════════════════════════
# PARSING ET ENRICHISSEMENT
# ══════════════════════════════════════════

@st.cache_data(show_spinner=False)
def _parse_inventaire(raw_bytes: bytes) -> pd.DataFrame:
    df = pd.read_csv(io.BytesIO(raw_bytes), sep=";", encoding="utf-8-sig", dtype=str)

    required = [
        "Num Piece", "Date", "Stock Origine", "Code client", "Nom client",
        "Ressource", "Code produit", "Libellé produit", "Quantité", "Montant HT",
    ]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Colonnes manquantes : {', '.join(missing)}")

    if "Type tâche" in df.columns:
        df = df[df["Type tâche"].str.strip().str.lower() == "inventaire"].copy()

    # Strip key identifier columns to avoid whitespace-induced mismatches
    for col in ("Ressource", "Code client", "Nom client", "Date", "Num Piece",
                "Stock Origine", "Code produit"):
        if col in df.columns:
            df[col] = df[col].str.strip()

    df["Montant HT"] = pd.to_numeric(
        df["Montant HT"].str.replace(",", ".", regex=False), errors="coerce"
    ).fillna(0.0)
    df["Quantité"] = pd.to_numeric(df["Quantité"], errors="coerce").fillna(0)

    double_pieces = set(df.loc[df["Code produit"].isin(PRODUITS_DOUBLE), "Num Piece"].unique())
    df["is_double"] = df["Num Piece"].isin(double_pieces)
    df["machine_type"] = df.apply(
        lambda r: _get_machine_type(r["Code client"], r["Ressource"], r["is_double"]),
        axis=1,
    )
    return df


def _get_machine_type(code_client: str, ressource: str, is_double: bool) -> str:
    code = str(code_client).upper()
    is_idf = str(ressource).upper() in {r.upper() for r in IDF_RESSOURCES}
    if code.startswith("BF") or code.startswith("CTF"):
        return "BF Double" if is_double else "BF Simple"
    if code.startswith(("FP", "FT")):
        return "FP IDF" if is_idf else "FP Province"
    if code.startswith("WU"):
        return "WUF"
    return "Autre"


def _get_status(total: float, machine_type: str) -> dict:
    s = SEUILS.get(machine_type)
    if not s:
        return {"emoji": "❓", "label": "Non classifié", "color": COLOR_GREY}
    if total < s["min"]:
        pct = round(total / s["min"] * 100, 1)
        return {"emoji": "🔵", "label": f"Sous seuil ({pct}% du min)", "color": COLOR_SEUIL}
    if total > s["max"]:
        return {"emoji": "🔵", "label": "Au-dessus du max", "color": COLOR_SEUIL}
    return {"emoji": "🟢", "label": "OK", "color": COLOR_OK}


def _build_summary(df: pd.DataFrame) -> pd.DataFrame:
    type_per_inv = df.groupby("Num Piece")["machine_type"].first()
    agg = df.groupby(
        ["Num Piece", "Ressource", "Code client", "Nom client", "Stock Origine", "Date"]
    ).agg(
        total             = ("Montant HT", "sum"),
        nb_produits_ref   = ("Code produit", "nunique"),
        nb_produits_vides = ("Quantité", lambda q: (q == 0).sum()),
    ).reset_index()

    agg["machine_type"] = agg["Num Piece"].map(type_per_inv)
    agg["seuil_min"]    = agg["machine_type"].map(lambda t: SEUILS.get(t, {}).get("min"))
    agg["seuil_max"]    = agg["machine_type"].map(lambda t: SEUILS.get(t, {}).get("max"))
    agg["type_label"]   = agg["machine_type"].map(lambda t: SEUILS.get(t, {}).get("label", t))
    agg["ecart_min"]    = (agg["total"] - agg["seuil_min"]).round(2)

    status_info = agg.apply(lambda r: _get_status(r["total"], r["machine_type"]), axis=1)
    agg["statut_emoji"] = status_info.apply(lambda d: d["emoji"])
    agg["statut_label"] = status_info.apply(lambda d: d["label"])
    agg["statut_color"] = status_info.apply(lambda d: d["color"])
    return agg.sort_values(["Ressource", "Nom client"])


# ══════════════════════════════════════════
# CROISEMENT PLANNING vs INVENTAIRES
# ══════════════════════════════════════════

def _parse_planning_for_reappro(planning_dict: dict) -> dict:
    """
    Transforme le planning MongoDB en {jour_fr: {client_code: {label, machine}}}.
    Format MongoDB : {"Lundi": [["BFAL50 - BF NANCY...", "1750M1"], ...], ...}
    """
    result = {}
    for jour, salles in planning_dict.items():
        if jour not in WEEKDAY_TO_JOUR.values():
            continue
        result[jour] = {}
        for salle in salles:
            if len(salle) < 2:
                continue
            client_full = str(salle[0]).strip()
            machine     = str(salle[1]).strip()
            if " - " in client_full:
                code  = client_full.split(" - ")[0].strip()
                label = client_full.split(" - ", 1)[1].strip()
            else:
                code  = client_full
                label = client_full
            result[jour][code] = {"label": label, "machine": machine}
    return result


def _build_joker_index(df_inv: pd.DataFrame, plannings_mongo: dict) -> dict:
    """
    Construit un index des jokers : inventaires faits par un réappro
    pour les salles planifiées par un autre réappro le même jour.

    Retourne :
    {
      reappro_prevu: {
        date_str: [
          {"code", "label", "machine", "fait_par": reappro_fait}, ...
        ]
      }
    }
    """
    # Index planning complet : {code: [(reappro, jour, label, machine), ...]}
    code_plan_idx: dict = {}
    for reappro, planning_raw in plannings_mongo.items():
        for jour, salles in planning_raw.items():
            if jour not in WEEKDAY_TO_JOUR.values():
                continue
            for salle in salles:
                if len(salle) < 2:
                    continue
                client_full = str(salle[0]).strip()
                machine     = str(salle[1]).strip()
                code  = client_full.split(" - ")[0].strip() if " - " in client_full else client_full
                label = client_full.split(" - ", 1)[1].strip() if " - " in client_full else client_full
                if code not in code_plan_idx:
                    code_plan_idx[code] = []
                code_plan_idx[code].append((reappro, jour, label, machine))

    # Pour chaque inventaire réalisé, vérifier si le code appartient
    # au planning d'un AUTRE réappro ce même jour
    inv_unique = (
        df_inv[["Code client", "Ressource", "Date", "Nom client"]]
        .drop_duplicates(subset=["Code client", "Ressource", "Date"])
    )

    joker_index: dict = {}  # {reappro_prevu: {date: [{code, label, machine, fait_par}]}}

    for _, row in inv_unique.iterrows():
        code         = row["Code client"]
        reappro_inv  = row["Ressource"]
        date_str     = row["Date"]

        if code not in code_plan_idx:
            continue

        try:
            dt = datetime.datetime.strptime(date_str, "%d/%m/%Y")
        except ValueError:
            continue
        jour_fait = WEEKDAY_TO_JOUR.get(dt.weekday())
        if not jour_fait:
            continue

        seen = set()  # deduplicate (planned_reappro, code) pairs
        for (planned_reappro, planned_jour, label, machine) in code_plan_idx[code]:
            dedup_key = (planned_reappro, code)
            if planned_reappro != reappro_inv and planned_jour == jour_fait and dedup_key not in seen:
                seen.add(dedup_key)
                if planned_reappro not in joker_index:
                    joker_index[planned_reappro] = {}
                if date_str not in joker_index[planned_reappro]:
                    joker_index[planned_reappro][date_str] = []
                joker_index[planned_reappro][date_str].append({
                    "code":     code,
                    "label":    label,
                    "machine":  machine,
                    "fait_par": reappro_inv,
                })

    return joker_index


def _croiser_plannings_inventaires(df_inv: pd.DataFrame, plannings_mongo: dict) -> dict:
    """
    Retourne :
    {
      reappro: {
        date_str: {
          "jour_fr":           str,
          "nb_planifie":       int,
          "nb_fait":           int,
          "planifie":          [{code, label, machine}, ...],
          "fait":              [code, ...],
          "manquants":         [{code, label, machine}, ...],
          "deja_fait_semaine": [{code, label, machine}, ...],
          "jokers":            [{code, label, machine, fait_par}, ...],
        }
      }
    }

    Règles :
      - Double passage : planifié ce jour mais inventaire fait un autre jour de la semaine ISO.
      - Joker : planifié par ce réappro mais inventaire fait par un autre réappro le même jour.
    """
    planning_index = {
        r: _parse_planning_for_reappro(p)
        for r, p in plannings_mongo.items()
    }

    # Index semaine : {(reappro, iso_year, iso_week): set(codes faits)}
    df_inv = df_inv.copy()
    df_inv["_dt"] = pd.to_datetime(df_inv["Date"], format="%d/%m/%Y", errors="coerce")
    df_inv["_iso_year"] = df_inv["_dt"].dt.isocalendar().year.astype("Int64")
    df_inv["_iso_week"] = df_inv["_dt"].dt.isocalendar().week.astype("Int64")

    codes_by_week = (
        df_inv.groupby(["Ressource", "_iso_year", "_iso_week"])["Code client"]
        .apply(set)
        .to_dict()
    )

    # {(reappro, iso_year, iso_week, code): date_str} — quel jour chaque code a réellement été fait
    code_date_by_week = (
        df_inv.groupby(["Ressource", "_iso_year", "_iso_week", "Code client"])["Date"]
        .first()
        .to_dict()
    )

    # {(iso_year, iso_week): set(codes)} — tous codes faits par N'IMPORTE QUEL réappro cette semaine
    all_codes_by_week = (
        df_inv.groupby(["_iso_year", "_iso_week"])["Code client"]
        .apply(set)
        .to_dict()
    )

    inv_done = (
        df_inv.groupby(["Ressource", "Date"])["Code client"]
        .apply(set)
        .reset_index()
    )

    # Joker index : {reappro_prevu: {date: [{code, label, machine, fait_par}]}}
    joker_index = _build_joker_index(df_inv, plannings_mongo)

    # All codes inventoried globally on each date (any reappro)
    # used to detect jokers for reappros that have no inventories themselves
    all_done_by_date: dict = {}
    for _, row in inv_done.iterrows():
        d = row["Date"]
        if d not in all_done_by_date:
            all_done_by_date[d] = set()
        all_done_by_date[d] |= row["Code client"]

    results = {}

    # Process reappros that have inventories
    for _, row in inv_done.iterrows():
        reappro    = row["Ressource"]
        date_str   = row["Date"]
        done_codes = row["Code client"]

        if reappro not in planning_index:
            continue
        try:
            dt = datetime.datetime.strptime(date_str, "%d/%m/%Y")
        except ValueError:
            continue

        jour_fr = WEEKDAY_TO_JOUR.get(dt.weekday())
        if not jour_fr or jour_fr not in planning_index[reappro]:
            continue

        _populate_result(
            results, reappro, date_str, jour_fr, dt,
            planning_index[reappro][jour_fr],
            done_codes, codes_by_week, code_date_by_week, all_codes_by_week,
            joker_index.get(reappro, {}).get(date_str, []),
        )

    # Also process reappros with NO inventories but who have jokers or are completely absent
    for reappro, planning in planning_index.items():
        for date_str in sorted(all_done_by_date.keys()):
            if reappro in results and date_str in results[reappro]:
                continue  # already processed
            try:
                dt = datetime.datetime.strptime(date_str, "%d/%m/%Y")
            except ValueError:
                continue
            jour_fr = WEEKDAY_TO_JOUR.get(dt.weekday())
            if not jour_fr or jour_fr not in planning:
                continue
            # Only add if there are jokers for this reappro on this date
            jokers_today = joker_index.get(reappro, {}).get(date_str, [])
            if jokers_today:
                _populate_result(
                    results, reappro, date_str, jour_fr, dt,
                    planning[jour_fr],
                    set(),  # no own inventories
                    codes_by_week, code_date_by_week, all_codes_by_week,
                    jokers_today,
                )

    return results


def _populate_result(
    results: dict,
    reappro: str,
    date_str: str,
    jour_fr: str,
    dt: datetime.datetime,
    jour_plan: dict,
    done_codes: set,
    codes_by_week: dict,
    code_date_by_week: dict,
    all_codes_by_week: dict,
    jokers_today: list,
) -> None:
    """Calcule et insère le résultat pour (reappro, date) dans results."""
    planned_codes   = set(jour_plan.keys())
    fait_codes      = done_codes & planned_codes
    joker_codes     = {j["code"] for j in jokers_today}
    manquants_bruts = planned_codes - done_codes - joker_codes

    iso_cal  = dt.isocalendar()
    codes_faits_semaine     = codes_by_week.get((reappro, iso_cal[0], iso_cal[1]), set())
    all_codes_this_week     = all_codes_by_week.get((iso_cal[0], iso_cal[1]), set())

    deja_fait_semaine_codes = sorted(manquants_bruts & codes_faits_semaine)
    restant_apres_deja      = manquants_bruts - codes_faits_semaine
    # Codes planifiés mais non faits par CE réappro, faits par UN AUTRE réappro dans la semaine
    fait_par_autre_codes    = sorted(restant_apres_deja & all_codes_this_week)
    manquants_reels         = restant_apres_deja - set(fait_par_autre_codes)

    # Enrichir chaque double passage avec le jour où il a réellement été fait
    deja_entries = []
    for c in deja_fait_semaine_codes:
        entry = {"code": c, **jour_plan[c]}
        actual_date = code_date_by_week.get((reappro, iso_cal[0], iso_cal[1], c))
        if actual_date:
            try:
                actual_jour = WEEKDAY_TO_JOUR.get(
                    datetime.datetime.strptime(actual_date, "%d/%m/%Y").weekday()
                )
                if actual_jour:
                    entry["jour_fait"] = actual_jour
            except ValueError:
                pass
        deja_entries.append(entry)

    if reappro not in results:
        results[reappro] = {}

    results[reappro][date_str] = {
        "jour_fr":               jour_fr,
        "nb_planifie":           len(planned_codes),
        "nb_fait":               len(fait_codes) + len(deja_fait_semaine_codes) + len(joker_codes) + len(fait_par_autre_codes),
        "planifie":              [{"code": c, **v} for c, v in jour_plan.items()],
        "fait":                  sorted(fait_codes),
        "manquants":             [{"code": c, **jour_plan[c]} for c in sorted(manquants_reels)],
        "deja_fait_semaine":     deja_entries,
        "jokers":                jokers_today,
        "fait_par_autre_semaine":[{"code": c, **jour_plan[c]} for c in fait_par_autre_codes],
    }


# ══════════════════════════════════════════
# PRODUITS MANQUANTS (qty = 0)
# ══════════════════════════════════════════

def _get_missing_products(df_raw: pd.DataFrame, num_piece: str) -> list:
    sub = df_raw[(df_raw["Num Piece"] == num_piece) & (df_raw["Quantité"] == 0)]
    return (
        sub[["Code produit", "Libellé produit"]]
        .drop_duplicates()
        .rename(columns={"Code produit": "code", "Libellé produit": "nom"})
        .to_dict("records")
    )


def _get_all_missing_products(df_raw: pd.DataFrame, summary: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for _, inv in summary.iterrows():
        for p in _get_missing_products(df_raw, inv["Num Piece"]):
            rows.append({
                "Réappro":            inv["Ressource"],
                "Salle":              inv["Nom client"],
                "Machine":            inv["Stock Origine"],
                "Type":               inv["type_label"],
                "Date":               inv["Date"],
                "Code produit":       p["code"],
                "Produit manquant":   p["nom"],
                "Statut inventaire":  inv["statut_emoji"],
            })
    return pd.DataFrame(rows) if rows else pd.DataFrame()


def _analyze_planno(
    piece: str,
    mtype_key: str,
    planno_index: dict,
    inv_qty_by_piece: dict,
    inv_lib_by_piece: dict,
    missing_code_idx: dict,
    missing_idx: dict,
) -> tuple:
    """
    Shared planno-vs-inventory analysis.
    Returns (prods, hors_planno, absents, detail_rows).
    """
    planno     = (planno_index or {}).get(mtype_key)
    inv_qtys   = inv_qty_by_piece.get(piece, {})
    inv_libs   = inv_lib_by_piece.get(piece, {})
    miss_codes = missing_code_idx.get(piece, set())

    if not planno:
        return missing_idx.get(piece, []), [], [], []

    prods       = [inv_libs.get(c, c) for c in miss_codes if c in planno]
    hors_planno = [inv_libs.get(c, c) for c in inv_qtys   if c not in planno]
    absents     = [planno[c]["libelle"] or c for c in planno if c not in inv_qtys]

    detail_rows = []
    for _c, _info in sorted(planno.items()):
        _qi    = inv_qtys.get(_c)
        _qp    = _info["niv_haut"]
        _ecart = (int(_qi) - _qp) if _qi is not None else -_qp
        _stat  = "Absent" if _qi is None else ("Épuisé" if _qi == 0 else "OK")
        detail_rows.append({
            "code": _c, "libelle": _info["libelle"],
            "qty_plan": _qp,
            "qty_inv":  int(_qi) if _qi is not None else "Absent",
            "ecart":    _ecart, "statut": _stat,
        })
    for _c in sorted(inv_qtys):
        if _c not in planno:
            detail_rows.append({
                "code": _c, "libelle": inv_libs.get(_c, ""),
                "qty_plan": "—", "qty_inv": int(inv_qtys[_c]),
                "ecart": "Hors planno", "statut": "Hors planno",
            })

    return prods, hors_planno, absents, detail_rows


# ══════════════════════════════════════════
# PAGE PRINCIPALE
# ══════════════════════════════════════════

def render():
    # Récupération des plannings depuis MongoDB
    plannings_mongo: dict = {}
    try:
        from mongo_storage import load_plannings_from_mongo

        @st.cache_data(show_spinner=False, ttl=300)
        def _get_plannings():
            return load_plannings_from_mongo()

        plannings_mongo, _ = _get_plannings()
    except Exception:
        pass

    # ── Planno index : {machine_type: {code: {libelle, niv_haut}}} ──
    def _build_planno_index():
        links   = load_machine_planno_links()          # {machine_type: planno_id}
        plannos = load_plannos_theoriques()            # liste de docs
        by_id   = {p["_id"]: p for p in plannos}
        index   = {}
        for mtype, pid in links.items():
            if pid and pid in by_id:
                prods = by_id[pid].get("produits", [])
                index[mtype] = {
                    pr["code"]: {"libelle": pr.get("libelle", ""), "niv_haut": int(pr.get("niv_haut") or 0)}
                    for pr in prods if pr.get("code")
                }
        return index

    planno_index: dict = {}
    try:
        planno_index = _build_planno_index()
    except Exception:
        pass

    # ── Upload fichier ──────────────────────
    st.markdown("### 📂 Déposer le fichier export inventaires (CSV)")
    uploaded = st.file_uploader(
        "Export inventaires CSV", type=["csv"],
        key="inv_uploader", label_visibility="collapsed",
    )

    if uploaded is None:
        st.info(
            "**Format attendu :** export CSV du logiciel de gestion "
            "(UTF-8 BOM, séparateur `;`). Colonnes nécessaires : "
            "`Num Piece`, `Date`, `Type tâche`, `Stock Origine`, `Code client`, "
            "`Nom client`, `Ressource`, `Code produit`, `Libellé produit`, "
            "`Quantité`, `Montant HT`."
        )
        _show_thresholds()
        _section_bilan_semaine(plannings_mongo)
        _section_connexion()
        return

    # ── Parsing ─────────────────────────────
    try:
        with st.spinner("Analyse du fichier…"):
            df_raw  = _parse_inventaire(uploaded.read())
            summary = _build_summary(df_raw)
        nb_doubles = (summary["machine_type"] == "BF Double").sum()
        st.success(
            f"✅ **{df_raw['Num Piece'].nunique()}** inventaires chargés — "
            f"**{df_raw['Ressource'].nunique()}** réappros — "
            f"**{nb_doubles}** machines doubles détectées (Volvic)"
        )
    except Exception as e:
        st.error(f"❌ Erreur : {e}")
        return

    # ── Croisement planning ──────────────────
    croisement: dict = {}
    if plannings_mongo:
        croisement = _croiser_plannings_inventaires(df_raw, plannings_mongo)
    else:
        st.warning(
            "⚠️ Plannings non disponibles — le suivi planning est désactivé. "
            "Vérifie la connexion MongoDB."
        )

    # ── KPIs globaux ────────────────────────
    st.divider()
    nb_ok    = (summary["statut_emoji"] == "🟢").sum()
    nb_over  = (summary["statut_emoji"] == "🟠").sum()

    total_inv_manquants = sum(
        len(d["manquants"])
        for r_data in croisement.values()
        for d in r_data.values()
    )
    nb_planifie = sum(
        d["nb_planifie"]
        for r_data in croisement.values()
        for d in r_data.values()
    )
    nb_fait = sum(
        d["nb_fait"]
        for r_data in croisement.values()
        for d in r_data.values()
    )
    taux_fait = round(nb_fait / nb_planifie * 100, 1) if nb_planifie else 0

    k1, k2, k3, k4, k5 = st.columns(5)
    k1.metric("📋 Planifiées / Faites", f"{nb_fait} / {nb_planifie}",
              delta=f"{taux_fait}%")
    k2.metric("✅ OK",                  nb_ok)
    k3.metric("🔵 Hors seuils",         nb_over)
    k4.metric("📭 Non faites",          total_inv_manquants,
              delta=f"-{total_inv_manquants}" if total_inv_manquants else None,
              delta_color="inverse")
    k5.metric("📊 Inventaires chargés", len(summary))

    st.divider()

    # ── Filtres ─────────────────────────────
    col_f1, col_f2, col_f3 = st.columns(3)
    with col_f1:
        filtre_reappro = st.selectbox(
            "🔍 Réappro",
            ["Tous"] + sorted(summary["Ressource"].unique()),
            key="inv_f_reappro",
        )
    with col_f2:
        filtre_statut = st.selectbox(
            "📊 Statut inventaire",
            ["Tous", "🟢 OK", "🔴 Mal fait", "🟠 Au-dessus max"],
            key="inv_f_statut",
        )
    with col_f3:
        filtre_type = st.selectbox(
            "🏋️ Type machine",
            ["Tous"] + sorted(summary["type_label"].unique()),
            key="inv_f_type",
        )

    filtered = summary.copy()
    if filtre_reappro != "Tous":
        filtered = filtered[filtered["Ressource"] == filtre_reappro]
    if filtre_statut != "Tous":
        filtered = filtered[filtered["statut_emoji"] == filtre_statut.split()[0]]
    if filtre_type != "Tous":
        filtered = filtered[filtered["type_label"] == filtre_type]

    # ── Onglets ─────────────────────────────
    tab_suivi, tab_detail = st.tabs([
        "📋 Suivi & inventaires",
        "🔍 Détail machine",
    ])

    # ════════════════════════════════════════
    # TAB — SUIVI & INVENTAIRES
    # ════════════════════════════════════════
    with tab_suivi:

        # KPIs planning
        if croisement:
            nb_r_ok = sum(
                1 for r_data in croisement.values()
                if all(len(d["manquants"]) == 0 for d in r_data.values())
            )
            nb_r_ko = len(croisement) - nb_r_ok
            p1, p2, p3 = st.columns(3)
            p1.metric("✅ Réappros tout OK",        nb_r_ok)
            p2.metric("🔴 Réappros avec manquants", nb_r_ko,
                      delta=f"-{nb_r_ko}" if nb_r_ko else None, delta_color="inverse")
            p3.metric("📭 Inventaires non faits",   total_inv_manquants,
                      delta=f"-{total_inv_manquants}" if total_inv_manquants else None,
                      delta_color="inverse")
            st.divider()

        # Bouton export
        col_exp, _ = st.columns([1, 4])
        with col_exp:
            if st.button("📥 Exporter en Excel", key="inv_export"):
                excel_bytes = _export_excel(filtered, df_raw, croisement, plannings_mongo, planno_index)
                date_str_dl = datetime.date.today().strftime("%Y%m%d")
                st.download_button(
                    "⬇️ Télécharger Excel", data=excel_bytes,
                    file_name=f"inventaires_{date_str_dl}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="inv_dl",
                )

        st.divider()

        # Accordéon par réappro
        # Include reappros with zero inventories but who had salles planned
        _reappros_with_inv = set(filtered["Ressource"].unique())
        _reappros_absent   = {
            r for r, data in croisement.items()
            if r not in _reappros_with_inv
            and any(len(d["manquants"]) > 0 for d in data.values())
        }
        reappros_liste = (
            [filtre_reappro] if filtre_reappro != "Tous"
            else sorted(_reappros_with_inv | _reappros_absent)
        )

        for reappro in reappros_liste:
            sub = filtered[filtered["Ressource"] == reappro]
            nb_ok_r   = (sub["statut_emoji"] == "🟢").sum()
            nb_over_r = (sub["statut_emoji"] == "🟠").sum()
            nb_bad_r  = 0
            for _, _r in sub.iterrows():
                _planno = (planno_index or {}).get(_r.get("machine_type", ""))
                if _planno:
                    _inv_qty = df_raw[df_raw["Num Piece"] == _r["Num Piece"]].groupby("Code produit")["Quantité"].sum().to_dict()
                    _has_issues = any(c not in _inv_qty for c in _planno) or any(c not in _planno for c in _inv_qty)
                else:
                    _has_issues = False
                if _has_issues or _r["statut_emoji"] == "🔴":
                    nb_bad_r += 1
            total_r   = sub["total"].sum()
            inv_manq_r = sum(len(d["manquants"]) for d in croisement.get(reappro, {}).values())
            jokers_r   = sum(len(d.get("jokers", [])) for d in croisement.get(reappro, {}).values())
            nb_planif_r = sum(d["nb_planifie"] for d in croisement.get(reappro, {}).values())
            nb_fait_r   = sum(d["nb_fait"] for d in croisement.get(reappro, {}).values())

            titre = (
                f"**{reappro}** — "
                f"📋 {nb_fait_r}/{nb_planif_r} · "
                f"🟢 {nb_ok_r} · 🔴 {nb_bad_r} · 🟠 {nb_over_r}"
                + (f" · 📭 {inv_manq_r} non faits" if inv_manq_r else "")
                + (f" · 🔀 {jokers_r} joker(s)" if jokers_r else "")
                + f" · 💰 {total_r:,.2f} €"
            )

            with st.expander(titre, expanded=False):

                # Section planning du réappro
                if reappro in croisement:
                    for date_str_c in sorted(croisement[reappro].keys()):
                        d        = croisement[reappro][date_str_c]
                        nb_manq  = len(d["manquants"])
                        nb_deja  = len(d.get("deja_fait_semaine", []))
                        nb_jok   = len(d.get("jokers", []))
                        nb_autre = len(d.get("fait_par_autre_semaine", []))
                        color    = COLOR_OK if nb_manq == 0 else COLOR_BAD
                        icon     = "✅" if nb_manq == 0 else "🔴"
                        bg       = "#f8fff8" if nb_manq == 0 else "#fff8f8"
                        extra    = ""
                        if nb_deja:  extra += f" · 🔄 {nb_deja} double(s)"
                        if nb_jok:   extra += f" · 🔀 {nb_jok} joker(s)"
                        if nb_autre: extra += f" · 👥 {nb_autre} fait(s) par autre"

                        st.markdown(
                            f"""<div style="border-left:4px solid {color};padding:7px 12px;
                                border-radius:4px;margin-bottom:6px;background:{bg}">
                            <b>{icon} {d['jour_fr']} {date_str_c}</b> —
                            {d['nb_fait']}/{d['nb_planifie']} inventaires réalisés{extra}
                            </div>""",
                            unsafe_allow_html=True,
                        )

                        if nb_manq > 0:
                            rows_m = [{"Code": m["code"], "Salle": m["label"], "Machine": m["machine"]}
                                      for m in d["manquants"]]
                            st.dataframe(
                                pd.DataFrame(rows_m).style.map(
                                    lambda _: f"background-color:{COLOR_BAD}; color:{WHITE}; font-weight:600"
                                ),
                                hide_index=True, use_container_width=True,
                                height=min(250, 38 + len(rows_m) * 35),
                            )

                        if nb_deja:
                            rows_deja = [{"Code": m["code"], "Salle": m["label"], "Machine": m["machine"], "Fait le": m.get("jour_fait", "?")}
                                         for m in d["deja_fait_semaine"]]
                            st.markdown(f"🔄 *Double passage ({nb_deja})*")
                            st.dataframe(
                                pd.DataFrame(rows_deja).style.map(
                                    lambda _: f"background-color:{COLOR_ORANGE}; color:{WHITE}; font-weight:600"
                                ),
                                hide_index=True, use_container_width=True,
                                height=min(200, 38 + len(rows_deja) * 35),
                            )

                        if nb_jok:
                            rows_j = [{"Code": j["code"], "Salle": j["label"],
                                       "Machine": j["machine"], "Fait par": j["fait_par"]}
                                      for j in d["jokers"]]
                            st.markdown(f"🔀 *Jokers ({nb_jok})*")
                            st.dataframe(
                                pd.DataFrame(rows_j).style.map(
                                    lambda _: f"background-color:#6C3483; color:{WHITE}; font-weight:600"
                                ),
                                hide_index=True, use_container_width=True,
                                height=min(200, 38 + len(rows_j) * 35),
                            )

                        if nb_autre:
                            rows_a = [{"Code": m["code"], "Salle": m["label"], "Machine": m["machine"]}
                                      for m in d["fait_par_autre_semaine"]]
                            st.markdown(f"👥 *Fait par un autre réappro cette semaine ({nb_autre})*")
                            st.dataframe(
                                pd.DataFrame(rows_a).style.map(
                                    lambda _: "background-color:#607D8B; color:white; font-weight:600"
                                ),
                                hide_index=True, use_container_width=True,
                                height=min(200, 38 + len(rows_a) * 35),
                            )

                    st.markdown("---")

                # Cartes inventaires réalisés
                for _, row in sub.iterrows():
                    _machine_card(row, df_raw, planno_index)

    # ════════════════════════════════════════
    # TAB — DÉTAIL MACHINE
    # ════════════════════════════════════════
    with tab_detail:
        machines_list = sorted(filtered["Nom client"].unique())
        if not machines_list:
            st.info("Aucune machine dans les filtres actifs.")
        else:
            sel = st.selectbox("Choisir une salle", machines_list, key="inv_sel_machine")
            for _, row in filtered[filtered["Nom client"] == sel].iterrows():
                st.markdown(
                    f"#### 🏋️ {row['Nom client']} — {row['Stock Origine']} ({row['type_label']})"
                )
                _machine_detail_full(row, df_raw, planno_index)

    _section_bilan_semaine(plannings_mongo)
    _section_connexion()



# ══════════════════════════════════════════
# BILAN SEMAINE
# ══════════════════════════════════════════

def _build_bilan_rows(df_raw: pd.DataFrame, plannings_mongo: dict) -> list:
    """
    Pour chaque (réappro, jour planifié, salle) sur toutes les semaines
    présentes dans le CSV, retourne une ligne avec fait=True/False.
    """
    # Total HT par visite : {(reappro, date_str, code_client): ht}
    done_ht = (
        df_raw.groupby(["Ressource", "Date", "Code client"])["Montant HT"]
        .sum().to_dict()
    )
    done_set = set(done_ht.keys())

    # ISO weeks dans le CSV
    df2 = df_raw.copy()
    df2["_dt"] = pd.to_datetime(df2["Date"], format="%d/%m/%Y", errors="coerce")
    iso_pairs = (
        df2.assign(
            _y=lambda d: d["_dt"].dt.isocalendar().year.astype("Int64"),
            _w=lambda d: d["_dt"].dt.isocalendar().week.astype("Int64"),
        )[["_y", "_w"]].dropna().drop_duplicates().astype(int).values.tolist()
    )

    rows = []
    for iso_year, iso_week in iso_pairs:
        monday = datetime.datetime.fromisocalendar(iso_year, iso_week, 1)
        for offset in range(5):
            day_dt   = monday + datetime.timedelta(days=offset)
            jour_fr  = WEEKDAY_TO_JOUR.get(day_dt.weekday())
            if not jour_fr:
                continue
            date_str = day_dt.strftime("%d/%m/%Y")

            for reappro, planning_raw in sorted(plannings_mongo.items()):
                planning = _parse_planning_for_reappro(planning_raw)
                if jour_fr not in planning:
                    continue
                for code, info in sorted(planning[jour_fr].items(),
                                         key=lambda x: x[1]["label"]):
                    key  = (reappro, date_str, code)
                    fait = key in done_set
                    rows.append({
                        "reappro":   reappro,
                        "iso_year":  iso_year,
                        "iso_week":  iso_week,
                        "date":      date_str,
                        "jour":      jour_fr,
                        "code":      code,
                        "salle":     info["label"],
                        "machine":   info["machine"],
                        "fait":      fait,
                        "valeur_ht": done_ht.get(key),
                    })
    return rows


def _export_bilan_excel(bilan_rows: list) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    def _fill(h): return PatternFill("solid", fgColor=h.lstrip("#"))
    def _font(bold=False, color="000000", size=9):
        return Font(bold=bold, color=color.lstrip("#"), size=size, name="Arial")
    def _align(h="left"):
        return Alignment(horizontal=h, vertical="center")
    _s   = Side(style="thin", color="CCCCCC")
    _brd = Border(left=_s, right=_s, top=_s, bottom=_s)

    C_DARK  = "1F4E79"; C_MID = "2E75B6"; C_SUB = "BDD7EE"
    C_OK_BG = "D4EDDA"; C_OK_FG = "1E7E34"
    C_ABS_BG = "BBDEFB"; C_ABS_FG = "1565C0"
    C_WHITE = "FFFFFF"

    HDRS   = ["Date", "Jour", "Salle", "Machine", "Statut", "Valeur HT"]
    WIDTHS = [11,     10,     36,      10,         12,       14]

    wb = Workbook()
    wb.remove(wb.active)

    for reappro in sorted({r["reappro"] for r in bilan_rows}):
        ws = wb.create_sheet(reappro[:31])
        ws.freeze_panes = "A3"

        N = len(HDRS)
        ws.merge_cells(f"A1:{get_column_letter(N)}1")
        tc = ws.cell(1, 1, f"Bilan semaine — {reappro}")
        tc.fill = _fill(C_DARK); tc.font = _font(True, C_WHITE, 12)
        tc.alignment = _align("center"); ws.row_dimensions[1].height = 24

        for ci, (h, w) in enumerate(zip(HDRS, WIDTHS), 1):
            c = ws.cell(2, ci, h)
            c.fill = _fill(C_MID); c.font = _font(True, C_WHITE, 9)
            c.alignment = _align("center")
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[2].height = 18

        sub  = [r for r in bilan_rows if r["reappro"] == reappro]
        ri   = 3
        cur  = None

        for r in sub:
            if r["date"] != cur:
                cur      = r["date"]
                day_rows = [x for x in sub if x["date"] == r["date"]]
                nb_f     = sum(1 for x in day_rows if x["fait"])
                nb_t     = len(day_rows)
                txt      = f"  {r['jour']}  {r['date']}   |   {nb_f} / {nb_t} faits"
                ws.merge_cells(f"A{ri}:{get_column_letter(N)}{ri}")
                sep = ws.cell(ri, 1, txt)
                sep.fill = _fill(C_SUB); sep.font = _font(True, C_DARK, 10)
                sep.alignment = _align("left"); ws.row_dimensions[ri].height = 20
                ri += 1

            bg  = C_OK_BG  if r["fait"] else C_ABS_BG
            fg  = C_OK_FG  if r["fait"] else C_ABS_FG
            sta = "✅ Fait" if r["fait"] else "❌ Non fait"
            ht  = round(r["valeur_ht"], 2) if r["valeur_ht"] is not None else ""
            vals = [r["date"], r["jour"], r["salle"], r["machine"], sta, ht]

            for ci, v in enumerate(vals, 1):
                c = ws.cell(ri, ci, v)
                c.fill = _fill(bg); c.border = _brd
                c.font = _font(bold=(ci == 5), color=(fg if ci == 5 else "000000"), size=9)
                c.alignment = _align("center" if ci in (1, 2, 4, 5, 6) else "left")
                if ci == 6 and isinstance(v, float):
                    c.number_format = "#,##0.00 €"
            ws.row_dimensions[ri].height = 16
            ri += 1

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf.read()


def _section_bilan_semaine(plannings_mongo: dict):
    st.markdown("---")
    st.markdown("### 📅 Bilan semaine")
    st.caption(
        "Uploadez un export CSV couvrant une semaine entière. "
        "Chaque salle planifiée apparaît — **verte** si faite, **bleue** si absente."
    )

    up = st.file_uploader(
        "Export inventaires CSV (semaine)", type=["csv"],
        key="bilan_uploader", label_visibility="collapsed",
    )
    if up is None:
        st.info("Déposez un fichier CSV pour générer le bilan.")
        return
    if not plannings_mongo:
        st.warning("⚠️ Plannings non disponibles.")
        return

    try:
        with st.spinner("Analyse…"):
            df_bilan = _parse_inventaire(up.read())
            bilan_rows = _build_bilan_rows(df_bilan, plannings_mongo)
    except Exception as e:
        st.error(f"❌ {e}")
        return

    if not bilan_rows:
        st.warning("Aucune correspondance entre le fichier et les plannings.")
        return

    # ── Avertissement si des ressources CSV ne sont pas dans les plannings ──
    csv_ressources = set(df_bilan["Ressource"].dropna().unique())
    mongo_employes = set(plannings_mongo.keys())
    unmatched = csv_ressources - mongo_employes
    if unmatched:
        st.warning(
            f"⚠️ Ces ressources du CSV n'ont **pas de planning** dans la base : "
            f"**{', '.join(sorted(unmatched))}**. "
            "Leurs inventaires ne seront pas comptabilisés dans le bilan. "
            "Vérifiez que le nom de la ressource dans le CSV correspond exactement "
            "au nom de l'employé dans les plannings MongoDB."
        )

    # ── KPIs globaux ─────────────────────────────────────────────────────────
    nb_planif = len(bilan_rows)
    nb_fait   = sum(1 for r in bilan_rows if r["fait"])
    nb_abs    = nb_planif - nb_fait
    pct       = round(nb_fait / nb_planif * 100, 1) if nb_planif else 0

    k1, k2, k3, k4 = st.columns(4)
    k1.metric("📋 Planifiées",  nb_planif)
    k2.metric("✅ Faites",      nb_fait,  delta=f"{pct}%")
    k3.metric("🔵 Absentes",   nb_abs,
              delta=f"-{nb_abs}" if nb_abs else None, delta_color="inverse")
    k4.metric("👥 Réappros",   len({r["reappro"] for r in bilan_rows}))

    # ── Export ────────────────────────────────────────────────────────────────
    col_exp, _ = st.columns([1, 4])
    with col_exp:
        st.download_button(
            "📥 Exporter Excel", data=_export_bilan_excel(bilan_rows),
            file_name=f"bilan_semaine_{datetime.date.today():%Y%m%d}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="bilan_dl",
        )

    st.divider()

    # ── Accordéon par réappro ─────────────────────────────────────────────────
    for reappro in sorted({r["reappro"] for r in bilan_rows}):
        sub   = [r for r in bilan_rows if r["reappro"] == reappro]
        nb_f  = sum(1 for r in sub if r["fait"])
        nb_t  = len(sub)
        nb_a  = nb_t - nb_f
        pct_r = round(nb_f / nb_t * 100, 1) if nb_t else 0

        titre = (
            f"**{reappro}** — 📋 {nb_f}/{nb_t}"
            f" · ✅ {nb_f} · 🔵 {nb_a} abs"
            f" · {pct_r}%"
        )
        with st.expander(titre, expanded=False):
            # Group by date in order
            seen_dates = []
            for r in sub:
                if r["date"] not in seen_dates:
                    seen_dates.append(r["date"])

            for date_str in seen_dates:
                day_rows = [r for r in sub if r["date"] == date_str]
                nb_df    = sum(1 for r in day_rows if r["fait"])
                nb_dt    = len(day_rows)
                all_done = nb_df == nb_dt
                color    = COLOR_OK if all_done else COLOR_SEUIL
                icon     = "✅" if all_done else "🔵"
                bg       = "#f8fff8" if all_done else "#f0f4ff"

                st.markdown(
                    f"""<div style="border-left:4px solid {color}; padding:7px 12px;
                        border-radius:4px; margin-bottom:4px; background:{bg}">
                    <b>{icon} {day_rows[0]['jour']} {date_str}</b>
                    — {nb_df}/{nb_dt} inventaires faits
                    </div>""",
                    unsafe_allow_html=True,
                )

                df_day = pd.DataFrame([{
                    "Salle":     r["salle"],
                    "Machine":   r["machine"],
                    "Statut":    "✅ Fait" if r["fait"] else "🔵 Abs",
                    "Valeur HT": f"{r['valeur_ht']:.2f} €" if r["valeur_ht"] else "—",
                } for r in day_rows])

                def _style_bilan(row):
                    if "Abs" in str(row["Statut"]):
                        return [f"background-color:{COLOR_SEUIL}; color:white; font-weight:600"] * len(row)
                    return ["background-color:#f0fff4"] * len(row)

                st.dataframe(
                    df_day.style.apply(_style_bilan, axis=1),
                    hide_index=True, use_container_width=True,
                    height=min(400, 38 + len(df_day) * 35),
                )


# ══════════════════════════════════════════
# COMPOSANTS UI
# ══════════════════════════════════════════

def _machine_card(row: pd.Series, df_raw: pd.DataFrame, planno_index: dict = None):
    total  = row["total"]
    s_min  = row["seuil_min"]
    s_max  = row["seuil_max"]
    ecart  = row["ecart_min"]

    ecart_str = f"{ecart:+.2f} €" if pd.notna(ecart) else "—"

    # ── Données planno (calculées en premier pour déterminer la couleur finale) ──
    mtype   = row.get("machine_type", "")
    planno  = (planno_index or {}).get(mtype)
    inv_sub = df_raw[df_raw["Num Piece"] == row["Num Piece"]]
    inv_qty = inv_sub.groupby("Code produit")["Quantité"].sum().to_dict()
    inv_lib = inv_sub.drop_duplicates("Code produit").set_index("Code produit")["Libellé produit"].to_dict()

    all_missing = _get_missing_products(df_raw, row["Num Piece"])

    if planno:
        missing_prods = [p for p in all_missing if p["code"] in planno]
        absents       = [(c, planno[c]["libelle"] or c) for c in planno if c not in inv_qty]
        hors_planno   = [inv_lib.get(c, c) for c in inv_qty if c not in planno]
    else:
        missing_prods = all_missing
        absents       = []
        hors_planno   = []

    # ── Couleur finale : rouge si problèmes planno, bleu si hors seuils, vert sinon ──
    has_planno_issues = bool(absents or hors_planno)
    color = COLOR_BAD if has_planno_issues else row["statut_color"]

    # ── HTML résumé produits ──
    lines_html = ""

    if missing_prods:
        noms = " &nbsp;·&nbsp; ".join(f"<b>{p['nom']}</b>" for p in missing_prods)
        lines_html += (
            f'<div style="margin-top:3px; font-size:0.8rem; color:{COLOR_SEUIL};">'
            f"🔵 Épuisés ({len(missing_prods)}) : {noms}</div>"
        )
    if absents:
        noms = " &nbsp;·&nbsp; ".join(f"<b>{lib}</b>" for _, lib in absents)
        lines_html += (
            f'<div style="margin-top:3px; font-size:0.8rem; color:{COLOR_BAD};">'
            f"❌ Absents du planno ({len(absents)}) : {noms}</div>"
        )
    if hors_planno:
        noms = " &nbsp;·&nbsp; ".join(f"<b>{n}</b>" for n in hors_planno)
        lines_html += (
            f'<div style="margin-top:3px; font-size:0.8rem; color:{COLOR_BAD};">'
            f"⚠️ Hors planno ({len(hors_planno)}) : {noms}</div>"
        )
    if not lines_html:
        ok_msg = "✅ Aucun produit épuisé" if not planno else "✅ Conforme au planno"
        lines_html = f'<div style="margin-top:3px; font-size:0.82rem; color:#2e7d32;">{ok_msg}</div>'

    vides_html = lines_html

    bg = "#fff8f8" if color == COLOR_BAD else "#f0f4ff" if color == COLOR_SEUIL else "#f8fff8"
    st.markdown(
        f"""<div style="border-left:5px solid {color}; background:{bg};
            border-radius:6px; padding:10px 14px; margin-bottom:4px;">
            <div style="display:flex; justify-content:space-between; align-items:center;">
                <span style="font-weight:700; font-size:0.95rem;">{row['statut_emoji']} {row['Nom client']}</span>
                <span style="font-size:0.8rem; color:#666;">{row['Stock Origine']} · {row['type_label']} · {row['Date']}</span>
            </div>
            <div style="display:flex; gap:20px; margin-top:5px; font-size:0.88rem;">
                <span><b>Montant :</b> <span style="color:{color}; font-weight:700;">{total:.2f} €</span></span>
                <span><b>Seuil :</b> {s_min:.0f} – {s_max:.2f} €</span>
                <span><b>Écart min :</b> {ecart_str}</span>
            </div>
            {vides_html}
        </div>""",
        unsafe_allow_html=True,
    )

    # ── Toggle planno théorique ────────────────────────────
    if planno:
        rows_table = []
        # Produits du planno
        for code, info in planno.items():
            qty_inv  = inv_qty.get(code)
            qty_plan = info["niv_haut"]
            if qty_inv is None:
                ecart_val = -qty_plan
                statut    = "❌ Absent"
            else:
                ecart_val = int(qty_inv) - qty_plan
                statut    = "✅ OK" if qty_inv > 0 else "🔴 Épuisé"
            rows_table.append({
                "Code":          code,
                "Libellé":       info["libelle"],
                "Qté planno":    qty_plan,
                "Qté inventaire": int(qty_inv) if qty_inv is not None else "—",
                "Écart":         ecart_val if qty_inv is not None else f"-{qty_plan}",
                "Statut":        statut,
                "_sort":         0 if qty_inv is None else (1 if qty_inv == 0 else 2),
            })

        # Produits hors-planno (dans inventaire mais pas dans le planno)
        for code, qty in inv_qty.items():
            if code not in planno:
                rows_table.append({
                    "Code":           code,
                    "Libellé":        inv_lib.get(code, ""),
                    "Qté planno":     "—",
                    "Qté inventaire": int(qty),
                    "Écart":          "Hors planno",
                    "Statut":         "⚠️ Hors planno",
                    "_sort":          3,
                })

        rows_table.sort(key=lambda x: x["_sort"])

        nb_absents    = sum(1 for r in rows_table if r["Statut"] == "❌ Absent")
        nb_epuises    = sum(1 for r in rows_table if r["Statut"] == "🔴 Épuisé")
        nb_hors       = sum(1 for r in rows_table if r["Statut"] == "⚠️ Hors planno")
        nb_ok         = sum(1 for r in rows_table if r["Statut"] == "✅ OK")

        label_exp = (
            f"📋 Planno — {nb_ok} OK"
            + (f" · {nb_epuises} épuisés" if nb_epuises else "")
            + (f" · {nb_absents} absents" if nb_absents else "")
            + (f" · {nb_hors} hors planno" if nb_hors else "")
        )

        toggle_key = f"planno_show_{row['Num Piece']}"
        if st.button(label_exp, key=f"planno_btn_{row['Num Piece']}", use_container_width=True):
            st.session_state[toggle_key] = not st.session_state.get(toggle_key, False)
            st.rerun()

        if st.session_state.get(toggle_key, False):
            with st.container(border=True):
                df_table = pd.DataFrame(rows_table).drop(columns=["_sort"])

                def _style_planno(r):
                    s = r["Statut"]
                    if s in ("❌ Absent", "⚠️ Hors planno"):
                        return [f"background-color:{COLOR_BAD}; color:{WHITE}; font-weight:600"] * len(r)
                    if s == "🔴 Épuisé":
                        return [f"background-color:{COLOR_SEUIL}; color:{WHITE}; font-weight:600"] * len(r)
                    return [""] * len(r)

                st.dataframe(
                    df_table.style.apply(_style_planno, axis=1),
                    use_container_width=True,
                    hide_index=True,
                    height=min(500, 38 + len(df_table) * 35),
                )

                col_chk, col_dl, _ = st.columns([2, 2, 2])
                with col_chk:
                    show_hors_c = st.checkbox(
                        "Inclure hors planno", value=True,
                        key=f"png_hors_card_{row['Num Piece']}",
                    )
                with col_dl:
                    _png_c = _generate_machine_png(row, df_raw, planno_index, show_hors_c)
                    _fname_c = f"{row['Stock Origine']}_{row['Nom client'].replace(' ', '_')}.png"
                    st.download_button(
                        "📥 Télécharger fiche PNG", data=_png_c, file_name=_fname_c,
                        mime="image/png", key=f"png_dl_card_{row['Num Piece']}",
                        use_container_width=True,
                    )
    else:
        # Pas de planno associé à ce type — petit message discret
        st.caption(f"⬜ Aucun planno théorique associé au type *{row.get('type_label', mtype)}*.")

    st.markdown("<div style='margin-bottom:8px'></div>", unsafe_allow_html=True)


def _machine_detail_full(row: pd.Series, df_raw: pd.DataFrame, planno_index: dict = None):
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("💰 Montant HT", f"{row['total']:.2f} €")
    c2.metric("📉 Seuil min",  f"{row['seuil_min']:.0f} €" if pd.notna(row["seuil_min"]) else "—")
    c3.metric("📈 Seuil max",  f"{row['seuil_max']:.2f} €" if pd.notna(row["seuil_max"]) else "—")
    c4.metric(
        "📊 Écart min",
        f"{row['ecart_min']:+.2f} €" if pd.notna(row["ecart_min"]) else "—",
        delta_color="normal" if pd.notna(row["ecart_min"]) and row["ecart_min"] >= 0 else "inverse",
    )

    sub = df_raw[df_raw["Num Piece"] == row["Num Piece"]].copy()
    sub = sub[["Code produit", "Libellé produit", "Quantité", "Montant HT"]].sort_values(
        "Montant HT", ascending=False
    )

    def _color_prod(r):
        if r["Quantité"] == 0:
            return [f"background-color:{COLOR_BAD}; color:{WHITE}; font-weight:700"] * len(r)
        return [""] * len(r)

    st.markdown("**Détail des produits** — les lignes rouges = produits épuisés (qté = 0)")
    st.dataframe(
        sub.style.apply(_color_prod, axis=1),
        use_container_width=True, hide_index=True,
        height=min(500, 38 + len(sub) * 35),
    )

    col_chk, col_dl, _ = st.columns([2, 2, 2])
    with col_chk:
        show_hors_d = st.checkbox(
            "Inclure hors planno", value=True,
            key=f"png_hors_det_{row['Num Piece']}",
        )
    with col_dl:
        _png_d = _generate_machine_png(row, df_raw, planno_index, show_hors_d)
        _fname_d = f"{row['Stock Origine']}_{row['Nom client'].replace(' ', '_')}.png"
        st.download_button(
            "📥 Télécharger fiche PNG", data=_png_d, file_name=_fname_d,
            mime="image/png", key=f"png_dl_det_{row['Num Piece']}",
            use_container_width=True,
        )


def _generate_machine_png(row: pd.Series, df_raw: pd.DataFrame, planno_index: dict = None, show_hors_planno: bool = True) -> bytes:
    """Génère une fiche PNG stylée (style planogramme) pour la machine."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.patches as mpatches

    piece  = row["Num Piece"]
    mtype  = row.get("machine_type", "")
    planno = (planno_index or {}).get(mtype, {})

    # Build per-product data (aggregate lines for same code)
    agg = (
        df_raw[df_raw["Num Piece"] == piece]
        .groupby("Code produit")
        .agg(libelle=("Libellé produit", "first"),
             qty=("Quantité", "sum"),
             ht=("Montant HT", "sum"))
        .reset_index()
    )
    inv_codes = set(agg["Code produit"])

    # Determine status
    def _statut(code, qty):
        if planno:
            if code not in planno: return "hors"
            return "epuise" if qty == 0 else "ok"
        return "epuise" if qty == 0 else "ok"

    table_data = []
    for _, r in agg.iterrows():
        code = r["Code produit"]
        qty  = int(r["qty"])
        ht   = float(r["ht"])
        pu   = ht / qty if qty > 0 else None
        qty_plan = planno[code]["niv_haut"] if planno and code in planno else None
        if qty_plan is not None:
            ecart = qty - qty_plan
            ecart_str = f"{ecart:+d}" if ecart != 0 else "✓"
        else:
            ecart_str = "—"
        table_data.append({
            "code":     code,
            "libelle":  str(r["libelle"])[:35],
            "qty_plan": str(qty_plan) if qty_plan is not None else "—",
            "qty":      str(qty),
            "ecart":    ecart_str,
            "statut":   _statut(code, qty),
        })

    # Absents from planno (not in inventory at all)
    for code, info in sorted(planno.items()):
        if code not in inv_codes:
            table_data.append({
                "code":     code,
                "libelle":  str(info.get("libelle", ""))[:35],
                "qty_plan": str(info.get("niv_haut", "—")),
                "qty":      "—",
                "ecart":    "—",
                "statut":   "absent",
            })

    if not show_hors_planno:
        table_data = [r for r in table_data if r["statut"] != "hors"]

    _order = {"absent": 0, "epuise": 1, "hors": 2, "ok": 3}
    table_data.sort(key=lambda x: (_order.get(x["statut"], 9), x["code"]))

    # ── Layout constants ──────────────────────────────────────────────────────
    COLS       = ["CODE PRODUIT", "LIBELLÉ", "QTÉ PLANNO", "QTÉ INV.", "ÉCART"]
    COL_FRACS  = [0.22, 0.38, 0.13, 0.13, 0.14]
    COL_ALIGN  = ["left", "left", "center", "center", "center"]

    ROW_H      = 0.36    # inches
    HEADER_H   = 0.50
    TITLE_H    = 0.60
    MARGIN     = 0.25
    FIG_W      = 11.0

    n       = len(table_data)
    FIG_H   = MARGIN + TITLE_H + HEADER_H + n * ROW_H + MARGIN

    # ── Colors ────────────────────────────────────────────────────────────────
    C_DARK   = (0x1F/255, 0x4E/255, 0x79/255)
    C_MID    = (0x2E/255, 0x75/255, 0xB6/255)
    C_ALT1   = (1.00, 1.00, 1.00)
    C_ALT2   = (0.93, 0.95, 0.98)
    C_ABSENT = (0.97, 0.84, 0.85)
    C_EPUISE = (0.73, 0.87, 0.98)
    C_HORS   = (1.00, 0.80, 0.80)

    BG = {"ok": None, "epuise": C_EPUISE, "hors": C_HORS, "absent": C_ABSENT}

    fig = plt.figure(figsize=(FIG_W, FIG_H), facecolor="white")
    ax  = fig.add_axes([0, 0, 1, 1])
    ax.set_xlim(0, FIG_W)
    ax.set_ylim(0, FIG_H)
    ax.axis("off")

    usable_w  = FIG_W - 2 * MARGIN
    col_widths = [f * usable_w for f in COL_FRACS]
    col_xs     = []
    cx = MARGIN
    for w in col_widths:
        col_xs.append(cx); cx += w

    def rect(x, y, w, h, fc, ec=(0.82, 0.82, 0.82)):
        ax.add_patch(mpatches.FancyBboxPatch(
            (x, y), w, h, boxstyle="square,pad=0",
            facecolor=fc, edgecolor=ec, linewidth=0.4,
        ))

    def txt(x, y, s, color="black", size=9, bold=False, ha="left"):
        ax.text(x, y, s, color=color, fontsize=size,
                fontweight="bold" if bold else "normal",
                ha=ha, va="center", clip_on=True,
                fontfamily="DejaVu Sans")

    y = FIG_H - MARGIN

    # Title bar
    rect(MARGIN, y - TITLE_H, usable_w, TITLE_H, C_DARK, ec=C_DARK)
    title = f"Machine : {row['Stock Origine']}    |    {row['Nom client']}"
    txt(FIG_W / 2, y - TITLE_H / 2, title, color="white", size=13, bold=True, ha="center")
    y -= TITLE_H

    # Header row
    rect(MARGIN, y - HEADER_H, usable_w, HEADER_H, C_MID, ec=C_MID)
    for hdr, cx2, cw, ha in zip(COLS, col_xs, col_widths, COL_ALIGN):
        tx = cx2 + 0.12 if ha == "left" else cx2 + cw / 2
        txt(tx, y - HEADER_H / 2, hdr, color="white", size=10, bold=True, ha=ha)
    y -= HEADER_H

    # Data rows
    for i, rd in enumerate(table_data):
        st_key = rd["statut"]
        bg = BG.get(st_key) or (C_ALT1 if i % 2 == 0 else C_ALT2)
        rect(MARGIN, y - ROW_H, usable_w, ROW_H, bg)

        vals = [rd["code"], rd["libelle"], rd["qty_plan"], rd["qty"], rd["ecart"]]
        for val, cx2, cw, ha in zip(vals, col_xs, col_widths, COL_ALIGN):
            tx    = cx2 + 0.12 if ha == "left" else cx2 + cw / 2
            bold  = st_key in ("absent", "hors") and val == rd["code"]
            fgcol = (0.12, 0.12, 0.12)
            txt(tx, y - ROW_H / 2, str(val), color=fgcol, size=9, bold=bold, ha=ha)

        y -= ROW_H

    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches=None, facecolor="white")
    plt.close(fig)
    buf.seek(0)
    return buf.read()


def _section_connexion():
    """Section 'Connexion' : associe chaque type de machine à un planno théorique."""
    st.markdown("---")
    st.markdown("### 🔗 Connexion — Planno théorique par type de machine")
    st.caption(
        "Associez un planno théorique à chaque type de machine. "
        "Cette association sera utilisée pour détecter les produits hors-planno lors des inventaires."
    )

    plannos    = load_plannos_theoriques()
    links      = load_machine_planno_links()

    if not plannos:
        st.warning("Aucun planno théorique enregistré. Ajoutez-en dans la page **Planogrammes** → section *Plannos théoriques*.")
        return

    # Options du selectbox : "Aucun" + liste des plannos
    planno_options = ["— Aucun —"] + [p["nom"] for p in plannos]
    planno_by_nom  = {p["nom"]: p for p in plannos}

    for machine_type, seuil in SEUILS.items():
        linked_id  = links.get(machine_type)
        linked_nom = next((p["nom"] for p in plannos if p["_id"] == linked_id), None)

        with st.container(border=True):
            col_info, col_select, col_btn = st.columns([3, 4, 2])

            with col_info:
                st.markdown(f"**{seuil['label']}**")
                st.caption(f"{seuil['min']:.0f} € — {seuil['max']:.2f} €")
                if linked_nom:
                    nb = len(planno_by_nom[linked_nom].get("produits", []))
                    st.markdown(
                        f"<span style='color:#2e7d32; font-size:0.82rem;'>✅ Lié à <b>{linked_nom}</b> ({nb} produits)</span>",
                        unsafe_allow_html=True,
                    )
                else:
                    st.markdown(
                        "<span style='color:#888; font-size:0.82rem;'>⬜ Aucun planno associé</span>",
                        unsafe_allow_html=True,
                    )

            with col_select:
                cur_idx = planno_options.index(linked_nom) if linked_nom in planno_options else 0
                chosen = st.selectbox(
                    "Planno théorique",
                    planno_options,
                    index=cur_idx,
                    key=f"conn_select_{machine_type}",
                    label_visibility="collapsed",
                )

            with col_btn:
                is_dissocier = (chosen != "— Aucun —" and chosen == linked_nom)
                btn_label    = "🔓 Dissocier" if is_dissocier else "🔗 Associer"
                btn_type     = "secondary" if is_dissocier else "primary"
                if st.button(btn_label, key=f"conn_btn_{machine_type}",
                             use_container_width=True, type=btn_type):
                    if chosen == "— Aucun —" or is_dissocier:
                        save_machine_planno_link(machine_type, None)
                        st.toast(f"🔓 Association supprimée pour {seuil['label']}.")
                    else:
                        new_id = planno_by_nom[chosen]["_id"]
                        save_machine_planno_link(machine_type, new_id)
                        nb = len(planno_by_nom[chosen].get("produits", []))
                        st.toast(f"✅ {seuil['label']} → {chosen} ({nb} produits).")
                    st.rerun()


def _show_thresholds():
    st.markdown("---")
    st.markdown("#### 📏 Seuils de validation par type de machine")
    rows = [
        {
            "Type":                        s["label"],
            "Seuil min (🔴 en dessous)":   f"{s['min']:.0f} €",
            "Seuil max (🟠 au-dessus)":    f"{s['max']:.2f} €",
            "Fourchette OK":               f"{s['min']:.0f} € — {s['max']:.2f} €",
        }
        for s in SEUILS.values()
    ]
    st.table(pd.DataFrame(rows))


# ══════════════════════════════════════════
# EXPORT EXCEL
# ══════════════════════════════════════════

def _build_export_rows(
    summary: pd.DataFrame,
    df_raw: pd.DataFrame,
    croisement: dict,
    plannings_mongo: dict,
    planno_index: dict = None,
) -> list:
    """
    Construit la liste complète des lignes d'export :
    une ligne par (réappro, jour planifié, salle) — qu'elle soit faite, non faite ou double passage.
    """
    from planning_parser import parse_planning_file
    import os

    ABBR = {"l": "Lundi", "ma": "Mardi", "me": "Mercredi", "j": "Jeudi", "v": "Vendredi"}

    # Index inventaires faits : {(reappro, date): set(codes)}
    inv_by_date = (
        df_raw.groupby(["Ressource", "Date"])["Code client"]
        .apply(set)
        .to_dict()
    )

    # Index semaine : {(reappro, iso_year, iso_week): set(codes)}
    df_raw2 = df_raw.copy()
    df_raw2["_dt"] = pd.to_datetime(df_raw2["Date"], format="%d/%m/%Y", errors="coerce")
    df_raw2["_iso_year"] = df_raw2["_dt"].dt.isocalendar().year.astype("Int64")
    df_raw2["_iso_week"] = df_raw2["_dt"].dt.isocalendar().week.astype("Int64")
    codes_by_week = (
        df_raw2.groupby(["Ressource", "_iso_year", "_iso_week"])["Code client"]
        .apply(set)
        .to_dict()
    )

    # Index inventaires résumés par (reappro, code, date)
    inv_idx = {}
    for _, r in summary.iterrows():
        inv_idx[(r["Ressource"], r["Code client"], r["Date"])] = r

    # Index produits manquants par Num Piece (libellés, tous)
    missing_idx = (
        df_raw[df_raw["Quantité"] == 0]
        .groupby("Num Piece")["Libellé produit"]
        .apply(lambda x: list(x.drop_duplicates()))
        .to_dict()
    )

    # Index qty par (Num Piece, Code produit) pour comparaison planno
    inv_qty_by_piece: dict = {}
    inv_lib_by_piece: dict = {}
    for _, _r in df_raw.iterrows():
        _np   = _r["Num Piece"]
        _code = _r["Code produit"]
        _lib  = _r["Libellé produit"]
        inv_qty_by_piece.setdefault(_np, {})
        inv_qty_by_piece[_np][_code] = inv_qty_by_piece[_np].get(_code, 0) + _r["Quantité"]
        inv_lib_by_piece.setdefault(_np, {})
        inv_lib_by_piece[_np][_code] = _lib

    # Index codes à qty=0 par Num Piece
    missing_code_idx: dict = {}
    for _, _r in df_raw[df_raw["Quantité"] == 0].iterrows():
        missing_code_idx.setdefault(_r["Num Piece"], set()).add(_r["Code produit"])

    # Dates présentes dans l'export
    all_dates = sorted(df_raw["Date"].unique())

    rows = []
    for date_str in all_dates:
        dt = datetime.datetime.strptime(date_str, "%d/%m/%Y")
        jour_fr = WEEKDAY_TO_JOUR.get(dt.weekday())
        if not jour_fr:
            continue
        iso = dt.isocalendar()

        for reappro, planning_raw in sorted(plannings_mongo.items()):
            planning = _parse_planning_for_reappro(planning_raw)
            if jour_fr not in planning:
                continue

            jour_plan = planning[jour_fr]
            codes_fait_today = inv_by_date.get((reappro, date_str), set())
            codes_fait_week  = codes_by_week.get((reappro, iso[0], iso[1]), set())

            for code, info in sorted(jour_plan.items(), key=lambda x: x[1]["label"]):
                key = (reappro, code, date_str)
                inv_row = inv_idx.get(key)

                if code in codes_fait_today and inv_row is not None:
                    total        = inv_row["total"]
                    smin         = inv_row["seuil_min"]
                    smax         = inv_row["seuil_max"]
                    machine_type = inv_row["type_label"]
                    ecart        = inv_row["ecart_min"]
                    statut_plan  = "Fait"
                    if inv_row["total"] < smin:
                        statut_inv = "Mal fait"
                    elif inv_row["total"] > smax:
                        statut_inv = "Au-dessus max"
                    else:
                        statut_inv = "OK"
                    fait_par = reappro

                    # ── Planno-aware product analysis ──────────────────────
                    piece     = inv_row["Num Piece"]
                    mtype_key = inv_row.get("machine_type", "")
                    prods, hors_planno, absents, detail_rows = _analyze_planno(
                        piece, mtype_key, planno_index,
                        inv_qty_by_piece, inv_lib_by_piece,
                        missing_code_idx, missing_idx,
                    )
                elif code not in codes_fait_today and code in codes_fait_week:
                    total = smin = smax = ecart = machine_type = None
                    statut_plan = "Double passage"
                    statut_inv  = "-"
                    prods       = []; hors_planno = []; absents = []; detail_rows = []
                    fait_par    = reappro
                else:
                    # Check joker: was it inventoried by another reappro on the same date?
                    joker_reappro = None
                    for (other_r, other_d), other_codes in inv_by_date.items():
                        if other_r != reappro and other_d == date_str and code in other_codes:
                            joker_reappro = other_r
                            break

                    if joker_reappro:
                        joker_key = (joker_reappro, code, date_str)
                        joker_inv = inv_idx.get(joker_key)
                        if joker_inv is not None:
                            total        = joker_inv["total"]
                            smin         = joker_inv["seuil_min"]
                            smax         = joker_inv["seuil_max"]
                            machine_type = joker_inv["type_label"]
                            ecart        = joker_inv["ecart_min"]
                            if joker_inv["total"] < smin:
                                statut_inv = "Mal fait"
                            elif joker_inv["total"] > smax:
                                statut_inv = "Au-dessus max"
                            else:
                                statut_inv = "OK"
                            piece     = joker_inv["Num Piece"]
                            mtype_key = joker_inv.get("machine_type", "")
                            prods, hors_planno, absents, detail_rows = _analyze_planno(
                                piece, mtype_key, planno_index,
                                inv_qty_by_piece, inv_lib_by_piece,
                                missing_code_idx, missing_idx,
                            )
                        else:
                            total = smin = smax = ecart = machine_type = None
                            statut_inv = "-"
                            prods = []; hors_planno = []; absents = []; detail_rows = []
                        statut_plan = f"Joker ({joker_reappro})"
                        fait_par    = joker_reappro
                    else:
                        total = smin = smax = ecart = machine_type = None
                        statut_plan = "Non fait"
                        statut_inv  = "Non fait"
                        prods       = []; hors_planno = []; absents = []; detail_rows = []
                        fait_par    = ""

                rows.append({
                    "reappro":        reappro,
                    "date":           date_str,
                    "jour":           jour_fr,
                    "code":           code,
                    "salle":          info["label"],
                    "machine":        info["machine"],
                    "statut_plan":    statut_plan,
                    "machine_type":   machine_type or "",
                    "total":          total,
                    "seuil_min":      smin,
                    "seuil_max":      smax,
                    "ecart":          ecart,
                    "statut_inv":     statut_inv,
                    "produits":       prods,
                    "hors_planno":    hors_planno,
                    "absents":        absents,
                    "detail_rows":    detail_rows,
                    "fait_par":       fait_par if statut_plan.startswith("Joker") else "",
                })

    return rows


def _export_excel(
    summary: pd.DataFrame,
    df_raw: pd.DataFrame,
    croisement: dict,
    plannings_mongo: dict = None,
    planno_index: dict = None,
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    plannings_mongo = plannings_mongo or {}
    rows = _build_export_rows(summary, df_raw, croisement, plannings_mongo, planno_index or {})

    # ── Style helpers ──────────────────────────────────────────────────────
    def _fill(hex_col):
        return PatternFill("solid", fgColor=hex_col.lstrip("#"))

    def _font(bold=False, color="000000", size=9, italic=False):
        return Font(bold=bold, color=color.lstrip("#"), size=size,
                    name="Arial", italic=italic)

    _side_thin  = Side(style="thin",   color="CCCCCC")
    _side_med   = Side(style="medium", color="1F4E79")
    _brd        = Border(left=_side_thin, right=_side_thin,
                         top=_side_thin,  bottom=_side_thin)
    _brd_sep    = Border(left=_side_med, right=_side_thin,
                         top=_side_med,  bottom=_side_med)

    def _align(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    # Color palette
    C_DARK      = "1F4E79"; C_MID  = "2E75B6"; C_SUB = "BDD7EE"
    C_OK_BG     = "D4EDDA"; C_BAD_BG = "F8D7DA"; C_ORA_BG = "FFF3CD"
    C_DOUB_BG   = "EDE7F6"; C_GREY_BG = "F5F5F5"; C_BLUE_BG = "BBDEFB"
    C_OK_FG     = "1E7E34"; C_BAD_FG  = "C0392B"
    C_ORA_FG    = "E67E22"; C_DOUB_FG = "6C3483"; C_BLUE_FG = "1565C0"
    C_WHITE     = "FFFFFF"

    STATUS_STYLE = {
        "Fait":          (C_OK_BG,   C_MID,     False),
        "Non fait":      (C_BAD_BG,  C_BAD_FG,  True),
        "Double passage":(C_DOUB_BG, C_DOUB_FG, True),
        "OK":            (C_OK_BG,   C_OK_FG,   True),
        "Mal fait":      (C_BLUE_BG, C_BLUE_FG, True),
        "Au-dessus max": (C_BLUE_BG, C_BLUE_FG, True),
        "-":             (C_GREY_BG, C_DARK,    False),
        "Non fait (inv)":(C_BAD_BG,  C_BAD_FG,  True),
    }
    # Joker entries added dynamically below since the reappro name is in the key

    # Column definitions for detail sheets: (header, width)
    DETAIL_COLS = [
        ("Date",             11), ("Jour",          10), ("Salle",            36),
        ("Machine",           9), ("Type machine",  16), ("Statut planning",  22),
        ("Fait par",         12), ("Montant HT",    13), ("Statut inv.",      14),
        ("Produits épuisés", 40), ("Hors planno",   40), ("Absents planno",   40),
    ]
    N = len(DETAIL_COLS)
    # Sub-header labels for product detail rows (cols 3-9 reused)
    # Col 3=Code, 4=Libellé, 5=Qté planno, 6=Qté inv., 7=Écart, 9=Statut produit

    wb = Workbook()
    wb.remove(wb.active)

    # ══════════════════════════════════════════
    # ONGLET RÉCAP GLOBAL
    # ══════════════════════════════════════════
    ws_g = wb.create_sheet("📊 Récap global")
    ws_g.freeze_panes = "A3"
    ws_g.sheet_properties.tabColor = C_DARK

    # Row 1 : big title
    RECAP_HDRS = [
        "Réappro", "Date", "Jour",
        "📋 Planifiées", "✅ Faites", "🔴 Non faites",
        "⚠️ Mal faits", "📈 Au-dessus max",
        "💰 Total € fait", "% réalisation",
    ]
    NR = len(RECAP_HDRS)
    ws_g.merge_cells(f"A1:{get_column_letter(NR)}1")
    tc = ws_g.cell(1, 1, "Suivi des inventaires — Récapitulatif")
    tc.fill = _fill(C_DARK); tc.font = _font(True, C_WHITE, 13)
    tc.alignment = _align("center"); ws_g.row_dimensions[1].height = 26

    # Row 2 : headers
    RECAP_WIDTHS = [13, 11, 11, 13, 12, 13, 12, 15, 16, 14]
    ws_g.row_dimensions[2].height = 20
    for ci, h in enumerate(RECAP_HDRS, 1):
        c = ws_g.cell(2, ci, h)
        c.fill = _fill(C_MID); c.font = _font(True, C_WHITE, 9)
        c.alignment = _align("center"); c.border = _brd
        ws_g.column_dimensions[get_column_letter(ci)].width = RECAP_WIDTHS[ci - 1]

    # Aggregate
    recap: dict = {}
    for r in rows:
        k = (r["reappro"], r["date"], r["jour"])
        if k not in recap:
            recap[k] = {"planif": 0, "fait": 0, "non_fait": 0,
                        "mal_fait": 0, "au_dessus": 0, "total_eur": 0.0,
                        "has_planno": False}
        d = recap[k]
        sp = r["statut_plan"]
        si = r["statut_inv"]
        d["planif"] += 1
        if sp in ("Fait",) or sp.startswith("Joker"):
            d["fait"] += 1; d["total_eur"] += r["total"] or 0
            if si == "Mal fait":        d["mal_fait"] += 1
            elif si == "Au-dessus max": d["au_dessus"] += 1
            if r.get("absents") or r.get("hors_planno"):
                d["has_planno"] = True
        elif sp == "Non fait":
            d["non_fait"] += 1

    ri = 3
    for (reappro, date_str, jour), d in sorted(recap.items()):
        pct = round(d["fait"] / d["planif"] * 100, 1) if d["planif"] else 0
        half_not_done = d["non_fait"] >= (d["planif"] / 2)
        if half_not_done:
            bg = C_GREY_BG                                        # >= moitié non faites → neutre
        elif d["has_planno"]:
            bg = C_BAD_BG                                         # produits absents/hors planno → rouge
        elif d["mal_fait"] > 0 or d["au_dessus"] > 0:
            bg = C_BLUE_BG                                        # hors seuils → bleu
        elif d["non_fait"] > 0:
            bg = C_ORA_BG                                         # quelques manquants → orange
        else:
            bg = C_OK_BG                                          # tout bon → vert
        vals = [
            reappro, date_str, jour,
            d["planif"], d["fait"], d["non_fait"],
            d["mal_fait"], d["au_dessus"],
            round(d["total_eur"], 2), f"{pct}%",
        ]
        ws_g.row_dimensions[ri].height = 17
        sheet_target = reappro[:31]
        for ci, v in enumerate(vals, 1):
            c = ws_g.cell(ri, ci, v)
            c.fill = _fill(bg); c.border = _brd
            c.alignment = _align("center" if ci > 3 else "left")
            if ci == 9 and isinstance(v, (int, float)):
                c.number_format = "#,##0.00 €"
            # Hyperlink sur toute la ligne → feuille du réappro
            c.hyperlink = f"#{sheet_target}!A1"
            c.font = Font(
                name="Arial", size=9, bold=(ci == 1),
                color=("1F4E79" if ci == 1 else "000000"),
                underline=("single" if ci == 1 else None),
            )
            # Couleurs spéciales colonnes Faites / Non faites
            if ci == 5 and d["fait"] == d["planif"]:
                c.font = Font(name="Arial", size=9, bold=True, color=C_OK_FG)
            if ci == 6 and d["non_fait"] > 0:
                c.font = Font(name="Arial", size=9, bold=True, color=C_BAD_FG)
        ri += 1

    # ══════════════════════════════════════════
    # UN ONGLET PAR RÉAPPRO
    # ══════════════════════════════════════════
    reappros_ordered = sorted({r["reappro"] for r in rows})
    for reappro in reappros_ordered:
        ws = wb.create_sheet(reappro[:31])
        ws.freeze_panes = "A3"
        ws.sheet_properties.outlinePr.summaryBelow = False  # + button above detail rows

        # Title row
        ws.merge_cells(f"A1:{get_column_letter(N)}1")
        t = ws.cell(1, 1, f"Suivi inventaires — {reappro}")
        t.fill = _fill(C_DARK); t.font = _font(True, C_WHITE, 12)
        t.alignment = _align("center"); ws.row_dimensions[1].height = 24

        # Header row
        ws.row_dimensions[2].height = 18
        for ci, (h, w) in enumerate(DETAIL_COLS, 1):
            c = ws.cell(2, ci, h)
            c.fill = _fill(C_MID); c.font = _font(True, C_WHITE, 9)
            c.alignment = _align("center"); c.border = _brd
            ws.column_dimensions[get_column_letter(ci)].width = w

        sub_rows = [r for r in rows if r["reappro"] == reappro]
        current_date = None
        ri = 3

        for r in sub_rows:
            # ── Date separator ─────────────────────────────────────────────
            if r["date"] != current_date:
                current_date = r["date"]
                day_rows   = [x for x in sub_rows if x["date"] == r["date"]]
                nb_fait    = sum(1 for x in day_rows if x["statut_plan"] == "Fait")
                nb_nf      = sum(1 for x in day_rows if x["statut_plan"] == "Non fait")
                nb_dbl     = sum(1 for x in day_rows if x["statut_plan"] == "Double passage")
                nb_mal     = sum(1 for x in day_rows if x["statut_inv"] == "Mal fait")
                total_jour = sum(x["total"] or 0 for x in day_rows if x["statut_plan"] == "Fait")
                sep_txt = (
                    f"  {r['jour']}  {r['date']}"
                    f"   |   ✅ {nb_fait} faits   🔴 {nb_nf} non faits"
                    + (f"   🔄 {nb_dbl} double(s)" if nb_dbl else "")
                    + (f"   ⚠️ {nb_mal} mal fait(s)" if nb_mal else "")
                    + f"   💰 {total_jour:,.2f} €"
                )
                ws.merge_cells(f"A{ri}:{get_column_letter(N)}{ri}")
                sep = ws.cell(ri, 1, sep_txt)
                sep.fill = _fill(C_SUB); sep.font = _font(True, C_DARK, 10)
                sep.alignment = _align("left")
                for ci in range(1, N + 1):
                    ws.cell(ri, ci).border = _brd_sep
                ws.row_dimensions[ri].height = 20
                ri += 1

            # ── Summary row (salle) ────────────────────────────────────────
            sp = r["statut_plan"]
            si = r["statut_inv"]
            has_planno_issues = bool(r.get("absents") or r.get("hors_planno"))

            if sp == "Non fait":           row_bg = C_BAD_BG
            elif sp == "Double passage":   row_bg = C_DOUB_BG
            elif sp.startswith("Joker"):   row_bg = "F3E5F5"
            elif has_planno_issues:        row_bg = C_BAD_BG
            elif si in ("Mal fait", "Au-dessus max"): row_bg = C_BLUE_BG
            elif si == "OK":               row_bg = C_OK_BG
            else:                          row_bg = C_GREY_BG

            prods_str   = " / ".join(r["produits"])           if r.get("produits")     else ""
            hors_str    = " / ".join(r["hors_planno"])        if r.get("hors_planno")  else ""
            absents_str = " / ".join(r["absents"])            if r.get("absents")      else ""
            has_prods   = bool(prods_str)
            has_hors    = bool(hors_str)
            has_absents = bool(absents_str)
            has_detail  = bool(r.get("detail_rows"))
            is_joker    = sp.startswith("Joker")

            ws.row_dimensions[ri].height = 20 if (has_prods or has_hors or has_absents) else 16

            values = [
                r["date"], r["jour"], r["salle"], r["machine"],
                r["machine_type"] or "-", sp,
                r.get("fait_par", "") or "",
                r["total"], si, prods_str, hors_str, absents_str,
            ]

            for ci, v in enumerate(values, 1):
                c = ws.cell(ri, ci, v)
                c.fill = _fill(row_bg); c.font = _font(size=9); c.border = _brd
                c.alignment = _align("center" if ci in (1, 2, 4, 8) else "left",
                                     wrap=(ci in (10, 11, 12)))
                if ci == 8 and isinstance(v, (int, float)):
                    c.number_format = "#,##0.00 €"
                if ci == 6:
                    if is_joker: bg2, fg2, bold2 = "E8D5F5", "6C3483", True
                    else:        bg2, fg2, bold2 = STATUS_STYLE.get(sp, (C_GREY_BG, C_DARK, False))
                    c.fill = _fill(bg2); c.font = _font(bold2, fg2, 9); c.alignment = _align("center")
                if ci == 7 and v:
                    c.fill = _fill("E8D5F5"); c.font = _font(True, "6C3483", 9); c.alignment = _align("center")
                if ci == 9:
                    bg2, fg2, bold2 = STATUS_STYLE.get(si, (C_GREY_BG, C_DARK, False))
                    c.fill = _fill(bg2); c.font = _font(bold2, fg2, 9); c.alignment = _align("center")
                if ci == 10 and has_prods:
                    c.fill = _fill(C_BLUE_BG); c.font = _font(True, C_BLUE_FG, 9)
                if ci == 11 and has_hors:
                    c.fill = _fill("FFCCCC"); c.font = _font(True, C_BAD_FG, 9)
                if ci == 12 and has_absents:
                    c.fill = _fill("FFCCCC"); c.font = _font(True, C_BAD_FG, 9)

            ri += 1

            # ── Product detail rows (collapsed, outline_level=1) ───────────
            PROD_STAT_COLORS = {
                "OK":          ("D4EDDA", "1E7E34"),
                "Épuisé":      (C_BLUE_BG, C_BLUE_FG),
                "Absent":      ("FFCCCC",  C_BAD_FG),
                "Hors planno": ("FFCCCC",  C_BAD_FG),
            }
            for pd_row in r.get("detail_rows", []):
                bg_d, fg_d = PROD_STAT_COLORS.get(pd_row["statut"], ("F5F5F5", "333333"))
                ws.row_dimensions[ri].outline_level = 1
                ws.row_dimensions[ri].hidden        = True
                ws.row_dimensions[ri].height        = 15
                # Col 3: indented code, col 4: libelle, col 5: qty planno,
                # col 6: qty inv, col 7: écart, col 9: statut
                detail_vals = {
                    3: f"  └ {pd_row['code']}",
                    4: pd_row["libelle"],
                    5: pd_row["qty_plan"],
                    6: pd_row["qty_inv"],
                    7: pd_row["ecart"],
                    9: pd_row["statut"],
                }
                for ci in range(1, N + 1):
                    v  = detail_vals.get(ci, "")
                    c  = ws.cell(ri, ci, v)
                    c.border = _brd
                    if ci in detail_vals:
                        c.fill = _fill(bg_d)
                        c.font = _font(bold=(ci == 3), color=fg_d, size=8,
                                       italic=(ci not in (3, 9)))
                        c.alignment = _align(
                            "center" if ci in (5, 6, 7) else "left"
                        )
                    else:
                        c.fill = _fill("FAFAFA")
                        c.font = _font(size=8)
                ri += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()
