"""
Module d'export Excel avec mise en forme couleur par onglet.
Inclut un onglet "Tournees Decalees" et la colonne Decalee dans le recap et les onglets individuels.
"""

import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Couleurs
CLR_VERT    = "1E7E34"
CLR_ROUGE   = "C0392B"
CLR_ORANGE  = "E67E22"
CLR_VIOLET  = "6C3483"
CLR_HEADER  = "1F4E79"
CLR_BLANC   = "FFFFFF"

_DARK_BG = {CLR_VERT, CLR_ROUGE, CLR_ORANGE, CLR_VIOLET, CLR_HEADER}


def _fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def _font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size)

def _border():
    side = Side(style="thin", color="BFBFBF")
    return Border(left=side, right=side, top=side, bottom=side)

def _write_header_row(ws, headers, row, bg_color=CLR_HEADER):
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.fill = _fill(bg_color)
        cell.font = _font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border()

def _write_data_row(ws, values, row, bg_color=CLR_BLANC):
    font_color = "FFFFFF" if bg_color in _DARK_BG else "000000"
    bold = bg_color in _DARK_BG
    for col_idx, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=v)
        cell.fill = _fill(bg_color)
        cell.font = _font(bold=bold, color=font_color)
        cell.alignment = Alignment(vertical="center")
        cell.border = _border()

def _auto_width(ws, min_w=12, max_w=45):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_w, max(min_w, max_len + 2))


def generer_excel(results: dict, jour: str) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    # ──────────────────────────────────────────────
    # Onglet RÉCAPITULATIF
    # ──────────────────────────────────────────────
    ws_recap = wb.create_sheet("Récapitulatif")
    ws_recap.freeze_panes = "A2"
    headers_recap = ["Réappro", "Prévues", "Faites", "Non Faites", "Jokers", "Tournée Décalée", "Taux (%)"]
    _write_header_row(ws_recap, headers_recap, 1)

    row = 2
    for reappro, data in sorted(results.items()):
        nb_prev  = len(data["salles_prevues"])
        nb_fait  = len(data["salles_faites"])
        nb_nf    = len(data["salles_non_faites"])
        nb_joker = sum(1 for s in data["salles_faites"] if s["is_joker"])
        taux     = round((nb_fait / nb_prev * 100) if nb_prev > 0 else 0, 1)
        td       = data.get("tournee_decalee")
        decale   = f"Tournée {td['jour_detecte']}" if td else ""

        if nb_nf == 0:
            bg = CLR_VERT
        elif td:
            bg = CLR_VIOLET
        elif nb_nf == nb_prev:
            bg = CLR_ROUGE
        else:
            bg = CLR_ORANGE

        _write_data_row(ws_recap,
            [reappro, nb_prev, nb_fait, nb_nf, nb_joker, decale, f"{taux}%"],
            row, bg)
        row += 1

    _auto_width(ws_recap)

    # ──────────────────────────────────────────────
    # Onglet SALLES NON FAITES
    # ──────────────────────────────────────────────
    ws_nf = wb.create_sheet("Salles Non Faites")
    ws_nf.freeze_panes = "A2"
    _write_header_row(ws_nf, ["Réappro", "Client / Salle", "Machine ID"], 1)

    row = 2
    for reappro, data in sorted(results.items()):
        for s in data["salles_non_faites"]:
            _write_data_row(ws_nf, [reappro, s["client"], s["machine"]], row, CLR_ROUGE)
            row += 1

    if row == 2:
        ws_nf.cell(row=2, column=1, value="Toutes les salles ont ete faites !")
    _auto_width(ws_nf)

    # ──────────────────────────────────────────────
    # Onglet JOKERS
    # ──────────────────────────────────────────────
    ws_joker = wb.create_sheet("Jokers - Remplacements")
    ws_joker.freeze_panes = "A2"
    _write_header_row(ws_joker, ["Réappro Prévu", "Client / Salle", "Machine ID", "Fait Par", "Valeur Ref"], 1)

    row = 2
    for reappro, data in sorted(results.items()):
        for s in data["salles_faites"]:
            if s["is_joker"]:
                _write_data_row(ws_joker,
                    [reappro, s["client"], s["machine"], s["employe_reel"], s["val_ref"]],
                    row, CLR_ORANGE)
                row += 1

    if row == 2:
        ws_joker.cell(row=2, column=1, value="Aucun remplacement detecte.")
    _auto_width(ws_joker)

    # ──────────────────────────────────────────────
    # Onglet TOURNÉES DÉCALÉES
    # ──────────────────────────────────────────────
    ws_td = wb.create_sheet("Tournees Decalees")
    ws_td.freeze_panes = "A2"

    # En-tête global
    headers_td = [
        "Réappro",
        "Jour Analysé",
        "Tournée Détectée",
        "Salles Matchées",
        "Total Machines Faites",
        "Machines Hors Planning",
    ]
    _write_header_row(ws_td, headers_td, 1)

    row = 2
    decale_detail_rows = []  # pour la section détail en dessous

    for reappro, data in sorted(results.items()):
        td = data.get("tournee_decalee")
        if not td:
            continue

        nb_match  = len(td["salles_decalees"])
        nb_faites = td["nb_machines_faites"]
        hors_plan = ", ".join(td["salles_hors_planning"]) if td["salles_hors_planning"] else "—"

        _write_data_row(ws_td,
            [reappro, jour, td["jour_detecte"], nb_match, nb_faites, hors_plan],
            row, CLR_VIOLET)
        row += 1

        # Stocker pour le détail
        for client, machine in td["salles_decalees"]:
            decale_detail_rows.append((reappro, td["jour_detecte"], client, machine))

    if row == 2:
        ws_td.cell(row=2, column=1, value="Aucune tournee decalee detectee.")
        _auto_width(ws_td)
    else:
        _auto_width(ws_td)

        # ── Section détail des salles décalées (séparée par une ligne vide)
        row += 1  # ligne vide
        ws_td.cell(row=row, column=1, value="DETAIL DES SALLES DECALEES")
        ws_td.cell(row=row, column=1).font = _font(bold=True, color="6C3483", size=12)
        row += 1

        _write_header_row(ws_td,
            ["Réappro", "Tournée (planning)", "Client / Salle", "Machine ID"],
            row, CLR_VIOLET)
        row += 1

        current_reappro = None
        for reappro_d, jour_d, client, machine in decale_detail_rows:
            # Alterner légèrement la teinte pour distinguer les blocs par réappro
            bg = CLR_VIOLET if reappro_d != current_reappro else "7D3C98"
            current_reappro = reappro_d
            _write_data_row(ws_td,
                [reappro_d, jour_d, client, machine],
                row, CLR_VIOLET)
            row += 1

        _auto_width(ws_td)

    # ──────────────────────────────────────────────
    # Un onglet par réappro
    # ──────────────────────────────────────────────
    headers_detail = ["Client / Salle", "Machine ID", "Statut", "Fait Par", "Valeur Ref"]

    for reappro, data in sorted(results.items()):
        safe_name = reappro[:31]
        ws = wb.create_sheet(safe_name)
        ws.freeze_panes = "A6"

        nb_prev  = len(data["salles_prevues"])
        nb_fait  = len(data["salles_faites"])
        nb_nf    = len(data["salles_non_faites"])
        nb_joker = sum(1 for s in data["salles_faites"] if s["is_joker"])
        taux     = round((nb_fait / nb_prev * 100) if nb_prev > 0 else 0, 1)
        td       = data.get("tournee_decalee")

        # Ligne 1 — Titre
        ws.merge_cells("A1:E1")
        c = ws["A1"]
        c.value = f"{reappro} — {jour}"
        c.font = _font(bold=True, color="1F4E79", size=13)
        c.alignment = Alignment(horizontal="center")

        # Ligne 2 — Résumé chiffres
        ws.merge_cells("A2:E2")
        taux_color = "00B050" if taux == 100 else ("C0392B" if taux == 0 else "E67E22")
        c2 = ws["A2"]
        c2.value = (
            f"Prévues: {nb_prev}   |   Faites: {nb_fait}   |   "
            f"Non Faites: {nb_nf}   |   Jokers: {nb_joker}   |   Taux: {taux}%"
        )
        c2.font = _font(bold=True, color=taux_color, size=11)
        c2.alignment = Alignment(horizontal="center")

        # Ligne 3 — Tournée décalée (si applicable)
        ws.merge_cells("A3:E3")
        c3 = ws["A3"]
        if td:
            c3.value = (
                f"TOURNEE DECALEE : a effectue sa tournee du {td['jour_detecte']} "
                f"({len(td['salles_decalees'])} salle(s) detectee(s) sur {td['nb_machines_faites']} faite(s))"
            )
            c3.fill = _fill(CLR_VIOLET)
            c3.font = _font(bold=True, color="FFFFFF", size=10)
        else:
            c3.value = ""
        c3.alignment = Alignment(horizontal="center")

        # Ligne 4 — vide (séparation)
        ws.row_dimensions[4].height = 6

        # Ligne 5 — header tableau
        _write_header_row(ws, headers_detail, 5)

        row = 6

        # Non faites
        for s in data["salles_non_faites"]:
            _write_data_row(ws, [s["client"], s["machine"], "Non Faite", "", ""], row, CLR_ROUGE)
            row += 1

        # Jokers
        for s in data["salles_faites"]:
            if s["is_joker"]:
                _write_data_row(ws,
                    [s["client"], s["machine"], f"Joker ({s['statut']})", s["employe_reel"], s["val_ref"]],
                    row, CLR_ORANGE)
                row += 1

        # Faites normalement
        for s in data["salles_faites"]:
            if not s["is_joker"]:
                _write_data_row(ws,
                    [s["client"], s["machine"], f"Fait ({s['statut']})", s["employe_reel"], s["val_ref"]],
                    row, CLR_VERT)
                row += 1

        # Si tournée décalée : bloc séparé avec les salles décalées
        if td:
            row += 1  # ligne vide
            ws.cell(row=row, column=1, value=f"SALLES DE LA TOURNEE DU {td['jour_detecte'].upper()} (DETECTEES)")
            ws.cell(row=row, column=1).font = _font(bold=True, color="6C3483", size=10)
            ws.cell(row=row, column=1).fill = _fill("EBD5FA")
            row += 1
            _write_header_row(ws, ["Client / Salle", "Machine ID", "Statut", "", ""], row, CLR_VIOLET)
            row += 1
            for client, machine in td["salles_decalees"]:
                _write_data_row(ws,
                    [client, machine, f"Decalee (tournee {td['jour_detecte']})", "", ""],
                    row, CLR_VIOLET)
                row += 1

        _auto_width(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
